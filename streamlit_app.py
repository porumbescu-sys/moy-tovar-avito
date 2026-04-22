
from __future__ import annotations

import html
import io
import json
import math
import re
import sqlite3
import hashlib
import shutil
import zipfile
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Optional
from collections import Counter, defaultdict
from urllib.parse import quote_plus, unquote, urljoin, urlparse

REQUESTS_IMPORT_ERROR: str | None = None
try:
    import requests
except ImportError as exc:
    requests = None
    REQUESTS_IMPORT_ERROR = str(exc)

import openpyxl
import pandas as pd
import streamlit as st
import streamlit.components.v1 as components

st.set_page_config(page_title="Мой Товар", page_icon="📦", layout="wide")

APP_TITLE = "Мой Товар"
APP_VERSION = "v54.8.0-crm-purchase-cost"


SERVER_DATA_DIRNAME = "data"
PERSISTED_PHOTO_FILENAME = "photo_catalog_latest.xlsx"
PERSISTED_AVITO_FILENAME = "avito_latest.xlsx"
PERSISTED_COMPARISON_FILENAME = "comparison_latest.xlsx"
PERSISTED_WATCHLIST_FILENAME = "hot_items_watchlist_latest.dat"
PERSISTED_PURCHASE_FILENAME = "weighted_purchase_latest.xlsx"
PERSISTED_META_SUFFIX = ".meta.json"

FALLBACK_PHOTO_DOMAINS = ["rashodniki.ru", "t-toner.ru", "interlink.ru", "mrimage.ru"]
FALLBACK_SEARCH_LIMIT = 2


def get_server_data_dir() -> Path:
    try:
        base = Path(__file__).resolve().with_name(SERVER_DATA_DIRNAME)
    except Exception:
        base = Path.cwd() / SERVER_DATA_DIRNAME
    base.mkdir(parents=True, exist_ok=True)
    return base


def get_persisted_photo_file_path() -> Path:
    return get_server_data_dir() / PERSISTED_PHOTO_FILENAME


def get_persisted_avito_file_path() -> Path:
    return get_server_data_dir() / PERSISTED_AVITO_FILENAME


def get_persisted_comparison_file_path() -> Path:
    return get_server_data_dir() / PERSISTED_COMPARISON_FILENAME


def get_persisted_watchlist_file_path() -> Path:
    return get_server_data_dir() / PERSISTED_WATCHLIST_FILENAME


def get_persisted_purchase_file_path() -> Path:
    return get_server_data_dir() / PERSISTED_PURCHASE_FILENAME


def get_persisted_meta_path(file_path: Path) -> Path:
    return file_path.with_suffix(file_path.suffix + PERSISTED_META_SUFFIX)


def read_persisted_original_name(file_path: Path, default_name: str) -> str:
    meta_path = get_persisted_meta_path(file_path)
    if meta_path.exists():
        try:
            meta = json.loads(meta_path.read_text(encoding="utf-8"))
            name = normalize_text(meta.get("original_name", ""))
            if name:
                return name
        except Exception:
            pass
    return default_name


def save_uploaded_source_file(target_path: Path, file_bytes: bytes, original_name: str) -> None:
    target_path.parent.mkdir(parents=True, exist_ok=True)
    target_path.write_bytes(file_bytes)
    meta_path = get_persisted_meta_path(target_path)
    meta_path.write_text(json.dumps({
        "original_name": normalize_text(original_name),
        "saved_at": datetime.utcnow().isoformat(timespec="seconds"),
        "size": len(file_bytes),
    }, ensure_ascii=False, indent=2), encoding="utf-8")

def clear_runtime_perf_caches() -> None:
    for key in [
        "_perf_crm_products_cache",
        "_perf_decision_cache",
        "_perf_hot_buy_cache",
        "_perf_analytics_bundle_cache",
        "_perf_hot_lookup_cache",
    ]:
        try:
            st.session_state.pop(key, None)
        except Exception:
            pass


def clear_loader_caches() -> None:
    """Сбрасываем только тяжёлые loader-кеши после обновления server-side файлов."""
    try:
        load_comparison_workbook.clear()
    except Exception:
        pass
    try:
        load_photo_map_file.clear()
    except Exception:
        pass
    try:
        load_avito_file.clear()
    except Exception:
        pass
    try:
        load_hot_watchlist_file.clear()
    except Exception:
        pass
    try:
        load_purchase_cost_file.clear()
    except Exception:
        pass
    clear_runtime_perf_caches()


def _perf_cache_bucket(name: str) -> dict[str, Any]:
    bucket = st.session_state.get(name)
    if not isinstance(bucket, dict):
        bucket = {}
        st.session_state[name] = bucket
    return bucket


def _perf_signature(*parts: Any) -> str:
    try:
        payload = json.dumps(parts, ensure_ascii=False, default=str, sort_keys=True)
    except Exception:
        payload = repr(parts)
    return hashlib.md5(payload.encode("utf-8")).hexdigest()


def _registry_runtime_signature() -> tuple[Any, Any]:
    task_df = load_task_registry_df()
    task_sig = (
        len(task_df) if isinstance(task_df, pd.DataFrame) else 0,
        normalize_text(task_df.iloc[0].get("created_at", "")) if isinstance(task_df, pd.DataFrame) and not task_df.empty else "",
        normalize_text(task_df.iloc[0].get("status", "")) if isinstance(task_df, pd.DataFrame) and not task_df.empty else "",
    )
    pipe_df = load_pipeline_registry_df()
    pipe_sig = (
        len(pipe_df) if isinstance(pipe_df, pd.DataFrame) else 0,
        normalize_text(pipe_df.iloc[0].get("updated_at", "")) if isinstance(pipe_df, pd.DataFrame) and not pipe_df.empty else "",
        normalize_text(pipe_df.iloc[0].get("pipeline_status", "")) if isinstance(pipe_df, pd.DataFrame) and not pipe_df.empty else "",
    )
    return task_sig, pipe_sig


def _base_runtime_data_signature(include_registry: bool = True) -> tuple[Any, ...]:
    parts = [
        st.session_state.get("comparison_version", ""),
        st.session_state.get("photo_last_sync_sig", ""),
        st.session_state.get("avito_last_sync_sig", ""),
        st.session_state.get("hot_items_last_sync_sig", ""),
        st.session_state.get("purchase_cost_last_sync_sig", ""),
    ]
    if include_registry:
        parts.extend(_registry_runtime_signature())
    return tuple(parts)


def get_cached_hot_watchlist_lookup(hot_df: pd.DataFrame | None, tab_label: str = "") -> dict[str, list[dict[str, Any]]]:
    if not isinstance(hot_df, pd.DataFrame) or hot_df.empty:
        return {}
    sig = _perf_signature(
        _base_runtime_data_signature(include_registry=False),
        normalize_text(tab_label),
        len(hot_df),
        tuple(hot_df.columns.tolist()),
    )
    bucket = _perf_cache_bucket("_perf_hot_lookup_cache")
    cached = bucket.get(sig)
    if isinstance(cached, dict):
        return cached
    lookup = build_hot_watchlist_lookup(hot_df, tab_label=tab_label)
    if len(bucket) > 12:
        bucket.clear()
    bucket[sig] = lookup
    return lookup


def get_cached_crm_workspace_products_df(
    sheet_df: pd.DataFrame | None,
    photo_df: pd.DataFrame | None,
    avito_df: pd.DataFrame | None,
    min_qty: float,
    sheet_name: str,
    sheet_label: str,
) -> pd.DataFrame:
    if not isinstance(sheet_df, pd.DataFrame) or sheet_df.empty:
        return pd.DataFrame()
    sig = _perf_signature(
        _base_runtime_data_signature(include_registry=True),
        normalize_text(sheet_name),
        normalize_text(sheet_label),
        round(float(min_qty or 0.0), 4),
        len(sheet_df),
        tuple(sheet_df.columns.tolist()),
    )
    bucket = _perf_cache_bucket("_perf_crm_products_cache")
    cached = bucket.get(sig)
    if isinstance(cached, pd.DataFrame):
        out = cached.copy(deep=False)
        out.attrs["runtime_sig"] = sig
        return out
    out = build_crm_workspace_products_df(sheet_df, photo_df, avito_df, min_qty, sheet_name, sheet_label)
    out.attrs["runtime_sig"] = sig
    if len(bucket) > 12:
        bucket.clear()
    bucket[sig] = out.copy(deep=False)
    return out


def get_cached_procurement_decision_df(products_df: pd.DataFrame) -> pd.DataFrame:
    if not isinstance(products_df, pd.DataFrame) or products_df.empty:
        return pd.DataFrame()
    sig = _perf_signature(
        products_df.attrs.get("runtime_sig", ""),
        len(products_df),
        float(st.session_state.get("distributor_threshold", 35.0) or 35.0),
    )
    bucket = _perf_cache_bucket("_perf_decision_cache")
    cached = bucket.get(sig)
    if isinstance(cached, pd.DataFrame):
        return cached.copy(deep=False)
    out = build_procurement_decision_df(products_df)
    if len(bucket) > 12:
        bucket.clear()
    bucket[sig] = out.copy(deep=False)
    return out


def get_cached_operational_analytics_bundle(
    sheet_df: pd.DataFrame | None,
    photo_df: pd.DataFrame | None,
    avito_df: pd.DataFrame | None,
    min_qty: float,
    sheet_name: str,
    hot_items_df: pd.DataFrame | None = None,
) -> dict[str, Any]:
    if not isinstance(sheet_df, pd.DataFrame) or sheet_df.empty:
        return {}
    sig = _perf_signature(
        _base_runtime_data_signature(include_registry=True),
        normalize_text(sheet_name),
        round(float(min_qty or 0.0), 4),
        len(sheet_df),
        tuple(sheet_df.columns.tolist()),
    )
    bucket = _perf_cache_bucket("_perf_analytics_bundle_cache")
    cached = bucket.get(sig)
    if isinstance(cached, dict):
        return cached
    registry_df = load_avito_registry_df()
    bundle = build_operational_analytics_bundle(sheet_df, photo_df, avito_df, registry_df, min_qty, sheet_name, hot_items_df)
    if len(bucket) > 8:
        bucket.clear()
    bucket[sig] = bundle
    return bundle


def get_cached_hot_buy_watchlist_table() -> pd.DataFrame:
    hot_df = st.session_state.get("hot_items_df")
    if not isinstance(hot_df, pd.DataFrame) or hot_df.empty:
        return pd.DataFrame()
    threshold_pct = float(st.session_state.get("distributor_threshold", 35.0) or 35.0)
    min_qty = float(st.session_state.get("distributor_min_qty", 1.0) or 1.0)
    sig = _perf_signature(_base_runtime_data_signature(include_registry=False), threshold_pct, min_qty)
    bucket = _perf_cache_bucket("_perf_hot_buy_cache")
    cached = bucket.get(sig)
    if isinstance(cached, pd.DataFrame):
        return cached.copy(deep=False)

    sheets = st.session_state.get("comparison_sheets", {})
    photo_df = st.session_state.get("photo_df")
    avito_df = st.session_state.get("avito_df")
    sheet_specs = [("Сравнение", "Оригинал"), ("Уценка", "Уценка"), ("Совместимые", "Совместимые")]
    all_parts: list[pd.DataFrame] = []
    if isinstance(sheets, dict) and sheets:
        for sheet_name, sheet_label in sheet_specs:
            sheet_df = sheets.get(sheet_name)
            if not isinstance(sheet_df, pd.DataFrame) or sheet_df.empty:
                continue
            try:
                products_df = get_cached_crm_workspace_products_df(sheet_df, photo_df, avito_df, min_qty, sheet_name, sheet_label)
                decision_df = get_cached_procurement_decision_df(products_df)
                buy_df = filter_procurement_queue(decision_df, "Можно брать")
                if isinstance(buy_df, pd.DataFrame) and not buy_df.empty:
                    all_parts.append(buy_df)
            except Exception:
                continue
        if all_parts:
            out = pd.concat(all_parts, ignore_index=True)
            if "Разница, %" in out.columns:
                out = out[pd.to_numeric(out["Разница, %"], errors="coerce").fillna(0.0).ge(float(threshold_pct))].copy()
            if not out.empty:
                sort_cols = [c for c in ["Приоритет", "Продажи, шт/мес", "Лист", "Артикул"] if c in out.columns]
                if sort_cols:
                    ascending = [False if c in {"Приоритет", "Продажи, шт/мес"} else True for c in sort_cols]
                    out = out.sort_values(sort_cols, ascending=ascending, kind="stable")
                out = out.reset_index(drop=True)
                if len(bucket) > 6:
                    bucket.clear()
                bucket[sig] = out.copy(deep=False)
                return out

    out = pd.DataFrame()
    if len(bucket) > 6:
        bucket.clear()
    bucket[sig] = out.copy(deep=False)
    return out


def log_operation(message: str, level: str = "info") -> None:
    try:
        if "operation_log" not in st.session_state or not isinstance(st.session_state.get("operation_log"), list):
            st.session_state["operation_log"] = []
        stamp = datetime.utcnow().strftime("%H:%M:%S")
        st.session_state["operation_log"].append({"time": stamp, "level": level, "message": normalize_text(message)})
        st.session_state["operation_log"] = st.session_state["operation_log"][-25:]
    except Exception:
        pass


def render_operation_log_sidebar() -> None:
    log_items = st.session_state.get("operation_log", [])
    with st.expander("История действий", expanded=False):
        if not log_items:
            st.caption("Пока пусто")
            return
        for item in reversed(log_items[-12:]):
            icon = "✅" if item.get("level") == "success" else ("⚠️" if item.get("level") == "warning" else "•")
            st.markdown(f"{icon} **{html.escape(str(item.get('time', '')))}** — {html.escape(str(item.get('message', '')))}", unsafe_allow_html=True)




SERVICE_SNAPSHOT_DIRNAME = "_service_snapshots"
SERVICE_EXPORT_DIRNAME = "_service_exports"
SERVICE_SAFE_BOOT_FLAG = "_service_safe_boot.flag"


def get_app_root_dir() -> Path:
    try:
        return Path(__file__).resolve().parent
    except Exception:
        return Path.cwd()


def get_service_snapshots_dir() -> Path:
    path = get_server_data_dir() / SERVICE_SNAPSHOT_DIRNAME
    path.mkdir(parents=True, exist_ok=True)
    return path


def get_service_exports_dir() -> Path:
    path = get_server_data_dir() / SERVICE_EXPORT_DIRNAME
    path.mkdir(parents=True, exist_ok=True)
    return path


def get_service_safe_boot_flag_path() -> Path:
    return get_server_data_dir() / SERVICE_SAFE_BOOT_FLAG


def is_service_safe_boot_enabled() -> bool:
    return get_service_safe_boot_flag_path().exists()


def enable_service_safe_boot() -> None:
    flag = get_service_safe_boot_flag_path()
    flag.parent.mkdir(parents=True, exist_ok=True)
    flag.write_text("1", encoding="utf-8")
    log_operation("Сервис: включён безопасный запуск", "warning")


def disable_service_safe_boot() -> None:
    flag = get_service_safe_boot_flag_path()
    if flag.exists():
        flag.unlink()
    log_operation("Сервис: безопасный запуск выключен", "success")


def _service_db_path(filename: str) -> Path:
    try:
        return Path(__file__).resolve().with_name(filename)
    except Exception:
        return Path.cwd() / filename


def _service_slug(value: str) -> str:
    txt = normalize_text(value) or "snapshot"
    txt = re.sub(r"[^A-Za-zА-Яа-я0-9._-]+", "_", txt)
    txt = re.sub(r"_+", "_", txt).strip("._-")
    return txt[:64] or "snapshot"


def _service_rel_path(path: Path) -> str:
    try:
        return str(path.resolve().relative_to(get_app_root_dir().resolve()))
    except Exception:
        return path.name


def _service_file_md5(path: Path) -> str:
    md5 = hashlib.md5()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(1024 * 1024), b""):
            if not chunk:
                break
            md5.update(chunk)
    return md5.hexdigest()


def get_service_live_file_entries() -> list[dict[str, Any]]:
    comparison = get_persisted_comparison_file_path()
    photo = get_persisted_photo_file_path()
    avito = get_persisted_avito_file_path()
    watchlist = get_persisted_watchlist_file_path()
    purchase = get_persisted_purchase_file_path()
    entries = [
        {"label": "comparison", "path": comparison},
        {"label": "comparison_meta", "path": get_persisted_meta_path(comparison)},
        {"label": "photo", "path": photo},
        {"label": "photo_meta", "path": get_persisted_meta_path(photo)},
        {"label": "avito", "path": avito},
        {"label": "avito_meta", "path": get_persisted_meta_path(avito)},
        {"label": "watchlist", "path": watchlist},
        {"label": "watchlist_meta", "path": get_persisted_meta_path(watchlist)},
        {"label": "purchase", "path": purchase},
        {"label": "purchase_meta", "path": get_persisted_meta_path(purchase)},
        {"label": "review_tasks_db", "path": _service_db_path("review_tasks.sqlite")},
        {"label": "avito_registry_db", "path": _service_db_path("avito_registry.sqlite")},
        {"label": "photo_registry_db", "path": _service_db_path("photo_registry.sqlite")},
        {"label": "price_patch_history_db", "path": _service_db_path("price_patch_history.sqlite")},
        {"label": "card_overrides_db", "path": _service_db_path("card_overrides.sqlite")},
    ]
    return entries


def maybe_create_service_snapshot_before_action(action_key: str, content_sig: str, reason: str) -> str:
    state_key = f"service_snapshot_sig__{action_key}"
    if st.session_state.get(state_key) == content_sig:
        return ""
    snap_dir = create_service_snapshot(reason=reason, source="auto")
    st.session_state[state_key] = content_sig
    return snap_dir.name if isinstance(snap_dir, Path) else ""


def create_service_snapshot(reason: str = "", source: str = "manual") -> Path:
    snap_root = get_service_snapshots_dir()
    timestamp = datetime.utcnow().strftime("%Y-%m-%d_%H-%M-%S")
    folder_name = f"{timestamp}__{_service_slug(source)}__{_service_slug(reason or 'snapshot')}"
    snap_dir = snap_root / folder_name
    files_dir = snap_dir / "files"
    files_dir.mkdir(parents=True, exist_ok=True)

    files_meta: list[dict[str, Any]] = []
    for entry in get_service_live_file_entries():
        path = entry["path"]
        if not path.exists() or not path.is_file():
            continue
        rel_path = _service_rel_path(path)
        target = files_dir / rel_path
        target.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(path, target)
        files_meta.append({
            "label": entry["label"],
            "relative_path": rel_path,
            "size": path.stat().st_size,
            "md5": _service_file_md5(path),
            "modified_at": datetime.utcfromtimestamp(path.stat().st_mtime).isoformat(timespec="seconds"),
        })

    manifest = {
        "snapshot_name": folder_name,
        "created_at": datetime.utcnow().isoformat(timespec="seconds"),
        "source": normalize_text(source),
        "reason": normalize_text(reason),
        "app_title": APP_TITLE,
        "app_version": APP_VERSION,
        "file_count": len(files_meta),
        "files": files_meta,
    }
    (snap_dir / "manifest.json").write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    log_operation(f"Сервис: создан snapshot {folder_name} ({len(files_meta)} файлов)", "success")
    return snap_dir


def _read_service_snapshot_manifest(snap_dir: Path) -> dict[str, Any]:
    manifest_path = snap_dir / "manifest.json"
    if manifest_path.exists():
        try:
            return json.loads(manifest_path.read_text(encoding="utf-8"))
        except Exception:
            pass
    return {
        "snapshot_name": snap_dir.name,
        "created_at": "",
        "source": "",
        "reason": "",
        "app_title": APP_TITLE,
        "app_version": APP_VERSION,
        "file_count": 0,
        "files": [],
    }


def list_service_snapshots(limit: int = 50) -> list[dict[str, Any]]:
    root = get_service_snapshots_dir()
    items: list[dict[str, Any]] = []
    for snap_dir in root.iterdir():
        if not snap_dir.is_dir():
            continue
        meta = _read_service_snapshot_manifest(snap_dir)
        meta["path"] = str(snap_dir)
        meta["name"] = snap_dir.name
        items.append(meta)
    items.sort(key=lambda x: (str(x.get("created_at", "")), str(x.get("name", ""))), reverse=True)
    return items[:limit]


def build_service_snapshot_compare_df(snapshot_name: str) -> pd.DataFrame:
    snap_dir = get_service_snapshots_dir() / snapshot_name
    files_dir = snap_dir / "files"
    rows: list[dict[str, Any]] = []
    for entry in get_service_live_file_entries():
        live_path = entry["path"]
        rel_path = _service_rel_path(live_path)
        snap_path = files_dir / rel_path
        live_exists = live_path.exists()
        snap_exists = snap_path.exists()
        changed = ""
        if live_exists and snap_exists:
            try:
                changed = "Да" if _service_file_md5(live_path) != _service_file_md5(snap_path) else "Нет"
            except Exception:
                changed = "?"
        elif live_exists != snap_exists:
            changed = "Да"
        rows.append({
            "Файл": rel_path,
            "В snapshot": "Да" if snap_exists else "—",
            "Сейчас": "Да" if live_exists else "—",
            "Snapshot, КБ": round(snap_path.stat().st_size / 1024, 1) if snap_exists else 0.0,
            "Сейчас, КБ": round(live_path.stat().st_size / 1024, 1) if live_exists else 0.0,
            "Изменён": changed,
        })
    return pd.DataFrame(rows)


def restore_service_snapshot(snapshot_name: str) -> dict[str, Any]:
    snap_dir = get_service_snapshots_dir() / snapshot_name
    files_dir = snap_dir / "files"
    if not snap_dir.exists():
        raise FileNotFoundError(f"Snapshot не найден: {snapshot_name}")

    emergency = create_service_snapshot(reason=f"before restore {snapshot_name}", source="pre_restore")

    restored: list[str] = []
    removed: list[str] = []
    for entry in get_service_live_file_entries():
        live_path = entry["path"]
        rel_path = _service_rel_path(live_path)
        snap_path = files_dir / rel_path
        if snap_path.exists():
            live_path.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(snap_path, live_path)
            restored.append(rel_path)
        elif live_path.exists():
            try:
                live_path.unlink()
                removed.append(rel_path)
            except Exception:
                pass

    clear_loader_caches()
    for key in [
        "comparison_sheets", "comparison_name", "comparison_version", "current_df",
        "photo_df", "photo_name", "photo_last_sync_sig", "photo_registry_message", "photo_registry_stats",
        "avito_df", "avito_name", "avito_last_sync_sig", "avito_registry_message", "avito_registry_stats",
        "hot_items_df", "hot_items_name", "hot_items_last_sync_sig",
        "purchase_cost_df", "purchase_cost_name", "purchase_cost_last_sync_sig",
        "patch_message",
        "last_result_original", "last_result_discount", "last_result_compatible",
        "last_result_sig_original", "last_result_sig_discount", "last_result_sig_compatible",
        "comparison_upload_applied_sig", "photo_upload_applied_sig", "avito_upload_applied_sig", "hot_upload_applied_sig", "purchase_upload_applied_sig",
        "service_snapshot_sig__comparison_upload", "service_snapshot_sig__photo_upload",
        "service_snapshot_sig__avito_upload", "service_snapshot_sig__watchlist_upload", "service_snapshot_sig__purchase_upload",
    ]:
        st.session_state.pop(key, None)

    notice = (
        f"Восстановлен snapshot: {snapshot_name}. "
        f"Файлов восстановлено: {len(restored)}, удалено по снимку: {len(removed)}. "
        f"Страховочный snapshot: {emergency.name}."
    )
    st.session_state["service_restore_notice"] = notice
    log_operation(f"Сервис: выполнено восстановление из {snapshot_name}", "warning")
    return {
        "restored": len(restored),
        "removed": len(removed),
        "emergency_snapshot": emergency.name,
        "notice": notice,
    }


def build_service_backup_zip_bytes(include_snapshots: bool = True) -> bytes:
    buf = io.BytesIO()
    export_meta = {
        "exported_at": datetime.utcnow().isoformat(timespec="seconds"),
        "app_title": APP_TITLE,
        "app_version": APP_VERSION,
        "include_snapshots": bool(include_snapshots),
    }
    with zipfile.ZipFile(buf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("export_manifest.json", json.dumps(export_meta, ensure_ascii=False, indent=2))
        for entry in get_service_live_file_entries():
            path = entry["path"]
            if path.exists() and path.is_file():
                zf.write(path, arcname=f"live/{_service_rel_path(path)}")
        if include_snapshots:
            snap_root = get_service_snapshots_dir()
            for path in snap_root.rglob("*"):
                if path.is_file():
                    zf.write(path, arcname=f"snapshots/{path.relative_to(snap_root)}")
    return buf.getvalue()


def _service_sqlite_status(label: str, path: Path) -> dict[str, Any]:
    if not path.exists():
        return {"name": label, "status": "warn", "details": "файл ещё не создан"}
    try:
        with sqlite3.connect(path) as conn:
            tables = pd.read_sql_query(
                "SELECT name FROM sqlite_master WHERE type='table' ORDER BY name",
                conn,
            )
        table_count = len(tables) if isinstance(tables, pd.DataFrame) else 0
        return {"name": label, "status": "ok", "details": f"ok • таблиц: {table_count}"}
    except Exception as exc:
        return {"name": label, "status": "fail", "details": f"ошибка открытия: {normalize_text(exc)}"}


def run_service_healthcheck() -> dict[str, Any]:
    checks: list[dict[str, Any]] = []

    comp_path = get_persisted_comparison_file_path()
    if comp_path.exists():
        try:
            wb = load_comparison_workbook(read_persisted_original_name(comp_path, comp_path.name), comp_path.read_bytes())
            sheets = list(wb.keys()) if isinstance(wb, dict) else []
            total_rows = sum(len(df) for df in wb.values()) if isinstance(wb, dict) else 0
            required = {"Сравнение", "Уценка", "Совместимые"}
            missing = sorted(required - set(sheets))
            status = "ok" if not missing else "warn"
            details = f"ok • листов: {len(sheets)}, строк: {total_rows}"
            if missing:
                details += f" • нет листов: {', '.join(missing)}"
            checks.append({"name": "comparison", "status": status, "details": details})
        except Exception as exc:
            checks.append({"name": "comparison", "status": "fail", "details": f"ошибка загрузки: {normalize_text(exc)}"})
    else:
        checks.append({"name": "comparison", "status": "fail", "details": "файл не найден"})

    photo_path = get_persisted_photo_file_path()
    if photo_path.exists():
        try:
            photo_df = load_photo_map_file(read_persisted_original_name(photo_path, photo_path.name), photo_path.read_bytes())
            checks.append({"name": "фото", "status": "ok", "details": f"ok • строк: {len(photo_df)}"})
        except Exception as exc:
            checks.append({"name": "фото", "status": "fail", "details": f"ошибка загрузки: {normalize_text(exc)}"})
    else:
        checks.append({"name": "фото", "status": "warn", "details": "файл не найден"})

    watch_path = get_persisted_watchlist_file_path()
    if watch_path.exists():
        try:
            watch_df = load_hot_watchlist_file(read_persisted_original_name(watch_path, watch_path.name), watch_path.read_bytes())
            status = "ok" if len(watch_df) > 0 else "warn"
            details = f"ok • строк: {len(watch_df)}" if len(watch_df) > 0 else "файл читается, но валидных строк нет"
            checks.append({"name": "watchlist", "status": status, "details": details})
        except Exception as exc:
            checks.append({"name": "watchlist", "status": "fail", "details": f"ошибка загрузки: {normalize_text(exc)}"})
    else:
        checks.append({"name": "watchlist", "status": "warn", "details": "файл не найден"})

    purchase_path = get_persisted_purchase_file_path()
    if purchase_path.exists():
        try:
            purchase_df = load_purchase_cost_file(read_persisted_original_name(purchase_path, purchase_path.name), purchase_path.read_bytes())
            checks.append({"name": "purchase cost", "status": "ok", "details": f"ok • строк: {len(purchase_df)}"})
        except Exception as exc:
            checks.append({"name": "purchase cost", "status": "fail", "details": f"ошибка загрузки: {normalize_text(exc)}"})
    else:
        checks.append({"name": "purchase cost", "status": "warn", "details": "файл не найден"})

    avito_path = get_persisted_avito_file_path()
    if avito_path.exists():
        try:
            avito_df = load_avito_file(read_persisted_original_name(avito_path, avito_path.name), avito_path.read_bytes())
            checks.append({"name": "Avito registry input", "status": "ok", "details": f"ok • строк: {len(avito_df)}"})
        except Exception as exc:
            checks.append({"name": "Avito registry input", "status": "fail", "details": f"ошибка загрузки: {normalize_text(exc)}"})
    else:
        checks.append({"name": "Avito registry input", "status": "warn", "details": "файл не найден"})

    checks.append(_service_sqlite_status("tasks DB", _service_db_path("review_tasks.sqlite")))
    checks.append(_service_sqlite_status("photo registry DB", _service_db_path("photo_registry.sqlite")))
    checks.append(_service_sqlite_status("Avito registry DB", _service_db_path("avito_registry.sqlite")))
    checks.append(_service_sqlite_status("price history DB", _service_db_path("price_patch_history.sqlite")))

    snaps = list_service_snapshots(limit=200)
    last_snapshot = snaps[0].get("created_at", "") if snaps else ""
    return {
        "checks": checks,
        "snapshots_count": len(snaps),
        "last_snapshot": last_snapshot,
        "safe_boot": is_service_safe_boot_enabled(),
    }


def render_service_mode_sidebar() -> None:
    status = run_service_healthcheck()
    service_open = st.checkbox(
        "Открыть сервисный режим",
        key="service_mode_open",
        help="Ленивая сервисная панель. Пока блок закрыт, проверки, сравнение snapshot и сбор backup.zip не запускаются.",
    )
    safe_boot_on = bool(status.get("safe_boot"))
    st.markdown(
        f"<div class='sidebar-mini'>Safe boot: <b>{'включён' if safe_boot_on else 'выключен'}</b></div>",
        unsafe_allow_html=True,
    )
    st.caption("ⓘ Safe boot — облегчённый запуск. Полезен, если после неудачного обновления нужно спокойно зайти в систему и сделать откат.")
    sb1, sb2 = st.columns(2)
    if sb1.button(
        "Включить safe boot",
        use_container_width=True,
        key="service_enable_safe_boot",
        help="Включает облегчённый запуск. Тяжёлые блоки можно не рендерить, чтобы быстрее восстановить систему.",
    ):
        enable_service_safe_boot()
        st.rerun()
    if sb2.button(
        "Выключить safe boot",
        use_container_width=True,
        key="service_disable_safe_boot",
        help="Возвращает обычный режим работы приложения.",
    ):
        disable_service_safe_boot()
        st.rerun()

    if notice := st.session_state.get("service_restore_notice"):
        st.success(notice)

    if not service_open:
        st.markdown(
            "<div class='sidebar-mini'>Пока блок закрыт — проверки, архивы и сравнение snapshot не строятся.</div>",
            unsafe_allow_html=True,
        )
        return

    st.markdown("**1. Статус системы**")
    st.caption("ⓘ Показывает, что именно сейчас читается без ошибок: основные файлы, реестры SQLite, snapshots и safe boot.")
    for rec in status.get("checks", []):
        icon = "✅" if rec.get("status") == "ok" else ("⚠️" if rec.get("status") == "warn" else "❌")
        st.markdown(f"{icon} **{html.escape(str(rec.get('name', '')))}** — {html.escape(str(rec.get('details', '')))}")
    st.markdown(
        f"<div class='sidebar-mini'>Snapshots: <b>{int(status.get('snapshots_count', 0))}</b> • последний: <b>{html.escape(str(status.get('last_snapshot') or '—'))}</b></div>",
        unsafe_allow_html=True,
    )

    st.markdown("**2. Сделать snapshot сейчас**")
    st.caption("ⓘ Snapshot — это точка отката. Перед рискованными действиями можно вручную сохранить текущее состояние системы.")
    st.text_input(
        "Причина snapshot",
        key="service_snapshot_reason",
        placeholder="Например: перед загрузкой нового comparison",
        help="Коротко подпиши, зачем создаётся снимок. Потом по этой причине легче найти нужную точку отката.",
    )
    if st.button(
        "Сделать snapshot сейчас",
        use_container_width=True,
        key="service_snapshot_now",
        help="Сохраняет текущие файлы /data и внутренние реестры в отдельную папку snapshot.",
    ):
        reason = st.session_state.get("service_snapshot_reason", "") or "manual snapshot"
        snap = create_service_snapshot(reason=reason, source="manual")
        st.success(f"Snapshot создан: {snap.name}")

    st.markdown("**3. Восстановление**")
    st.caption("ⓘ Здесь можно сравнить текущее состояние с выбранным snapshot и вернуть систему назад. Перед восстановлением автоматически создаётся страховочный snapshot.")
    snapshots = list_service_snapshots(limit=100)
    if snapshots:
        options = [item["name"] for item in snapshots]
        st.selectbox(
            "Выбери snapshot",
            options=options,
            key="service_selected_snapshot",
            help="Список доступных точек отката. В названии и описании видно дату и причину создания.",
            format_func=lambda x: next(
                (
                    f"{item.get('created_at', '')} • {item.get('reason', '') or item.get('name', '')}"
                    for item in snapshots if item.get("name") == x
                ),
                x,
            ),
        )
        selected_snapshot = st.session_state.get("service_selected_snapshot")
        compare_df = build_service_snapshot_compare_df(selected_snapshot)
        if not compare_df.empty:
            st.dataframe(compare_df, use_container_width=True, height=180)
        st.checkbox(
            "Я понимаю, что текущее состояние будет заменено выбранным snapshot",
            key="service_restore_confirm",
            help="Защита от случайного отката. Без этого подтверждения восстановление не запустится.",
        )
        if st.button(
            "Восстановить snapshot",
            use_container_width=True,
            key="service_restore_snapshot",
            help="Копирует файлы из выбранного snapshot обратно в рабочее состояние приложения.",
        ):
            if not st.session_state.get("service_restore_confirm", False):
                st.warning("Подтверди восстановление чекбоксом выше.")
            else:
                restore_info = restore_service_snapshot(selected_snapshot)
                st.success(restore_info.get("notice", "Snapshot восстановлен."))
                st.rerun()
    else:
        st.info("Snapshot пока нет.")

    st.markdown("**4. Скачать резервную копию**")
    st.caption("ⓘ Backup.zip — это архив всего важного состояния, который можно скачать наружу и хранить отдельно от сервера.")
    st.checkbox(
        "Включить snapshots в backup.zip",
        key="service_backup_include_snapshots",
        value=True,
        help="Если включено, в архив попадут и live-файлы, и папка snapshots. Архив будет больше, но надёжнее.",
    )
    if st.button(
        "Собрать backup.zip",
        use_container_width=True,
        key="service_build_backup_zip",
        help="Собирает полный архив резервной копии для скачивания.",
    ):
        backup_bytes = build_service_backup_zip_bytes(
            include_snapshots=bool(st.session_state.get("service_backup_include_snapshots", True))
        )
        ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        export_path = get_service_exports_dir() / f"moy_tovar_backup_{ts}.zip"
        export_path.write_bytes(backup_bytes)
        st.session_state["service_backup_zip_path"] = str(export_path)
        st.session_state["service_backup_zip_name"] = export_path.name
        log_operation(f"Сервис: собран backup.zip ({export_path.name})", "success")
    backup_path = Path(st.session_state["service_backup_zip_path"]) if st.session_state.get("service_backup_zip_path") else None
    if backup_path and backup_path.exists():
        st.download_button(
            "⬇️ Скачать backup.zip",
            data=backup_path.read_bytes(),
            file_name=st.session_state.get("service_backup_zip_name", backup_path.name),
            mime="application/zip",
            use_container_width=True,
            key="service_download_backup_zip",
        )

    st.markdown("**5. Лог сервиса**")
    st.caption("ⓘ Здесь видны последние сервисные действия: snapshot, backup, restore, safe boot и предупреждения проверки.")
    service_keywords = ("Сервис:", "snapshot", "restore", "backup", "safe boot")
    service_log_items = [
        item for item in st.session_state.get("operation_log", [])
        if any(keyword.lower() in str(item.get("message", "")).lower() for keyword in service_keywords)
    ]
    if not service_log_items:
        st.caption("Сервисный лог пока пуст.")
    else:
        for item in reversed(service_log_items[-10:]):
            icon = "✅" if item.get("level") == "success" else ("⚠️" if item.get("level") == "warning" else "•")
            st.markdown(f"{icon} **{html.escape(str(item.get('time', '')))}** — {html.escape(str(item.get('message', '')))}", unsafe_allow_html=True)



def load_persisted_photo_source_into_state() -> bool:
    target = get_persisted_photo_file_path()
    if not target.exists():
        return False
    try:
        raw = target.read_bytes()
        df = load_photo_map_file(read_persisted_original_name(target, target.name), raw)
        st.session_state.photo_df = df
        st.session_state.photo_name = read_persisted_original_name(target, target.name) + " • из /data"
        return True
    except Exception:
        return False


def load_persisted_avito_source_into_state() -> bool:
    target = get_persisted_avito_file_path()
    if not target.exists():
        return False
    try:
        raw = target.read_bytes()
        st.session_state.avito_df = load_avito_file(read_persisted_original_name(target, target.name), raw)
        st.session_state.avito_name = read_persisted_original_name(target, target.name) + " • из /data"
        return True
    except Exception:
        return False


def load_persisted_comparison_source_into_state() -> bool:
    target = get_persisted_comparison_file_path()
    if not target.exists():
        return False
    try:
        raw = target.read_bytes()
        wb = load_comparison_workbook(read_persisted_original_name(target, target.name), raw)
        st.session_state.comparison_sheets = wb
        st.session_state.comparison_name = read_persisted_original_name(target, target.name) + " • из /data"
        st.session_state.comparison_version = datetime.utcnow().isoformat()
        available = list(wb.keys())
        if available and st.session_state.get("selected_sheet", "Сравнение") not in available:
            st.session_state.selected_sheet = available[0]
        rebuild_current_df()
        refresh_all_search_results()
        return True
    except Exception:
        return False



def normalize_watchlist_sheet_name(value: Any) -> str:
    txt = contains_text(value)
    if "ОРИГИН" in txt or "СРАВН" in txt:
        return "Оригинал"
    if "УЦЕН" in txt:
        return "Уценка"
    if "СОВМЕСТ" in txt:
        return "Совместимые"
    return normalize_text(value)


@st.cache_data(show_spinner=False, ttl=3600, max_entries=4)
def load_hot_watchlist_file(file_name: str, file_bytes: bytes) -> pd.DataFrame:
    suffix = Path(file_name).suffix.lower()

    def _read_csv_bytes(raw_bytes: bytes) -> pd.DataFrame:
        bio = io.BytesIO(raw_bytes)
        for enc in [None, "utf-8-sig", "cp1251", "windows-1251"]:
            try:
                bio.seek(0)
                if enc is None:
                    return pd.read_csv(bio)
                return pd.read_csv(bio, encoding=enc)
            except UnicodeDecodeError:
                continue
            except Exception:
                continue
        bio.seek(0)
        return pd.read_csv(bio, encoding="cp1251")

    def _read_excel_bytes(raw_bytes: bytes, ext: str) -> pd.DataFrame:
        bio = io.BytesIO(raw_bytes)
        if ext in {".xlsx", ".xlsm"} or raw_bytes[:2] == b"PK":
            bio.seek(0)
            return pd.read_excel(bio, engine="openpyxl")
        if ext == ".xls":
            try:
                bio.seek(0)
                return pd.read_excel(bio, engine="openpyxl")
            except Exception:
                pass
            try:
                html_tables = pd.read_html(io.BytesIO(raw_bytes))
                if html_tables:
                    return html_tables[0]
            except Exception:
                pass
            raise ValueError(
                "Статистика/Watchlist в старом формате .xls. На сервере он не читается без xlrd. "
                "Сохрани файл как .xlsx или .csv и загрузи снова."
            )
        try:
            bio.seek(0)
            return pd.read_excel(bio, engine="openpyxl")
        except Exception:
            pass
        try:
            html_tables = pd.read_html(io.BytesIO(raw_bytes))
            if html_tables:
                return html_tables[0]
        except Exception:
            pass
        raise ValueError("Не удалось распознать формат файла статистики. Используй .xlsx или .csv.")

    def _derive_velocity_band(month_value: float) -> str:
        if month_value >= 30:
            return "Очень быстро"
        if month_value >= 10:
            return "Быстро"
        if month_value >= 3:
            return "Стабильно"
        if month_value > 0:
            return "Медленно"
        return "Нет продаж"

    def _derive_abc_class(month_value: float) -> str:
        if month_value >= 20:
            return "A"
        if month_value >= 5:
            return "B"
        if month_value > 0:
            return "C"
        return "D"

    if suffix == ".csv":
        raw = _read_csv_bytes(file_bytes)
    else:
        raw = _read_excel_bytes(file_bytes, suffix)

    raw = raw.dropna(how="all").copy()
    if raw.empty:
        return pd.DataFrame(columns=[
            "watch_article", "watch_key", "watch_name", "current_sheet", "comparison_article",
            "sales_qty_15m", "sales_per_month", "abc_class", "velocity_band",
            "best_supplier", "best_supplier_gap_pct", "buy_signal_30pct", "days_of_cover",
            "priority_score", "action_today", "watch_article_norm", "watch_key_norm",
            "comparison_article_norm", "match_keys_text", "stats_source_kind",
            "sales_per_day", "sales_per_week", "sales_per_year", "deals_count",
            "first_sale", "last_sale", "days_without_sales", "market_min_price",
            "market_min_supplier", "supplier_presence_text", "our_price_now",
            "best_supplier_price_now", "best_supplier_stock_now", "our_stock_now",
        ])

    raw.columns = [normalize_text(c) for c in raw.columns]
    rows = []

    is_velocity_format = {"Артикул", "Наименование", "В месяц"}.issubset(set(raw.columns))

    if is_velocity_format:
        for _, r in raw.iterrows():
            watch_article = normalize_text(r.get("Артикул", ""))
            watch_name = normalize_text(r.get("Наименование", ""))
            comparison_article = watch_article
            sales_per_month = safe_float(r.get("В месяц"), 0.0)
            sales_per_day = safe_float(r.get("В день"), 0.0)
            sales_per_week = safe_float(r.get("В неделю"), 0.0)
            sales_per_year = safe_float(r.get("В год"), 0.0)
            total_sold_qty = safe_float(r.get("Всего шт."), 0.0)
            deals_count = safe_int(r.get("Сделок", 0), 0)
            first_sale = normalize_text(r.get("Первая продажа", ""))
            last_sale = normalize_text(r.get("Последняя продажа", ""))
            days_without_sales = safe_float(r.get("Дней без продаж"), 0.0)
            market_min_price = safe_float(r.get("Мин. цена конкурентов"), 0.0)
            market_min_supplier = normalize_text(r.get("Поставщик (мин.)", ""))
            supplier_presence_text = normalize_text(r.get("Наличие у поставщиков", ""))
            our_price_now = safe_float(r.get("Наша цена"), 0.0)

            keys = unique_preserve_order([
                normalize_article(watch_article),
                normalize_article(comparison_article),
            ])
            if not any(keys):
                name_codes = build_row_compare_codes("", watch_name)
                keys.extend([x for x in name_codes if x])
            if not any(keys):
                continue

            priority_score = max(sales_per_month, 0.0) * 10.0 + max(deals_count, 0) * 0.2 - min(max(days_without_sales, 0.0), 180.0) * 0.1
            rows.append({
                "watch_article": watch_article,
                "watch_key": watch_article,
                "watch_name": watch_name,
                "current_sheet": "",
                "comparison_article": comparison_article,
                "sales_qty_15m": total_sold_qty,
                "sales_per_month": sales_per_month,
                "abc_class": _derive_abc_class(sales_per_month),
                "velocity_band": _derive_velocity_band(sales_per_month),
                "ledger_end_qty": 0.0,
                "our_price_now": our_price_now,
                "our_stock_now": 0.0,
                "best_supplier": market_min_supplier,
                "best_supplier_price_now": market_min_price,
                "best_supplier_stock_now": 0.0,
                "best_supplier_gap_pct": 0.0,
                "buy_signal_30pct": "",
                "days_of_cover": 0.0,
                "priority_score": round(priority_score, 2),
                "action_today": "",
                "watch_article_norm": normalize_article(watch_article),
                "watch_key_norm": normalize_article(watch_article),
                "comparison_article_norm": normalize_article(comparison_article),
                "match_keys_text": "|".join([k for k in keys if k]),
                "stats_source_kind": "velocity",
                "sales_per_day": sales_per_day,
                "sales_per_week": sales_per_week,
                "sales_per_year": sales_per_year,
                "deals_count": deals_count,
                "first_sale": first_sale,
                "last_sale": last_sale,
                "days_without_sales": days_without_sales,
                "market_min_price": market_min_price,
                "market_min_supplier": market_min_supplier,
                "supplier_presence_text": supplier_presence_text,
            })
    else:
        for _, r in raw.iterrows():
            watch_article = normalize_text(r.get("watch_article", ""))
            watch_key = normalize_text(r.get("watch_key", ""))
            watch_name = normalize_text(r.get("watch_name", ""))
            comparison_article = normalize_text(r.get("comparison_article", ""))
            keys = unique_preserve_order([
                normalize_article(watch_article),
                normalize_article(watch_key),
                normalize_article(comparison_article),
            ])
            if not any(keys):
                continue
            rows.append({
                "watch_article": watch_article,
                "watch_key": watch_key,
                "watch_name": watch_name,
                "current_sheet": normalize_watchlist_sheet_name(r.get("current_sheet", "")),
                "comparison_article": comparison_article,
                "sales_qty_15m": safe_float(r.get("sales_qty_15m"), 0.0),
                "sales_per_month": safe_float(r.get("sales_per_month"), 0.0),
                "abc_class": normalize_text(r.get("abc_class", "")),
                "velocity_band": normalize_text(r.get("velocity_band", "")),
                "ledger_end_qty": safe_float(r.get("ledger_end_qty"), 0.0),
                "our_price_now": safe_float(r.get("our_price_now"), 0.0),
                "our_stock_now": safe_float(r.get("our_stock_now"), 0.0),
                "best_supplier": normalize_text(r.get("best_supplier", "")),
                "best_supplier_price_now": safe_float(r.get("best_supplier_price_now"), 0.0),
                "best_supplier_stock_now": safe_float(r.get("best_supplier_stock_now"), 0.0),
                "best_supplier_gap_pct": safe_float(r.get("best_supplier_gap_pct"), 0.0),
                "buy_signal_30pct": normalize_text(r.get("buy_signal_30pct", "")),
                "days_of_cover": safe_float(r.get("days_of_cover"), 0.0),
                "priority_score": safe_float(r.get("priority_score"), 0.0),
                "action_today": normalize_text(r.get("action_today", "")),
                "watch_article_norm": normalize_article(watch_article),
                "watch_key_norm": normalize_article(watch_key),
                "comparison_article_norm": normalize_article(comparison_article),
                "match_keys_text": "|".join([k for k in keys if k]),
                "stats_source_kind": "legacy_watchlist",
                "sales_per_day": safe_float(r.get("sales_per_day"), 0.0),
                "sales_per_week": safe_float(r.get("sales_per_week"), 0.0),
                "sales_per_year": safe_float(r.get("sales_per_year"), 0.0),
                "deals_count": safe_int(r.get("deals_count", 0), 0),
                "first_sale": normalize_text(r.get("first_sale", "")),
                "last_sale": normalize_text(r.get("last_sale", "")),
                "days_without_sales": safe_float(r.get("days_without_sales"), 0.0),
                "market_min_price": safe_float(r.get("market_min_price"), 0.0),
                "market_min_supplier": normalize_text(r.get("market_min_supplier", "")),
                "supplier_presence_text": normalize_text(r.get("supplier_presence_text", "")),
            })

    out = pd.DataFrame(rows)
    return out.reset_index(drop=True)


def build_hot_watchlist_lookup(hot_df: pd.DataFrame | None, tab_label: str = "") -> dict[str, list[dict[str, Any]]]:
    if not isinstance(hot_df, pd.DataFrame) or hot_df.empty:
        return {}
    work = hot_df.copy()
    tab_label = normalize_text(tab_label)
    if tab_label:
        filtered = work[(work["current_sheet"] == "") | (work["current_sheet"] == tab_label)].copy()
        if not filtered.empty:
            work = filtered
    lookup: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for _, row in work.iterrows():
        rec = row.to_dict()
        keys = [normalize_article(x) for x in normalize_text(rec.get("match_keys_text", "")).split("|") if normalize_article(x)]
        if not keys:
            continue
        for key in keys:
            lookup[key].append(rec)
    return lookup


def pick_hot_watch_rec(row: pd.Series, lookup: dict[str, list[dict[str, Any]]]) -> dict[str, Any] | None:
    if not lookup:
        return None
    candidate_keys = []
    article_norm = normalize_article(row.get("article_norm", row.get("article", "")))
    if article_norm:
        candidate_keys.append(article_norm)
    row_codes = row.get("row_codes")
    if isinstance(row_codes, list):
        candidate_keys.extend([normalize_article(x) for x in row_codes if normalize_article(x)])
    best = None
    best_score = -10**18
    seen = set()
    for key in candidate_keys:
        if not key or key in seen:
            continue
        seen.add(key)
        for rec in lookup.get(key, []):
            score = safe_float(rec.get("priority_score"), 0.0)
            if normalize_text(rec.get("buy_signal_30pct", "")).upper() == "BUY":
                score += 100000.0
            if key == normalize_article(rec.get("comparison_article_norm", "")):
                score += 1000.0
            elif key == normalize_article(rec.get("watch_article_norm", "")):
                score += 500.0
            if score > best_score:
                best = rec
                best_score = score
    return best


def apply_hot_watchlist(df: pd.DataFrame | None, hot_df: pd.DataFrame | None, tab_label: str = "") -> pd.DataFrame | None:
    if not isinstance(df, pd.DataFrame) or df.empty:
        return df
    if not isinstance(hot_df, pd.DataFrame) or hot_df.empty:
        return df
    lookup = build_hot_watchlist_lookup(hot_df, tab_label=tab_label)
    if not lookup:
        return df
    work = df.copy()
    matches = [pick_hot_watch_rec(row, lookup) for _, row in work.iterrows()]
    work["hot_flag"] = [bool(m) for m in matches]
    work["hot_sales_per_month"] = [safe_float((m or {}).get("sales_per_month"), 0.0) for m in matches]
    work["hot_priority_score"] = [safe_float((m or {}).get("priority_score"), 0.0) for m in matches]
    work["hot_abc_class"] = [normalize_text((m or {}).get("abc_class", "")) for m in matches]
    work["hot_velocity_band"] = [normalize_text((m or {}).get("velocity_band", "")) for m in matches]
    work["hot_action_today"] = [normalize_text((m or {}).get("action_today", "")) for m in matches]
    work["hot_buy_signal"] = [normalize_text((m or {}).get("buy_signal_30pct", "")) for m in matches]
    work["hot_best_supplier"] = [normalize_text((m or {}).get("best_supplier", "")) for m in matches]
    work["hot_best_supplier_gap_pct"] = [safe_float((m or {}).get("best_supplier_gap_pct"), 0.0) for m in matches]
    work["hot_watch_article"] = [normalize_text((m or {}).get("watch_article", "")) for m in matches]
    return work




def dataframe_to_excel_bytes(df: pd.DataFrame) -> bytes:
    out = io.BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        (df if isinstance(df, pd.DataFrame) else pd.DataFrame()).to_excel(writer, index=False, sheet_name="Watchlist")
    return out.getvalue()


def hot_watchlist_summary_text() -> str:
    hot_df = st.session_state.get("hot_items_df")
    if not isinstance(hot_df, pd.DataFrame) or hot_df.empty:
        return "статистика ещё не загружена"
    buy_count = len(get_cached_hot_buy_watchlist_table())
    strong_count = int(pd.to_numeric(hot_df.get("sales_per_month", pd.Series(dtype=float)), errors="coerce").fillna(0.0).ge(2.0).sum())
    source_kind = normalize_text(hot_df.get("stats_source_kind", pd.Series(dtype=object)).iloc[0] if "stats_source_kind" in hot_df.columns and not hot_df.empty else "")
    if source_kind == "velocity":
        return f"Строк статистики: {len(hot_df)} • сильный спрос: {strong_count} • можно брать: {buy_count}"
    return f"Ходовых: {len(hot_df)} • сильный спрос: {strong_count} • можно брать: {buy_count}"


def build_hot_buy_watchlist_table() -> pd.DataFrame:
    return get_cached_hot_buy_watchlist_table()


def render_hot_buy_watchlist_lazy_panel() -> None:
    global_open = bool(st.session_state.get("show_hot_buy_watchlist_table", False))
    crm_open = any(bool(v) for k, v in st.session_state.items() if str(k).startswith("crm_show_buy_"))
    if not (global_open or crm_open):
        return

    buy_df = build_hot_buy_watchlist_table()
    st.markdown('<div class="result-wrap">', unsafe_allow_html=True)
    render_block_header(
        "Ходовые позиции — сейчас можно брать",
        "Ленивая таблица только по тем позициям, где статистика продаж подтверждена и поставщик проходит твой порог выгоды.",
        icon="🔥",
        help_text="Показывает только позиции, где продажи подтверждены статистикой, а лучший поставщик реально проходит твой порог выгоды и имеет остаток.",
    )
    if buy_df.empty:
        st.info("Сейчас нет позиций, где поставщик проходит твой порог выгоды и есть подтверждённый сигнал по статистике.")
    else:
        c1, c2 = st.columns([1.2, 1])
        c1.metric("Позиций можно брать", len(buy_df))
        c2.download_button(
            "⬇️ Скачать таблицу «можно брать» в Excel",
            dataframe_to_excel_bytes(buy_df),
            file_name="hot_buy_watchlist.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="download_hot_buy_watchlist",
            use_container_width=True,
        )
        grid_key = f"hot_buy_grid_{int(global_open)}_{int(crm_open)}_{len(buy_df)}"
        st.dataframe(buy_df, use_container_width=True, height=min(560, 120 + max(1, len(buy_df)) * 28), key=grid_key)
    st.markdown('</div>', unsafe_allow_html=True)


def hot_supplier_note(row: pd.Series | dict | None, best: dict | None, threshold_pct: float = 35.0) -> tuple[str, str]:
    help_text = "Товар ходовой → товар хорошо продавался за выбранный период"
    if not best:
        help_text += f"\nСейчас брать невыгодно → нет поставщика с ценой минимум на {threshold_pct:.0f}% ниже нашей цены"
        return "Сейчас брать невыгодно", help_text

    source = normalize_text((best or {}).get("source", ""))
    delta_pct = safe_float((best or {}).get("delta_percent"), 0.0)
    if delta_pct >= float(threshold_pct):
        action_text = f"Сейчас можно брать у {source}" if source else "Сейчас можно брать"
        help_text += f"\nСейчас можно брать → лучший поставщик сейчас минимум на {threshold_pct:.0f}% дешевле нашей цены"
        return action_text, help_text

    help_text += f"\nСейчас брать невыгодно → нет поставщика с ценой минимум на {threshold_pct:.0f}% ниже нашей цены"
    return "Сейчас брать невыгодно", help_text

def load_persisted_watchlist_source_into_state() -> bool:
    target = get_persisted_watchlist_file_path()
    if not target.exists():
        return False
    try:
        raw = target.read_bytes()
        st.session_state.hot_items_df = load_hot_watchlist_file(read_persisted_original_name(target, target.name), raw)
        st.session_state.hot_items_name = read_persisted_original_name(target, target.name) + " • из /data"
        return True
    except Exception:
        return False


@st.cache_data(show_spinner=False, ttl=3600, max_entries=4)
def load_purchase_cost_file(file_name: str, file_bytes: bytes) -> pd.DataFrame:
    suffix = Path(file_name).suffix.lower()
    bio = io.BytesIO(file_bytes)
    engine = "openpyxl" if suffix in {".xlsx", ".xlsm"} or file_bytes[:2] == b"PK" else None
    try:
        xls = pd.ExcelFile(bio, engine=engine)
    except Exception:
        bio.seek(0)
        xls = pd.ExcelFile(bio)

    target_sheet = None
    for sheet in xls.sheet_names:
        sheet_txt = contains_text(sheet)
        if "ИТОГ" in sheet_txt and "ВЗВЕШ" in sheet_txt:
            target_sheet = sheet
            break
    if target_sheet is None:
        target_sheet = xls.sheet_names[0]

    raw = pd.read_excel(io.BytesIO(file_bytes), sheet_name=target_sheet, engine="openpyxl" if engine == "openpyxl" else None)
    raw = raw.dropna(how="all").copy()
    if raw.empty:
        return pd.DataFrame(columns=[
            "purchase_name", "purchase_name_compact", "purchase_codes", "purchase_codes_text",
            "purchase_avg_cost", "purchase_total_qty", "purchase_total_cost", "purchase_source_sheet",
            "purchase_match_hint"
        ])

    raw.columns = [normalize_text(c) for c in raw.columns]
    col_name = next((c for c in raw.columns if compact_text(c) in {"НОМЕНКЛАТУРА(B)", "НОМЕНКЛАТУРА", "НАИМЕНОВАНИЕ", "НАЗВАНИЕ"}), "")
    col_avg = next((c for c in raw.columns if "СРЕДНЯЯ" in contains_text(c) and "1" in contains_text(c)), "")
    if not col_avg:
        col_avg = next((c for c in raw.columns if "ЗАКУПК" in contains_text(c) and "1" in contains_text(c)), "")
    col_qty = next((c for c in raw.columns if "ОБЩЕЕ" in contains_text(c) and ("КОЛ" in contains_text(c) or "ШТ" in contains_text(c))), "")
    col_total = next((c for c in raw.columns if ("СКОРР" in contains_text(c) or "СУММ" in contains_text(c)) and "ЗАКУПК" in contains_text(c)), "")

    if not col_name or not col_avg:
        raise ValueError("В файле закупки не найдены колонки номенклатуры и средней закупки.")

    rows = []
    for _, r in raw.iterrows():
        purchase_name = normalize_text(r.get(col_name, ""))
        purchase_avg_cost = safe_float(r.get(col_avg), 0.0)
        if not purchase_name or purchase_avg_cost <= 0:
            continue
        codes = build_row_compare_codes("", purchase_name)
        rows.append({
            "purchase_name": purchase_name,
            "purchase_name_compact": compact_text(purchase_name),
            "purchase_codes": codes,
            "purchase_codes_text": "|" + "|".join(codes) + "|" if codes else "",
            "purchase_avg_cost": purchase_avg_cost,
            "purchase_total_qty": safe_float(r.get(col_qty), 0.0) if col_qty else 0.0,
            "purchase_total_cost": safe_float(r.get(col_total), 0.0) if col_total else 0.0,
            "purchase_source_sheet": target_sheet,
            "purchase_match_hint": "title_or_code",
            "purchase_name_tokens": tokenize_text(purchase_name),
        })
    return pd.DataFrame(rows)


def load_persisted_purchase_source_into_state() -> bool:
    target = get_persisted_purchase_file_path()
    if not target.exists():
        return False
    try:
        raw = target.read_bytes()
        st.session_state.purchase_cost_df = load_purchase_cost_file(read_persisted_original_name(target, target.name), raw)
        st.session_state.purchase_cost_name = read_persisted_original_name(target, target.name) + " • из /data"
        return True
    except Exception:
        return False


def purchase_cost_summary_text() -> str:
    purchase_df = st.session_state.get("purchase_cost_df")
    if not isinstance(purchase_df, pd.DataFrame) or purchase_df.empty:
        return "Средняя закупка: файл ещё не загружен"
    sheet_name = normalize_text(purchase_df.get("purchase_source_sheet", pd.Series(dtype=object)).iloc[0]) if "purchase_source_sheet" in purchase_df.columns and len(purchase_df) else ""
    return f"Средняя закупка: {len(purchase_df)} строк • лист: {sheet_name or '—'}"


def _purchase_match_score(product_name: str, candidate: dict[str, Any]) -> tuple[float, float, int]:
    target_tokens = set(tokenize_text(product_name))
    cand_tokens = set(candidate.get("purchase_name_tokens", []) or [])
    inter = len(target_tokens & cand_tokens)
    union = len(target_tokens | cand_tokens) or 1
    jaccard = inter / union
    target_comp = compact_text(product_name)
    cand_comp = normalize_text(candidate.get("purchase_name_compact", ""))
    prefix = 1.0 if (target_comp and (target_comp in cand_comp or cand_comp in target_comp)) else 0.0
    return (prefix, jaccard, inter)


def build_purchase_cost_indexes(purchase_df: pd.DataFrame | None) -> tuple[dict[str, list[dict[str, Any]]], dict[str, list[dict[str, Any]]]]:
    if not isinstance(purchase_df, pd.DataFrame) or purchase_df.empty:
        return {}, {}
    by_name: dict[str, list[dict[str, Any]]] = {}
    by_code: dict[str, list[dict[str, Any]]] = {}
    for _, r in purchase_df.iterrows():
        rec = r.to_dict()
        name_key = normalize_text(rec.get("purchase_name_compact", ""))
        if name_key:
            by_name.setdefault(name_key, []).append(rec)
        for code in rec.get("purchase_codes", []) or []:
            code_norm = normalize_article(code)
            if code_norm:
                by_code.setdefault(code_norm, []).append(rec)
    return by_name, by_code


def resolve_purchase_cost_for_product(article: object, name: object, purchase_by_name: dict[str, list[dict[str, Any]]], purchase_by_code: dict[str, list[dict[str, Any]]]) -> dict[str, Any]:
    article_norm = normalize_article(article)
    name_compact = compact_text(name)
    default = {
        "purchase_avg_cost": None,
        "purchase_total_qty": None,
        "purchase_total_cost": None,
        "purchase_match_source": "",
        "purchase_source_name": "",
        "purchase_source_sheet": "",
    }

    exact_name = purchase_by_name.get(name_compact, [])
    if len(exact_name) == 1:
        rec = exact_name[0]
        return {
            "purchase_avg_cost": safe_float(rec.get("purchase_avg_cost"), 0.0),
            "purchase_total_qty": safe_float(rec.get("purchase_total_qty"), 0.0),
            "purchase_total_cost": safe_float(rec.get("purchase_total_cost"), 0.0),
            "purchase_match_source": "name_exact",
            "purchase_source_name": normalize_text(rec.get("purchase_name", "")),
            "purchase_source_sheet": normalize_text(rec.get("purchase_source_sheet", "")),
        }

    code_matches = purchase_by_code.get(article_norm, []) if article_norm else []
    if len(code_matches) == 1:
        rec = code_matches[0]
        return {
            "purchase_avg_cost": safe_float(rec.get("purchase_avg_cost"), 0.0),
            "purchase_total_qty": safe_float(rec.get("purchase_total_qty"), 0.0),
            "purchase_total_cost": safe_float(rec.get("purchase_total_cost"), 0.0),
            "purchase_match_source": "code_from_name",
            "purchase_source_name": normalize_text(rec.get("purchase_name", "")),
            "purchase_source_sheet": normalize_text(rec.get("purchase_source_sheet", "")),
        }
    if len(code_matches) > 1:
        scored = sorted(
            code_matches,
            key=lambda rec: _purchase_match_score(normalize_text(name), rec),
            reverse=True,
        )
        best = scored[0] if scored else None
        if best is not None:
            prefix, jaccard, inter = _purchase_match_score(normalize_text(name), best)
            if prefix > 0 or jaccard >= 0.45 or inter >= 3:
                return {
                    "purchase_avg_cost": safe_float(best.get("purchase_avg_cost"), 0.0),
                    "purchase_total_qty": safe_float(best.get("purchase_total_qty"), 0.0),
                    "purchase_total_cost": safe_float(best.get("purchase_total_cost"), 0.0),
                    "purchase_match_source": "code_best_name",
                    "purchase_source_name": normalize_text(best.get("purchase_name", "")),
                    "purchase_source_sheet": normalize_text(best.get("purchase_source_sheet", "")),
                }
    return default


def apply_purchase_cost_map(df: pd.DataFrame | None, purchase_df: pd.DataFrame | None) -> pd.DataFrame | None:
    if df is None:
        return None
    out = df.copy()
    for col in ["purchase_avg_cost", "purchase_total_qty", "purchase_total_cost", "purchase_match_source", "purchase_source_name", "purchase_source_sheet"]:
        if col not in out.columns:
            out[col] = None if "cost" in col or "qty" in col else ""
    if not isinstance(purchase_df, pd.DataFrame) or purchase_df.empty:
        return out
    purchase_by_name, purchase_by_code = build_purchase_cost_indexes(purchase_df)
    resolved = out.apply(
        lambda r: resolve_purchase_cost_for_product(
            r.get("article", ""),
            r.get("name", ""),
            purchase_by_name,
            purchase_by_code,
        ),
        axis=1,
    )
    resolved_df = pd.DataFrame(list(resolved))
    for col in ["purchase_avg_cost", "purchase_total_qty", "purchase_total_cost", "purchase_match_source", "purchase_source_name", "purchase_source_sheet"]:
        if col in resolved_df.columns:
            out[col] = resolved_df[col]
    return out

DEFAULT_DISCOUNT_1 = 12.0
DEFAULT_DISCOUNT_2 = 20.0
DEFAULT_TEMPLATE1_FOOTER = (
    "Цeна с НДC : +17%\n\n"
    "Работaeм по будням, c 10 дo 18:00. Самовывоз по адресу: Москва, ул. Сущёвский Вал, 5с20\n\n"
    "Еcли пoтрeбуeтся пepeсылкa - oтпpaвляeм толькo Авитo-Яндeкc, Авито-СДЭК или Авито-Авито. Отправляем без наценки."
)

CATALOG_COLUMN_ALIASES = {
    "article": ["Артикул", "артикул", "Код", "код", "sku", "article"],
    "name": ["Наименование", "Номенклатура", "Название", "name"],
    "price": ["Наша цена", "Цена", "Цена продажи", "price"],
    "qty": ["Наш склад", "Свободно", "Остаток", "Количество", "qty"],
    "total_qty": ["Всего", "Итого", "Общий остаток", "Всего шт", "Итого шт"],
    "transit_qty": ["Транзит", "В транзите", "В пути", "Поступает", "Транзит шт"],
}

PHOTO_COLUMN_ALIASES = {
    "article": ["Артикул", "артикул", "Код", "код", "sku", "article"],
    "photo_url": [
        "Фото", "Ссылка на фото", "URL фото", "photo", "image", "image_url",
        "photo_url", "url", "link", "picture", "картинка", "ссылка",
        "imag", "images"
    ],
    "brand": ["brend", "бренд", "brand"],
    "color": ["czvet", "цвет", "color"],
    "capacity": ["emkost-kartridzha", "емкость-картриджа", "емкость картриджа", "capacity"],
    "manufacturer_code": ["kod-proizvoditelya", "код-производителя", "код производителя"],
    "model": ["model", "модель"],
    "description": ["originalinosti", "описание", "description"],
    "fits_models": ["podhodit-k-modelyam", "подходит-к-моделям", "подходит к моделям"],
    "iso_pages": ["resurs-po-iso-str", "ресурс-по-iso-стр", "ресурс", "iso", "pages"],
    "print_technology": ["tehnologiya-pechati", "технология-печати", "технология печати"],
    "item_type": ["tip", "тип"],
    "print_type": ["tip-pechati", "тип-печати", "тип печати"],
    "weight": ["weight", "вес"],
    "length": ["length", "длина"],
    "width": ["width", "ширина"],
    "height": ["height", "высота"],
}

AVITO_COLUMN_ALIASES = {
    "ad_id": ["Номер объявления", "ID объявления", "Номер"],
    "title": ["Название объявления", "Заголовок", "Название"],
    "price": ["Цена"],
    "url": ["Ссылка", "URL", "Ссылка на объявление", "Link"],
    "account": ["Аккаунт", "account", "Кабинет", "Профиль"],
}

CYRILLIC_ARTICLE_TRANSLATION = str.maketrans({
    "А": "A", "В": "B", "Е": "E", "К": "K", "М": "M", "Н": "H", "О": "O", "Р": "P", "С": "C", "Т": "T", "У": "Y", "Х": "X",
    "а": "A", "в": "B", "е": "E", "к": "K", "м": "M", "н": "H", "о": "O", "р": "P", "с": "C", "т": "T", "у": "Y", "х": "X",
    "Ё": "E", "ё": "E",
})

MERLION_PANTUM_EXTRA_ZERO_CODES = {
    "DL420P", "DL5120P", "DLR5220",
    "TL420HP", "TL420XP", "TL5120HP", "TL5120P", "TL5120XP",
}


def normalize_merlion_source_price(row: pd.Series, source: str, price: float) -> float:
    """
    Безопасная правка только для проблемной группы Pantum у Мерлиона.
    Делим на 10 только когда:
    - источник именно Мерлион
    - в строке есть Pantum
    - найден один из известных OEM-кодов
    - цена выглядит как явно раздутая: 51000 / 73500 / 122000 и т.п.
    """
    try:
        price_val = float(price)
    except Exception:
        return float(price)

    if compact_text(source) != "МЕРЛИОН":
        return price_val
    if price_val < 10000:
        return price_val

    row_text = contains_text(f"{row.get('article', '')} {row.get('name', '')}")
    if "PANTUM" not in row_text:
        return price_val

    row_codes = row.get("row_codes")
    if not isinstance(row_codes, list) or not row_codes:
        row_codes = build_row_compare_codes(row.get("article", ""), row.get("name", ""))
    if not any(normalize_article(code) in MERLION_PANTUM_EXTRA_ZERO_CODES for code in (row_codes or [])):
        return price_val

    rounded = int(round(price_val))
    if abs(price_val - rounded) > 1e-9:
        return price_val
    if rounded % 100 != 0:
        return price_val

    return price_val / 10.0


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    text = re.sub(r"\s+", " ", str(value).strip())
    if text.lower() in {"nan", "nat", "none"}:
        return ""
    return text


def normalize_article(value: object) -> str:
    text = normalize_text(value)
    if not text:
        return ""
    text = text.translate(CYRILLIC_ARTICLE_TRANSLATION)
    return re.sub(r"[^A-Za-z0-9]", "", text).upper()


def extract_first_url(value: object) -> str:
    text = normalize_text(value)
    if not text:
        return ""

    # Поддержка Excel-формулы HYPERLINK(...)
    if text.startswith("="):
        m = re.match(r'^=\s*(?:HYPERLINK|ГИПЕРССЫЛКА)\(\s*"([^"]+)"', text, flags=re.IGNORECASE)
        if m:
            return normalize_text(m.group(1))

    m = re.search(r'https?://[^\s"\'<>]+', text, flags=re.IGNORECASE)
    if not m:
        return ""

    url = m.group(0).strip()

    # Иногда после ссылки в ячейке идёт хвост вида: " ! alt : ..."
    url = re.sub(r'[!;,]+$', '', url).strip()

    # Обрезаем совсем явный мусор после картинки/файла, если он попал в строку.
    for stopper in [' ! ', ' alt :', ' title :', ' desc :', ' caption :']:
        pos = url.lower().find(stopper.strip().lower())
        if pos > 0:
            url = url[:pos].strip()

    return url


COLOR_TEMPLATE_KEYWORDS = [
    ("пурп", "пурпурный"),
    ("magenta", "пурпурный (magenta)"),
    ("голуб", "голубой"),
    ("cyan", "голубой (cyan)"),
    ("желт", "желтый"),
    ("yellow", "желтый (yellow)"),
    ("черн", "черный"),
    ("чёрн", "чёрный"),
    ("black", "чёрный (black)"),
    ("син", "синий"),
    ("blue", "синий (blue)"),
    ("красн", "красный"),
    ("red", "красный (red)"),
    ("зел", "зеленый"),
    ("green", "зеленый (green)"),
    ("сер", "серый"),
    ("grey", "серый"),
    ("gray", "серый"),
]


def extract_color_from_text(value: object) -> str:
    text = contains_text(value)
    if not text:
        return ""
    for needle, label in COLOR_TEMPLATE_KEYWORDS:
        if contains_text(needle) in text:
            return label
    return ""


def extract_iso_pages_from_text(value: object) -> str:
    raw = normalize_text(value)
    if not raw:
        return ""
    m = re.search(r"(\d[\d\s]{1,})\s*(?:стр|страниц)", raw, flags=re.IGNORECASE)
    if not m:
        m = re.search(r"(\d[\d\s]{1,})\s*стр", raw, flags=re.IGNORECASE)
    if not m:
        return ""
    digits = re.sub(r"\s+", "", m.group(1))
    if not digits:
        return ""
    try:
        return f"{int(digits):,}".replace(",", " ")
    except Exception:
        return digits


def normalize_pages_value(value: object) -> str:
    raw = normalize_text(value)
    if not raw:
        return ""
    # Берём только первое числовое значение и игнорируем хвосты вроде "(A4)" / "(А4)".
    # Иначе строка "2500 (A4)" превращается в ошибочные "25004".
    m = re.search(r"(\d[\d\s]{1,})", raw)
    if not m:
        m = re.search(r"(\d+)", raw)
    if m:
        digits = re.sub(r"\s+", "", m.group(1))
        if digits:
            try:
                return f"{int(digits):,}".replace(",", " ")
            except Exception:
                return digits
    return raw


def normalize_meta_measure(value: object) -> str:
    raw = normalize_text(value)
    if not raw:
        return ""
    try:
        num = float(str(raw).replace(" ", "").replace(",", "."))
        if abs(num - int(num)) < 1e-9:
            return str(int(num))
        txt = f"{num:.2f}".rstrip("0").rstrip(".")
        return txt.replace(".", ",")
    except Exception:
        return raw


def format_meta_dimensions(length: object, width: object, height: object) -> str:
    l = normalize_meta_measure(length)
    w = normalize_meta_measure(width)
    h = normalize_meta_measure(height)
    if not (l or w or h):
        return ""
    parts = [x for x in [l, w, h] if x]
    return " × ".join(parts) + " см"


def format_meta_weight(weight: object) -> str:
    w = normalize_meta_measure(weight)
    if not w:
        return ""
    if re.search(r"[A-Za-zА-Яа-яЁё]", w):
        return w
    return f"{w} кг"


def simplify_template_color(value: object) -> str:
    raw = normalize_text(value)
    if not raw:
        return ""
    low = raw.lower().strip()

    english_only_map = {
        "black": "чёрный",
        "cyan": "голубой",
        "yellow": "жёлтый",
        "magenta": "пурпурный",
        "blue": "синий",
        "red": "красный",
        "green": "зелёный",
        "grey": "серый",
        "gray": "серый",
    }

    # Сначала нормализуем случаи, когда цвет пришёл только как английское слово
    # или только в скобках: magenta / (magenta) -> пурпурный
    raw_no_parens = re.sub(r"^[\s(]+|[\s)]+$", "", low)
    if raw_no_parens in english_only_map:
        return english_only_map[raw_no_parens]

    # Если цвет уже содержит русское название и английский перевод в скобках,
    # оставляем только один, более короткий и привычный вариант.
    replacements = [
        (r"^(ч[её]рн(?:ый|ая|ое)?)\s*\((?:black)\)$", "чёрный"),
        (r"^(черный)\s*\((?:black)\)$", "чёрный"),
        (r"^(голуб(?:ой|ая|ое)?)\s*\((?:cyan)\)$", "голубой"),
        (r"^(ж[её]лт(?:ый|ая|ое)?)\s*\((?:yellow)\)$", "жёлтый"),
        (r"^(пурпурн(?:ый|ая|ое)?)\s*\((?:magenta)\)$", "пурпурный"),
        (r"^(син(?:ий|яя|ее)?)\s*\((?:blue)\)$", "синий"),
        (r"^(красн(?:ый|ая|ое)?)\s*\((?:red)\)$", "красный"),
        (r"^(зел[её]н(?:ый|ая|ое)?)\s*\((?:green)\)$", "зелёный"),
        (r"^(сер(?:ый|ая|ое)?)\s*\((?:grey|gray)\)$", "серый"),
    ]
    for pattern, repl in replacements:
        if re.match(pattern, low, flags=re.IGNORECASE):
            return repl

    # Если есть русский цвет + английский перевод в конце, убираем перевод.
    if re.search(r"\(([A-Za-z]+)\)$", raw) and re.search(r"[А-Яа-яЁё]", raw):
        raw = re.sub(r"\s*\([A-Za-z]+\)$", "", raw).strip()
        return raw

    return raw


def compose_article_template_label(row: pd.Series) -> str:
    article = normalize_text(row.get("article", ""))
    color = simplify_template_color(normalize_text(row.get("meta_color", "")) or extract_color_from_text(row.get("name", "")))
    pages = normalize_pages_value(row.get("meta_iso_pages", "")) or extract_iso_pages_from_text(row.get("name", ""))
    details = []
    if color:
        details.append(color)
    if pages:
        details.append(f"{pages} стр")
    return f"{article} ({', '.join(details)})" if details else article


def unique_text_values(values: list[object]) -> list[str]:
    seen: set[str] = set()
    out: list[str] = []
    for value in values:
        txt = normalize_text(value)
        if not txt:
            continue
        key = contains_text(txt)
        if key in seen:
            continue
        seen.add(key)
        out.append(txt)
    return out


def build_template_shared_lines(result_df: pd.DataFrame) -> list[str]:
    if result_df is None or result_df.empty:
        return []
    article_norms = {normalize_article(x) for x in result_df.get("article", []).tolist() if normalize_article(x)}
    model_values = unique_text_values(result_df.get("meta_model", pd.Series(dtype=object)).tolist())
    manufacturer_values = unique_text_values(result_df.get("meta_manufacturer_code", pd.Series(dtype=object)).tolist())
    fits_values = unique_text_values(result_df.get("meta_fits_models", pd.Series(dtype=object)).tolist())

    filtered_manufacturer = [v for v in manufacturer_values if normalize_article(v) not in article_norms]

    lines: list[str] = []
    if model_values:
        lines.append(f"Модель - {' / '.join(model_values)}")
    if filtered_manufacturer:
        lines.append(f"Код производителя - {' / '.join(filtered_manufacturer)}")
    if fits_values:
        lines.append(f"Подходит к моделям - {' / '.join(fits_values)}")
    return lines


def compact_text(value: object) -> str:
    return re.sub(r"\s+", "", normalize_text(value)).upper()


def contains_text(value: object) -> str:
    return normalize_text(value).upper()


def safe_float(value: Any, default: float = 0.0) -> float:
    if value is None:
        return float(default)
    try:
        if pd.isna(value):
            return float(default)
    except Exception:
        pass
    if isinstance(value, str):
        txt = normalize_text(value).replace(" ", "").replace(",", ".")
        if not txt:
            return float(default)
        try:
            return float(txt)
        except Exception:
            return float(default)
    try:
        return float(value)
    except Exception:
        return float(default)


def normalize_gap_percent(value: Any) -> float:
    x = safe_float(value, 0.0)
    if 0 < abs(x) < 1.0:
        return x * 100.0
    return x


def fmt_price(value: Any) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    val = safe_float(value, 0.0)
    if float(val).is_integer():
        return f"{int(val):,}".replace(",", " ")
    return f"{val:,.2f}".replace(",", " ").replace(".", ",")




def fmt_price_with_rub(value: Any) -> str:
    txt = fmt_price(value)
    return f"{txt} руб." if txt else ""

def fmt_qty(value: Any) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    val = safe_float(value, 0.0)
    if float(val).is_integer():
        return str(int(val))
    return f"{val:,.2f}".replace(",", " ").replace(".", ",")


def round_up_to_100(value: float) -> int:
    return int(math.ceil(float(value) / 100.0) * 100)


def round_to_nearest_100(value: float) -> int:
    return int(math.floor(float(value) / 100.0 + 0.5) * 100)


def current_discount(price_mode: str, custom_discount: float) -> float:
    if price_mode == "-12%":
        return DEFAULT_DISCOUNT_1
    if price_mode == "-20%":
        return DEFAULT_DISCOUNT_2
    return max(0.0, float(custom_discount))


def current_price_label(price_mode: str, custom_discount: float) -> str:
    disc = current_discount(price_mode, custom_discount)
    if float(disc).is_integer():
        return f"Цена -{int(disc)}%"
    return f"Цена -{str(round(disc, 2)).replace('.', ',')}%"


def get_selected_price_raw(row: pd.Series, price_mode: str, round100: bool, custom_discount: float) -> float:
    disc = current_discount(price_mode, custom_discount)
    value = safe_float(row.get("sale_price", 0.0), 0.0) * (1 - disc / 100)
    return float(round_up_to_100(value)) if round100 else float(round(value, 2))


def tokenize_text(value: object) -> list[str]:
    text = normalize_text(value)
    if not text:
        return []
    return [t for t in re.split(r"[^A-Za-zА-Яа-я0-9]+", text.upper()) if t]


ARTICLE_PIECE_RE = re.compile(r"[A-Za-zА-Яа-я0-9._/-]{3,}")


def is_candidate_article_norm(norm: str) -> bool:
    if not norm:
        return False
    if len(norm) < 5:
        return False
    has_digit = any(ch.isdigit() for ch in norm)
    has_alpha = any(ch.isalpha() for ch in norm)
    return has_digit and has_alpha


def extract_article_candidates_from_text(text: object) -> list[str]:
    raw = normalize_text(text)
    if not raw:
        return []
    chunks = ARTICLE_PIECE_RE.findall(raw)
    out: list[str] = []
    seen: set[str] = set()
    for chunk in chunks:
        norm = normalize_article(chunk)
        if not is_candidate_article_norm(norm) or norm in seen:
            continue
        seen.add(norm)
        out.append(norm)
    return out


def unique_norm_codes(items: list[object]) -> list[str]:
    seen: set[str] = set()
    out: list[str] = []
    for item in items:
        norm = normalize_article(item)
        if not is_candidate_article_norm(norm) or norm in seen:
            continue
        seen.add(norm)
        out.append(norm)
    return out


def build_row_compare_codes(article: object, name: object) -> list[str]:
    return unique_norm_codes([article, *extract_article_candidates_from_text(name)])


COMPATIBLE_BRAND_MARKERS: list[tuple[str, str]] = [
    ("NV PRINT", "NV Print"),
    ("STATIC CONTROL", "Static Control"),
    ("HI-BLACK", "Hi-Black"),
    ("NETPRODUCT", "NetProduct"),
    ("PROFILINE", "ProfiLine"),
    ("G&G", "G&G"),
    ("MYTONER", "MyToner"),
    ("MYTONE", "MyTone"),
    ("BLOSSOM", "Blossom"),
    ("CACTUS", "Cactus"),
    ("SAKURA", "Sakura"),
    ("KATUN", "Katun"),
    ("UNITON", "Uniton"),
    ("COPYRITE", "Copyrite"),
    ("COLORING", "Coloring"),
    ("BULAT", "Bulat"),
    ("CET", "CET"),
]

COMPATIBLE_ARTICLE_PREFIX_BRANDS = {
    "BS": "Blossom",
    "CS": "Cactus",
    "BSL": "Bulat",
    "CET": "CET",
    "NV": "NV Print",
    "SC": "Static Control",
    "STA": "Static Control",
    "KAT": "Katun",
    "KTN": "Katun",
    "SKR": "Sakura",
}


def extract_compatible_brand(name: object, article: object = "") -> str:
    text = contains_text(name)
    for needle, label in COMPATIBLE_BRAND_MARKERS:
        if needle in text:
            return label
    article_text = normalize_text(article)
    if article_text:
        prefix = article_text.split('-', 1)[0].strip().upper()
        if prefix in COMPATIBLE_ARTICLE_PREFIX_BRANDS:
            return COMPATIBLE_ARTICLE_PREFIX_BRANDS[prefix]
    return ""


SERIES_SUFFIX_ORDER = {"A": 0, "AC": 1, "X": 2, "XC": 3, "Y": 4, "YC": 5, "M": 6, "MC": 7, "C": 8, "K": 9}


def split_article_family_suffix(article_norm: str) -> tuple[str, str]:
    m = re.match(r"^(.*?\d)([A-ZА-Я]{1,3})$", article_norm)
    if m:
        return m.group(1), m.group(2)
    return article_norm, ""


def natural_chunks(value: str) -> list[object]:
    parts = re.split(r"(\d+)", value)
    result: list[object] = []
    for part in parts:
        if not part:
            continue
        result.append(int(part) if part.isdigit() else part)
    return result


def series_sort_key(candidate: dict[str, object]) -> tuple[object, ...]:
    article_norm = str(candidate.get("article_norm", ""))
    family, suffix = split_article_family_suffix(article_norm)
    rank = SERIES_SUFFIX_ORDER.get(suffix, 50)
    return (*natural_chunks(family), rank, suffix, article_norm)


def get_series_candidates(df: pd.DataFrame, raw_query: str) -> dict[str, object]:
    tokens = split_query_parts(raw_query)
    if len(tokens) != 1:
        return {"prefix": "", "candidates": []}
    token = tokens[0]
    token_norm = normalize_article(token)
    if len(token_norm) < 4:
        return {"prefix": token, "candidates": []}

    candidates_by_key: dict[str, dict[str, object]] = {}

    direct_df = df[df["article_norm"].str.startswith(token_norm, na=False)].copy()
    linked_mask = df["row_codes"].apply(lambda codes: any(str(code).startswith(token_norm) for code in (codes or [])) if isinstance(codes, list) else False)
    linked_df = df[linked_mask].copy()

    for source_df in [direct_df, linked_df]:
        for _, row in source_df.iterrows():
            candidate = {
                "article": str(row.get("article", "")),
                "article_norm": str(row.get("article_norm", "")),
                "name": str(row.get("name", "")),
                "free_qty": safe_float(row.get("free_qty", 0), 0.0),
                "sale_price": safe_float(row.get("sale_price", 0), 0.0),
                "original_block_reasons": list(row.get("original_block_reasons", []) or []),
            }
            if candidate["article_norm"] and candidate["article_norm"] not in candidates_by_key:
                candidates_by_key[candidate["article_norm"]] = candidate

    candidates = list(candidates_by_key.values())
    candidates.sort(key=series_sort_key)
    if len(candidates) < 2:
        return {"prefix": token, "candidates": []}
    return {"prefix": token, "candidates": candidates}


def build_sheet_code_reason_lookup(df: pd.DataFrame | None, reason: str) -> dict[str, set[str]]:
    lookup: dict[str, set[str]] = {}
    if df is None or df.empty:
        return lookup
    for _, row in df.iterrows():
        codes = row.get("row_codes")
        if not isinstance(codes, list) or not codes:
            codes = build_row_compare_codes(row.get("article", ""), row.get("name", ""))
        for code in codes:
            norm = normalize_article(code)
            if not norm:
                continue
            lookup.setdefault(norm, set()).add(reason)
    return lookup


def merge_code_reason_lookups(*lookups: dict[str, set[str]]) -> dict[str, list[str]]:
    merged: dict[str, set[str]] = {}
    for lookup in lookups:
        for code, reasons in (lookup or {}).items():
            if not code or not reasons:
                continue
            merged.setdefault(code, set()).update(set(reasons))
    return {code: sorted(reasons) for code, reasons in merged.items() if reasons}


def get_original_block_reasons(codes: list[str], block_lookup: dict[str, list[str]]) -> list[str]:
    # Ограничения по Уценке/Совместимым отключены: comparison-файл считается уже поправленным.
    return []

def original_reason_badge_text(reasons: list[str]) -> str:
    if not isinstance(reasons, list) or not reasons:
        return ""
    order = []
    if "Уценка" in reasons:
        order.append("🟠 Уценка")
    if "Совместимые" in reasons:
        order.append("🟣 Совместимые")
    other = [r for r in reasons if r not in {"Уценка", "Совместимые"}]
    order.extend([f"⚪ {r}" for r in other])
    return " · ".join(order)


def original_reason_short_tag(reasons: list[str]) -> str:
    if not isinstance(reasons, list) or not reasons:
        return ""
    has_discount = "Уценка" in reasons
    has_compatible = "Совместимые" in reasons
    if has_discount and has_compatible:
        return "[скрыт: Уценка + Совместимые]"
    if has_discount:
        return "[скрыт: Уценка]"
    if has_compatible:
        return "[скрыт: Совместимые]"
    return "[скрыт]"


def original_reason_summary_html(hidden_reasons: dict[str, list[str]]) -> str:
    if not hidden_reasons:
        return ""
    only_discount = 0
    only_compatible = 0
    both = 0
    for reasons in hidden_reasons.values():
        has_discount = "Уценка" in reasons
        has_compatible = "Совместимые" in reasons
        if has_discount and has_compatible:
            both += 1
        elif has_discount:
            only_discount += 1
        elif has_compatible:
            only_compatible += 1
    chips = []
    if only_discount:
        chips.append(f"<span class='series-reason-chip chip-discount'>🟠 только Уценка: {only_discount}</span>")
    if only_compatible:
        chips.append(f"<span class='series-reason-chip chip-compatible'>🟣 только Совместимые: {only_compatible}</span>")
    if both:
        chips.append(f"<span class='series-reason-chip chip-both'>🔒 обе причины: {both}</span>")
    if not chips:
        return ""
    return "<div class='series-reason-row'>" + "".join(chips) + "</div>"


def build_compatible_price_lookup(compatible_df: pd.DataFrame | None) -> dict[str, dict[str, set[float]]]:
    lookup: dict[str, dict[str, set[float]]] = {}
    if compatible_df is None or compatible_df.empty:
        return lookup
    for _, row in compatible_df.iterrows():
        codes = build_row_compare_codes(row.get("article", ""), row.get("name", ""))
        if not codes:
            continue
        for pair in row.get("source_pairs", []) or []:
            source = str(pair.get("source", "") or "")
            price = safe_float(row.get(pair.get("price_col", "")), 0.0)
            price = normalize_merlion_source_price(row, source, price)
            qty = parse_qty_generic(row.get(pair.get("qty_col", "")))
            if not source or price <= 0 or qty <= 0:
                continue
            price_key = round(float(price), 2)
            for code in codes:
                lookup.setdefault(code, {}).setdefault(source, set()).add(price_key)
    return lookup


def merge_source_price_lookups(*lookups: dict[str, dict[str, set[float]]]) -> dict[str, dict[str, set[float]]]:
    merged: dict[str, dict[str, set[float]]] = {}
    for lookup in lookups:
        for code, source_map in (lookup or {}).items():
            code_norm = normalize_article(code)
            if not code_norm or not isinstance(source_map, dict):
                continue
            bucket = merged.setdefault(code_norm, {})
            for source, prices in source_map.items():
                source_name = str(source or "")
                if not source_name:
                    continue
                for price in prices or []:
                    price_val = safe_float(price, 0.0)
                    if price_val > 0:
                        bucket.setdefault(source_name, set()).add(round(float(price_val), 2))
    return merged


def merge_blocked_source_prices(codes: list[str], compatible_lookup: dict[str, dict[str, set[float]]]) -> dict[str, list[float]]:
    out: dict[str, set[float]] = {}
    for code in codes or []:
        for source, prices in compatible_lookup.get(code, {}).items():
            out.setdefault(source, set()).update(prices)
    return {source: sorted(values) for source, values in out.items() if values}


def is_blocked_by_compatible_price(row: pd.Series, source: str, price: float) -> bool:
    # Ограничения по совпадениям с Уценкой/Совместимыми сняты.
    return False


def filter_suspicious_low_offers(row: pd.Series, offers: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[str]]:
    # Дополнительный outlier-фильтр отключён: после правки comparison-файла показываем все цены как есть.
    return offers, []


def unique_preserve_order(items: list[str]) -> list[str]:
    seen: set[str] = set()
    out: list[str] = []
    for item in items:
        norm = normalize_text(item)
        if not norm or norm in seen:
            continue
        seen.add(norm)
        out.append(item)
    return out


def split_query_parts(query: str) -> list[str]:
    parts: list[str] = []
    raw_chunks = re.split(r"[\n,;]+", query)
    for chunk in raw_chunks:
        chunk = normalize_text(chunk)
        if not chunk:
            continue
        if "/" in chunk:
            slash_parts = [normalize_text(x) for x in re.split(r"\s*/\s*", chunk) if normalize_text(x)]
            if len(slash_parts) > 1:
                parts.extend(slash_parts)
                continue
        space_parts = [normalize_text(x) for x in re.split(r"\s+", chunk) if normalize_text(x)]
        if len(space_parts) > 1 and all(len(normalize_article(x)) >= 3 for x in space_parts):
            parts.extend(space_parts)
        else:
            parts.append(chunk)
    return parts


def normalize_query_for_display(query: str) -> str:
    return "\n".join(split_query_parts(query))


def find_column(columns: list[str], aliases: list[str]) -> Optional[str]:
    lower_map = {str(col).strip().lower(): col for col in columns}
    for alias in aliases:
        hit = lower_map.get(alias.strip().lower())
        if hit:
            return hit
    for alias in aliases:
        a = alias.strip().lower()
        for col in columns:
            c = str(col).strip().lower()
            if a in c or c in a:
                return col
    return None


def detect_mapping(df: pd.DataFrame, aliases_map: dict[str, list[str]]) -> dict[str, Optional[str]]:
    return {key: find_column(list(df.columns), aliases) for key, aliases in aliases_map.items()}


def parse_qty_generic(value: Any) -> float:
    raw = normalize_text(value)
    compact = compact_text(value)
    if not raw:
        return 0.0
    try:
        return max(0.0, float(raw.replace(" ", "").replace(",", ".")))
    except Exception:
        pass

    mapping = {
        "+": 1.0,
        "++": 5.0,
        "+++": 10.0,
        "МАЛО": 1.0,
        "ЕСТЬ": 1.0,
        "СРЕДНЕ": 5.0,
        "СРЕДНЕЕ": 5.0,
        "СРЕДНИЙ": 5.0,
        "СРЕДНЯЯ": 5.0,
        "МНОГО": 10.0,
        "CALL": 0.0,
        "НЕТ": 0.0,
        "ПОДЗАКАЗ": 0.0,
        "ПОДЗАКАЗ": 0.0,
        "ЗАКАЗ": 0.0,
        "ОЖИДАЕТСЯ": 0.0,
    }
    for key, val in mapping.items():
        if key in compact:
            return val
    m = re.search(r"(\d+[\.,]?\d*)", raw)
    if m:
        try:
            return max(0.0, float(m.group(1).replace(",", ".")))
        except Exception:
            return 0.0
    return 0.0


def parse_excel_hyperlink_formula(value: object) -> tuple[str, str]:
    text = str(value or "").strip()
    if not text.startswith("="):
        return "", ""
    m = re.match(r'^=\s*(?:HYPERLINK|ГИПЕРССЫЛКА)\(\s*"([^"]+)"\s*[;,]\s*"([^"]*)"\s*\)$', text, flags=re.IGNORECASE)
    if not m:
        return "", ""
    return m.group(1).strip(), m.group(2).strip()


def cell_display_and_url(cell) -> tuple[str, str]:
    url = ""
    display = ""
    if cell is None:
        return display, url
    try:
        if getattr(cell, "hyperlink", None):
            url = str(cell.hyperlink.target or "").strip()
    except Exception:
        pass
    formula_url, formula_display = parse_excel_hyperlink_formula(cell.value)
    if formula_url:
        url = formula_url
        display = formula_display
    else:
        display = normalize_text(cell.value)
    return display, url


@st.cache_data(show_spinner=False, ttl=3600, max_entries=6)
def load_comparison_workbook(file_name: str, file_bytes: bytes) -> dict[str, pd.DataFrame]:
    wb = pd.ExcelFile(io.BytesIO(file_bytes))
    sheets: dict[str, pd.DataFrame] = {}
    for sheet in wb.sheet_names:
        raw = pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet)
        raw = raw.dropna(how="all")
        mapping = detect_mapping(raw, CATALOG_COLUMN_ALIASES)
        required = ["article", "name", "price", "qty"]
        missing = [k for k in required if not mapping.get(k)]
        if missing:
            continue

        df = raw.copy()
        df["article"] = df[mapping["article"]].map(normalize_text)
        df["article_norm"] = df[mapping["article"]].map(normalize_article)
        df["name"] = df[mapping["name"]].map(normalize_text)
        df["sale_price"] = df[mapping["price"]].apply(safe_float)
        df["free_qty"] = df[mapping["qty"]].apply(parse_qty_generic)
        total_col = mapping.get("total_qty")
        transit_col = mapping.get("transit_qty")
        df["total_qty"] = df[total_col].apply(parse_qty_generic) if total_col else df["free_qty"]
        df["transit_qty"] = df[transit_col].apply(parse_qty_generic) if transit_col else 0.0
        df["has_extended_stock"] = bool(total_col or transit_col)
        df["search_blob"] = (df["article"] + " " + df["name"]).map(contains_text)
        df["search_blob_compact"] = (df["article"] + " " + df["name"]).map(compact_text)
        df["name_tokens"] = df["name"].map(tokenize_text)
        df["sheet_name"] = sheet

        columns = list(raw.columns)
        source_pairs: list[dict[str, str]] = []
        seen_sources: set[str] = set()
        for col in columns:
            col_txt = normalize_text(col)
            m = re.match(r"^(.*?)\s+цена$", col_txt, flags=re.IGNORECASE)
            if not m:
                continue
            source = normalize_text(m.group(1))
            if not source or compact_text(source) in {"НАША"}:
                continue
            qty_col = None
            for candidate in columns:
                candidate_txt = normalize_text(candidate)
                if compact_text(candidate_txt) == compact_text(f"{source} шт"):
                    qty_col = candidate
                    break
            if not qty_col or source in seen_sources:
                continue
            seen_sources.add(source)
            source_pairs.append({"source": source, "price_col": col, "qty_col": qty_col})

        df["source_pairs"] = [source_pairs for _ in range(len(df))]
        df["photo_url"] = ""
        df["photo_name"] = ""
        df["row_codes"] = df.apply(lambda row: build_row_compare_codes(row.get("article", ""), row.get("name", "")), axis=1)
        df["blocked_source_prices"] = [{} for _ in range(len(df))]
        df = df[(df["article_norm"] != "") & (df["name"] != "")].copy()
        df = df.reset_index(drop=True)
        sheets[sheet] = df
    if not sheets:
        raise ValueError("Не удалось прочитать comparison-файл: на листах не найдены обязательные колонки Артикул / Наименование / Наша цена / Наш склад.")

    compatible_df = sheets.get("Совместимые")
    discount_df = sheets.get("Уценка")
    compatible_lookup = build_compatible_price_lookup(compatible_df)
    discount_lookup = build_compatible_price_lookup(discount_df)
    blocked_price_lookup = merge_source_price_lookups(compatible_lookup, discount_lookup)
    original_df = sheets.get("Сравнение")
    original_block_lookup = merge_code_reason_lookups(
        build_sheet_code_reason_lookup(discount_df, "Уценка"),
        build_sheet_code_reason_lookup(compatible_df, "Совместимые"),
    )
    if isinstance(original_df, pd.DataFrame) and not original_df.empty:
        original_df = original_df.copy()
        if blocked_price_lookup:
            original_df["blocked_source_prices"] = original_df["row_codes"].apply(lambda codes: merge_blocked_source_prices(codes, blocked_price_lookup))
        original_df["original_block_reasons"] = original_df["row_codes"].apply(lambda codes: get_original_block_reasons(codes, original_block_lookup))
        original_df["blocked_in_original"] = original_df["original_block_reasons"].map(lambda x: isinstance(x, list) and len(x) > 0)
        sheets["Сравнение"] = original_df
    return sheets


@st.cache_data(show_spinner=False, ttl=3600, max_entries=6)
def load_photo_map_file(file_name: str, file_bytes: bytes) -> pd.DataFrame:
    suffix = Path(file_name).suffix.lower()

    def _sheet_priority(sheet_name: str) -> int:
        name = contains_text(sheet_name)
        if "ФОТО" in name or "СЫЛ" in name:
            return 0
        if "WORKSHEET" in name:
            return 20
        return 10

    def _empty_df() -> pd.DataFrame:
        return pd.DataFrame(columns=[
            "article", "article_norm", "photo_url", "source_sheet", "sheet_priority",
            "meta_brand", "meta_color", "meta_capacity", "meta_manufacturer_code", "meta_model",
            "meta_description", "meta_fits_models", "meta_iso_pages", "meta_print_technology",
            "meta_item_type", "meta_print_type", "meta_weight", "meta_length", "meta_width", "meta_height",
        ])

    def _from_raw(raw: pd.DataFrame, sheet_name: str = "") -> pd.DataFrame:
        raw = raw.dropna(how="all")
        if raw.empty:
            return _empty_df()
        raw = raw.copy()
        raw.columns = [normalize_text(c) for c in raw.columns]
        mapping = detect_mapping(raw, PHOTO_COLUMN_ALIASES)

        if not mapping.get("article"):
            for col in raw.columns:
                if compact_text(col) == "АРТИКУЛ":
                    mapping["article"] = col
                    break

        if "images" in raw.columns and raw["images"].map(lambda x: bool(extract_first_url(x))).sum() > 0:
            mapping["photo_url"] = "images"
        elif not mapping.get("photo_url") and "imag" in raw.columns and raw["imag"].map(lambda x: bool(extract_first_url(x))).sum() > 0:
            mapping["photo_url"] = "imag"

        if not mapping.get("photo_url") and len(raw.columns) >= 2:
            first_col = mapping.get("article") or raw.columns[0]
            best_col = None
            best_hits = 0
            for col in raw.columns:
                if col == first_col:
                    continue
                hits = raw[col].map(lambda x: bool(extract_first_url(x))).sum()
                if hits > best_hits:
                    best_hits = hits
                    best_col = col
            if best_col is not None and best_hits > 0:
                mapping["photo_url"] = best_col

        if not mapping.get("article"):
            return _empty_df()

        out = pd.DataFrame()
        out["article"] = raw[mapping["article"]].map(normalize_text)
        out["article_norm"] = raw[mapping["article"]].map(normalize_article)
        out["photo_url"] = raw[mapping["photo_url"]].map(extract_first_url) if mapping.get("photo_url") else ""
        out["source_sheet"] = sheet_name
        out["sheet_priority"] = _sheet_priority(sheet_name)
        out["meta_brand"] = raw[mapping["brand"]].map(normalize_text) if mapping.get("brand") else ""
        out["meta_color"] = raw[mapping["color"]].map(normalize_text) if mapping.get("color") else ""
        out["meta_capacity"] = raw[mapping["capacity"]].map(normalize_text) if mapping.get("capacity") else ""
        out["meta_manufacturer_code"] = raw[mapping["manufacturer_code"]].map(normalize_text) if mapping.get("manufacturer_code") else ""
        out["meta_model"] = raw[mapping["model"]].map(normalize_text) if mapping.get("model") else ""
        out["meta_description"] = raw[mapping["description"]].map(normalize_text) if mapping.get("description") else ""
        out["meta_fits_models"] = raw[mapping["fits_models"]].map(normalize_text) if mapping.get("fits_models") else ""
        out["meta_iso_pages"] = raw[mapping["iso_pages"]].map(normalize_text) if mapping.get("iso_pages") else ""
        out["meta_print_technology"] = raw[mapping["print_technology"]].map(normalize_text) if mapping.get("print_technology") else ""
        out["meta_item_type"] = raw[mapping["item_type"]].map(normalize_text) if mapping.get("item_type") else ""
        out["meta_print_type"] = raw[mapping["print_type"]].map(normalize_text) if mapping.get("print_type") else ""
        out["meta_weight"] = raw[mapping["weight"]].map(normalize_text) if mapping.get("weight") else ""
        out["meta_length"] = raw[mapping["length"]].map(normalize_text) if mapping.get("length") else ""
        out["meta_width"] = raw[mapping["width"]].map(normalize_text) if mapping.get("width") else ""
        out["meta_height"] = raw[mapping["height"]].map(normalize_text) if mapping.get("height") else ""
        out = out[out["article_norm"] != ""].reset_index(drop=True)
        return out if not out.empty else _empty_df()

    if suffix == ".csv":
        bio = io.BytesIO(file_bytes)
        try:
            raw = pd.read_csv(bio)
        except UnicodeDecodeError:
            bio.seek(0)
            raw = pd.read_csv(bio, encoding="cp1251")
        out = _from_raw(raw, "CSV")
        if out.empty:
            raise ValueError("В файле фото нужны колонки с артикулом и хотя бы с фото или полезными полями.")
        out = out.sort_values(["sheet_priority", "article_norm"]).drop_duplicates(subset=["article_norm"], keep="first").reset_index(drop=True)
        return out[[
            "article", "article_norm", "photo_url", "source_sheet",
            "meta_brand", "meta_color", "meta_capacity", "meta_manufacturer_code", "meta_model",
            "meta_description", "meta_fits_models", "meta_iso_pages", "meta_print_technology",
            "meta_item_type", "meta_print_type", "meta_weight", "meta_length", "meta_width", "meta_height",
        ]]

    sheets = pd.read_excel(io.BytesIO(file_bytes), sheet_name=None)
    parts: list[pd.DataFrame] = []
    for sheet_name, raw in sheets.items():
        part = _from_raw(raw, sheet_name)
        if not part.empty:
            parts.append(part)

    if not parts:
        raise ValueError("В файле фото нужны колонки с артикулом и хотя бы с фото или полезными полями из Worksheet.")

    combined = pd.concat(parts, ignore_index=True)
    combined = combined.sort_values(["article_norm", "sheet_priority"]).reset_index(drop=True)

    def _first_non_empty(series: pd.Series) -> str:
        for value in series.tolist():
            txt = normalize_text(value)
            if txt:
                return txt
        return ""

    def _best_photo(series: pd.Series) -> str:
        for value in series.tolist():
            txt = normalize_text(value)
            if txt:
                return txt
        return ""

    rows: list[dict[str, str]] = []
    for article_norm, grp in combined.groupby("article_norm", sort=False):
        grp = grp.sort_values(["sheet_priority", "source_sheet"])
        row = {
            "article": _first_non_empty(grp["article"]),
            "article_norm": article_norm,
            "photo_url": _best_photo(grp["photo_url"]),
            "source_sheet": _first_non_empty(grp["source_sheet"]),
            "meta_brand": _first_non_empty(grp["meta_brand"]),
            "meta_color": _first_non_empty(grp["meta_color"]),
            "meta_capacity": _first_non_empty(grp["meta_capacity"]),
            "meta_manufacturer_code": _first_non_empty(grp["meta_manufacturer_code"]),
            "meta_model": _first_non_empty(grp["meta_model"]),
            "meta_description": _first_non_empty(grp["meta_description"]),
            "meta_fits_models": _first_non_empty(grp["meta_fits_models"]),
            "meta_iso_pages": _first_non_empty(grp["meta_iso_pages"]),
            "meta_print_technology": _first_non_empty(grp["meta_print_technology"]),
            "meta_item_type": _first_non_empty(grp["meta_item_type"]),
            "meta_print_type": _first_non_empty(grp["meta_print_type"]),
            "meta_weight": _first_non_empty(grp["meta_weight"]),
            "meta_length": _first_non_empty(grp["meta_length"]),
            "meta_width": _first_non_empty(grp["meta_width"]),
            "meta_height": _first_non_empty(grp["meta_height"]),
        }
        rows.append(row)

    combined = pd.DataFrame(rows)
    return combined[[
        "article", "article_norm", "photo_url", "source_sheet",
        "meta_brand", "meta_color", "meta_capacity", "meta_manufacturer_code", "meta_model",
        "meta_description", "meta_fits_models", "meta_iso_pages", "meta_print_technology",
        "meta_item_type", "meta_print_type", "meta_weight", "meta_length", "meta_width", "meta_height",
    ]]


def apply_photo_map(df: pd.DataFrame | None, photo_df: pd.DataFrame | None) -> pd.DataFrame | None:
    if df is None:
        return None
    out = df.copy()
    for col in [
        "photo_url", "photo_name", "meta_brand", "meta_color", "meta_capacity",
        "meta_manufacturer_code", "meta_model", "meta_description", "meta_fits_models",
        "meta_iso_pages", "meta_print_technology", "meta_item_type", "meta_print_type",
        "meta_weight", "meta_length", "meta_width", "meta_height",
    ]:
        if col not in out.columns:
            out[col] = ""
    if photo_df is None or photo_df.empty:
        out["photo_name"] = out.get("name", "")
        return out
    lookup = photo_df.set_index("article_norm").to_dict(orient="index")
    def _meta(norm: str, key: str) -> str:
        row = lookup.get(norm, {})
        return normalize_text(row.get(key, ""))
    out["photo_url"] = out["article_norm"].map(lambda x: _meta(x, "photo_url"))
    out["photo_name"] = out["name"]
    out["meta_brand"] = out["article_norm"].map(lambda x: _meta(x, "meta_brand"))
    out["meta_color"] = out["article_norm"].map(lambda x: _meta(x, "meta_color"))
    out["meta_capacity"] = out["article_norm"].map(lambda x: _meta(x, "meta_capacity"))
    out["meta_manufacturer_code"] = out["article_norm"].map(lambda x: _meta(x, "meta_manufacturer_code"))
    out["meta_model"] = out["article_norm"].map(lambda x: _meta(x, "meta_model"))
    out["meta_description"] = out["article_norm"].map(lambda x: _meta(x, "meta_description"))
    out["meta_fits_models"] = out["article_norm"].map(lambda x: _meta(x, "meta_fits_models"))
    out["meta_iso_pages"] = out["article_norm"].map(lambda x: _meta(x, "meta_iso_pages"))
    out["meta_print_technology"] = out["article_norm"].map(lambda x: _meta(x, "meta_print_technology"))
    out["meta_item_type"] = out["article_norm"].map(lambda x: _meta(x, "meta_item_type"))
    out["meta_print_type"] = out["article_norm"].map(lambda x: _meta(x, "meta_print_type"))
    out["meta_weight"] = out["article_norm"].map(lambda x: _meta(x, "meta_weight"))
    out["meta_length"] = out["article_norm"].map(lambda x: _meta(x, "meta_length"))
    out["meta_width"] = out["article_norm"].map(lambda x: _meta(x, "meta_width"))
    out["meta_height"] = out["article_norm"].map(lambda x: _meta(x, "meta_height"))
    return out


@st.cache_data(show_spinner=False, ttl=3600, max_entries=6)
def load_avito_file(file_name: str, file_bytes: bytes) -> pd.DataFrame:
    suffix = Path(file_name).suffix.lower()
    if suffix == ".csv":
        bio = io.BytesIO(file_bytes)
        try:
            raw = pd.read_csv(bio)
        except UnicodeDecodeError:
            bio.seek(0)
            raw = pd.read_csv(bio, encoding="cp1251")
        mapping = detect_mapping(raw, AVITO_COLUMN_ALIASES)
        if not mapping.get("title"):
            raise ValueError("Не удалось определить колонку 'Название объявления' в файле Авито.")
        rows = []
        for _, r in raw.iterrows():
            rows.append({
                "ad_id": normalize_text(r[mapping["ad_id"]]) if mapping.get("ad_id") else "",
                "title": normalize_text(r[mapping["title"]]) if mapping.get("title") else "",
                "price": normalize_text(r[mapping["price"]]) if mapping.get("price") else "",
                "url": normalize_text(r[mapping["url"]]) if mapping.get("url") else "",
                "account": normalize_text(r[mapping["account"]]) if mapping.get("account") else "",
            })
        out = pd.DataFrame(rows)
        out["title_norm"] = out["title"].map(contains_text)
        out["title_codes"] = out["title"].map(extract_article_candidates_from_text)
        out["registry_key"] = out.apply(build_avito_registry_key, axis=1)
        return out

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=False)
    ws = wb.active
    headers = [normalize_text(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]

    def find_header_index(candidates: list[str]) -> Optional[int]:
        for idx, header in enumerate(headers, start=1):
            for cand in candidates:
                if header.lower() == cand.lower():
                    return idx
        for idx, header in enumerate(headers, start=1):
            h = header.lower()
            for cand in candidates:
                c = cand.lower()
                if c in h or h in c:
                    return idx
        return None

    ad_id_col = find_header_index(AVITO_COLUMN_ALIASES["ad_id"])
    title_col = find_header_index(AVITO_COLUMN_ALIASES["title"])
    price_col = find_header_index(AVITO_COLUMN_ALIASES["price"])
    url_col = find_header_index(AVITO_COLUMN_ALIASES["url"])
    account_col = find_header_index(AVITO_COLUMN_ALIASES["account"])
    if not title_col:
        raise ValueError("Не удалось определить колонку 'Название объявления' в файле Авито.")

    rows = []
    for r in range(2, ws.max_row + 1):
        ad_display, ad_url = cell_display_and_url(ws.cell(r, ad_id_col)) if ad_id_col else ("", "")
        title_display, title_url = cell_display_and_url(ws.cell(r, title_col))
        explicit_url = normalize_text(ws.cell(r, url_col).value) if url_col else ""
        price_value = normalize_text(ws.cell(r, price_col).value) if price_col else ""
        account_value = normalize_text(ws.cell(r, account_col).value) if account_col else ""
        final_url = explicit_url or title_url or ad_url
        if not ad_display and not title_display:
            continue
        rows.append({
            "ad_id": ad_display,
            "title": title_display,
            "price": price_value,
            "url": final_url,
            "account": account_value,
        })
    out = pd.DataFrame(rows)
    out["title_norm"] = out["title"].map(contains_text)
    out["title_codes"] = out["title"].map(extract_article_candidates_from_text)
    out["registry_key"] = out.apply(build_avito_registry_key, axis=1)
    return out


def get_avito_registry_path() -> Path:
    try:
        return Path(__file__).resolve().with_name("avito_registry.sqlite")
    except Exception:
        return Path.cwd() / "avito_registry.sqlite"


def avito_now_str() -> str:
    return datetime.now().replace(microsecond=0).isoformat(sep=" ")


def build_avito_registry_key(row: pd.Series | dict[str, Any]) -> str:
    ad_id = normalize_text(row.get("ad_id", ""))
    if ad_id:
        return f"ad:{ad_id}"
    seed = "|".join([
        normalize_text(row.get("title", "")),
        normalize_text(row.get("url", "")),
    ])
    return "hash:" + hashlib.md5(seed.encode("utf-8", errors="ignore")).hexdigest()


def ensure_avito_registry() -> None:
    path = get_avito_registry_path()
    conn = sqlite3.connect(path)
    try:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS avito_registry (
                registry_key TEXT PRIMARY KEY,
                ad_id TEXT,
                title TEXT,
                title_norm TEXT,
                price_raw TEXT,
                url TEXT,
                account TEXT,
                first_seen TEXT,
                last_seen TEXT,
                last_changed_at TEXT,
                previous_price_raw TEXT,
                change_count INTEGER DEFAULT 0,
                status TEXT,
                last_import_name TEXT
            )
            """
        )
        conn.commit()
    finally:
        conn.close()


def sync_avito_registry(avito_df: pd.DataFrame, import_name: str) -> dict[str, Any]:
    ensure_avito_registry()
    path = get_avito_registry_path()
    now = avito_now_str()
    stats = {"new": 0, "changed": 0, "unchanged": 0, "missing": 0, "total": 0}
    if avito_df is None or avito_df.empty:
        return stats

    work = avito_df.copy()
    work["registry_key"] = work.apply(build_avito_registry_key, axis=1)
    work = work.drop_duplicates(subset=["registry_key"], keep="first").reset_index(drop=True)
    stats["total"] = len(work)

    conn = sqlite3.connect(path)
    conn.row_factory = sqlite3.Row
    try:
        current_keys = work["registry_key"].tolist()
        placeholders = ",".join(["?"] * len(current_keys)) if current_keys else "''"
        existing = {}
        if current_keys:
            for row in conn.execute(f"SELECT * FROM avito_registry WHERE registry_key IN ({placeholders})", current_keys):
                existing[row["registry_key"]] = dict(row)

        for _, row in work.iterrows():
            key = row["registry_key"]
            payload = {
                "registry_key": key,
                "ad_id": normalize_text(row.get("ad_id", "")),
                "title": normalize_text(row.get("title", "")),
                "title_norm": contains_text(row.get("title", "")),
                "price_raw": normalize_text(row.get("price", "")),
                "url": normalize_text(row.get("url", "")),
                "account": normalize_text(row.get("account", "")),
                "last_import_name": normalize_text(import_name),
            }
            old = existing.get(key)
            if old is None:
                conn.execute(
                    """
                    INSERT INTO avito_registry
                    (registry_key, ad_id, title, title_norm, price_raw, url, account, first_seen, last_seen, last_changed_at, previous_price_raw, change_count, status, last_import_name)
                    VALUES (:registry_key, :ad_id, :title, :title_norm, :price_raw, :url, :account, :first_seen, :last_seen, :last_changed_at, :previous_price_raw, :change_count, :status, :last_import_name)
                    """,
                    {
                        **payload,
                        "first_seen": now,
                        "last_seen": now,
                        "last_changed_at": now,
                        "previous_price_raw": "",
                        "change_count": 0,
                        "status": "active",
                    },
                )
                stats["new"] += 1
            else:
                changed = any([
                    payload["title"] != normalize_text(old.get("title", "")),
                    payload["price_raw"] != normalize_text(old.get("price_raw", "")),
                    payload["url"] != normalize_text(old.get("url", "")),
                    payload["account"] != normalize_text(old.get("account", "")),
                ])
                if changed:
                    conn.execute(
                        """
                        UPDATE avito_registry SET
                            ad_id=:ad_id,
                            title=:title,
                            title_norm=:title_norm,
                            previous_price_raw=:previous_price_raw,
                            price_raw=:price_raw,
                            url=:url,
                            account=:account,
                            last_seen=:last_seen,
                            last_changed_at=:last_changed_at,
                            change_count=:change_count,
                            status='active',
                            last_import_name=:last_import_name
                        WHERE registry_key=:registry_key
                        """,
                        {
                            **payload,
                            "previous_price_raw": normalize_text(old.get("price_raw", "")),
                            "last_seen": now,
                            "last_changed_at": now,
                            "change_count": int(old.get("change_count", 0) or 0) + 1,
                        },
                    )
                    stats["changed"] += 1
                else:
                    conn.execute(
                        """
                        UPDATE avito_registry SET
                            ad_id=:ad_id,
                            title=:title,
                            title_norm=:title_norm,
                            price_raw=:price_raw,
                            url=:url,
                            account=:account,
                            last_seen=:last_seen,
                            status='active',
                            last_import_name=:last_import_name
                        WHERE registry_key=:registry_key
                        """,
                        {**payload, "last_seen": now},
                    )
                    stats["unchanged"] += 1

        if current_keys:
            placeholders = ",".join(["?"] * len(current_keys))
            cur = conn.execute(
                f"UPDATE avito_registry SET status='missing_in_latest_export' WHERE registry_key NOT IN ({placeholders}) AND status='active'",
                current_keys,
            )
            stats["missing"] = cur.rowcount if cur.rowcount is not None else 0
        conn.commit()
    finally:
        conn.close()
    return stats


def load_avito_registry_df() -> pd.DataFrame:
    path = get_avito_registry_path()
    if not path.exists():
        return pd.DataFrame()
    conn = sqlite3.connect(path)
    try:
        df = pd.read_sql_query("SELECT * FROM avito_registry", conn)
    except Exception:
        conn.close()
        return pd.DataFrame()
    finally:
        conn.close()
    if df.empty:
        return df
    for col in ["ad_id", "title", "price_raw", "url", "account", "status", "first_seen", "last_seen", "last_changed_at", "previous_price_raw", "last_import_name"]:
        if col in df.columns:
            df[col] = df[col].fillna("").map(normalize_text)
    if "title" in df.columns:
        df["title_norm"] = df["title"].map(contains_text)
    return df


def registry_summary_text() -> str:
    df = load_avito_registry_df()
    if df.empty:
        return "Реестр пуст"
    active = int((df.get("status", pd.Series(dtype=object)) == "active").sum()) if "status" in df.columns else len(df)
    changed = int((pd.to_numeric(df.get("change_count", pd.Series(dtype=float)), errors="coerce").fillna(0) > 0).sum()) if "change_count" in df.columns else 0
    return f"В реестре: {len(df)} • активных: {active} • менялись: {changed}"



def get_photo_registry_path() -> Path:
    try:
        return Path(__file__).resolve().with_name("photo_registry.sqlite")
    except Exception:
        return Path.cwd() / "photo_registry.sqlite"


def ensure_photo_registry() -> None:
    path = get_photo_registry_path()
    required_columns = {
        "article_norm": "TEXT PRIMARY KEY",
        "article": "TEXT",
        "photo_url": "TEXT",
        "source_sheet": "TEXT",
        "meta_brand": "TEXT",
        "meta_color": "TEXT",
        "meta_capacity": "TEXT",
        "meta_manufacturer_code": "TEXT",
        "meta_model": "TEXT",
        "meta_description": "TEXT",
        "meta_fits_models": "TEXT",
        "meta_iso_pages": "TEXT",
        "meta_print_technology": "TEXT",
        "meta_item_type": "TEXT",
        "meta_print_type": "TEXT",
        "meta_weight": "TEXT",
        "meta_length": "TEXT",
        "meta_width": "TEXT",
        "meta_height": "TEXT",
        "first_seen": "TEXT",
        "last_seen": "TEXT",
        "last_changed_at": "TEXT",
        "import_name": "TEXT",
        "change_count": "INTEGER DEFAULT 0",
    }
    with sqlite3.connect(path) as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS photo_registry (
                article_norm TEXT PRIMARY KEY,
                article TEXT,
                photo_url TEXT,
                source_sheet TEXT,
                meta_brand TEXT,
                meta_color TEXT,
                meta_capacity TEXT,
                meta_manufacturer_code TEXT,
                meta_model TEXT,
                meta_description TEXT,
                meta_fits_models TEXT,
                meta_iso_pages TEXT,
                meta_print_technology TEXT,
                meta_item_type TEXT,
                meta_print_type TEXT,
                meta_weight TEXT,
                meta_length TEXT,
                meta_width TEXT,
                meta_height TEXT,
                first_seen TEXT,
                last_seen TEXT,
                last_changed_at TEXT,
                import_name TEXT,
                change_count INTEGER DEFAULT 0
            )
            """
        )
        existing_cols = {row[1] for row in conn.execute("PRAGMA table_info(photo_registry)")}
        for col_name, col_type in required_columns.items():
            if col_name not in existing_cols:
                conn.execute(f"ALTER TABLE photo_registry ADD COLUMN {col_name} {col_type}")
        conn.commit()


def load_photo_registry_df() -> pd.DataFrame:
    ensure_photo_registry()
    path = get_photo_registry_path()
    if not path.exists():
        return pd.DataFrame()
    with sqlite3.connect(path) as conn:
        df = pd.read_sql_query("SELECT * FROM photo_registry", conn)
    if df.empty:
        return df
    expected_cols = [
        "article", "article_norm", "photo_url", "source_sheet",
        "meta_brand", "meta_color", "meta_capacity", "meta_manufacturer_code",
        "meta_model", "meta_description", "meta_fits_models", "meta_iso_pages",
        "meta_print_technology", "meta_item_type", "meta_print_type",
        "meta_weight", "meta_length", "meta_width", "meta_height",
        "first_seen", "last_seen", "last_changed_at", "import_name",
    ]
    for col in expected_cols:
        if col not in df.columns:
            df[col] = ""
        df[col] = df[col].fillna("").map(normalize_text)
    # Откатили веб-парсер: старые web-fallback записи не подмешиваем в рабочий реестр.
    if "import_name" in df.columns:
        df = df[df["import_name"].fillna("") != "web-fallback"].copy()
    if "source_sheet" in df.columns:
        df = df[~df["source_sheet"].fillna("").str.startswith("web:")].copy()
    return df.reset_index(drop=True)


def photo_registry_summary_text() -> str:
    df = load_photo_registry_df()
    if df.empty:
        return "Реестр фото пуст"
    with_photo = int(df.get("photo_url", pd.Series(dtype=object)).fillna("").map(lambda x: 1 if normalize_text(x) else 0).sum())
    with_meta = int((
        df.get("meta_model", pd.Series(dtype=object)).fillna("").map(bool)
        | df.get("meta_brand", pd.Series(dtype=object)).fillna("").map(bool)
        | df.get("meta_fits_models", pd.Series(dtype=object)).fillna("").map(bool)
        | df.get("meta_color", pd.Series(dtype=object)).fillna("").map(bool)
        | df.get("meta_iso_pages", pd.Series(dtype=object)).fillna("").map(bool)
        | df.get("meta_description", pd.Series(dtype=object)).fillna("").map(bool)
    ).sum())
    return f"В реестре: {len(df)} • с фото: {with_photo} • с метаданными: {with_meta}"


def sync_photo_registry(photo_df: pd.DataFrame, import_name: str) -> dict[str, Any]:
    ensure_photo_registry()
    path = get_photo_registry_path()
    now = datetime.utcnow().replace(microsecond=0).isoformat() + "Z"
    work = photo_df.copy() if isinstance(photo_df, pd.DataFrame) else pd.DataFrame()
    if work.empty:
        return {"new": 0, "changed": 0, "unchanged": 0, "total": 0}

    use_cols = [
        "article", "article_norm", "photo_url", "source_sheet",
        "meta_brand", "meta_color", "meta_capacity", "meta_manufacturer_code",
        "meta_model", "meta_description", "meta_fits_models", "meta_iso_pages",
        "meta_print_technology", "meta_item_type", "meta_print_type",
        "meta_weight", "meta_length", "meta_width", "meta_height",
    ]
    for col in use_cols:
        if col not in work.columns:
            work[col] = ""
    work = work[use_cols].copy()
    work = work[work["article_norm"].map(normalize_text) != ""].copy()
    work = work.drop_duplicates(subset=["article_norm"], keep="first").reset_index(drop=True)

    stats = {"new": 0, "changed": 0, "unchanged": 0, "total": len(work)}
    tracked_cols = [
        "article", "photo_url", "source_sheet",
        "meta_brand", "meta_color", "meta_capacity", "meta_manufacturer_code",
        "meta_model", "meta_description", "meta_fits_models", "meta_iso_pages",
        "meta_print_technology", "meta_item_type", "meta_print_type",
        "meta_weight", "meta_length", "meta_width", "meta_height",
    ]

    with sqlite3.connect(path) as conn:
        existing = {}
        keys = work["article_norm"].tolist()
        if keys:
            placeholders = ",".join(["?"] * len(keys))
            for row in conn.execute(f"SELECT * FROM photo_registry WHERE article_norm IN ({placeholders})", keys):
                cols = [d[0] for d in conn.execute("SELECT * FROM photo_registry LIMIT 0").description]
                existing[row[0]] = dict(zip(cols, row))

        for _, rec in work.iterrows():
            key = normalize_text(rec.get("article_norm", ""))
            if not key:
                continue
            payload = {col: normalize_text(rec.get(col, "")) for col in tracked_cols}
            old = existing.get(key)
            if old is None:
                conn.execute(
                    """
                    INSERT INTO photo_registry (
                        article_norm, article, photo_url, source_sheet,
                        meta_brand, meta_color, meta_capacity, meta_manufacturer_code, meta_model,
                        meta_description, meta_fits_models, meta_iso_pages, meta_print_technology,
                        meta_item_type, meta_print_type, meta_weight, meta_length, meta_width, meta_height,
                        first_seen, last_seen, last_changed_at, import_name, change_count
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 0)
                    """,
                    (
                        key, payload["article"], payload["photo_url"], payload["source_sheet"],
                        payload["meta_brand"], payload["meta_color"], payload["meta_capacity"], payload["meta_manufacturer_code"], payload["meta_model"],
                        payload["meta_description"], payload["meta_fits_models"], payload["meta_iso_pages"], payload["meta_print_technology"],
                        payload["meta_item_type"], payload["meta_print_type"], payload["meta_weight"], payload["meta_length"], payload["meta_width"], payload["meta_height"],
                        now, now, now, normalize_text(import_name),
                    ),
                )
                stats["new"] += 1
            else:
                changed = any(normalize_text(old.get(col, "")) != payload[col] for col in tracked_cols)
                if changed:
                    change_count = int(old.get("change_count") or 0) + 1
                    conn.execute(
                        """
                        UPDATE photo_registry SET
                            article=?,
                            photo_url=?,
                            source_sheet=?,
                            meta_brand=?,
                            meta_color=?,
                            meta_capacity=?,
                            meta_manufacturer_code=?,
                            meta_model=?,
                            meta_description=?,
                            meta_fits_models=?,
                            meta_iso_pages=?,
                            meta_print_technology=?,
                            meta_item_type=?,
                            meta_print_type=?,
                            meta_weight=?,
                            meta_length=?,
                            meta_width=?,
                            meta_height=?,
                            last_seen=?,
                            last_changed_at=?,
                            import_name=?,
                            change_count=?
                        WHERE article_norm=?
                        """,
                        (
                            payload["article"], payload["photo_url"], payload["source_sheet"],
                            payload["meta_brand"], payload["meta_color"], payload["meta_capacity"], payload["meta_manufacturer_code"], payload["meta_model"],
                            payload["meta_description"], payload["meta_fits_models"], payload["meta_iso_pages"], payload["meta_print_technology"],
                            payload["meta_item_type"], payload["meta_print_type"], payload["meta_weight"], payload["meta_length"], payload["meta_width"], payload["meta_height"],
                            now, now, normalize_text(import_name), change_count, key,
                        ),
                    )
                    stats["changed"] += 1
                else:
                    conn.execute(
                        "UPDATE photo_registry SET last_seen=?, import_name=? WHERE article_norm=?",
                        (now, normalize_text(import_name), key),
                    )
                    stats["unchanged"] += 1
        conn.commit()
    return stats


def ensure_photo_registry_loaded() -> None:
    if isinstance(st.session_state.get("photo_df"), pd.DataFrame) and not st.session_state.get("photo_df").empty:
        return
    reg = load_photo_registry_df()
    if isinstance(reg, pd.DataFrame) and not reg.empty:
        required_cols = [
            "article", "article_norm", "photo_url", "source_sheet",
            "meta_brand", "meta_color", "meta_capacity", "meta_manufacturer_code",
            "meta_model", "meta_description", "meta_fits_models", "meta_iso_pages",
            "meta_print_technology", "meta_item_type", "meta_print_type",
            "meta_weight", "meta_length", "meta_width", "meta_height",
        ]
        for col in required_cols:
            if col not in reg.columns:
                reg[col] = ""
        st.session_state.photo_df = reg[required_cols].copy()
        if normalize_text(st.session_state.get("photo_name", "")) in {"", "ещё не загружен"}:
            st.session_state.photo_name = "из реестра сервера"


def ensure_persisted_source_files_loaded() -> None:
    if (not isinstance(st.session_state.get("comparison_sheets"), dict) or not st.session_state.get("comparison_sheets")) and get_persisted_comparison_file_path().exists():
        load_persisted_comparison_source_into_state()
    if (not isinstance(st.session_state.get("photo_df"), pd.DataFrame) or st.session_state.get("photo_df").empty) and get_persisted_photo_file_path().exists():
        load_persisted_photo_source_into_state()
    if (not isinstance(st.session_state.get("avito_df"), pd.DataFrame) or st.session_state.get("avito_df").empty) and get_persisted_avito_file_path().exists():
        load_persisted_avito_source_into_state()
    if (not isinstance(st.session_state.get("hot_items_df"), pd.DataFrame) or st.session_state.get("hot_items_df").empty) and get_persisted_watchlist_file_path().exists():
        load_persisted_watchlist_source_into_state()






def get_card_override_path() -> Path:
    try:
        return Path(__file__).resolve().with_name("card_overrides.sqlite")
    except NameError:
        return Path.cwd() / "card_overrides.sqlite"


def ensure_card_override_db() -> None:
    path = get_card_override_path()
    path.parent.mkdir(parents=True, exist_ok=True)
    required_columns = {
        "sheet_name": "TEXT NOT NULL",
        "article_norm": "TEXT NOT NULL",
        "article": "TEXT",
        "photo_url": "TEXT",
        "name_override": "TEXT",
        "meta_brand": "TEXT",
        "meta_model": "TEXT",
        "meta_manufacturer_code": "TEXT",
        "meta_print_type": "TEXT",
        "meta_color": "TEXT",
        "meta_capacity": "TEXT",
        "meta_iso_pages": "TEXT",
        "meta_item_type": "TEXT",
        "meta_print_technology": "TEXT",
        "meta_description": "TEXT",
        "meta_fits_models": "TEXT",
        "meta_weight": "TEXT",
        "meta_length": "TEXT",
        "meta_width": "TEXT",
        "meta_height": "TEXT",
        "note": "TEXT",
        "updated_at": "TEXT",
    }
    with sqlite3.connect(path) as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS card_overrides (
                sheet_name TEXT NOT NULL,
                article_norm TEXT NOT NULL,
                article TEXT,
                photo_url TEXT,
                name_override TEXT,
                meta_brand TEXT,
                meta_model TEXT,
                meta_manufacturer_code TEXT,
                meta_print_type TEXT,
                meta_color TEXT,
                meta_capacity TEXT,
                meta_iso_pages TEXT,
                meta_item_type TEXT,
                meta_print_technology TEXT,
                meta_description TEXT,
                meta_fits_models TEXT,
                meta_weight TEXT,
                meta_length TEXT,
                meta_width TEXT,
                meta_height TEXT,
                note TEXT,
                updated_at TEXT,
                PRIMARY KEY (sheet_name, article_norm)
            )
            """
        )
        existing_cols = {row[1] for row in conn.execute("PRAGMA table_info(card_overrides)")}
        for col_name, col_type in required_columns.items():
            if col_name not in existing_cols:
                conn.execute(f"ALTER TABLE card_overrides ADD COLUMN {col_name} {col_type}")
        conn.commit()


@st.cache_data(ttl=1800, max_entries=5)
def load_card_overrides_df() -> pd.DataFrame:
    path = get_card_override_path()
    if not path.exists():
        return pd.DataFrame()
    with sqlite3.connect(path) as conn:
        df = pd.read_sql_query("SELECT * FROM card_overrides", conn)
    if df.empty:
        return df
    for col in [
        "sheet_name", "article_norm", "article", "photo_url", "name_override",
        "meta_brand", "meta_model", "meta_manufacturer_code", "meta_print_type",
        "meta_color", "meta_capacity", "meta_iso_pages", "meta_item_type",
        "meta_print_technology", "meta_description", "meta_fits_models",
        "meta_weight", "meta_length", "meta_width", "meta_height",
        "note", "updated_at",
    ]:
        if col in df.columns:
            df[col] = df[col].fillna("").map(normalize_text)
    return df


def clear_card_override_cache() -> None:
    try:
        load_card_overrides_df.clear()
    except Exception:
        pass


def save_card_override(sheet_name: str, article: str, article_norm: str, payload: dict[str, Any]) -> None:
    ensure_card_override_db()
    now = datetime.utcnow().replace(microsecond=0).isoformat() + "Z"
    clean = {k: normalize_text(v) for k, v in (payload or {}).items()}
    with sqlite3.connect(get_card_override_path()) as conn:
        conn.execute(
            """
            INSERT INTO card_overrides (
                sheet_name, article_norm, article, photo_url, name_override,
                meta_brand, meta_model, meta_manufacturer_code, meta_print_type,
                meta_color, meta_capacity, meta_iso_pages, meta_item_type,
                meta_print_technology, meta_description, meta_fits_models,
                meta_weight, meta_length, meta_width, meta_height,
                note, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(sheet_name, article_norm) DO UPDATE SET
                article=excluded.article,
                photo_url=excluded.photo_url,
                name_override=excluded.name_override,
                meta_brand=excluded.meta_brand,
                meta_model=excluded.meta_model,
                meta_manufacturer_code=excluded.meta_manufacturer_code,
                meta_print_type=excluded.meta_print_type,
                meta_color=excluded.meta_color,
                meta_capacity=excluded.meta_capacity,
                meta_iso_pages=excluded.meta_iso_pages,
                meta_item_type=excluded.meta_item_type,
                meta_print_technology=excluded.meta_print_technology,
                meta_description=excluded.meta_description,
                meta_fits_models=excluded.meta_fits_models,
                meta_weight=excluded.meta_weight,
                meta_length=excluded.meta_length,
                meta_width=excluded.meta_width,
                meta_height=excluded.meta_height,
                note=excluded.note,
                updated_at=excluded.updated_at
            """,
            (
                normalize_text(sheet_name), normalize_text(article_norm), normalize_text(article),
                clean.get("photo_url", ""), clean.get("name_override", ""), clean.get("meta_brand", ""),
                clean.get("meta_model", ""), clean.get("meta_manufacturer_code", ""), clean.get("meta_print_type", ""),
                clean.get("meta_color", ""), clean.get("meta_capacity", ""), clean.get("meta_iso_pages", ""),
                clean.get("meta_item_type", ""), clean.get("meta_print_technology", ""), clean.get("meta_description", ""),
                clean.get("meta_fits_models", ""), clean.get("meta_weight", ""), clean.get("meta_length", ""),
                clean.get("meta_width", ""), clean.get("meta_height", ""), clean.get("note", ""), now
            ),
        )
        conn.commit()
    clear_card_override_cache()


def delete_card_override(sheet_name: str, article_norm: str) -> None:
    ensure_card_override_db()
    with sqlite3.connect(get_card_override_path()) as conn:
        conn.execute(
            "DELETE FROM card_overrides WHERE sheet_name=? AND article_norm=?",
            (normalize_text(sheet_name), normalize_text(article_norm)),
        )
        conn.commit()
    clear_card_override_cache()



def get_task_registry_path() -> Path:
    try:
        return Path(__file__).resolve().with_name("review_tasks.sqlite")
    except NameError:
        return Path.cwd() / "review_tasks.sqlite"


def ensure_task_registry_db() -> None:
    path = get_task_registry_path()
    path.parent.mkdir(parents=True, exist_ok=True)
    with sqlite3.connect(path) as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS review_tasks (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                article_norm TEXT NOT NULL,
                article TEXT,
                sheet_name TEXT,
                name_snapshot TEXT,
                created_at TEXT,
                due_date TEXT,
                status TEXT,
                reason TEXT,
                note TEXT,
                completed_at TEXT,
                source TEXT
            )
            """
        )
        conn.commit()


@st.cache_data(ttl=300, max_entries=4)
def load_task_registry_df() -> pd.DataFrame:
    path = get_task_registry_path()
    if not path.exists():
        return pd.DataFrame(
            columns=[
                "id", "article_norm", "article", "sheet_name", "name_snapshot",
                "created_at", "due_date", "status", "reason", "note", "completed_at", "source"
            ]
        )
    ensure_task_registry_db()
    with sqlite3.connect(path) as conn:
        df = pd.read_sql_query("SELECT * FROM review_tasks ORDER BY id DESC", conn)
    if df.empty:
        return df
    text_cols = [
        "article_norm", "article", "sheet_name", "name_snapshot",
        "created_at", "due_date", "status", "reason", "note", "completed_at", "source"
    ]
    for col in text_cols:
        if col in df.columns:
            df[col] = df[col].fillna("").map(normalize_text)
    return df


def clear_task_registry_cache() -> None:
    try:
        load_task_registry_df.clear()
    except Exception:
        pass
    clear_runtime_perf_caches()


def create_review_task(
    article: str,
    article_norm: str,
    sheet_name: str,
    name_snapshot: str,
    due_date: Any,
    reason: str = "",
    note: str = "",
    source: str = "manual_review",
) -> None:
    ensure_task_registry_db()
    now = datetime.utcnow().replace(microsecond=0).isoformat() + "Z"
    due_txt = ""
    if due_date:
        try:
            if hasattr(due_date, "isoformat"):
                due_txt = due_date.isoformat()
            else:
                due_txt = str(due_date)
        except Exception:
            due_txt = str(due_date)
    with sqlite3.connect(get_task_registry_path()) as conn:
        conn.execute(
            """
            INSERT INTO review_tasks (
                article_norm, article, sheet_name, name_snapshot,
                created_at, due_date, status, reason, note, completed_at, source
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                normalize_text(article_norm),
                normalize_text(article),
                normalize_text(sheet_name),
                normalize_text(name_snapshot),
                now,
                normalize_text(due_txt),
                "NEW",
                normalize_text(reason),
                normalize_text(note),
                "",
                normalize_text(source),
            ),
        )
        conn.commit()
    clear_task_registry_cache()


def safe_int(value: Any, default: int = 0) -> int:
    try:
        if value is None:
            return int(default)
        if isinstance(value, bool):
            return int(value)
        text_value = str(value).strip()
        if not text_value:
            return int(default)
        return int(float(text_value.replace(",", ".")))
    except Exception:
        return int(default)

def update_review_task_status(task_id: Any, status: str) -> None:
    ensure_task_registry_db()
    task_id = safe_int(task_id, 0)
    if task_id <= 0:
        return
    new_status = normalize_text(status).upper()
    completed_at = ""
    if new_status == "DONE":
        completed_at = datetime.utcnow().replace(microsecond=0).isoformat() + "Z"
    with sqlite3.connect(get_task_registry_path()) as conn:
        conn.execute(
            "UPDATE review_tasks SET status=?, completed_at=? WHERE id=?",
            (new_status, completed_at, int(task_id)),
        )
        conn.commit()
    clear_task_registry_cache()


def task_effective_status(row: dict[str, Any] | pd.Series) -> str:
    raw_status = normalize_text((row or {}).get("status", "")).upper() if isinstance(row, dict) else normalize_text(row.get("status", "")).upper()
    due_txt = normalize_text((row or {}).get("due_date", "")) if isinstance(row, dict) else normalize_text(row.get("due_date", ""))
    if raw_status in {"DONE", "CANCELLED"}:
        return raw_status
    if due_txt:
        try:
            due_dt = datetime.fromisoformat(due_txt).date()
            if due_dt < datetime.utcnow().date():
                return "OVERDUE"
        except Exception:
            pass
    if raw_status == "ACTIVE":
        return "ACTIVE"
    return "NEW"


def task_status_ru(status: str) -> str:
    mapping = {
        "NEW": "Новая",
        "ACTIVE": "Активная",
        "OVERDUE": "Просрочена",
        "DONE": "Выполнена",
        "CANCELLED": "Отменена",
    }
    return mapping.get(normalize_text(status).upper(), normalize_text(status))


def build_task_view_df(sheet_filter: str | None = None) -> pd.DataFrame:
    df = load_task_registry_df()
    if not isinstance(df, pd.DataFrame) or df.empty:
        return pd.DataFrame(
            columns=[
                "ID", "Артикул", "Название", "Лист", "Создана", "Срок",
                "Статус", "Причина", "Комментарий", "Источник"
            ]
        )
    work = df.copy()
    work["effective_status"] = work.apply(lambda r: task_effective_status(r), axis=1)
    if sheet_filter and normalize_text(sheet_filter):
        work = work[work.get("sheet_name", pd.Series(dtype=object)).fillna("").map(normalize_text).eq(normalize_text(sheet_filter))]
    work["ID"] = pd.to_numeric(work.get("id", 0), errors="coerce").fillna(0).astype(int)
    work["Артикул"] = work.get("article", "")
    work["Название"] = work.get("name_snapshot", "")
    work["Лист"] = work.get("sheet_name", "")
    work["Создана"] = work.get("created_at", "")
    work["Срок"] = work.get("due_date", "")
    work["Статус"] = work["effective_status"].map(task_status_ru)
    work["Причина"] = work.get("reason", "")
    work["Комментарий"] = work.get("note", "")
    work["Источник"] = work.get("source", "")
    return work[["ID", "Артикул", "Название", "Лист", "Создана", "Срок", "Статус", "Причина", "Комментарий", "Источник"]].copy()


def task_summary_counts() -> dict[str, int]:
    df = load_task_registry_df()
    if not isinstance(df, pd.DataFrame) or df.empty:
        return {"new": 0, "active": 0, "overdue": 0, "done": 0, "open": 0}
    eff = [task_effective_status(r) for _, r in df.iterrows()]
    return {
        "new": sum(1 for x in eff if x == "NEW"),
        "active": sum(1 for x in eff if x == "ACTIVE"),
        "overdue": sum(1 for x in eff if x == "OVERDUE"),
        "done": sum(1 for x in eff if x == "DONE"),
        "open": sum(1 for x in eff if x in {"NEW", "ACTIVE", "OVERDUE"}),
    }


def tasks_summary_text() -> str:
    c = task_summary_counts()
    if c["open"] <= 0:
        return "Задач нет"
    return f"новых: {c['new']} • активных: {c['active']} • просрочено: {c['overdue']}"


def trigger_search_from_task(article: str, sheet_label: str) -> None:
    sheet_label = normalize_text(sheet_label) or "Оригинал"
    label_map = {
        "Сравнение": "Оригинал",
        "Оригинал": "Оригинал",
        "Уценка": "Уценка",
        "Совместимые": "Совместимые",
    }
    resolved_label = label_map.get(sheet_label, sheet_label)
    tab_mapping = {"Оригинал": "original", "Уценка": "discount", "Совместимые": "compatible"}
    if resolved_label in tab_mapping:
        st.session_state["active_workspace_label"] = resolved_label
        trigger_search_from_article(article, tab_mapping[resolved_label])
    else:
        trigger_search_from_article(article, st.session_state.get("active_workspace_label", "original"))


def apply_task_filters(task_df: pd.DataFrame, status_filter: str, period_filter: str, sheet_filter: str) -> pd.DataFrame:
    if not isinstance(task_df, pd.DataFrame) or task_df.empty:
        return pd.DataFrame(columns=task_df.columns if isinstance(task_df, pd.DataFrame) else [])
    out = task_df.copy()

    if sheet_filter and sheet_filter != "Все листы":
        out = out[out["Лист"].astype(str) == sheet_filter]

    if status_filter == "Новые":
        out = out[out["Статус"].eq("Новая")]
    elif status_filter == "Активные":
        out = out[out["Статус"].eq("Активная")]
    elif status_filter == "Просроченные":
        out = out[out["Статус"].eq("Просрочена")]
    elif status_filter == "Выполненные":
        out = out[out["Статус"].eq("Выполнена")]
    elif status_filter == "Не выполненные":
        out = out[~out["Статус"].isin(["Выполнена", "Отменена"])]

    today = datetime.utcnow().date()
    if period_filter and period_filter != "Все":
        due = pd.to_datetime(out["Срок"], errors="coerce").dt.date
        if period_filter == "Сегодня":
            out = out[due == today]
        elif period_filter == "7 дней":
            out = out[(due.notna()) & (due >= today) & (due <= (today + timedelta(days=7)))]
        elif period_filter == "14 дней":
            out = out[(due.notna()) & (due >= today) & (due <= (today + timedelta(days=14)))]
        elif period_filter == "30 дней":
            out = out[(due.notna()) & (due >= today) & (due <= (today + timedelta(days=30)))]
        elif period_filter == "Просроченные":
            out = out[out["Статус"].eq("Просрочена")]
    return out


def render_tasks_table_ui(task_df: pd.DataFrame, key_prefix: str, default_sheet: str | None = None) -> None:
    if not isinstance(task_df, pd.DataFrame) or task_df.empty:
        st.info("Задач пока нет.")
        return

    filters = st.columns([1.25, 1.25, 1.2])
    status_filter = filters[0].selectbox(
        "Статус",
        ["Все", "Новые", "Активные", "Просроченные", "Выполненные", "Не выполненные"],
        key=f"task_status_filter_{key_prefix}",
        help="Фильтрует задачи по статусу: новые, активные, просроченные, выполненные или все сразу.",
    )
    period_filter = filters[1].selectbox(
        "Период",
        ["Все", "Сегодня", "7 дней", "14 дней", "30 дней", "Просроченные"],
        key=f"task_period_filter_{key_prefix}",
        help="Фильтр по сроку задачи: на сегодня, на ближайшие дни или только просроченные.",
    )
    sheet_options = ["Все листы"] + sorted([x for x in task_df["Лист"].fillna("").astype(str).unique().tolist() if str(x).strip()])
    default_index = sheet_options.index(default_sheet) if default_sheet in sheet_options else 0
    sheet_filter = filters[2].selectbox(
        "Лист",
        sheet_options,
        index=default_index,
        key=f"task_sheet_filter_{key_prefix}",
        help="Ограничивает список задач выбранным листом: Оригинал, Уценка или Совместимые.",
    )

    st.caption(
        "Статус — показывает, на каком этапе задача: новая, активная, просроченная или выполненная. "
        "Период — помогает увидеть срочные и просроченные задачи. "
        "Лист — ограничивает список выбранным разделом comparison."
    )

    filtered = apply_task_filters(task_df, status_filter, period_filter, sheet_filter)
    if filtered.empty:
        st.info("По выбранным фильтрам задач не найдено.")
        return

    st.dataframe(filtered, use_container_width=True, hide_index=True, height=min(520, 120 + len(filtered) * 36))

    labels = []
    row_map = {}
    for _, row in filtered.iterrows():
        label = f"{row['Артикул']} • {row['Статус']} • срок: {row['Срок'] or '—'}"
        labels.append(label)
        row_map[label] = row
    pick_col, b1, b2, b3 = st.columns([4, 1.1, 1.1, 1.1])
    selected = pick_col.selectbox(
        "Открыть или изменить задачу",
        labels,
        key=f"task_selected_{key_prefix}",
        help="Выбери задачу, чтобы открыть карточку товара, отметить её выполненной или вернуть в работу.",
    )
    if not selected:
        return
    row = row_map[selected]
    if b1.button("Открыть", key=f"task_open_{key_prefix}", use_container_width=True):
        trigger_search_from_task(str(row.get("Артикул", "")), str(row.get("Лист", "")))
    if b2.button("Выполнено", key=f"task_done_{key_prefix}", use_container_width=True):
        update_review_task_status(row.get("ID", 0), "DONE")
        st.success(f"Задача по {row.get('Артикул', '')} отмечена как выполненная.")
        st.rerun()
    if b3.button("Вернуть в работу", key=f"task_active_{key_prefix}", use_container_width=True):
        update_review_task_status(row.get("ID", 0), "ACTIVE")
        st.success(f"Задача по {row.get('Артикул', '')} возвращена в работу.")
        st.rerun()


def render_task_center_lazy_panel() -> None:
    global_open = bool(st.session_state.get("show_task_center_global", False))
    crm_open = any(
        bool(v) for k, v in st.session_state.items()
        if str(k).startswith("crm_show_tasks_")
    )
    if not (global_open or crm_open):
        return
    counts = task_summary_counts()
    task_df = build_task_view_df()
    st.markdown('<div class="result-wrap">', unsafe_allow_html=True)
    render_block_header(
        "Задачи / напоминания",
        "Список задач по карточкам: что проверить, к какому сроку и по какой причине.",
        icon="🔔",
        help_text="Это отдельный слой задач поверх карточек. Задачи не меняют comparison-файл и не пропадают при загрузке нового файла.",
    )
    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Открытых задач", counts.get("open", 0))
    m2.metric("Новые", counts.get("new", 0))
    m3.metric("Активные", counts.get("active", 0))
    m4.metric("Просроченные", counts.get("overdue", 0))
    st.caption("ⓘ Новые — ещё не разобраны. Активные — в работе. Просроченные — срок уже вышел. Этот блок нужен, чтобы не терять ручные проверки по карточкам.")
    st.caption(
        "Что показывает: задачи на пересмотр карточек и цен. "
        "Как пользоваться: смотри срочные задачи, открывай карточку и после проверки отмечай задачу выполненной."
    )
    render_tasks_table_ui(task_df, "global_tasks")
    st.markdown('</div>', unsafe_allow_html=True)

def apply_card_overrides(df: pd.DataFrame | None, sheet_name: str) -> pd.DataFrame | None:
    if df is None or not isinstance(df, pd.DataFrame) or df.empty:
        return df
    overrides = load_card_overrides_df()
    if overrides.empty:
        return df
    sheet_name_norm = normalize_text(sheet_name)
    work = overrides[overrides.get("sheet_name", pd.Series(dtype=object)).fillna("").map(normalize_text).eq(sheet_name_norm)].copy()
    if work.empty:
        return df
    work = work.drop_duplicates(subset=["article_norm"], keep="last")
    by_key = {normalize_text(r["article_norm"]): r for _, r in work.iterrows() if normalize_text(r.get("article_norm", ""))}
    out = df.copy()

    text_cols = [
        "photo_url",
        "source_sheet",
        "name",
        "meta_brand",
        "meta_model",
        "meta_manufacturer_code",
        "meta_print_type",
        "meta_color",
        "meta_capacity",
        "meta_iso_pages",
        "meta_item_type",
        "meta_print_technology",
        "meta_description",
        "meta_fits_models",
        "meta_weight",
        "meta_length",
        "meta_width",
        "meta_height",
        "manual_note",
    ]
    for col in text_cols:
        if col not in out.columns:
            out[col] = ""
        else:
            out[col] = out[col].astype(object)

    for idx, row in out.iterrows():
        key = normalize_text(row.get("article_norm", ""))
        if not key:
            continue
        ov = by_key.get(key)
        if ov is None:
            continue
        photo_url = normalize_text(ov.get("photo_url", ""))
        if photo_url:
            out.at[idx, "photo_url"] = photo_url
            out.at[idx, "source_sheet"] = "manual_override"
        name_override = normalize_text(ov.get("name_override", ""))
        if name_override:
            out.at[idx, "name"] = name_override
        for field_name in [
            "meta_brand",
            "meta_model",
            "meta_manufacturer_code",
            "meta_print_type",
            "meta_color",
            "meta_capacity",
            "meta_iso_pages",
            "meta_item_type",
            "meta_print_technology",
            "meta_description",
            "meta_fits_models",
            "meta_weight",
            "meta_length",
            "meta_width",
            "meta_height",
        ]:
            field_value = normalize_text(ov.get(field_name, ""))
            if field_value:
                out.at[idx, field_name] = field_value
        note = normalize_text(ov.get("note", ""))
        if note:
            out.at[idx, "manual_note"] = note
    return out


def trigger_search_from_article(article: str, tab_key: str) -> None:
    query = normalize_query_for_display(article)
    if not query:
        return
    st.session_state[f"search_input_{tab_key}"] = query
    st.session_state[f"submitted_query_{tab_key}"] = query
    st.session_state[f"search_input_widget_pending_{tab_key}"] = query
    st.session_state[f"last_result_{tab_key}"] = None
    st.session_state[f"last_result_sig_{tab_key}"] = None
    st.rerun()


def render_analytics_jump_helper(df: pd.DataFrame | None, tab_key: str, box_key: str) -> None:
    if not isinstance(df, pd.DataFrame) or df.empty or "Артикул" not in df.columns:
        return
    articles = [normalize_text(x) for x in df["Артикул"].tolist() if normalize_text(x)]
    articles = unique_preserve_order(articles)
    if not articles:
        return
    c1, c2 = st.columns([4, 1.2])
    selected_article = c1.selectbox(
        "Открыть позицию в обычном поиске",
        articles,
        key=f"analytics_open_select_{box_key}_{tab_key}",
        help="Быстрый переход из аналитики в обычную карточку товара.",
    )
    if c2.button("Открыть", key=f"analytics_open_btn_{box_key}_{tab_key}", use_container_width=True):
        trigger_search_from_article(selected_article, tab_key)


def render_crm_issue_open_helper(
    df: pd.DataFrame | None,
    tab_key: str,
    box_key: str,
    button_text: str = "Открыть",
    open_editor: bool = False,
) -> None:
    if not isinstance(df, pd.DataFrame) or df.empty or "Артикул" not in df.columns:
        return
    tabkey_to_label = {v: k for k, v in CRM_SHEET_LABEL_TO_TABKEY.items()}
    working_df = df.copy()
    working_df["Артикул"] = working_df["Артикул"].astype(str)
    labels = []
    row_map = {}
    for _, row in working_df.iterrows():
        article = normalize_text(row.get("Артикул", ""))
        if not article:
            continue
        name = normalize_text(row.get("Название", "") or row.get("Товар", ""))
        label = f"{article} • {name[:90]}" if name else article
        labels.append(label)
        row_map[label] = row.to_dict()
    labels = unique_preserve_order(labels)
    if not labels:
        return
    c1, c2 = st.columns([4, 1.3])
    selected_label = c1.selectbox(
        "Открыть позицию в CRM",
        labels,
        key=f"crm_issue_open_select_{box_key}_{tab_key}",
        help=(
            "Открывает позицию сразу в новой CRM-карточке. "
            "Для блока без фото можно сразу перейти в редактор фото."
        ),
    )
    if c2.button(button_text, key=f"crm_issue_open_btn_{box_key}_{tab_key}", use_container_width=True):
        row = row_map.get(selected_label, {})
        article = normalize_text(row.get("Артикул", ""))
        sheet_label = normalize_text(row.get("Лист", "")) or tabkey_to_label.get(normalize_text(tab_key), "Оригинал")
        open_product_in_crm(article, sheet_label=sheet_label, open_photo_editor=bool(open_editor))


def render_crm_quality_issue_lazy_panels(
    sheet_df: pd.DataFrame | None,
    photo_df: pd.DataFrame | None,
    avito_df: pd.DataFrame | None,
    min_qty: float,
    sheet_name: str,
    tab_label: str,
    tab_key: str,
) -> None:
    show_no_photo = bool(st.session_state.get(f"crm_show_no_photo_{sheet_name}", False))
    show_no_avito = bool(st.session_state.get(f"crm_show_no_avito_{sheet_name}", False))
    if not (show_no_photo or show_no_avito):
        return
    if not isinstance(sheet_df, pd.DataFrame) or sheet_df.empty:
        return

    registry_df = load_avito_registry_df()
    bundle = build_operational_analytics_bundle(
        sheet_df,
        photo_df,
        avito_df,
        registry_df,
        min_qty,
        tab_label,
        st.session_state.get("hot_items_df"),
    )
    meta_df = bundle.get("meta_df", pd.DataFrame()) if isinstance(bundle, dict) else pd.DataFrame()
    if not isinstance(meta_df, pd.DataFrame) or meta_df.empty:
        return

    def _render_issue_panel(
        issue_df: pd.DataFrame,
        title: str,
        subtitle: str,
        icon: str,
        box_key: str,
        button_text: str,
        open_editor: bool,
    ) -> None:
        st.markdown('<div class="result-wrap">', unsafe_allow_html=True)
        render_block_header(
            title,
            subtitle,
            icon=icon,
            help_text="Это ленивый поверхностный инструмент CRM. Он ничего не меняет в ядре, а только помогает открыть нужные позиции сразу в новой CRM-карточке.",
        )
        if issue_df.empty:
            st.info("По текущему листу строк не найдено.")
        else:
            view_cols = [
                "Артикул", "Название", "Наш остаток", "Причины", "Объявлений Авито",
                "Фото", "Шаблон", "Лучший поставщик", "Разница, %",
            ]
            view_cols = [c for c in view_cols if c in issue_df.columns]
            view = issue_df[view_cols].copy()
            st.dataframe(
                view,
                use_container_width=True,
                hide_index=True,
                height=min(520, 120 + len(view) * 35),
            )
            render_crm_issue_open_helper(issue_df, tab_key, box_key, button_text=button_text, open_editor=open_editor)
        st.markdown('</div>', unsafe_allow_html=True)

    if show_no_photo:
        no_photo_df = meta_df[meta_df.get("Фото", pd.Series(dtype=object)).fillna("").eq("Нет")].copy()
        if not no_photo_df.empty:
            no_photo_df = no_photo_df.sort_values(["Наш остаток", "Название"], ascending=[False, True], na_position="last").reset_index(drop=True)
        _render_issue_panel(
            no_photo_df,
            f"Нет фото — позиции для доработки ({len(no_photo_df)})",
            "Показывает только позиции текущего листа без фото. Можно сразу открыть товар в новой CRM-карточке и быстро добавить ссылку на фото.",
            icon="🖼️",
            box_key="crm_no_photo",
            button_text="Открыть и редактировать",
            open_editor=True,
        )

    if show_no_avito:
        no_avito_mask = pd.to_numeric(meta_df.get("Объявлений Авито", 0), errors="coerce").fillna(0).eq(0)
        no_avito_df = meta_df[no_avito_mask].copy()
        if not no_avito_df.empty:
            no_avito_df = no_avito_df.sort_values(["Наш остаток", "Название"], ascending=[False, True], na_position="last").reset_index(drop=True)
        _render_issue_panel(
            no_avito_df,
            f"Без Avito — позиции для размещения ({len(no_avito_df)})",
            "Показывает только позиции текущего листа без объявлений Avito. Можно сразу открыть позицию в новой CRM-карточке и перейти к размещению.",
            icon="🛒",
            box_key="crm_no_avito",
            button_text="Открыть",
            open_editor=False,
        )


def render_card_editor_panel(result_df: pd.DataFrame | None, sheet_name: str, tab_key: str) -> None:
    if not isinstance(result_df, pd.DataFrame) or result_df.empty:
        return
    if not st.checkbox("✏️ Редактировать карточку", key=f"show_card_editor_{tab_key}", help="Открывает безопасный редактор карточки. Правки сохраняются поверх файла и не пропадают после новой загрузки comparison."):
        return

    rows = result_df.copy()
    options = []
    option_map = {}
    for _, row in rows.iterrows():
        art = normalize_text(row.get("article", ""))
        key = normalize_text(row.get("article_norm", ""))
        name = normalize_text(row.get("name", ""))
        label = f"{art} — {name[:120]}"
        options.append(label)
        option_map[label] = row

    default_label = options[0] if options else None
    selected_label = st.selectbox("Позиция для редактирования", options, index=0 if default_label else None, key=f"card_editor_select_{tab_key}")
    if not selected_label:
        return
    row = option_map[selected_label]
    art = normalize_text(row.get("article", ""))
    art_norm = normalize_text(row.get("article_norm", ""))
    current_photo = normalize_text(row.get("photo_url", ""))
    current_name = normalize_text(row.get("name", ""))
    current_brand = normalize_text(row.get("meta_brand", ""))
    current_model = normalize_text(row.get("meta_model", ""))
    current_code = normalize_text(row.get("meta_manufacturer_code", ""))
    current_print_type = normalize_text(row.get("meta_print_type", ""))
    current_color = normalize_text(row.get("meta_color", ""))
    current_capacity = normalize_text(row.get("meta_capacity", ""))
    current_iso_pages = normalize_text(row.get("meta_iso_pages", ""))
    current_item_type = normalize_text(row.get("meta_item_type", ""))
    current_print_technology = normalize_text(row.get("meta_print_technology", ""))
    current_description = normalize_text(row.get("meta_description", ""))
    current_fits = normalize_text(row.get("meta_fits_models", ""))
    current_weight = normalize_text(row.get("meta_weight", ""))
    current_length = normalize_text(row.get("meta_length", ""))
    current_width = normalize_text(row.get("meta_width", ""))
    current_height = normalize_text(row.get("meta_height", ""))
    current_note = normalize_text(row.get("manual_note", ""))

    st.caption("Правки сохраняются как ручные overrides и накладываются поверх comparison-файла после каждой новой загрузки.")
    with st.form(f"card_editor_form_{tab_key}_{art_norm}", clear_on_submit=False):
        col1, col2 = st.columns([1.2, 1.8])
        with col1:
            photo_url = st.text_input("Фото (ссылка)", value=current_photo, key=f"card_edit_photo_{tab_key}_{art_norm}")
            if current_photo:
                st.link_button("Открыть текущее фото", current_photo, use_container_width=True)
        with col2:
            name_override = st.text_area("Название", value=current_name, height=90, key=f"card_edit_name_{tab_key}_{art_norm}")
        cmeta1, cmeta2, cmeta3 = st.columns(3)
        meta_brand = cmeta1.text_input("Бренд", value=current_brand, key=f"card_edit_brand_{tab_key}_{art_norm}")
        meta_model = cmeta2.text_input("Модель", value=current_model, key=f"card_edit_model_{tab_key}_{art_norm}")
        meta_code = cmeta3.text_input("Код производителя", value=current_code, key=f"card_edit_code_{tab_key}_{art_norm}")

        cmeta4, cmeta5, cmeta6 = st.columns(3)
        meta_print_type = cmeta4.text_input("Тип печати", value=current_print_type, key=f"card_edit_print_type_{tab_key}_{art_norm}")
        meta_color = cmeta5.text_input("Цвет", value=current_color, key=f"card_edit_color_{tab_key}_{art_norm}")
        meta_capacity = cmeta6.text_input("Емкость", value=current_capacity, key=f"card_edit_capacity_{tab_key}_{art_norm}")

        cmeta7, cmeta8, cmeta9 = st.columns(3)
        meta_iso_pages = cmeta7.text_input("Ресурс, стр.", value=current_iso_pages, key=f"card_edit_iso_pages_{tab_key}_{art_norm}")
        meta_item_type = cmeta8.text_input("Тип", value=current_item_type, key=f"card_edit_item_type_{tab_key}_{art_norm}")
        meta_print_technology = cmeta9.text_input("Технология", value=current_print_technology, key=f"card_edit_print_technology_{tab_key}_{art_norm}")

        cmeta10, cmeta11, cmeta12, cmeta13 = st.columns(4)
        meta_weight = cmeta10.text_input("Вес", value=current_weight, key=f"card_edit_weight_{tab_key}_{art_norm}")
        meta_length = cmeta11.text_input("Длина", value=current_length, key=f"card_edit_length_{tab_key}_{art_norm}")
        meta_width = cmeta12.text_input("Ширина", value=current_width, key=f"card_edit_width_{tab_key}_{art_norm}")
        meta_height = cmeta13.text_input("Высота", value=current_height, key=f"card_edit_height_{tab_key}_{art_norm}")

        meta_fits = st.text_area("Подходит к моделям", value=current_fits, height=70, key=f"card_edit_fits_{tab_key}_{art_norm}")
        meta_description = st.text_area("Описание", value=current_description, height=65, key=f"card_edit_description_{tab_key}_{art_norm}")
        note = st.text_area("Заметка", value=current_note, height=65, key=f"card_edit_note_{tab_key}_{art_norm}")

        st.markdown("### 🔔 Напоминание / задача")
        st.caption("Можно создать задачу на пересмотр позиции через несколько дней. Задача сохранится отдельно и не пропадёт после загрузки нового файла.")
        t1, t2, t3 = st.columns([1.1, 1.4, 1.7])
        create_task_flag = t1.checkbox(
            "Создать задачу",
            key=f"card_edit_make_task_{tab_key}_{art_norm}",
            help="Создаёт напоминание по этой карточке с датой проверки и комментарием.",
        )
        task_due_date = t2.date_input(
            "Когда проверить",
            value=(datetime.utcnow().date() + timedelta(days=14)),
            key=f"card_edit_task_due_{tab_key}_{art_norm}",
        )
        task_reason = t3.selectbox(
            "Причина",
            ["Пересмотреть цену", "Проверить после правки", "Нет продаж", "Проверить фото/карточку", "Проверить спрос", "Другое"],
            key=f"card_edit_task_reason_{tab_key}_{art_norm}",
            help="Коротко описывает, зачем создана задача.",
        )
        task_note = st.text_area(
            "Комментарий к задаче",
            value="",
            height=65,
            key=f"card_edit_task_note_{tab_key}_{art_norm}",
            placeholder="Например: снизили цену, проверить продажи через 14 дней.",
        )

        b1, b2 = st.columns(2)
        save_clicked = b1.form_submit_button("💾 Сохранить карточку", use_container_width=True, type="primary")
        reset_clicked = b2.form_submit_button("↺ Сбросить ручные правки", use_container_width=True)

    if save_clicked:
        save_card_override(
            sheet_name,
            art,
            art_norm,
            {
                "photo_url": photo_url,
                "name_override": name_override,
                "meta_brand": meta_brand,
                "meta_model": meta_model,
                "meta_manufacturer_code": meta_code,
                "meta_print_type": meta_print_type,
                "meta_color": meta_color,
                "meta_capacity": meta_capacity,
                "meta_iso_pages": meta_iso_pages,
                "meta_item_type": meta_item_type,
                "meta_print_technology": meta_print_technology,
                "meta_description": meta_description,
                "meta_fits_models": meta_fits,
                "meta_weight": meta_weight,
                "meta_length": meta_length,
                "meta_width": meta_width,
                "meta_height": meta_height,
                "note": note,
            },
        )
        if create_task_flag:
            create_review_task(
                article=art,
                article_norm=art_norm,
                sheet_name=sheet_name,
                name_snapshot=name_override or current_name,
                due_date=task_due_date,
                reason=task_reason,
                note=task_note or note,
                source="card_editor",
            )
        st.success(f"Карточка {art} сохранена.")
        if create_task_flag:
            st.info(f"Задача по {art} создана до {task_due_date}.")
        st.rerun()

    if reset_clicked:
        delete_card_override(sheet_name, art_norm)
        st.success(f"Ручные правки для {art} сброшены.")
        st.rerun()

def ensure_photo_web_cache_table() -> None:
    path = get_photo_registry_path()
    path.parent.mkdir(parents=True, exist_ok=True)
    with sqlite3.connect(path) as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS photo_web_cache (
                article_norm TEXT PRIMARY KEY,
                article TEXT,
                photo_url TEXT,
                source_page TEXT,
                source_domain TEXT,
                status TEXT,
                checked_at TEXT
            )
            """
        )
        conn.commit()


def get_photo_web_cache(article_norm: str) -> dict[str, Any] | None:
    article_norm = normalize_article(article_norm)
    if not article_norm:
        return None
    ensure_photo_web_cache_table()
    with sqlite3.connect(get_photo_registry_path()) as conn:
        conn.row_factory = sqlite3.Row
        row = conn.execute(
            "SELECT * FROM photo_web_cache WHERE article_norm=?",
            (article_norm,),
        ).fetchone()
    return dict(row) if row else None


def save_photo_web_cache(article_norm: str, article: str, photo_url: str, source_page: str, source_domain: str, status: str) -> None:
    article_norm = normalize_article(article_norm)
    if not article_norm:
        return
    ensure_photo_web_cache_table()
    with sqlite3.connect(get_photo_registry_path()) as conn:
        conn.execute(
            """
            INSERT INTO photo_web_cache (article_norm, article, photo_url, source_page, source_domain, status, checked_at)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(article_norm) DO UPDATE SET
                article=excluded.article,
                photo_url=excluded.photo_url,
                source_page=excluded.source_page,
                source_domain=excluded.source_domain,
                status=excluded.status,
                checked_at=excluded.checked_at
            """,
            (article_norm, normalize_text(article), normalize_text(photo_url), normalize_text(source_page), normalize_text(source_domain), normalize_text(status), datetime.utcnow().isoformat(timespec="seconds")),
        )
        conn.commit()


def fetch_url_text(url: str, timeout: int = 12) -> str:
    if not normalize_text(url):
        return ""
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0 Safari/537.36",
        "Accept-Language": "ru,en;q=0.9",
    }
    try:
        if requests is not None:
            r = requests.get(url, headers=headers, timeout=timeout)
            if r.ok:
                r.encoding = r.encoding or r.apparent_encoding or "utf-8"
                return r.text
            return ""
    except Exception:
        return ""
    return ""


def extract_image_candidates_from_html(html_text: str, page_url: str, article_norm: str = "") -> list[str]:
    if not normalize_text(html_text):
        return []
    attrs = ["content", "src", "data-src", "data-large_image", "data-zoom-image", "data-original", "href"]
    patterns = []
    for attr in attrs:
        patterns.append(rf"{attr}=[\"']([^\"']+)[\"']")
    urls: list[str] = []
    for pat in patterns:
        for match in re.findall(pat, html_text, flags=re.IGNORECASE):
            url = normalize_text(match)
            if not url:
                continue
            abs_url = urljoin(page_url, url)
            low = abs_url.lower()
            if not any(ext in low for ext in [".jpg", ".jpeg", ".png", ".webp", ".gif"]) and not any(k in low for k in ["image", "images", "product", "uploads", "cache"]):
                continue
            if any(k in low for k in ["logo", "icon", "favicon", "sprite", "placeholder", "blank.gif", "1x1", "captcha"]):
                continue
            urls.append(abs_url)
    seen = set()
    scored: list[tuple[int, str]] = []
    for url in urls:
        if url in seen:
            continue
        seen.add(url)
        score = 0
        low = url.lower()
        if article_norm and article_norm.lower() in re.sub(r'[^a-z0-9]', '', low):
            score += 10
        if any(k in low for k in ["product", "goods", "item", "kartrid", "cartridge", "uploads"]):
            score += 3
        if low.endswith((".jpg", ".jpeg", ".png", ".webp")):
            score += 2
        scored.append((score, url))
    scored.sort(key=lambda x: (-x[0], x[1]))
    return [u for _, u in scored]


def discover_product_pages(article: str, domain: str) -> list[str]:
    article = normalize_text(article)
    if not article or requests is None:
        return []
    queries = [
        f"site:{domain} {article}",
        f'site:{domain} "{article}"',
    ]
    found: list[str] = []
    for q in queries:
        for search_url in [
            f"https://duckduckgo.com/html/?q={quote_plus(q)}",
            f"https://html.duckduckgo.com/html/?q={quote_plus(q)}",
        ]:
            html_text = fetch_url_text(search_url, timeout=10)
            if not html_text:
                continue
            for raw in re.findall(r"href=[\"']([^\"']+)[\"']", html_text, flags=re.IGNORECASE):
                href = html.unescape(raw)
                href = urljoin(search_url, href)
                if 'duckduckgo.com/l/?uddg=' in href:
                    m = re.search(r'uddg=([^&]+)', href)
                    if m:
                        href = unquote(m.group(1))
                if domain not in href:
                    continue
                if href not in found:
                    found.append(href)
                if len(found) >= FALLBACK_SEARCH_LIMIT:
                    return found
    return found


def discover_photo_url_for_article(article: str) -> tuple[str, str, str]:
    article_norm = normalize_article(article)
    cached = get_photo_web_cache(article_norm)
    if cached and normalize_text(cached.get("status", "")) in {"found", "not_found"}:
        return (normalize_text(cached.get("photo_url", "")), normalize_text(cached.get("source_page", "")), normalize_text(cached.get("source_domain", "")))

    if requests is None:
        save_photo_web_cache(article_norm, article, "", "", "", "not_found")
        return "", "", ""

    for domain in FALLBACK_PHOTO_DOMAINS:
        for page_url in discover_product_pages(article, domain):
            html_text = fetch_url_text(page_url, timeout=12)
            if not html_text:
                continue
            imgs = extract_image_candidates_from_html(html_text, page_url, article_norm=article_norm)
            if imgs:
                best = imgs[0]
                save_photo_web_cache(article_norm, article, best, page_url, domain, "found")
                return best, page_url, domain
    save_photo_web_cache(article_norm, article, "", "", "", "not_found")
    return "", "", ""


def inject_web_photos_into_registry(found_rows: list[dict[str, str]], import_name: str = "web-fallback") -> None:
    if not found_rows:
        return
    payload = pd.DataFrame(found_rows)
    for col in ["article", "article_norm", "photo_url", "source_sheet", "meta_color", "meta_iso_pages", "meta_manufacturer_code", "meta_model", "meta_fits_models"]:
        if col not in payload.columns:
            payload[col] = ""
    sync_photo_registry(payload[["article", "article_norm", "photo_url", "source_sheet", "meta_color", "meta_iso_pages", "meta_manufacturer_code", "meta_model", "meta_fits_models"]], import_name)


def try_fill_missing_photos(df: pd.DataFrame | None, enabled: bool = False, limit: int = 12) -> pd.DataFrame | None:
    if df is None or df.empty or not enabled:
        return df
    work = df.copy()
    missing = work[work.get("photo_url", pd.Series(dtype=object)).fillna("").map(lambda x: not bool(normalize_text(x)))].head(limit)
    if missing.empty:
        return work
    found_rows: list[dict[str, str]] = []
    article_to_url: dict[str, str] = {}
    for _, row in missing.iterrows():
        article = normalize_text(row.get("article", ""))
        article_norm = normalize_article(row.get("article_norm", article))
        if not article_norm:
            continue
        url, source_page, domain = discover_photo_url_for_article(article)
        if url:
            article_to_url[article_norm] = url
            found_rows.append({
                "article": article or article_norm,
                "article_norm": article_norm,
                "photo_url": url,
                "source_sheet": f"web:{domain}",
                "meta_color": normalize_text(row.get("meta_color", "")),
                "meta_iso_pages": normalize_text(row.get("meta_iso_pages", "")),
                "meta_manufacturer_code": normalize_text(row.get("meta_manufacturer_code", "")),
                "meta_model": normalize_text(row.get("meta_model", "")),
                "meta_fits_models": normalize_text(row.get("meta_fits_models", "")),
            })
    if found_rows:
        inject_web_photos_into_registry(found_rows)
        work["photo_url"] = work.apply(lambda r: article_to_url.get(normalize_article(r.get("article_norm", r.get("article", ""))), normalize_text(r.get("photo_url", ""))), axis=1)
        reg_df = load_photo_registry_df()
        if isinstance(reg_df, pd.DataFrame) and not reg_df.empty:
            st.session_state.photo_df = reg_df[[
                "article", "article_norm", "photo_url", "source_sheet",
                "meta_color", "meta_iso_pages", "meta_manufacturer_code",
                "meta_model", "meta_fits_models",
            ]].copy()
    return work

def init_state() -> None:
    defaults = {
        "comparison_sheets": {},
        "comparison_name": "ещё не загружен",
        "comparison_version": "",
        "selected_sheet": "Сравнение",
        "current_df": None,
        "photo_df": None,
        "photo_name": "ещё не загружен",
        "photo_registry_message": "",
        "photo_registry_stats": {},
        "photo_last_sync_sig": "",
        "avito_df": None,
        "avito_name": "ещё не загружен",
        "avito_registry_message": "",
        "avito_registry_stats": {},
        "avito_last_sync_sig": "",
        "hot_items_df": None,
        "hot_items_name": "ещё не загружен",
        "hot_items_last_sync_sig": "",
        "search_input": "",
        "submitted_query": "",
        "last_result": None,
        "price_mode": "-12%",
        "custom_discount": 10.0,
        "round100": True,
        "search_mode": "Умный",
        "template1_footer": DEFAULT_TEMPLATE1_FOOTER,
        "price_patch_input": "",
        "patch_message": "",
        "distributor_threshold": 35.0,
        "distributor_min_qty": 1.0,
        "operation_log": [],
        "app_mode_main": "Каталог",
        "crm_queue_filter": "Все",
        "crm_workspace_article_norm": "",
    }
    for key, value in defaults.items():
        if key not in st.session_state:
            st.session_state[key] = value


init_state()
ensure_photo_registry_loaded()
ensure_persisted_source_files_loaded()


def rebuild_current_df() -> None:
    sheets = st.session_state.get("comparison_sheets", {})
    selected = st.session_state.get("selected_sheet", "")
    photo_df = st.session_state.get("photo_df")
    base = sheets.get(selected)
    if isinstance(base, pd.DataFrame):
        st.session_state.current_df = apply_photo_map(base, photo_df)
    else:
        st.session_state.current_df = None


def refresh_all_search_results() -> None:
    sheets = st.session_state.get("comparison_sheets", {})
    search_mode = st.session_state.get("search_mode", "Умный")
    tab_specs = [
        ("Сравнение", "original"),
        ("Уценка", "discount"),
        ("Совместимые", "compatible"),
    ]
    for sheet_name, tab_key in tab_specs:
        base_df = sheets.get(sheet_name) if isinstance(sheets, dict) else None
        submitted_key = f"submitted_query_{tab_key}"
        result_key = f"last_result_{tab_key}"
        sig_key = f"last_result_sig_{tab_key}"
        query = normalize_text(st.session_state.get(submitted_key, ""))
        desired_sig = (query, search_mode, sheet_name, st.session_state.get("comparison_version", ""))
        if query and isinstance(base_df, pd.DataFrame):
            st.session_state[result_key] = search_in_df(base_df, query, search_mode, sheet_name=sheet_name)
            st.session_state[sig_key] = desired_sig
        else:
            st.session_state[result_key] = None
            st.session_state[sig_key] = None


def search_in_df(df: pd.DataFrame, query: str, search_mode: str, sheet_name: str = "") -> pd.DataFrame:
    tokens = split_query_parts(query)
    if not tokens:
        return df.iloc[0:0].copy()

    exact_hits = []
    linked_hits = []
    relaxed_hits = []
    contains_hits = []
    seen: set[str] = set()
    relaxed_sheet = sheet_name in {"Уценка", "Совместимые"}

    for token in tokens:
        token_norm = normalize_article(token)
        token_upper = contains_text(token)

        exact = df[df["article_norm"] == token_norm]
        for _, row in exact.iterrows():
            key = str(row["article_norm"])
            if key in seen:
                continue
            seen.add(key)
            row_dict = row.to_dict()
            row_dict["match_type"] = "exact"
            row_dict["match_query"] = token
            exact_hits.append(row_dict)

        if search_mode in {"Умный", "Артикул + коды из названия", "Артикул + название + бренд"} and token_norm:
            linked = df[df["row_codes"].apply(lambda codes: token_norm in (codes or []) if isinstance(codes, list) else False)]
            for _, row in linked.iterrows():
                key = str(row["article_norm"])
                if key in seen:
                    continue
                seen.add(key)
                row_dict = row.to_dict()
                row_dict["match_type"] = "linked"
                row_dict["match_query"] = token
                linked_hits.append(row_dict)

        # Для Уценки и Совместимых разрешаем более мягкий OEM-поиск:
        # пользователь может ввести оригинальный код, а мы найдём его внутри
        # префиксного артикула/названия (например BS-HPCE505A или CE505A-UCENKA).
        if relaxed_sheet and token_norm:
            relaxed_mask = (
                df["article_norm"].str.contains(re.escape(token_norm), na=False, regex=True)
                | df["search_blob_compact"].str.contains(re.escape(token_norm), na=False, regex=True)
            )
            relaxed = df[relaxed_mask]
            for _, row in relaxed.iterrows():
                key = str(row["article_norm"])
                if key in seen:
                    continue
                seen.add(key)
                row_dict = row.to_dict()
                row_dict["match_type"] = "relaxed"
                row_dict["match_query"] = token
                relaxed_hits.append(row_dict)

        if search_mode == "Артикул + название + бренд":
            mask = (
                df["search_blob"].str.contains(re.escape(token_upper), na=False, regex=True)
                | df["search_blob_compact"].str.contains(re.escape(token_norm), na=False, regex=True)
            )
            contains = df[mask]
            for _, row in contains.iterrows():
                key = str(row["article_norm"])
                if key in seen:
                    continue
                seen.add(key)
                row_dict = row.to_dict()
                row_dict["match_type"] = "contains"
                row_dict["match_query"] = token
                contains_hits.append(row_dict)

    rows = exact_hits + linked_hits + relaxed_hits + contains_hits
    if not rows:
        return df.iloc[0:0].copy()
    out = pd.DataFrame(rows)
    rank_map = {"exact": 0, "linked": 1, "relaxed": 2, "contains": 3}
    out["_rank"] = out["match_type"].map(lambda x: rank_map.get(str(x), 99))
    out = out.sort_values(["_rank", "article"]).drop(columns=["_rank"]).reset_index(drop=True)
    return out


def parse_price_updates(text: str) -> list[tuple[str, float]]:
    updates: list[tuple[str, float]] = []
    for line in text.splitlines():
        line = normalize_text(line)
        if not line:
            continue
        cleaned = line.replace("🔽", " ").replace("🔼", " ").replace("—", "-")
        m = re.search(r"([A-Za-zА-Яа-я0-9./_-]+)\s*-?\s*([0-9][0-9\s.,]*)", cleaned)
        if not m:
            continue
        article = normalize_article(m.group(1))
        price_txt = re.sub(r"[^0-9,\.]", "", m.group(2)).replace(",", ".")
        try:
            price = float(price_txt)
        except ValueError:
            continue
        if article:
            updates.append((article, price))
    return updates


def apply_price_updates(df: pd.DataFrame, updates_text: str) -> tuple[pd.DataFrame, str]:
    updates = parse_price_updates(updates_text)
    if not updates:
        return df, "Не нашёл строк для правки цен."
    out = df.copy()
    updated = 0
    missed: list[str] = []
    for article_norm, new_price in updates:
        mask = out["article_norm"] == article_norm
        if mask.any():
            out.loc[mask, "sale_price"] = float(new_price)
            updated += 1
        else:
            missed.append(article_norm)
    msg = f"Обновлено цен: {updated}"
    if missed:
        msg += " | Не найдено: " + ", ".join(missed[:10])
    return out, msg


def apply_price_updates_to_sheets(sheets: dict[str, pd.DataFrame], updates_text: str) -> tuple[dict[str, pd.DataFrame], str]:
    updates = parse_price_updates(updates_text)
    if not updates:
        return sheets, "Не нашёл строк для правки цен."
    if not isinstance(sheets, dict) or not sheets:
        return sheets, "Сначала загрузите comparison-файл."

    updated_sheets: dict[str, pd.DataFrame] = {}
    total_updated = 0
    hits_by_sheet: list[str] = []
    found_articles: set[str] = set()

    for sheet_name, df in sheets.items():
        if not isinstance(df, pd.DataFrame) or df.empty:
            updated_sheets[sheet_name] = df
            continue
        out = df.copy()
        sheet_hits = 0
        if "row_codes" not in out.columns:
            out["row_codes"] = out.apply(lambda row: build_row_compare_codes(row.get("article", ""), row.get("name", "")), axis=1)
        for article_norm, new_price in updates:
            mask = out["row_codes"].apply(lambda codes: article_norm in (codes or []) if isinstance(codes, list) else False)
            if mask.any():
                out.loc[mask, "sale_price"] = float(new_price)
                sheet_hits += int(mask.sum())
                found_articles.add(article_norm)
        updated_sheets[sheet_name] = out
        if sheet_hits:
            total_updated += sheet_hits
            hits_by_sheet.append(f"{sheet_name}: {sheet_hits}")

    missed = [article for article, _ in updates if article not in found_articles]
    msg = f"Обновлено цен: {total_updated}"
    if hits_by_sheet:
        msg += " | По листам: " + "; ".join(hits_by_sheet)
    if missed:
        msg += " | Не найдено: " + ", ".join(missed[:10])
    return updated_sheets, msg


def patch_comparison_workbook_bytes(file_bytes: bytes, updates_text: str) -> tuple[bytes | None, str]:
    updates = parse_price_updates(updates_text)
    if not updates:
        return None, "Не нашёл строк для правки цен."

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
    total_updated = 0
    hits_by_sheet: list[str] = []
    found_articles: set[str] = set()

    for ws in wb.worksheets:
        headers = [normalize_text(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]
        mapping = {key: find_column(headers, aliases) for key, aliases in CATALOG_COLUMN_ALIASES.items()}
        required = ["article", "name", "price"]
        if any(not mapping.get(key) for key in required):
            continue

        def _col_idx(header_value: str | None) -> int | None:
            if not header_value:
                return None
            for idx, header in enumerate(headers, start=1):
                if normalize_text(header) == normalize_text(header_value):
                    return idx
            return None

        article_idx = _col_idx(mapping.get("article"))
        name_idx = _col_idx(mapping.get("name"))
        price_idx = _col_idx(mapping.get("price"))
        if not article_idx or not name_idx or not price_idx:
            continue

        sheet_hits = 0
        for r in range(2, ws.max_row + 1):
            article = normalize_text(ws.cell(r, article_idx).value)
            name = normalize_text(ws.cell(r, name_idx).value)
            row_codes = build_row_compare_codes(article, name)
            matched_price = None
            for article_norm, new_price in updates:
                if article_norm in row_codes:
                    matched_price = float(new_price)
                    found_articles.add(article_norm)
                    break
            if matched_price is None:
                continue
            cell = ws.cell(r, price_idx)
            cell.value = int(matched_price) if float(matched_price).is_integer() else float(matched_price)
            sheet_hits += 1

        if sheet_hits:
            total_updated += sheet_hits
            hits_by_sheet.append(f"{ws.title}: {sheet_hits}")

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)

    missed = [article for article, _ in updates if article not in found_articles]
    msg = f"Обновлено цен: {total_updated}"
    if hits_by_sheet:
        msg += " | По листам: " + "; ".join(hits_by_sheet)
    if missed:
        msg += " | Не найдено: " + ", ".join(missed[:10])
    return bio.read(), msg


def get_source_pairs(df: pd.DataFrame) -> list[dict[str, str]]:
    if df is None or df.empty:
        return []
    pairs = df["source_pairs"].iloc[0]
    if isinstance(pairs, list):
        return pairs
    return []


def get_row_offers(row: pd.Series, min_qty: float = 1.0) -> list[dict[str, Any]]:
    offers: list[dict[str, Any]] = []
    hidden_compatible_sources: list[str] = []
    for pair in row.get("source_pairs", []) or []:
        source = pair["source"]
        price = safe_float(row.get(pair["price_col"]), 0.0)
        price = normalize_merlion_source_price(row, source, price)
        qty = parse_qty_generic(row.get(pair["qty_col"]))
        if price <= 0 or qty < float(min_qty):
            continue
        if is_blocked_by_compatible_price(row, source, price):
            hidden_compatible_sources.append(f"{source} {fmt_price(price)}")
            continue
        offers.append({
            "source": source,
            "price": price,
            "qty": qty,
            "price_fmt": fmt_price(price),
            "qty_fmt": fmt_qty(qty),
        })

    offers, hidden_outlier_sources = filter_suspicious_low_offers(row, offers)

    try:
        row["hidden_compatible_sources"] = unique_preserve_order(hidden_compatible_sources)
        row["hidden_outlier_sources"] = unique_preserve_order(hidden_outlier_sources)
    except Exception:
        pass

    offers.sort(key=lambda x: (x["price"], -x["qty"], x["source"]))
    return offers


def get_best_offer(row: pd.Series, min_qty: float = 1.0) -> dict[str, Any] | None:
    own_price = safe_float(row.get("sale_price"), 0.0)
    offers = get_row_offers(row, min_qty=min_qty)
    if not offers:
        return None
    best = offers[0]
    best = dict(best)
    if own_price > 0:
        delta = own_price - best["price"]
        best["delta"] = delta
        best["delta_fmt"] = fmt_price(abs(delta))
        best["delta_percent"] = (delta / own_price) * 100.0 if own_price else 0.0
        best["delta_percent_fmt"] = f"{best['delta_percent']:.1f}".replace(".0", "")
        if abs(delta) < 1e-9:
            best["status"] = "цена равна"
        elif delta > 0:
            best["status"] = "лучше нас"
        else:
            best["status"] = "дороже нас"
    else:
        best["status"] = "найдено"
    return best


def get_best_offer_if_cheaper(row: pd.Series | dict[str, Any], min_qty: float = 1.0) -> dict[str, Any] | None:
    best = get_best_offer(row, min_qty=min_qty)
    if not best:
        return None
    if safe_float(best.get("delta"), 0.0) <= 0:
        return None
    return best


def get_best_offer_if_profitable(row: pd.Series | dict[str, Any], min_qty: float = 1.0, threshold_pct: float = 35.0) -> dict[str, Any] | None:
    best = get_best_offer_if_cheaper(row, min_qty=min_qty)
    if not best:
        return None
    if safe_float(best.get("delta_percent"), 0.0) < float(threshold_pct):
        return None
    return best


def build_distributor_compare(result_df: pd.DataFrame, min_qty: float = 1.0) -> dict[str, dict[str, Any]]:
    out: dict[str, dict[str, Any]] = {}
    if result_df is None or result_df.empty:
        return out
    for _, row in result_df.iterrows():
        row_key = str(row.get("article_norm", ""))
        out[row_key] = {
            "row_key": row_key,
            "article": row.get("article", ""),
            "name": row.get("name", ""),
            "sale_price": safe_float(row.get("sale_price"), 0.0),
            "best_offer": get_best_offer(row, min_qty=min_qty),
        }
    return out


def build_all_prices_df(result_df: pd.DataFrame, min_qty: float, price_mode: str, round100: bool, custom_discount: float) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    if result_df is None or result_df.empty:
        return pd.DataFrame()
    for _, row in result_df.iterrows():
        article = str(row.get("article", ""))
        name = str(row.get("name", ""))
        own_price = safe_float(row.get("sale_price"), 0.0)
        own_qty = safe_float(row.get("free_qty"), 0.0)
        selected_price = get_selected_price_raw(row, price_mode, round100, custom_discount)

        rows.append({
            "Артикул": article,
            "Название": name,
            "Источник": "Мы",
            "Цена": own_price,
            "Остаток": own_qty,
            "Наша цена": own_price,
            "Наша цена выбранная": selected_price,
            "Разница к нам, руб": 0.0,
            "Разница к нам, %": 0.0,
            "Статус": "наша позиция",
        })

        for offer in get_row_offers(row, min_qty=min_qty):
            delta = own_price - offer["price"]
            delta_pct = (delta / own_price) * 100.0 if own_price else None
            status = "лучше нас" if delta > 0 else "дороже нас" if delta < 0 else "цена равна"
            rows.append({
                "Артикул": article,
                "Название": name,
                "Источник": offer["source"],
                "Цена": offer["price"],
                "Остаток": offer["qty"],
                "Наша цена": own_price,
                "Наша цена выбранная": selected_price,
                "Разница к нам, руб": delta,
                "Разница к нам, %": delta_pct,
                "Статус": status,
            })
    out = pd.DataFrame(rows)
    out["_is_own"] = out["Источник"].map(lambda x: 0 if str(x) == "Мы" else 1)
    out = out.sort_values(["Артикул", "_is_own", "Цена", "Источник"], ascending=[True, True, True, True], na_position="last").drop(columns=["_is_own"]).reset_index(drop=True)
    return out


def all_prices_to_excel_bytes(df: pd.DataFrame) -> bytes:
    bio = io.BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Все цены")
    bio.seek(0)
    return bio.read()



def translate_watch_action(action: Any, threshold_pct: float = 35.0) -> str:
    raw = normalize_text(action)
    if not raw:
        return ""
    tokens = [t.strip().upper() for t in raw.replace("|", ";").split(";") if t.strip()]
    translated = []
    for token in tokens:
        if token == "BUY":
            translated.append(f"Можно брать (-{int(threshold_pct)}%+)")
        elif token == "RESTOCK":
            translated.append("Пополнить запас")
        elif token == "WATCH":
            translated.append("Наблюдать")
        elif token in {"NO_MATCH", "NO_MATCH_IN_COMPARISON"}:
            translated.append("Нет в сравнении")
        else:
            translated.append(token)
    uniq = []
    for item in translated:
        if item not in uniq:
            uniq.append(item)
    return "; ".join(uniq)

def build_report_df(
    df: pd.DataFrame,
    threshold_percent: float,
    min_qty: float,
    tab_label: str = "",
    hot_lookup: dict[str, list[dict[str, Any]]] | None = None,
) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    if df is None or df.empty:
        return pd.DataFrame()

    hot_lookup = hot_lookup or {}

    for _, row in df.iterrows():
        own_price = safe_float(row.get("sale_price"), 0.0)
        own_qty = safe_float(row.get("free_qty"), 0.0)
        if own_price <= 0:
            continue

        hot_rec = pick_hot_watch_rec(row, hot_lookup) if hot_lookup else None
        best = get_best_offer(row, min_qty=min_qty)
        profitable_best = get_best_offer_if_profitable(row, min_qty=min_qty, threshold_pct=float(threshold_percent))

        best_source = ""
        best_price = None
        best_qty = None
        delta = None
        delta_pct = None
        profitable_offer = False

        if profitable_best:
            best_source = normalize_text(profitable_best.get("source", ""))
            best_price_val = safe_float(profitable_best.get("price"), 0.0)
            best_qty_val = safe_float(profitable_best.get("qty"), 0.0)
            delta_val = safe_float(profitable_best.get("delta"), 0.0)
            delta_pct_val = safe_float(profitable_best.get("delta_percent"), 0.0)

            if best_price_val > 0:
                best_price = best_price_val
            if best_qty_val >= 0:
                best_qty = best_qty_val
            delta = delta_val
            delta_pct = round(delta_pct_val, 2)

            profitable_offer = True

        # Управленческий отчёт:
        # строка попадает в отчёт, если она есть в watchlist
        # ИЛИ если поставщик реально выгоднее нас на нужный порог.
        if not hot_rec and not profitable_offer:
            continue

        action_text = ""
        if hot_rec:
            action_text, _ = hot_supplier_note(row, profitable_best, threshold_pct=float(threshold_percent))
        elif profitable_offer:
            action_text = f"Сейчас можно брать у {best_source}" if best_source else "Сейчас можно брать"

        rows.append({
            "Лист": tab_label,
            "Артикул": normalize_text(row.get("article", "")),
            "Товар": normalize_text(row.get("name", "")),
            "Спрос, шт/мес": safe_float((hot_rec or {}).get("sales_per_month"), 0.0) if hot_rec else None,
            "Наша цена": own_price,
            "Наш остаток": own_qty,
            "Лучший поставщик": best_source,
            "Цена поставщика": best_price,
            "Остаток поставщика": best_qty,
            "Ниже нашей цены, %": delta_pct,
            "Дней запаса": safe_float((hot_rec or {}).get("days_of_cover"), 0.0) if hot_rec else None,
            "Приоритет": safe_float((hot_rec or {}).get("priority_score"), 0.0) if hot_rec else None,
            "Действие": action_text,
            "Разница, руб": delta,
            "Ходовая": "Да" if hot_rec else "",
        })

    if not rows:
        return pd.DataFrame()

    out = pd.DataFrame(rows)
    out["_sort_hot"] = out["Ходовая"].astype(str).eq("Да").astype(int)
    out["_sort_priority"] = pd.to_numeric(out.get("Приоритет"), errors="coerce").fillna(-1.0)
    out["_sort_gap"] = pd.to_numeric(out.get("Ниже нашей цены, %"), errors="coerce").fillna(-1.0)

    out = out.sort_values(
        ["_sort_hot", "_sort_priority", "_sort_gap", "Артикул"],
        ascending=[False, False, False, True],
        kind="stable",
    ).drop(columns=["_sort_hot", "_sort_priority", "_sort_gap"]).reset_index(drop=True)

    preferred_columns = [
        "Лист",
        "Артикул",
        "Товар",
        "Спрос, шт/мес",
        "Наша цена",
        "Наш остаток",
        "Лучший поставщик",
        "Цена поставщика",
        "Остаток поставщика",
        "Ниже нашей цены, %",
        "Дней запаса",
        "Приоритет",
        "Действие",
        "Разница, руб",
        "Ходовая",
    ]
    ordered_columns = [col for col in preferred_columns if col in out.columns] + [col for col in out.columns if col not in preferred_columns]
    out = out[ordered_columns]

    return out

def report_to_excel_bytes(df: pd.DataFrame) -> bytes:
    bio = io.BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Отчёт")
    bio.seek(0)
    return bio.read()

def build_product_analysis_df(result_df: pd.DataFrame, min_qty: float = 1.0) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    if result_df is None or result_df.empty:
        return pd.DataFrame()

    enriched_df = apply_purchase_cost_map(result_df, st.session_state.get("purchase_cost_df")) if isinstance(result_df, pd.DataFrame) else result_df
    if not isinstance(enriched_df, pd.DataFrame) or enriched_df.empty:
        return pd.DataFrame()

    seen: set[str] = set()
    for _, row in enriched_df.iterrows():
        row_key = str(row.get("article_norm") or normalize_article(row.get("article", "")))
        if row_key in seen:
            continue
        seen.add(row_key)

        best_offer = get_best_offer_if_cheaper(row, min_qty=min_qty)
        purchase_avg_cost = safe_float(row.get("purchase_avg_cost"), 0.0)
        rows.append({
            "Артикул": str(row.get("article", "") or ""),
            "Название": str(row.get("name", "") or ""),
            "КОЛ.": safe_float(row.get("free_qty", 0), 0.0),
            "тек прод": safe_float(row.get("sale_price", 0), 0.0),
            "дистр": safe_float(best_offer.get("price", 0), 0.0) if best_offer else None,
            "Дистрибьютор": str(best_offer.get("source", "") or "") if best_offer else "",
            "Остаток дистрибьютора": safe_float(best_offer.get("qty", 0), 0.0) if best_offer else None,
            "сред. Зак.": purchase_avg_cost if purchase_avg_cost > 0 else None,
            "Источник закупки": normalize_text(row.get("purchase_match_source", "")),
            "Название закупки": normalize_text(row.get("purchase_source_name", "")),
            "Лист закупки": normalize_text(row.get("purchase_source_sheet", "")),
        })

    return pd.DataFrame(rows)


def build_product_analysis_workbook_bytes(result_df: pd.DataFrame, min_qty: float = 1.0) -> bytes:
    analysis_df = build_product_analysis_df(result_df, min_qty=min_qty)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Анализ товара"

    headers = [
        "Артикул", "", "КОЛ.", "тек прод", "дистр", "МИ", "ВЦМ", "Ятовары", "Мы на авито",
        "авито мин", "сред. Зак.", "Прод пред", "пред на Авито", "", "% прод", "% Авито"
    ]
    ws.append(headers)

    column_widths = {
        "A": 14, "B": 4, "C": 10, "D": 12, "E": 12, "F": 10, "G": 10, "H": 12,
        "I": 13, "J": 12, "K": 12, "L": 12, "M": 14, "N": 4, "O": 10, "P": 10,
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    header_fill = openpyxl.styles.PatternFill(fill_type="solid", fgColor="D9E2F3")
    thin_gray = openpyxl.styles.Side(style="thin", color="D0D7E2")
    border = openpyxl.styles.Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)
    header_font = openpyxl.styles.Font(bold=True)
    center = openpyxl.styles.Alignment(horizontal="center", vertical="center")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center

    currency_format = '#,##0.00'
    percent_format = '0.00%'

    for excel_row, rec in enumerate(analysis_df.to_dict(orient="records"), start=2):
        ws.cell(excel_row, 1).value = rec.get("Артикул", "")
        ws.cell(excel_row, 3).value = rec.get("КОЛ.", None)
        ws.cell(excel_row, 4).value = rec.get("тек прод", None)
        ws.cell(excel_row, 5).value = rec.get("дистр", None)
        ws.cell(excel_row, 6).value = None
        ws.cell(excel_row, 7).value = None
        ws.cell(excel_row, 8).value = None
        ws.cell(excel_row, 9).value = None
        ws.cell(excel_row, 10).value = None
        ws.cell(excel_row, 11).value = rec.get("сред. Зак.", None)
        ws.cell(excel_row, 12).value = f'=IF(E{excel_row}="","",E{excel_row}-E{excel_row}*5%)'
        ws.cell(excel_row, 13).value = f'=IF(L{excel_row}="","",L{excel_row}-L{excel_row}*20%)'
        ws.cell(excel_row, 15).value = f'=IF(OR(K{excel_row}="",K{excel_row}=0,L{excel_row}=""),"",L{excel_row}/K{excel_row}-1)'
        ws.cell(excel_row, 16).value = f'=IF(OR(K{excel_row}="",K{excel_row}=0,M{excel_row}=""),"",M{excel_row}/K{excel_row}-1)'

        if rec.get("дистр") not in (None, ""):
            comment_lines = []
            dist_name = normalize_text(rec.get("Дистрибьютор", ""))
            if dist_name:
                comment_lines.append(f"Лучшее предложение: {dist_name}")
            dist_qty = rec.get("Остаток дистрибьютора")
            if dist_qty not in (None, ""):
                comment_lines.append(f"Остаток: {fmt_qty(dist_qty)} шт.")
            if comment_lines:
                ws.cell(excel_row, 5).comment = openpyxl.comments.Comment("\n".join(comment_lines), "ChatGPT")

        purchase_cost = rec.get("сред. Зак.")
        if purchase_cost not in (None, ""):
            purchase_comment_lines = [f"Средняя закупка: {fmt_price(purchase_cost)}"]
            purchase_source = normalize_text(rec.get("Источник закупки", ""))
            if purchase_source:
                purchase_comment_lines.append(f"Источник маппинга: {purchase_source}")
            purchase_name = normalize_text(rec.get("Название закупки", ""))
            if purchase_name:
                purchase_comment_lines.append(f"Номенклатура: {purchase_name}")
            purchase_sheet = normalize_text(rec.get("Лист закупки", ""))
            if purchase_sheet:
                purchase_comment_lines.append(f"Лист: {purchase_sheet}")
            ws.cell(excel_row, 11).comment = openpyxl.comments.Comment("\n".join(purchase_comment_lines), "ChatGPT")

        for col_idx in [4, 5, 6, 7, 8, 9, 10, 11, 12, 13]:
            ws.cell(excel_row, col_idx).number_format = currency_format
        for col_idx in [15, 16]:
            ws.cell(excel_row, col_idx).number_format = percent_format

    max_row = max(ws.max_row, 2)
    for row in ws.iter_rows(min_row=2, max_row=max_row, min_col=1, max_col=16):
        for cell in row:
            cell.border = border
            if cell.column in (3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 15, 16):
                cell.alignment = center

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:P{max_row}"

    info = wb.create_sheet("Справка")
    info["A1"] = "Как читать файл"
    info["A1"].font = openpyxl.styles.Font(bold=True, size=12)
    info["A3"] = "Артикул / КОЛ. / тек прод"
    info["B3"] = "Заполняются автоматически из результата поиска и текущего листа comparison-файла."
    info["A4"] = "дистр"
    info["B4"] = "Подставляется лучшая валидная цена поставщика. В комментарии к ячейке есть поставщик и остаток."
    info["A5"] = "МИ / ВЦМ / Ятовары / Мы на авито / авито мин / сред. Зак."
    info["B5"] = "Сред. Зак. теперь подставляется автоматически из загруженного файла средней закупки, если найден безопасный матч. Остальные поля этой группы можно дозаполнять вручную перед обсуждением."
    info["A6"] = "Прод пред"
    info["B6"] = "Считается как дистр - 5%."
    info["A7"] = "пред на Авито"
    info["B7"] = "Считается как Прод пред - 20%."
    info["A8"] = "% прод / % Авито"
    info["B8"] = "Считаются относительно среднего закупа."
    info.column_dimensions["A"].width = 26
    info.column_dimensions["B"].width = 90
    info.freeze_panes = "A3"

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.read()



def build_offer_template(df: pd.DataFrame, query: str, round100: bool, footer_text: str, search_mode: str) -> str:
    result_df = search_in_df(df, query, search_mode, sheet_name=st.session_state.get("selected_sheet", ""))
    if result_df.empty:
        return ""
    lines: list[str] = []
    hashtags: list[str] = []
    for _, row in result_df.iterrows():
        article_head = compose_article_template_label(row)
        if safe_float(row.get("free_qty"), 0.0) > 0:
            avito_raw = safe_float(row.get("sale_price"), 0.0) * (1 - DEFAULT_DISCOUNT_1 / 100)
            cash_raw = avito_raw * 0.90
            avito = round_up_to_100(avito_raw) if round100 else round(avito_raw)
            cash = round_to_nearest_100(cash_raw) if round100 else round(cash_raw)
            lines.append(f"{article_head} --- {fmt_price(avito)} руб. - Авито / {fmt_price(cash)} руб. за наличный расчет")
        else:
            lines.append(f"{article_head} --- продан")
        hashtags.append(f"#{normalize_article(row['article'])}")

    shared_lines = build_template_shared_lines(result_df)
    footer = [normalize_text(x) for x in str(footer_text).splitlines() if normalize_text(x)]
    out_lines: list[str] = []
    out_lines.extend(lines)
    if shared_lines:
        out_lines.append("")
        out_lines.extend(shared_lines)
    if footer:
        out_lines.extend(footer)
    if hashtags:
        out_lines.append(",".join(unique_preserve_order(hashtags)))
    return "\n".join(out_lines)


def build_selected_price_template(df: pd.DataFrame, query: str, price_mode: str, round100: bool, custom_discount: float, search_mode: str) -> str:
    result_df = search_in_df(df, query, search_mode, sheet_name=st.session_state.get("selected_sheet", ""))
    if result_df.empty:
        return ""
    parts = []
    for _, row in result_df.iterrows():
        if safe_float(row.get("free_qty"), 0.0) <= 0:
            continue
        selected_price = get_selected_price_raw(row, price_mode, round100, custom_discount)
        parts.append(f"{normalize_text(row['name'])} --- {fmt_price(selected_price)} руб.")
    return "\n\n".join(parts)




def build_offer_template_from_result_df(result_df: pd.DataFrame, round100: bool, footer_text: str) -> str:
    if result_df is None or result_df.empty:
        return ""
    lines: list[str] = []
    hashtags: list[str] = []
    for _, row in result_df.iterrows():
        article_head = compose_article_template_label(row)
        if safe_float(row.get("free_qty"), 0.0) > 0:
            avito_raw = safe_float(row.get("sale_price"), 0.0) * (1 - DEFAULT_DISCOUNT_1 / 100)
            cash_raw = avito_raw * 0.90
            avito = round_up_to_100(avito_raw) if round100 else round(avito_raw)
            cash = round_to_nearest_100(cash_raw) if round100 else round(cash_raw)
            lines.append(f"{article_head} --- {fmt_price(avito)} руб. - Авито / {fmt_price(cash)} руб. за наличный расчет")
        else:
            lines.append(f"{article_head} --- продан")
        hashtags.append(f"#{normalize_article(row['article'])}")

    shared_lines = build_template_shared_lines(result_df)
    footer = [normalize_text(x) for x in str(footer_text).splitlines() if normalize_text(x)]
    out_lines: list[str] = []
    out_lines.extend(lines)
    if shared_lines:
        out_lines.append("")
        out_lines.extend(shared_lines)
    if footer:
        out_lines.extend(footer)
    if hashtags:
        out_lines.append(",".join(unique_preserve_order(hashtags)))
    return "\n".join(out_lines)


def build_selected_price_template_from_result_df(result_df: pd.DataFrame, price_mode: str, round100: bool, custom_discount: float) -> str:
    if result_df is None or result_df.empty:
        return ""
    parts: list[str] = []
    for _, row in result_df.iterrows():
        if safe_float(row.get("free_qty"), 0.0) <= 0:
            continue
        selected_price = get_selected_price_raw(row, price_mode, round100, custom_discount)
        parts.append(f"{normalize_text(row['name'])} --- {fmt_price(selected_price)} руб.")
    return "\n\n".join(parts)


def find_avito_ads(avito_df: pd.DataFrame, result_df: pd.DataFrame) -> pd.DataFrame:
    registry_df = load_avito_registry_df()
    if (avito_df is None or avito_df.empty) and registry_df.empty:
        return pd.DataFrame()
    if result_df is None or result_df.empty:
        return pd.DataFrame()

    tokens: list[str] = []
    raw_tokens: list[str] = []
    for _, row in result_df.iterrows():
        row_codes = row.get("row_codes", [])
        if isinstance(row_codes, list) and row_codes:
            tokens.extend(unique_norm_codes(row_codes))
        else:
            tokens.extend(build_row_compare_codes(row.get("article", ""), row.get("name", "")))
        article_raw = normalize_text(row.get("article", ""))
        if article_raw:
            raw_tokens.append(article_raw)
        for code in extract_article_candidates_from_text(row.get("name", "")):
            raw_tokens.append(code)
    tokens = unique_preserve_order([normalize_article(x) for x in tokens if normalize_article(x)])
    raw_tokens = unique_preserve_order([normalize_article(x) for x in raw_tokens if normalize_article(x)])
    if not tokens and not raw_tokens:
        return pd.DataFrame()

    base_df = avito_df.copy() if isinstance(avito_df, pd.DataFrame) and not avito_df.empty else registry_df.copy()
    if base_df.empty:
        return pd.DataFrame()
    if "title_norm" not in base_df.columns:
        base_df["title_norm"] = base_df["title"].map(contains_text)
    if "title_codes" not in base_df.columns:
        base_df["title_codes"] = base_df["title"].map(extract_article_candidates_from_text)
    if "registry_key" not in base_df.columns:
        base_df["registry_key"] = base_df.apply(build_avito_registry_key, axis=1)

    matches = []
    for _, row in base_df.iterrows():
        title_codes = unique_norm_codes(row.get("title_codes", []) if isinstance(row.get("title_codes", []), list) else extract_article_candidates_from_text(row.get("title", "")))
        title_compact = compact_text(row.get("title", ""))
        exact_hits = [t for t in tokens if t in title_codes]
        substring_hits = [t for t in tokens if t and t in title_compact and t not in exact_hits]
        raw_hits = [t for t in raw_tokens if t and t in title_compact and t not in exact_hits and t not in substring_hits]
        if exact_hits or substring_hits or raw_hits:
            item = row.to_dict()
            item["matched_tokens"] = ", ".join(unique_preserve_order(exact_hits + substring_hits + raw_hits))
            item["match_score"] = len(exact_hits) * 100 + len(substring_hits) * 10 + len(raw_hits)
            item["match_kind"] = "точное" if exact_hits else ("связанное" if substring_hits else "по названию")
            matches.append(item)
    if not matches:
        return pd.DataFrame()

    out = pd.DataFrame(matches)
    out = out.sort_values(["match_score", "registry_key"], ascending=[False, True]).drop_duplicates(subset=["registry_key"], keep="first").reset_index(drop=True)
    if not registry_df.empty:
        reg = registry_df.copy()
        if "registry_key" not in reg.columns:
            reg["registry_key"] = reg.apply(build_avito_registry_key, axis=1)
        reg_cols = [c for c in ["registry_key", "first_seen", "last_seen", "last_changed_at", "previous_price_raw", "change_count", "status", "account", "last_import_name"] if c in reg.columns]
        out = out.merge(reg[reg_cols], on="registry_key", how="left", suffixes=("", "_reg"))
        if "account_reg" in out.columns:
            out["account"] = out["account"].where(out["account"].astype(str).str.len() > 0, out["account_reg"])
            out = out.drop(columns=["account_reg"])
    return out


def render_sidebar_card_header(title: str, icon: str = "📁", help_text: str = "") -> None:
    tooltip_html = ""
    if normalize_text(help_text):
        tooltip_html = (
            '<div class="sidebar-card-help-wrap">'
            '<div class="sidebar-card-help">?</div>'
            f'<div class="sidebar-card-tooltip">{html.escape(help_text)}</div>'
            '</div>'
        )
    st.markdown(
        f"""
        <div class="sidebar-card-header">
          <div class="sidebar-card-header-main">
            <div class="sidebar-card-icon">{html.escape(icon)}</div>
            <div class="sidebar-card-title-wrap">
              <div class="sidebar-card-kicker">Быстрый доступ</div>
              <div class="sidebar-card-title">{html.escape(title)}</div>
            </div>
          </div>
          {tooltip_html}
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_block_header(title: str, subtitle: str = "", icon: str = "📦", help_text: str = "") -> None:
    tooltip_html = ""
    if normalize_text(help_text):
        tooltip_html = (
            '<div class="block-help-wrap">'
            '<div class="block-help">?</div>'
            f'<div class="block-tooltip">{html.escape(help_text)}</div>'
            '</div>'
        )
    st.markdown(
        f"""
        <div class="block-header">
          <div class="block-header-main">
            <div class="block-icon">{html.escape(icon)}</div>
            <div class="block-title-wrap">
              <div class="block-kicker">Раздел интерфейса</div>
              <div class="section-title">{html.escape(title)}</div>
              <div class="section-sub">{html.escape(subtitle)}</div>
            </div>
          </div>
          <div class="block-header-right">
            <div class="block-sparkles">✦ ✦ ✦</div>
            {tooltip_html}
          </div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_info_banner(title: str, text: str, icon: str = "💡", chips: Optional[list[str]] = None, tone: str = "blue") -> None:
    chips_html = ""
    if chips:
        chips_html = "<div class='banner-chip-row'>" + "".join(
            f"<span class='banner-chip'>{html.escape(chip)}</span>" for chip in chips if normalize_text(chip)
        ) + "</div>"
    st.markdown(
        f"""
        <div class="info-banner tone-{html.escape(tone)}">
          <div class="info-banner-icon">{html.escape(icon)}</div>
          <div class="info-banner-body">
            <div class="info-banner-title">{html.escape(title)}</div>
            <div class="info-banner-text">{html.escape(text)}</div>
            {chips_html}
          </div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def status_visual_class(status: str) -> str:
    status_text = contains_text(status)
    if "ЛУЧШЕ" in status_text:
        return "offer-good"
    if "ДОРОЖЕ" in status_text:
        return "offer-bad"
    if "РАВНА" in status_text:
        return "offer-neutral"
    if "НАША ПОЗИЦИЯ" in status_text:
        return "offer-own"
    return "offer-muted"



def render_results_table(df: pd.DataFrame, price_mode: str, round100: bool, custom_discount: float, distributor_map: Optional[dict[str, dict[str, Any]]] = None, show_photos: bool = True) -> None:
    selected_label = current_price_label(price_mode, custom_discount)
    distributor_map = distributor_map or {}
    rows_html = []

    def state_label_from_class(status_class: str) -> str:
        return {
            "offer-good": "выгоднее",
            "offer-bad": "дороже",
            "offer-neutral": "цена равна",
            "offer-own": "наша позиция",
        }.get(status_class, "найдено")

    for _, row in df.iterrows():
        row_key = str(row.get("article_norm", ""))
        selected_raw = get_selected_price_raw(row, price_mode, round100, custom_discount)
        best = distributor_map.get(row_key, {}).get("best_offer") if row_key in distributor_map else None
        match_type = str(row.get("match_type", ""))

        if match_type == "exact":
            badge_html = "<div class='match-badge match-badge-exact'>Точное совпадение</div>"
        elif match_type == "linked":
            badge_html = "<div class='match-badge match-badge-linked'>Код из названия</div>"
        else:
            badge_html = "<div class='match-badge match-badge-soft'>По названию / бренду</div>"

        hot_html = ""
        manual_note_html = ""
        manual_note = normalize_text(row.get("manual_note", ""))
        if manual_note:
            manual_note_html = f"<div class='manual-note'>{html.escape(manual_note)}</div>"
        if bool(row.get("hot_flag", False)):
            abc = normalize_text(row.get("hot_abc_class", "")).upper()
            sales_pm = safe_float(row.get("hot_sales_per_month"), 0.0)
            action_today = normalize_text(row.get("hot_action_today", "")).upper()
            buy_signal = normalize_text(row.get("hot_buy_signal", "")).upper()

            demand_map = {
                "A": "Спрос высокий",
                "B": "Спрос хороший",
                "C": "Спрос умеренный",
            }
            primary = "🔥 Товар ходовой"
            note_parts = []
            if abc:
                note_parts.append(demand_map.get(abc, f"Класс {abc}"))
            if sales_pm > 0:
                note_parts.append(f"≈ {sales_pm:.1f} шт/мес")
            note = " • ".join(note_parts)

            action_text, _hot_help = hot_supplier_note(row, best, threshold_pct=35.0)
            action_html = f"<div class='hot-sub-badge'>{html.escape(action_text)}</div>" if action_text else ""
            hot_html = f"<div class='hot-badge'>{html.escape(primary)}</div>{('<div class="hot-meta">' + html.escape(note) + '</div>') if note else ''}{action_html}"

        if best:
            status_class = status_visual_class(str(best.get("status", "")))
            state_label = state_label_from_class(status_class)
            pct_txt = str(best.get("delta_percent_fmt", "")).strip()
            pct_html = f"-{html.escape(pct_txt)}%" if pct_txt else "—"
            compare_html = f"""
            <div class='best-box {status_class}'>
              <div class='best-top'>
                <span class='best-source-pill'>{html.escape(str(best.get('source', '')))}</span>
                <span class='best-state-pill'>{html.escape(state_label)}</span>
              </div>
              <div class='best-price'>{html.escape(str(best.get('price_fmt', '')))} руб.</div>
              <div class='best-meta-row'>Остаток: <b>{html.escape(str(best.get('qty_fmt', '')))}</b></div>
              <div class='best-delta-row'>Разница к нам: {html.escape(str(best.get('delta_fmt', '')))} руб. • {pct_html}</div>
            </div>
            """
        else:
            compare_html = """
            <div class='best-box best-box-empty'>
              <div class='best-empty-title'>Нет цены лучше</div>
            </div>
            """

        photo_url = normalize_text(row.get("photo_url", "")) if show_photos else ""
        if show_photos and photo_url:
            photo_html = f"""
            <a href="{html.escape(photo_url, quote=True)}" target="_blank" class="photo-wrap">
              <img src="{html.escape(photo_url, quote=True)}" class="result-photo" loading="lazy" onerror="this.style.display='none'; this.parentNode.innerHTML='<div class=&quot;photo-empty photo-empty-small&quot;>нет фото</div>';">
            </a>
            """
        elif show_photos:
            photo_html = "<div class='photo-empty photo-empty-small'>нет фото</div>"
        else:
            photo_html = ""

        item_photo_html = f"<div class='item-photo'>{photo_html}</div>" if show_photos else ""

        free_qty = safe_float(row.get("free_qty"), 0.0)
        total_qty = safe_float(row.get("total_qty"), free_qty)
        transit_qty = safe_float(row.get("transit_qty"), 0.0)
        stock_html = f"""
        <div class='stock-main'>{fmt_qty(free_qty)}</div>
        <div class='stock-sub'>Свободно: {fmt_qty(free_qty)}</div>
        <div class='stock-sub'>Всего: {fmt_qty(total_qty)}</div>
        <div class='stock-sub'>Транзит: {fmt_qty(transit_qty)}</div>
        """

        rows_html.append(
            f"""
            <tr>
              <td class='item-col'>
                <div class='item-wrap'>
                  {item_photo_html}
                  <div class='item-main'>
                    <div class='item-top'><span class='article-pill'>{html.escape(str(row['article']))}</span></div>
                    <div class='name-cell'>{html.escape(str(row['name']))}</div>
                    {badge_html}
                    {hot_html}
                    {manual_note_html}
                  </div>
                </div>
              </td>
              <td class='stock-cell'>{stock_html}</td>
              <td class='sale-col'>{fmt_price(row['sale_price'])} руб.</td>
              <td class='selected-col'>{fmt_price(selected_raw)} руб.</td>
              <td class='compare-col'>{compare_html}</td>
            </tr>
            """
        )

    table_html = f"""
    <!doctype html>
    <html><head><meta charset='utf-8'/>
    <style>
      body {{ margin:0; font-family: Inter, Arial, sans-serif; background: transparent; }}
      .wrap {{ background:linear-gradient(180deg, #ffffff 0%, #fbfdff 100%); border:1px solid #dbe5f1; border-radius:22px; overflow:hidden; box-shadow: 0 10px 26px rgba(15,23,42,.06); }}
      table {{ width:100%; border-collapse:separate; border-spacing:0; font-size:14px; }}
      thead th {{ position: sticky; top: 0; z-index: 2; background:linear-gradient(180deg, #f4f8ff 0%, #eef3fb 100%); color:#334155; text-align:left; padding:12px 12px; font-weight:800; border-bottom:1px solid #d7e1ef; }}
      tbody td {{ padding:12px; border-bottom:1px solid #e5edf6; vertical-align:top; color:#1e293b; background: rgba(255,255,255,.96); }}
      tbody tr:nth-child(even) td {{ background: #fcfdff; }}
      tbody tr:hover td {{ background: #f7faff; }}
      .article-pill {{ display:inline-block; padding:6px 10px; border-radius:999px; background:#edf2ff; color:#315efb; font-weight:800; white-space:nowrap; }}
      .name-cell {{ font-weight:800; line-height:1.33; color:#1e293b; margin-bottom:6px; max-width: 560px; }}
      .match-badge {{ display:inline-block; padding:4px 9px; border-radius:999px; font-size:12px; font-weight:800; }}
      .match-badge-exact {{ background:#e8f7ee; color:#15803d; }}
      .match-badge-linked {{ background:#e8f1ff; color:#1d4ed8; }}
      .match-badge-soft {{ background:#fff4e5; color:#b45309; }}
      .manual-note {{ margin-top:8px; font-size:12px; color:#475569; background:#f8fafc; border:1px dashed #cbd5e1; padding:6px 8px; border-radius:10px; max-width:560px; }}
      .sale-col {{ font-weight:800; white-space:nowrap; }}
      .selected-col {{ background: linear-gradient(180deg, #f4f8ff 0%, #eef4ff 100%); border-left:1px solid #c7d7ff; border-right:1px solid #c7d7ff; font-weight:900; color:#315efb; white-space:nowrap; }}
      .compare-col {{ min-width:230px; }}
      .stock-cell {{ min-width:110px; }}
      .stock-main {{ font-weight:900; font-size:18px; color:#0f172a; line-height:1.05; margin-bottom:6px; }}
      .stock-sub {{ font-size:12px; color:#64748b; line-height:1.4; }}
      .best-box {{ border-radius:18px; padding:11px 12px; min-width:190px; border:1px solid #dce6f7; background:linear-gradient(180deg, #f8fbff 0%, #f3f8ff 100%); box-shadow: inset 0 1px 0 rgba(255,255,255,.72); }}
      .best-box-empty {{ text-align:center; background:linear-gradient(180deg, #fafcff 0%, #f5f7fb 100%); border-color:#e2e8f0; min-height:76px; display:flex; align-items:center; justify-content:center; }}
      .best-empty-title {{ color:#64748b; font-weight:800; }}
      .best-top {{ display:flex; justify-content:space-between; gap:8px; align-items:center; margin-bottom:7px; }}
      .best-source-pill, .best-state-pill {{ display:inline-flex; align-items:center; padding:5px 10px; border-radius:999px; font-size:12px; font-weight:800; line-height:1; }}
      .best-price {{ font-size:18px; font-weight:900; color:#12348a; line-height:1.15; margin-bottom:6px; }}
      .best-meta-row {{ font-size:12px; color:#475569; margin-bottom:5px; }}
      .best-delta-row {{ font-size:12px; color:#64748b; line-height:1.45; }}
      .offer-good {{ border-color:#cfead6; background:linear-gradient(180deg, #fbfffc 0%, #f2fff6 100%); }}
      .offer-good .best-source-pill {{ background:#e9efff; color:#315efb; }}
      .offer-good .best-state-pill {{ background:#e8f7ee; color:#15803d; }}
      .offer-good .best-price {{ color:#103a8c; }}
      .offer-bad {{ border-color:#f7d7dd; background:linear-gradient(180deg, #fffafb 0%, #fff3f4 100%); }}
      .offer-bad .best-source-pill {{ background:#ffe8ec; color:#be123c; }}
      .offer-bad .best-state-pill {{ background:#ffe8ec; color:#be123c; }}
      .offer-bad .best-price {{ color:#991b1b; }}
      .offer-neutral {{ border-color:#d7e2ff; background:linear-gradient(180deg, #fbfdff 0%, #f3f7ff 100%); }}
      .offer-neutral .best-source-pill {{ background:#e9efff; color:#315efb; }}
      .offer-neutral .best-state-pill {{ background:#eef4ff; color:#315efb; }}
      .offer-own .best-source-pill {{ background:#eef2ff; color:#315efb; }}
      .offer-own .best-state-pill {{ background:#f1f5f9; color:#475569; }}
      .photo-col {{ width:92px; text-align:center; }}
      .photo-wrap {{ display:inline-flex; align-items:center; justify-content:center; width:72px; height:72px; border-radius:14px; overflow:hidden; border:1px solid #dbe5f1; background:#f8fbff; text-decoration:none; }}
      .result-photo {{ width:100%; height:100%; object-fit:cover; display:block; }}
      .photo-empty {{ border-radius:14px; display:flex; align-items:center; justify-content:center; background:#f8fafc; border:1px dashed #d6deea; color:#94a3b8; font-size:11px; font-weight:800; text-transform:uppercase; }}
      .photo-empty-small {{ width:72px; height:72px; }}
    </style></head><body>
      <div class='wrap'><table>
        <thead><tr><th>Товар</th><th>Наш склад</th><th>Наша цена</th><th>{html.escape(selected_label)}</th><th>Где лучше нас</th></tr></thead>
        <tbody>{''.join(rows_html)}</tbody>
      </table></div>
    </body></html>
    """
    height = min(max(220, 72 + len(df) * 72), 1050)
    components.html(table_html, height=height, scrolling=True)


def render_all_prices_block(result_df: pd.DataFrame, min_qty: float, price_mode: str, round100: bool, custom_discount: float, widget_key_prefix: str = "main") -> None:
    all_prices_df = build_all_prices_df(result_df, min_qty, price_mode, round100, custom_discount)
    if all_prices_df.empty:
        st.info("Для текущего результата нет данных по всем ценам.")
        return
    for article, group_df in all_prices_df.groupby("Артикул", sort=False):
        base_name = normalize_text(group_df.iloc[0].get("Название", ""))
        own_row = group_df[group_df["Источник"] == "Мы"].head(1)
        own_price_line = ""
        if not own_row.empty:
            own_price_line = f"Наша цена: {fmt_price(own_row.iloc[0]['Цена'])} руб. • Остаток: {fmt_qty(own_row.iloc[0]['Остаток'])}"
        st.markdown(
            f"""
            <div class='all-prices-head'>
              <div>
                <div class='all-prices-article'>{html.escape(article)}</div>
                <div class='all-prices-name'>{html.escape(base_name)}</div>
                <div class='all-prices-own'>{own_price_line}</div>
              </div>
            </div>
            """,
            unsafe_allow_html=True,
        )
        cols = st.columns(max(len(group_df), 1))
        for i, (_, rec) in enumerate(group_df.iterrows()):
            with cols[i]:
                source = str(rec.get("Источник", ""))
                status = str(rec.get("Статус", ""))
                status_class = status_visual_class(status)
                badge = {
                    "offer-good": "🟢 выгоднее",
                    "offer-bad": "🔴 дороже",
                    "offer-neutral": "🟡 цена равна",
                    "offer-own": "🔵 наша позиция",
                    "offer-muted": "⚪ найдено",
                }.get(status_class, status)
                st.markdown(
                    "<div class='offer-card-simple'>"
                    f"<div class='offer-card-source'>{html.escape(source)}</div>"
                    f"<span class='offer-status-badge {status_class}'>{html.escape(badge)}</span>"
                    f"<div class='offer-card-price'>{html.escape(fmt_price(rec.get('Цена')))} руб.</div>"
                    f"<div class='offer-card-meta'>Остаток: <b>{html.escape(fmt_qty(rec.get('Остаток')))}</b></div>"
                    + (
                        f"<div class='offer-card-meta'>Разница к нам: {html.escape(fmt_price(rec.get('Разница к нам, руб')))} руб. • "
                        f"{html.escape(str(round(float(rec.get('Разница к нам, %')), 2)).replace('.0', ''))}%</div>"
                        if source != "Мы" and pd.notna(rec.get("Разница к нам, %"))
                        else ""
                    )
                    + "</div>",
                    unsafe_allow_html=True,
                )
        with st.expander(f"Таблица по {article}"):
            show = group_df.copy()
            for col in ["Цена", "Остаток", "Наша цена", "Наша цена выбранная", "Разница к нам, руб"]:
                if col in show.columns:
                    show[col] = show[col].apply(lambda v: fmt_price(v) if "цена" in col.lower() or "руб" in col.lower() else fmt_qty(v))
            show["Разница к нам, %"] = show["Разница к нам, %"].apply(lambda v: (str(round(float(v), 2)).replace(".0", "") + "%") if pd.notna(v) else "")
            st.dataframe(show, use_container_width=True, hide_index=True)
        st.markdown("<div style='height:10px'></div>", unsafe_allow_html=True)

    st.download_button(
        "⬇️ Скачать все цены в Excel",
        all_prices_to_excel_bytes(all_prices_df),
        file_name=f"moy_tovar_all_prices_{widget_key_prefix}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        key=f"download_all_prices_{widget_key_prefix}",
    )


def render_results_insight_dashboard(result_df: pd.DataFrame, compare_map: dict[str, dict[str, Any]], source_pairs: list[dict[str, str]]) -> None:
    found_count = len(result_df) if isinstance(result_df, pd.DataFrame) else 0
    better_rows = 0
    gains: list[float] = []
    for item in compare_map.values():
        offer = item.get("best_offer")
        if offer and safe_float(offer.get("delta"), 0.0) > 0:
            better_rows += 1
            gains.append(safe_float(offer.get("delta_percent"), 0.0))
    avg_gain = sum(gains) / len(gains) if gains else 0.0
    hot_count = 0
    hot_buy_count = 0
    if isinstance(result_df, pd.DataFrame) and not result_df.empty and "hot_flag" in result_df.columns:
        hot_count = int(result_df["hot_flag"].fillna(False).map(bool).sum())
        if "hot_buy_signal" in result_df.columns:
            hot_buy_count = int(result_df["hot_buy_signal"].fillna("").map(normalize_text).str.upper().eq("BUY").sum())
    hot_label = "Ходовые в поиске"
    hot_note = "Watchlist не совпал с результатами"
    hot_help = "Товар ходовой → товар хорошо продавался за выбранный период"
    if hot_count:
        hot_label = "Товар ходовой"
        if hot_buy_count > 0:
            hot_note = "Сейчас можно брать"
            hot_help += "\nСейчас можно брать → лучший поставщик сейчас минимум на 35% дешевле нашей цены"
        else:
            hot_note = "Сейчас брать невыгодно"
            hot_help += "\nСейчас брать невыгодно → нет поставщика с ценой минимум на 35% ниже нашей цены"
    cards = [
        ("🔎", "Найдено позиций", str(found_count), "Сколько строк вошло в текущий поиск", ""),
        ("💚", "Есть цена лучше", str(better_rows), "Сколько позиций реально дешевле у поставщиков", ""),
        ("📈", "Средняя выгода", (f"{avg_gain:.1f}%" if gains else "—"), "Считается приложением, не берётся из готовых колонок Excel", ""),
        ("🔥", hot_label, str(hot_count), hot_note, hot_help),
        ("🧩", "Источников найдено", str(len(source_pairs)), ", ".join([x["source"] for x in source_pairs]) if source_pairs else "Нет колонок источников", ""),
    ]
    html_parts = []
    for icon, label, value, note, help_text in cards:
        help_html = f"<span class='insight-help' title='{html.escape(help_text)}'>?</span>" if help_text else ""
        html_parts.append(
            f"<div class='insight-card'><div class='insight-top'><span class='insight-icon'>{icon}</span><span class='insight-label'>{label}</span>{help_html}</div><div class='insight-value'>{value}</div><div class='insight-note'>{note}</div></div>"
        )
    html_cards = "".join(html_parts)
    st.markdown(f"<div class='insight-grid'>{html_cards}</div>", unsafe_allow_html=True)


def render_avito_block(avito_df: pd.DataFrame, result_df: pd.DataFrame) -> None:
    ads = find_avito_ads(avito_df, result_df)
    if ads.empty:
        st.caption("Объявления Авито по этим артикулам не найдены.")
        return
    st.caption(f"Найдено объявлений Авито: {len(ads)}")
    for _, row in ads.head(20).iterrows():
        title = normalize_text(row.get("title", ""))
        url = normalize_text(row.get("url", ""))
        account = normalize_text(row.get("account", ""))
        left, right = st.columns([6, 2])
        with left:
            st.markdown(f"**{html.escape(title)}**", unsafe_allow_html=True)
        with right:
            if account:
                st.markdown(
                    f"<div style='text-align:right;'><span style='display:inline-block;padding:6px 10px;border-radius:999px;background:#eef4ff;border:1px solid #d8e5ff;color:#315efb;font-weight:800;font-size:12px;'>{html.escape(account)}</span></div>",
                    unsafe_allow_html=True,
                )
        meta = []
        if normalize_text(row.get("ad_id", "")):
            meta.append(f"ID: {normalize_text(row.get('ad_id'))}")
        if normalize_text(row.get("price", "")):
            meta.append(f"Цена: {normalize_text(row.get('price'))}")
        if normalize_text(row.get("matched_tokens", "")):
            meta.append(f"Совпадения: {normalize_text(row.get('matched_tokens'))}")
        if meta:
            st.caption(" • ".join(meta))
        hist = []
        if normalize_text(row.get("first_seen", "")):
            hist.append(f"Впервые: {normalize_text(row.get('first_seen'))}")
        if normalize_text(row.get("last_seen", "")):
            hist.append(f"Последняя выгрузка: {normalize_text(row.get('last_seen'))}")
        if normalize_text(row.get("last_changed_at", "")):
            hist.append(f"Изменение: {normalize_text(row.get('last_changed_at'))}")
        if normalize_text(row.get("previous_price_raw", "")) and normalize_text(row.get("price", "")) and normalize_text(row.get("previous_price_raw", "")) != normalize_text(row.get("price", "")):
            hist.append(f"Было: {normalize_text(row.get('previous_price_raw'))}")
        if hist:
            st.caption(" • ".join(hist))
        if url:
            st.link_button("Открыть объявление", url, use_container_width=False)
        st.markdown("---")


def to_excel_bytes(df: pd.DataFrame, price_mode: str, round100: bool, custom_discount: float, min_qty: float) -> bytes:
    export_df = df.copy()
    export_df[current_price_label(price_mode, custom_discount)] = export_df.apply(lambda row: fmt_price(get_selected_price_raw(row, price_mode, round100, custom_discount)), axis=1)
    export_df["Лучшая цена поставщика"] = export_df.apply(lambda row: (get_best_offer_if_cheaper(row, min_qty=min_qty) or {}).get("price_fmt", ""), axis=1)
    export_df["Лучший поставщик"] = export_df.apply(lambda row: (get_best_offer_if_cheaper(row, min_qty=min_qty) or {}).get("source", ""), axis=1)
    export_df["Фото"] = export_df.get("photo_url", "")
    export_df = export_df[["article", "name", "free_qty", "sale_price", current_price_label(price_mode, custom_discount), "Лучший поставщик", "Лучшая цена поставщика", "Фото"]].rename(columns={
        "article": "Артикул",
        "name": "Название",
        "free_qty": "Наш склад",
        "sale_price": "Наша цена",
    })
    bio = io.BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        export_df.to_excel(writer, index=False, sheet_name="Результаты")
    bio.seek(0)
    return bio.read()


st.markdown(
    """
    <style>
    .stApp { background: #eef3f9; }
    header[data-testid="stHeader"] { background: rgba(0,0,0,0); }
    [data-testid="stDecoration"] { display: none; }
    .block-container { max-width: 1560px; padding-top: 3.4rem; padding-bottom: 1.2rem; }
    [data-testid="stSidebar"] { background: linear-gradient(180deg, #08122f 0%, #102358 55%, #172a63 100%); border-right: 1px solid rgba(255,255,255,.08); }
    [data-testid="stSidebar"] * { color: #e9efff !important; }
    .sidebar-brand { display:flex; align-items:center; gap:12px; margin: 0.15rem 0 0.95rem 0; padding: 0.15rem 0.1rem 0.35rem 0.1rem; }
    .sidebar-brand-logo { width:44px; height:44px; border-radius:14px; background: linear-gradient(180deg, rgba(255,255,255,.18), rgba(255,255,255,.08)); display:flex; align-items:center; justify-content:center; box-shadow: inset 0 1px 0 rgba(255,255,255,.15); font-size:22px; }
    .sidebar-brand-title { font-size: 1.22rem; font-weight: 900; line-height:1.05; color:#ffffff !important; }
    .sidebar-brand-sub { font-size: .82rem; color: #c7d6ff !important; margin-top: 4px; }
    .sidebar-card { background: linear-gradient(180deg, rgba(255,255,255,.08), rgba(255,255,255,.045)); border: 1px solid rgba(255,255,255,.13); border-radius: 22px; padding: 1rem 0.95rem 0.95rem 0.95rem; margin: 0.95rem 0 1.05rem 0; box-shadow: 0 12px 26px rgba(2, 8, 23, .24), inset 0 1px 0 rgba(255,255,255,.06); position: relative; overflow: hidden; }
    .sidebar-card::before { content: ''; position: absolute; inset: 0 auto auto 0; width: 100%; height: 3px; background: linear-gradient(90deg, rgba(111,163,255,.95) 0%, rgba(49,94,251,.95) 100%); opacity: .95; }
    .sidebar-card-header { display:flex; align-items:flex-start; justify-content:space-between; gap:10px; margin-bottom: .6rem; padding-bottom: .55rem; border-bottom: 1px solid rgba(255,255,255,.10); }
    .sidebar-card-header-main { display:flex; align-items:center; gap:10px; min-width:0; }
    .sidebar-card-title-wrap { min-width: 0; }
    .sidebar-card-kicker { color:#cfe0ff !important; font-size:10px; text-transform: uppercase; letter-spacing:.06em; font-weight:900; margin-bottom:2px; }
    .sidebar-card-icon { width:34px; height:34px; border-radius:12px; background: linear-gradient(180deg, rgba(255,255,255,.18), rgba(255,255,255,.08)); display:flex; align-items:center; justify-content:center; font-size:17px; box-shadow: inset 0 1px 0 rgba(255,255,255,.12); flex: 0 0 34px; }
    .sidebar-card-title { font-size: 1.01rem; font-weight: 900; color:#ffffff !important; line-height:1.15; margin:0; }
    .sidebar-card-help-wrap { position: relative; flex: 0 0 auto; }
    .sidebar-card-help { display:flex; align-items:center; justify-content:center; width:24px; height:24px; border-radius:999px; border:1px solid rgba(255,255,255,.18); background: rgba(255,255,255,.08); color:#ffffff !important; font-size:12px; font-weight:900; cursor:help; user-select:none; }
    .sidebar-card-tooltip { position:absolute; right:0; top:30px; width:250px; max-width:min(250px, 66vw); padding:10px 11px; border-radius:12px; background:#f8fbff; color:#0f172a !important; font-size:12px; line-height:1.45; box-shadow:0 16px 34px rgba(2, 8, 23, .30); opacity:0; transform:translateY(6px); pointer-events:none; transition:opacity .18s ease, transform .18s ease; z-index:35; }
    .sidebar-card-tooltip::before { content:''; position:absolute; top:-6px; right:9px; width:12px; height:12px; background:#f8fbff; transform:rotate(45deg); }
    .sidebar-card-help-wrap:hover .sidebar-card-tooltip { opacity:1; transform:translateY(0); }
    .sidebar-card-note { font-size: .79rem; line-height: 1.52; color:#c7d6ff !important; margin-bottom: .65rem; }
    .sidebar-status { background: rgba(7, 31, 74, .92); border: 1px solid rgba(255,255,255,.06); border-radius: 14px; padding: .76rem .82rem; color:#ffffff !important; font-weight: 800; margin-top: .58rem; }
    .sidebar-mini { font-size:.78rem; color:#c7d6ff !important; line-height:1.5; margin-top:.65rem; }
    [data-testid="stSidebar"] .stFileUploader section,
    [data-testid="stSidebar"] [data-testid="stFileUploaderDropzone"],
    [data-testid="stSidebar"] section[data-testid="stFileUploaderDropzone"] {
        background: linear-gradient(180deg, rgba(10,24,67,.92), rgba(9,20,56,.88)) !important;
        border: 1px dashed rgba(140,173,255,.34) !important;
        border-radius: 18px !important;
        box-shadow: inset 0 1px 0 rgba(255,255,255,.05) !important;
    }
    [data-testid="stSidebar"] [data-testid="stFileUploaderDropzone"] * {
        color: #dfe8ff !important;
        -webkit-text-fill-color: #dfe8ff !important;
        opacity: 1 !important;
    }
    [data-testid="stSidebar"] [data-testid="stFileUploaderDropzone"] button,
    [data-testid="stSidebar"] .stFileUploader button,
    [data-testid="stSidebar"] .stFileUploader [data-baseweb="button"] {
        background: linear-gradient(180deg, #3b6dff 0%, #2758ef 100%) !important;
        color: #ffffff !important;
        -webkit-text-fill-color: #ffffff !important;
        border: none !important;
        border-radius: 14px !important;
        font-weight: 800 !important;
        box-shadow: 0 10px 20px rgba(49,94,251,.28) !important;
    }
    [data-testid="stSidebar"] [data-testid="stFileUploaderFileData"],
    [data-testid="stSidebar"] [data-testid="stFileUploaderFile"],
    [data-testid="stSidebar"] [data-testid="stFileUploaderFileName"] {
        color: #f8fbff !important;
        -webkit-text-fill-color: #f8fbff !important;
        opacity: 1 !important;
    }
    [data-testid="stSidebar"] .stTextArea textarea,
    [data-testid="stSidebar"] .stTextInput input,
    [data-testid="stSidebar"] .stNumberInput input,
    [data-testid="stSidebar"] [data-baseweb="input"] input,
    [data-testid="stSidebar"] [data-baseweb="base-input"] input,
    [data-testid="stSidebar"] [data-baseweb="textarea"] textarea,
    [data-testid="stSidebar"] [data-baseweb="select"] > div {
        background: #ffffff !important;
        color: #0f172a !important;
        -webkit-text-fill-color: #0f172a !important;
        caret-color: #0f172a !important;
        border: 1px solid #d7e2f2 !important;
        border-radius: 16px !important;
        box-shadow: inset 0 1px 0 rgba(255,255,255,.85), 0 4px 10px rgba(2,8,23,.08) !important;
    }
    [data-testid="stSidebar"] .stTextArea textarea::placeholder,
    [data-testid="stSidebar"] .stTextInput input::placeholder,
    [data-testid="stSidebar"] .stNumberInput input::placeholder,
    [data-testid="stSidebar"] [data-baseweb="textarea"] textarea::placeholder {
        color: #7b8798 !important;
        -webkit-text-fill-color: #7b8798 !important;
        opacity: 1 !important;
    }
    [data-testid="stSidebar"] .stNumberInput button,
    [data-testid="stSidebar"] .stNumberInput [data-baseweb="button"],
    [data-testid="stSidebar"] .stNumberInput svg {
        background: #eef3ff !important;
        color: #315efb !important;
        fill: #315efb !important;
        stroke: #315efb !important;
        border-color: #d7e2f2 !important;
        opacity: 1 !important;
    }

    /* stronger sidebar field styling to override white baseweb wrappers */
    [data-testid="stSidebar"] .stNumberInput,
    [data-testid="stSidebar"] .stTextInput,
    [data-testid="stSidebar"] .stTextArea,
    [data-testid="stSidebar"] .stSelectbox {
        background: transparent !important;
    }
    [data-testid="stSidebar"] .stNumberInput > div,
    [data-testid="stSidebar"] .stTextInput > div,
    [data-testid="stSidebar"] .stTextArea > div,
    [data-testid="stSidebar"] .stSelectbox > div {
        background: transparent !important;
    }
    [data-testid="stSidebar"] .stNumberInput [data-baseweb="base-input"],
    [data-testid="stSidebar"] .stTextInput [data-baseweb="base-input"],
    [data-testid="stSidebar"] .stNumberInput [data-baseweb="input"],
    [data-testid="stSidebar"] .stTextInput [data-baseweb="input"],
    [data-testid="stSidebar"] .stTextArea [data-baseweb="textarea"],
    [data-testid="stSidebar"] .stSelectbox [data-baseweb="select"] {
        background: #ffffff !important;
        border-radius: 16px !important;
        box-shadow: inset 0 1px 0 rgba(255,255,255,.85), 0 4px 10px rgba(2,8,23,.08) !important;
        border: 1px solid #d7e2f2 !important;
    }
    [data-testid="stSidebar"] .stNumberInput [data-baseweb="base-input"] > div,
    [data-testid="stSidebar"] .stTextInput [data-baseweb="base-input"] > div,
    [data-testid="stSidebar"] .stNumberInput [data-baseweb="input"] > div,
    [data-testid="stSidebar"] .stTextInput [data-baseweb="input"] > div,
    [data-testid="stSidebar"] .stTextArea [data-baseweb="textarea"] > div,
    [data-testid="stSidebar"] .stSelectbox [data-baseweb="select"] > div {
        background: transparent !important;
        border-radius: 16px !important;
        border: none !important;
        box-shadow: none !important;
    }
    [data-testid="stSidebar"] .stNumberInput [data-baseweb="base-input"] input,
    [data-testid="stSidebar"] .stTextInput [data-baseweb="base-input"] input,
    [data-testid="stSidebar"] .stNumberInput [data-baseweb="input"] input,
    [data-testid="stSidebar"] .stTextInput [data-baseweb="input"] input,
    [data-testid="stSidebar"] .stTextArea [data-baseweb="textarea"] textarea,
    [data-testid="stSidebar"] .stSelectbox [data-baseweb="select"] input,
    [data-testid="stSidebar"] .stSelectbox [data-baseweb="select"] div[role="combobox"] {
        background: transparent !important;
        color: #0f172a !important;
        -webkit-text-fill-color: #0f172a !important;
        border: none !important;
        box-shadow: none !important;
    }
    [data-testid="stSidebar"] .stNumberInput [data-baseweb="base-input"] button,
    [data-testid="stSidebar"] .stNumberInput [data-baseweb="input"] button,
    [data-testid="stSidebar"] .stNumberInput [data-baseweb="base-input"] > div > button,
    [data-testid="stSidebar"] .stNumberInput [data-baseweb="input"] > div > button {
        background: #eef3ff !important;
        color: #315efb !important;
        border: none !important;
        border-radius: 12px !important;
        box-shadow: none !important;
    }
    [data-testid="stSidebar"] .stTextArea [data-baseweb="textarea"] textarea {
        min-height: 110px;
    }
    /* final hard override for white sidebar fields */
    [data-testid="stSidebar"] input:not([type="checkbox"]):not([type="radio"]),
    [data-testid="stSidebar"] textarea,
    [data-testid="stSidebar"] select,
    [data-testid="stSidebar"] div[role="combobox"],
    [data-testid="stSidebar"] [data-baseweb="input"],
    [data-testid="stSidebar"] [data-baseweb="base-input"],
    [data-testid="stSidebar"] [data-baseweb="textarea"],
    [data-testid="stSidebar"] [data-baseweb="select"],
    [data-testid="stSidebar"] [data-baseweb="input"] > div,
    [data-testid="stSidebar"] [data-baseweb="base-input"] > div,
    [data-testid="stSidebar"] [data-baseweb="textarea"] > div,
    [data-testid="stSidebar"] [data-baseweb="select"] > div {
        background: #ffffff !important;
        color: #0f172a !important;
        -webkit-text-fill-color: #0f172a !important;
        border-color: #d7e2f2 !important;
        border: 1px solid #d7e2f2 !important;
        box-shadow: inset 0 1px 0 rgba(255,255,255,.85), 0 4px 10px rgba(2,8,23,.08) !important;
        border-radius: 16px !important;
    }
    [data-testid="stSidebar"] input::placeholder,
    [data-testid="stSidebar"] textarea::placeholder {
        color: #7b8798 !important;
        -webkit-text-fill-color: #7b8798 !important;
        opacity: 1 !important;
    }
    [data-testid="stSidebar"] [data-baseweb="select"] svg,
    [data-testid="stSidebar"] div[role="combobox"] svg {
        fill: #6b7ea6 !important;
        color: #6b7ea6 !important;
    }
    [data-testid="stSidebar"] [data-baseweb="input"] button,
    [data-testid="stSidebar"] [data-baseweb="base-input"] button,
    [data-testid="stSidebar"] .stNumberInput button {
        background: #eef3ff !important;
        color: #315efb !important;
        border: none !important;
        box-shadow: none !important;
    }
    [data-testid="stSidebar"] .stTextArea textarea,
    [data-testid="stSidebar"] [data-baseweb="textarea"] textarea {
        padding: 12px 14px !important;
        line-height: 1.55 !important;
    }
    /* ultra-specific fix for lingering white textarea/select backgrounds in sidebar */
    [data-testid="stSidebar"] section[data-testid="stSidebarContent"] textarea,
    [data-testid="stSidebar"] section[data-testid="stSidebarContent"] textarea:focus,
    [data-testid="stSidebar"] section[data-testid="stSidebarContent"] textarea:hover,
    [data-testid="stSidebar"] section[data-testid="stSidebarContent"] .stTextArea textarea,
    [data-testid="stSidebar"] section[data-testid="stSidebarContent"] .stTextArea textarea:focus,
    [data-testid="stSidebar"] section[data-testid="stSidebarContent"] .stTextArea textarea:hover,
    [data-testid="stSidebar"] section[data-testid="stSidebarContent"] [data-baseweb="textarea"],
    [data-testid="stSidebar"] section[data-testid="stSidebarContent"] [data-baseweb="textarea"] > div,
    [data-testid="stSidebar"] section[data-testid="stSidebarContent"] [data-baseweb="textarea"] textarea,
    [data-testid="stSidebar"] section[data-testid="stSidebarContent"] [data-baseweb="select"],
    [data-testid="stSidebar"] section[data-testid="stSidebarContent"] [data-baseweb="select"] > div,
    [data-testid="stSidebar"] section[data-testid="stSidebarContent"] div[role="combobox"] {
        background-color: #ffffff !important;
        background: #ffffff !important;
        color: #0f172a !important;
        -webkit-text-fill-color: #0f172a !important;
        border: 1px solid #d7e2f2 !important;
        box-shadow: inset 0 1px 0 rgba(255,255,255,.85), 0 4px 10px rgba(2,8,23,.08) !important;
        border-radius: 16px !important;
    }
    [data-testid="stSidebar"] section[data-testid="stSidebarContent"] textarea::placeholder {
        color: #7b8798 !important;
        -webkit-text-fill-color: #7b8798 !important;
        opacity: 1 !important;
    }
    [data-testid="stSidebar"] .stButton > button,
    [data-testid="stSidebar"] .stDownloadButton > button {
        width: 100% !important;
        min-height: 48px !important;
        background: linear-gradient(180deg, #3b6dff 0%, #2758ef 100%) !important;
        color: #ffffff !important;
        -webkit-text-fill-color: #ffffff !important;
        border: none !important;
        border-radius: 16px !important;
        font-weight: 900 !important;
        box-shadow: 0 10px 22px rgba(49,94,251,.30) !important;
    }
    [data-testid="stSidebar"] .stButton > button:hover,
    [data-testid="stSidebar"] .stDownloadButton > button:hover {
        background: linear-gradient(180deg, #4a79ff 0%, #2d61f2 100%) !important;
        color: #ffffff !important;
        -webkit-text-fill-color: #ffffff !important;
    }
    [data-testid="stSidebar"] .stButton > button:disabled,
    [data-testid="stSidebar"] .stDownloadButton > button:disabled {
        background: linear-gradient(180deg, rgba(96,114,167,.72), rgba(80,95,143,.72)) !important;
        color: rgba(255,255,255,.86) !important;
        -webkit-text-fill-color: rgba(255,255,255,.86) !important;
        box-shadow: none !important;
        opacity: 1 !important;
    }
    [data-testid="stSidebar"] label,
    [data-testid="stSidebar"] .stRadio p,
    [data-testid="stSidebar"] .stCheckbox p,
    [data-testid="stSidebar"] .stSelectbox p {
        color: #eef4ff !important;
        -webkit-text-fill-color: #eef4ff !important;
    }
    .topbar { position: relative; background: linear-gradient(110deg, #0f172a 0%, #1742a8 56%, #2d6bff 100%); color: white; padding: 18px 20px; border-radius: 24px; margin-top: 0.55rem; margin-bottom: 14px; box-shadow: 0 18px 38px rgba(15, 23, 42, .22); overflow: hidden; }
    .topbar-grid { display:grid; grid-template-columns: 1.6fr 1fr 1fr 1fr; gap: 12px; align-items:center; position:relative; z-index:1; }
    .brand-box { display:flex; gap:14px; align-items:center; }
    .logo { width:58px;height:58px;border-radius:18px;background:rgba(255,255,255,.16); display:flex;align-items:center;justify-content:center;font-size:28px;font-weight:700; }
    .brand-title { font-size: 25px; font-weight: 900; line-height: 1; letter-spacing: -.02em; }
    .brand-sub { font-size: 13px; opacity: .92; margin-top: 6px; }
    .stat-box { background: rgba(255,255,255,.12); border: 1px solid rgba(255,255,255,.14); border-radius: 18px; padding: 12px 13px; min-height: 76px; backdrop-filter: blur(3px); }
    .stat-cap { font-size: 12px; opacity: .82; margin-bottom: 6px; }
    .stat-val { font-size: 16px; font-weight: 800; line-height: 1.3; }
    .toolbar, .result-wrap { position: relative; background: linear-gradient(180deg, #ffffff 0%, #fbfdff 100%); border: 1px solid #dbe5f1; border-radius: 22px; padding: 16px 18px 18px 18px; margin-bottom: 14px; box-shadow: 0 10px 26px rgba(15, 23, 42, .06); overflow: hidden; }
    .block-header { display:flex; align-items:flex-start; justify-content:space-between; gap:16px; padding: 2px 0 14px 0; margin-bottom: 14px; border-bottom: 1px solid #e7eef9; position: relative; }
    .block-header-main { display:flex; align-items:flex-start; gap:14px; min-width: 0; }
    .block-header-right { display:flex; align-items:flex-start; gap:10px; flex: 0 0 auto; }
    .block-icon { width: 48px; height: 48px; border-radius: 16px; background: linear-gradient(180deg, #3767ff 0%, #2455ef 100%); color: #ffffff; display:flex; align-items:center; justify-content:center; font-size: 24px; flex: 0 0 48px; }
    .block-title-wrap { min-width: 0; }
    .block-kicker { display:inline-flex; align-items:center; padding: 4px 9px; margin-bottom: 7px; border-radius: 999px; background: #eef4ff; border: 1px solid #d8e5ff; color: #315efb; font-size: 11px; font-weight: 900; letter-spacing: .04em; text-transform: uppercase; }
    .section-title { font-size: 22px; font-weight: 900; color:#0f172a; margin:0 0 5px 0; line-height:1.12; letter-spacing:-0.02em; }
    .section-sub { font-size: 13px; color:#64748b; margin:0; line-height:1.55; max-width: 980px; }
    .block-sparkles { display:flex; align-items:center; gap: 3px; color:#89a9ff; font-size: 12px; font-weight: 900; letter-spacing: .04em; opacity: .9; margin-top: 5px; }
    .block-help-wrap { position: relative; flex: 0 0 auto; }
    .block-help { display:flex; align-items:center; justify-content:center; width: 32px; height: 32px; border-radius: 999px; border: 1px solid #cfe0ff; background: linear-gradient(180deg, #f6f9ff 0%, #eef4ff 100%); color: #315efb; font-size: 15px; font-weight: 900; cursor: help; user-select: none; }
    .block-tooltip { position: absolute; right: 0; top: 40px; width: 340px; max-width: min(340px, 82vw); padding: 13px 14px; border-radius: 16px; background: #0f172a; color: #f8fbff; font-size: 12.8px; line-height: 1.5; box-shadow: 0 18px 36px rgba(15, 23, 42, .28); opacity: 0; transform: translateY(6px); pointer-events: none; transition: opacity .18s ease, transform .18s ease; z-index: 20; }
    .block-help-wrap:hover .block-tooltip { opacity: 1; transform: translateY(0); }
    .info-banner { display:flex; gap:14px; align-items:flex-start; padding:15px 16px; margin: 6px 0 14px 0; border-radius: 18px; border: 1px solid #dbe7fb; background: linear-gradient(180deg, #fbfdff 0%, #f5f9ff 100%); box-shadow: 0 8px 18px rgba(15,23,42,.05); }
    .info-banner-icon { width:42px; height:42px; flex:0 0 42px; border-radius: 14px; display:flex; align-items:center; justify-content:center; font-size: 20px; background: linear-gradient(180deg, #3767ff 0%, #2455ef 100%); color:#fff; }
    .info-banner-title { font-size: 15px; font-weight: 900; color:#0f172a; margin-bottom: 4px; }
    .info-banner-text { font-size: 13px; line-height: 1.55; color:#64748b; }
    .banner-chip-row { display:flex; flex-wrap:wrap; gap:8px; margin-top: 10px; }
    .banner-chip { display:inline-flex; align-items:center; gap:6px; padding: 6px 10px; border-radius: 999px; background:#eef4ff; border:1px solid #d8e5ff; color:#315efb; font-size: 12px; font-weight: 800; }
    .tone-green { background: linear-gradient(180deg, #fbfffd 0%, #f2fff7 100%); border-color: #d2f1dd; }
    .tone-purple { background: linear-gradient(180deg, #fcfbff 0%, #f6f3ff 100%); border-color: #e6dcff; }
    .insight-grid { display:grid; grid-template-columns: repeat(auto-fit, minmax(220px, 1fr)); gap: 12px; margin: 14px 0 16px 0; }
    .insight-card { background: linear-gradient(180deg, #ffffff 0%, #f8fbff 100%); border: 1px solid #dbe7fb; border-radius: 20px; padding: 14px 15px; box-shadow: 0 8px 18px rgba(15,23,42,.05); }
    .insight-top { display:flex; align-items:center; gap:8px; margin-bottom: 10px; }
    .insight-icon { width:32px; height:32px; display:flex; align-items:center; justify-content:center; border-radius: 12px; background:#eef4ff; font-size:16px; }
    .insight-label { color:#64748b; font-size:12px; font-weight:800; }
    .insight-value { color:#0f172a; font-size: 28px; font-weight: 900; line-height:1.1; margin-bottom: 6px; }
    .insight-note { color:#6b7c93; font-size:12px; line-height:1.45; }
    .all-prices-head { display:flex; align-items:flex-start; justify-content:space-between; gap:10px; margin: 14px 0 10px 0; padding: 14px 16px; border-radius: 18px; background: linear-gradient(180deg, #fbfdff 0%, #f5f9ff 100%); border:1px solid #dbe7fb; }
    .all-prices-article { color:#315efb; font-size: 18px; font-weight: 900; margin-bottom: 4px; }
    .all-prices-name { color:#0f172a; font-size: 14px; font-weight: 800; line-height: 1.45; }
    .all-prices-own { margin-top: 6px; color:#64748b; font-size: 12.5px; }
    .offer-card-simple { border-radius: 18px; padding: 14px; border:1px solid #dbe7fb; background: linear-gradient(180deg, #ffffff 0%, #f9fbff 100%); min-height: 140px; box-shadow: 0 8px 18px rgba(15,23,42,.05); }
    .offer-card-source { color:#0f172a; font-size: 15px; font-weight: 900; margin-bottom: 10px; }
    .offer-status-badge { display:inline-flex; align-items:center; justify-content:center; padding:5px 9px; border-radius:999px; font-size:11px; font-weight:900; margin-bottom: 8px; }
    .offer-good { background:#e9f9ef; color:#15803d; }
    .offer-bad { background:#fff1f2; color:#be123c; }
    .offer-neutral { background:#eef4ff; color:#315efb; }
    .offer-own { background:#f3f4f6; color:#475569; }
    .offer-muted { background:#f8fafc; color:#64748b; }
    .offer-card-price { color:#0f2f83; font-size: 24px; font-weight: 900; line-height: 1.15; margin-bottom: 6px; }
    .offer-card-meta { color:#64748b; font-size: 12.5px; line-height:1.45; margin-bottom: 4px; }
    </style>
    """,
    unsafe_allow_html=True,
)


with st.sidebar:
    st.markdown(
        """
        <div class="sidebar-brand">
          <div class="sidebar-brand-logo">📦</div>
          <div>
            <div class="sidebar-brand-title">Мой Товар</div>
            <div class="sidebar-brand-sub">comparison-файл + фото + поиск 💙</div>
          </div>
        </div>
        """,
        unsafe_allow_html=True,
    )
    if REQUESTS_IMPORT_ERROR:
        st.caption("requests не установлен: веб-функции отключены")

    st.markdown('<div class="sidebar-card">', unsafe_allow_html=True)
    render_sidebar_card_header("Comparison-файл", "📘", "Главный файл приложения. Содержит листы Сравнение, Уценка, Совместимые. Можно хранить последний файл на сервере в /data.")
    uploaded = st.file_uploader("Загрузить comparison-файл", type=["xlsx", "xlsm"], label_visibility="collapsed")
    if uploaded is not None:
        try:
            comp_bytes = uploaded.getvalue()
            comp_sig = hashlib.md5(comp_bytes).hexdigest()
            if st.session_state.get("comparison_upload_applied_sig", "") != comp_sig:
                maybe_create_service_snapshot_before_action("comparison_upload", comp_sig, f"before comparison upload: {uploaded.name}")
                save_uploaded_source_file(get_persisted_comparison_file_path(), comp_bytes, uploaded.name)
                clear_loader_caches()
                st.session_state.comparison_sheets = load_comparison_workbook(uploaded.name, comp_bytes)
                st.session_state.comparison_name = uploaded.name + " • сохранён в /data"
                st.session_state.comparison_version = datetime.utcnow().isoformat()
                available = list(st.session_state.comparison_sheets.keys())
                if available and st.session_state.selected_sheet not in available:
                    st.session_state.selected_sheet = available[0]
                rebuild_current_df()
                refresh_all_search_results()
                st.session_state["comparison_upload_applied_sig"] = comp_sig
                log_operation(f"Обновлён comparison-файл: {uploaded.name}", "success")
        except Exception as exc:
            log_operation(f"Ошибка comparison-файла: {exc}", "warning")
            st.error(f"Ошибка файла: {exc}")
    else:
        if not (isinstance(st.session_state.get("comparison_sheets"), dict) and st.session_state.get("comparison_sheets")):
            load_persisted_comparison_source_into_state()
    st.markdown(f'<div class="sidebar-status">Файл: {html.escape(st.session_state.get("comparison_name", "ещё не загружен"))}</div>', unsafe_allow_html=True)
    st.markdown(f'<div class="sidebar-mini">Файл на сервере: {html.escape(str(get_persisted_comparison_file_path()))}</div>', unsafe_allow_html=True)
    st.markdown('<div class="sidebar-mini">Рабочие разделы переключаются сверху: <b>Оригинал</b>, <b>Уценка</b>, <b>Совместимые</b>. Рендерится только активный раздел — это быстрее.</div>', unsafe_allow_html=True)
    st.markdown('</div>', unsafe_allow_html=True)

    st.markdown('<div class="sidebar-card">', unsafe_allow_html=True)
    render_sidebar_card_header("Фото товаров", "🖼️", "Файл с артикулами и ссылками на фото. Можно держать его в реестре на сервере и дозагружать только новинки/изменения.")
    photo_uploaded = st.file_uploader("Загрузить файл фото", type=["xlsx", "xls", "xlsm", "csv"], key="photo_uploader", label_visibility="collapsed")
    if photo_uploaded is not None:
        try:
            photo_bytes = photo_uploaded.getvalue()
            photo_sig = hashlib.md5(photo_bytes).hexdigest()
            if st.session_state.get("photo_upload_applied_sig", "") != photo_sig:
                maybe_create_service_snapshot_before_action("photo_upload", photo_sig, f"before photo upload: {photo_uploaded.name}")
                save_uploaded_source_file(get_persisted_photo_file_path(), photo_bytes, photo_uploaded.name)
                clear_loader_caches()
                loaded_photo_df = load_photo_map_file(photo_uploaded.name, photo_bytes)
                if st.session_state.get("photo_last_sync_sig", "") != photo_sig:
                    photo_stats = sync_photo_registry(loaded_photo_df, photo_uploaded.name)
                    st.session_state.photo_registry_stats = photo_stats
                    st.session_state.photo_registry_message = (
                        f"Синхронизация фото: новых {photo_stats.get('new', 0)}, обновлённых {photo_stats.get('changed', 0)}, без изменений {photo_stats.get('unchanged', 0)}. Исходник сохранён в /data"
                    )
                    st.session_state.photo_last_sync_sig = photo_sig
                reg_df = load_photo_registry_df()
                if isinstance(reg_df, pd.DataFrame) and not reg_df.empty:
                    st.session_state.photo_df = reg_df[[
                        "article", "article_norm", "photo_url", "source_sheet",
                        "meta_brand", "meta_color", "meta_capacity", "meta_manufacturer_code",
                        "meta_model", "meta_description", "meta_fits_models", "meta_iso_pages",
                        "meta_print_technology", "meta_item_type", "meta_print_type",
                        "meta_weight", "meta_length", "meta_width", "meta_height",
                    ]].copy()
                else:
                    st.session_state.photo_df = loaded_photo_df
                st.session_state.photo_name = photo_uploaded.name + " • сохранён в /data"
                rebuild_current_df()
                refresh_all_search_results()
                st.session_state["photo_upload_applied_sig"] = photo_sig
                log_operation(f"Обновлён каталог фото: {photo_uploaded.name}", "success")
        except Exception as exc:
            log_operation(f"Ошибка файла фото: {exc}", "warning")
            st.error(f"Ошибка файла фото: {exc}")
    else:
        if not isinstance(st.session_state.get("photo_df"), pd.DataFrame):
            if not load_persisted_photo_source_into_state():
                ensure_photo_registry_loaded()
    st.markdown(f'<div class="sidebar-status">Фото: {html.escape(st.session_state.get("photo_name", "ещё не загружен"))}</div>', unsafe_allow_html=True)
    st.markdown(f'<div class="sidebar-mini">Файл на сервере: {html.escape(str(get_persisted_photo_file_path()))}</div>', unsafe_allow_html=True)
    st.markdown('<div class="sidebar-mini">Автопоиск фото на внешних сайтах отключён. Используются только данные из Каталога расходки и локального реестра.</div>', unsafe_allow_html=True)
    if st.session_state.get("photo_registry_message"):
        st.markdown(f'<div class="sidebar-mini">{html.escape(st.session_state.get("photo_registry_message"))}</div>', unsafe_allow_html=True)
    st.markdown(f'<div class="sidebar-mini">{html.escape(photo_registry_summary_text())}</div>', unsafe_allow_html=True)
    st.markdown('</div>', unsafe_allow_html=True)

    st.markdown('<div class="sidebar-card">', unsafe_allow_html=True)
    render_sidebar_card_header("Авито", "🛒", "Загруженный файл Авито помогает найти действующие объявления. Параллельно ведём локальный реестр: новые объявления добавляются, изменившиеся обновляются.")
    avito_uploaded = st.file_uploader("Загрузить файл Авито", type=["xlsx", "xlsm", "csv"], key="avito_uploader", label_visibility="collapsed")
    if avito_uploaded is not None:
        try:
            avito_bytes = avito_uploaded.getvalue()
            avito_sig = hashlib.md5(avito_bytes).hexdigest()
            if st.session_state.get("avito_upload_applied_sig", "") != avito_sig:
                maybe_create_service_snapshot_before_action("avito_upload", avito_sig, f"before avito upload: {avito_uploaded.name}")
                save_uploaded_source_file(get_persisted_avito_file_path(), avito_bytes, avito_uploaded.name)
                clear_loader_caches()
                st.session_state.avito_df = load_avito_file(avito_uploaded.name, avito_bytes)
                st.session_state.avito_name = avito_uploaded.name + " • сохранён в /data"
                if st.session_state.get("avito_last_sync_sig", "") != avito_sig:
                    sync_stats = sync_avito_registry(st.session_state.avito_df, avito_uploaded.name)
                    st.session_state.avito_registry_stats = sync_stats
                    st.session_state.avito_registry_message = (
                        f"Синхронизация: новых {sync_stats.get('new', 0)}, изменённых {sync_stats.get('changed', 0)}, без изменений {sync_stats.get('unchanged', 0)}. Исходник сохранён в /data"
                    )
                    st.session_state.avito_last_sync_sig = avito_sig
                st.session_state["avito_upload_applied_sig"] = avito_sig
                log_operation(f"Обновлён файл Авито: {avito_uploaded.name}", "success")
        except Exception as exc:
            log_operation(f"Ошибка файла Авито: {exc}", "warning")
            st.error(f"Ошибка файла Авито: {exc}")
    else:
        if not isinstance(st.session_state.get("avito_df"), pd.DataFrame):
            load_persisted_avito_source_into_state()
    st.markdown(f'<div class="sidebar-status">Авито: {html.escape(st.session_state.get("avito_name", "ещё не загружен"))}</div>', unsafe_allow_html=True)
    st.markdown(f'<div class="sidebar-mini">Файл на сервере: {html.escape(str(get_persisted_avito_file_path()))}</div>', unsafe_allow_html=True)
    if st.session_state.get("avito_registry_message"):
        st.markdown(f'<div class="sidebar-mini">{html.escape(st.session_state.get("avito_registry_message"))}</div>', unsafe_allow_html=True)
    st.markdown(f'<div class="sidebar-mini">{html.escape(registry_summary_text())}</div>', unsafe_allow_html=True)
    st.markdown('</div>', unsafe_allow_html=True)

    st.markdown('<div class="sidebar-card">', unsafe_allow_html=True)
    render_sidebar_card_header("Статистика продаж / ходовые", "🔥", "Новый файл статистики продаж. Берём продажи/мес из столбца «В месяц» и используем их в CRM, аналитике и решениях по складу.")
    hot_uploaded = st.file_uploader(
        "Загрузить статистику",
        type=["xlsx", "xls", "csv"],
        key="hot_items_uploader",
        label_visibility="collapsed",
        help="Можно загрузить новый файл статистики продаж или старый watchlist. Для нового файла продажи/мес берутся из столбца «В месяц».",
    )
    st.caption("ⓘ Статистика продаж отвечает за продажи/мес, дни без продаж, спрос и приоритет. Для нового файла продажи/мес берутся из столбца «В месяц».")
    if hot_uploaded is not None:
        try:
            hot_bytes = hot_uploaded.getvalue()
            hot_sig = hashlib.md5(hot_bytes).hexdigest()
            if st.session_state.get("hot_upload_applied_sig", "") != hot_sig:
                maybe_create_service_snapshot_before_action("watchlist_upload", hot_sig, f"before watchlist upload: {hot_uploaded.name}")
                save_uploaded_source_file(get_persisted_watchlist_file_path(), hot_bytes, hot_uploaded.name)
                clear_loader_caches()
                st.session_state.hot_items_df = load_hot_watchlist_file(hot_uploaded.name, hot_bytes)
                st.session_state.hot_items_name = hot_uploaded.name + " • сохранён в /data"
                st.session_state.hot_items_last_sync_sig = hot_sig
                st.session_state["hot_upload_applied_sig"] = hot_sig
                log_operation(f"Обновлена статистика продаж: {hot_uploaded.name}", "success")
        except Exception as exc:
            log_operation(f"Ошибка файла ходовых: {exc}", "warning")
            st.error(f"Ошибка статистики: {exc}")
    else:
        if not isinstance(st.session_state.get("hot_items_df"), pd.DataFrame):
            load_persisted_watchlist_source_into_state()
    st.markdown(f'<div class="sidebar-status">Статистика: {html.escape(st.session_state.get("hot_items_name", "ещё не загружена"))}</div>', unsafe_allow_html=True)
    st.markdown(f'<div class="sidebar-mini">Файл на сервере: {html.escape(str(get_persisted_watchlist_file_path()))}</div>', unsafe_allow_html=True)
    st.markdown(f'<div class="sidebar-mini">{html.escape(hot_watchlist_summary_text())}</div>', unsafe_allow_html=True)
    hot_df_state = st.session_state.get("hot_items_df")
    hot_buy_total = 0
    if isinstance(hot_df_state, pd.DataFrame) and not hot_df_state.empty:
        hot_buy_total = len(get_cached_hot_buy_watchlist_table())
    st.checkbox(
        f"Показать таблицу «можно брать» ({hot_buy_total})",
        key="show_hot_buy_watchlist_table",
        help="Лениво открывает таблицу только по ходовым позициям, где поставщик проходит твой порог выгоды (по умолчанию 35%). Пока чекбокс выключен, таблица не строится.",
    )
    st.caption("ⓘ Таблица «можно брать» показывает только те ходовые позиции, где поставщик проходит твой порог выгоды и есть остаток.")
    st.markdown('</div>', unsafe_allow_html=True)

    st.markdown('<div class="sidebar-card">', unsafe_allow_html=True)
    render_sidebar_card_header("Средняя закупка", "💳", "Отдельный файл со средневзвешенной закупкой за 1 шт. Используется только поверх CRM-карточки и не ломает старое ядро comparison.")
    purchase_uploaded = st.file_uploader(
        "Загрузить файл средней закупки",
        type=["xlsx", "xlsm"],
        key="purchase_cost_uploader",
        label_visibility="collapsed",
        help="Файл закупки, где есть лист Итог_взвешенный и колонка «Средняя закупка за 1 шт». Маппинг идёт по названию и коду, извлечённому из номенклатуры.",
    )
    st.caption("ⓘ Берём лист «Итог_взвешенный». Артикулов в файле нет, поэтому маппинг идёт безопасно по названию и коду внутри номенклатуры.")
    if purchase_uploaded is not None:
        try:
            purchase_bytes = purchase_uploaded.getvalue()
            purchase_sig = hashlib.md5(purchase_bytes).hexdigest()
            if st.session_state.get("purchase_upload_applied_sig", "") != purchase_sig:
                maybe_create_service_snapshot_before_action("purchase_upload", purchase_sig, f"before purchase upload: {purchase_uploaded.name}")
                save_uploaded_source_file(get_persisted_purchase_file_path(), purchase_bytes, purchase_uploaded.name)
                clear_loader_caches()
                st.session_state.purchase_cost_df = load_purchase_cost_file(purchase_uploaded.name, purchase_bytes)
                st.session_state.purchase_cost_name = purchase_uploaded.name + " • сохранён в /data"
                st.session_state.purchase_cost_last_sync_sig = purchase_sig
                st.session_state["purchase_upload_applied_sig"] = purchase_sig
                log_operation(f"Обновлён файл средней закупки: {purchase_uploaded.name}", "success")
        except Exception as exc:
            log_operation(f"Ошибка файла средней закупки: {exc}", "warning")
            st.error(f"Ошибка файла закупки: {exc}")
    else:
        if not isinstance(st.session_state.get("purchase_cost_df"), pd.DataFrame):
            load_persisted_purchase_source_into_state()
    st.markdown(f'<div class="sidebar-status">Средняя закупка: {html.escape(st.session_state.get("purchase_cost_name", "ещё не загружен"))}</div>', unsafe_allow_html=True)
    st.markdown(f'<div class="sidebar-mini">Файл на сервере: {html.escape(str(get_persisted_purchase_file_path()))}</div>', unsafe_allow_html=True)
    st.markdown(f'<div class="sidebar-mini">{html.escape(purchase_cost_summary_text())}</div>', unsafe_allow_html=True)

    st.markdown('</div>', unsafe_allow_html=True)

    st.markdown('<div class="sidebar-card">', unsafe_allow_html=True)
    render_sidebar_card_header("Отчёт и цены", "📊", "Порог выгоды и минимальный остаток для пересчёта лучшей цены.")
    st.number_input(
        "Порог отчёта, %",
        min_value=0.0,
        max_value=95.0,
        step=1.0,
        key="distributor_threshold",
        help="Минимальный процент выгоды от нашей цены. Ниже этого порога поставщик не считается интересным для отчёта и части подсказок.",
    )
    st.number_input(
        "Мин. остаток у поставщика",
        min_value=1.0,
        max_value=999999.0,
        step=1.0,
        key="distributor_min_qty",
        help="Минимальный остаток у поставщика, ниже которого предложение считается слишком слабым и не участвует в сравнении.",
    )
    st.markdown('<div class="sidebar-mini">Колонки Мин. у конкурентов / Разница из Excel не используются. Всё считаем заново прямо в приложении.</div>', unsafe_allow_html=True)
    st.markdown('</div>', unsafe_allow_html=True)

    st.markdown('<div class="sidebar-card">', unsafe_allow_html=True)
    render_sidebar_card_header("Быстрая правка цен", "✏️", "Меняет Наша цена по всем листам comparison-файла. Полезно для локальной проверки без правки исходного Excel.")
    st.text_area("Правка цен", key="price_patch_input", height=110, label_visibility="collapsed", placeholder="CE278A 8900\nCF364A - 29700")
    if st.button("Править цены в файле", use_container_width=True):
        sheets_state = st.session_state.get("comparison_sheets")
        persisted_path = get_persisted_comparison_file_path()
        original_name = read_persisted_original_name(persisted_path, persisted_path.name) if persisted_path.exists() else st.session_state.get("comparison_name", "comparison_latest.xlsx")
        if persisted_path.exists():
            try:
                create_service_snapshot(reason="before price patch", source="auto")
                before_snapshot = build_price_snapshot_for_updates(st.session_state.get("comparison_sheets"), st.session_state.price_patch_input)
                updated_bytes, patch_message = patch_comparison_workbook_bytes(persisted_path.read_bytes(), st.session_state.price_patch_input)
                if updated_bytes is not None:
                    save_uploaded_source_file(persisted_path, updated_bytes, original_name.replace(" • из /data", "").replace(" • сохранён в /data", ""))
                    clear_loader_caches()
                    st.session_state.comparison_sheets = load_comparison_workbook(original_name, updated_bytes)
                    st.session_state.comparison_name = original_name + " • из /data"
                    st.session_state.comparison_version = datetime.utcnow().isoformat()
                    rebuild_current_df()
                    refresh_all_search_results()
                    after_snapshot = build_price_snapshot_for_updates(st.session_state.get("comparison_sheets"), st.session_state.price_patch_input)
                    history_logged = log_price_patch_history_diff(before_snapshot, after_snapshot, source="manual", note="Быстрая правка цен")
                    if history_logged:
                        patch_message += f" | История: {history_logged}"
                st.session_state.patch_message = patch_message
                log_operation(f"Быстрая правка цен: {patch_message}", "success")
            except Exception as exc:
                st.session_state.patch_message = f"Ошибка правки файла: {exc}"
                log_operation(f"Ошибка быстрой правки цен: {exc}", "warning")
        elif isinstance(sheets_state, dict) and sheets_state:
            before_snapshot = build_price_snapshot_for_updates(sheets_state, st.session_state.price_patch_input)
            updated_sheets, patch_message = apply_price_updates_to_sheets(sheets_state, st.session_state.price_patch_input)
            st.session_state.comparison_sheets = updated_sheets
            st.session_state.comparison_version = datetime.utcnow().isoformat()
            rebuild_current_df()
            after_snapshot = build_price_snapshot_for_updates(updated_sheets, st.session_state.price_patch_input)
            history_logged = log_price_patch_history_diff(before_snapshot, after_snapshot, source="manual", note="Быстрая правка цен")
            if history_logged:
                patch_message += f" | История: {history_logged}"
            st.session_state.patch_message = patch_message
            refresh_all_search_results()
            log_operation(f"Быстрая правка цен: {patch_message}", "success")
        else:
            st.session_state.patch_message = "Сначала загрузите comparison-файл."
            log_operation("Быстрая правка цен: comparison-файл не загружен", "warning")
    if st.session_state.get("patch_message"):
        st.markdown(f'<div class="sidebar-mini">{html.escape(st.session_state.patch_message)}</div>', unsafe_allow_html=True)
    else:
        st.markdown('<div class="sidebar-mini">Прайс сохраняется локально. После правок цены не пропадут до загрузки нового файла.</div>', unsafe_allow_html=True)
    st.markdown('</div>', unsafe_allow_html=True)


    st.markdown('<div class="sidebar-card">', unsafe_allow_html=True)
    render_sidebar_card_header("Сервисный режим", "🛡️", "Ленивый блок для проверки системы, snapshot, восстановления и backup.zip.")
    render_service_mode_sidebar()
    st.markdown('</div>', unsafe_allow_html=True)

    st.markdown('<div class="sidebar-card">', unsafe_allow_html=True)
    render_sidebar_card_header("Настройки", "⚙️", "Управляет режимом поиска, главной ценой, пользовательской скидкой и округлением.")
    st.radio("Режим поиска", ["Только артикул", "Умный", "Артикул + название + бренд"], key="search_mode")
    st.radio("Какая цена главная", ["-12%", "-20%", "Своя скидка"], key="price_mode")
    st.number_input("Своя скидка, %", min_value=0.0, max_value=99.0, step=1.0, key="custom_discount")
    st.checkbox("Округлять вверх до 100", key="round100")
    st.markdown('</div>', unsafe_allow_html=True)

    st.markdown('<div class="sidebar-card">', unsafe_allow_html=True)
    render_sidebar_card_header("Текст шаблона 1", "🧾", "Этот текст добавляется один раз в конце шаблона 1.")
    st.markdown('<div class="sidebar-card-note">Этот текст добавляется один раз в конце шаблона 1. Хэштеги по артикулам подставляются автоматически.</div>', unsafe_allow_html=True)
    st.text_area("Текст шаблона 1", key="template1_footer", height=170, label_visibility="collapsed")
    st.markdown('<div class="sidebar-mini">Текст сохраняется локально и останется до следующего изменения.</div>', unsafe_allow_html=True)
    st.markdown('</div>', unsafe_allow_html=True)

    render_operation_log_sidebar()

comparison_name = st.session_state.get("comparison_name", "ещё не загружен")
sheets = st.session_state.get("comparison_sheets", {})
loaded_sheet_count = len(sheets) if isinstance(sheets, dict) else 0
rows_count = sum(len(df) for df in sheets.values()) if isinstance(sheets, dict) and sheets else 0
price_mode = st.session_state.price_mode
round100 = st.session_state.round100
custom_discount = float(st.session_state.custom_discount)
search_mode = st.session_state.search_mode
price_label = current_price_label(price_mode, custom_discount)

st.markdown(f"""
<div class="topbar"><div class="topbar-grid">
<div class="brand-box"><div class="logo">📦</div><div><div class="brand-title">{APP_TITLE}</div><div class="brand-sub">Один comparison-файл • поиск • фото • пересчёт цен поставщиков • {APP_VERSION}</div></div></div>
<div class="stat-box"><div class="stat-cap">Файл</div><div class="stat-val">{html.escape(comparison_name)}</div></div>
<div class="stat-box"><div class="stat-cap">Вкладок</div><div class="stat-val">{loaded_sheet_count if loaded_sheet_count else '—'}</div></div>
<div class="stat-box"><div class="stat-cap">Всего строк</div><div class="stat-val">{rows_count}</div></div>
</div></div>
""", unsafe_allow_html=True)





def get_price_patch_history_path() -> Path:
    try:
        return Path(__file__).resolve().with_name("price_patch_history.sqlite")
    except Exception:
        return Path.cwd() / "price_patch_history.sqlite"


def ensure_price_patch_history() -> None:
    path = get_price_patch_history_path()
    conn = sqlite3.connect(path)
    try:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS price_patch_history (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                changed_at TEXT,
                article_norm TEXT,
                article TEXT,
                sheet_name TEXT,
                old_price REAL,
                new_price REAL,
                change_source TEXT,
                note TEXT
            )
            """
        )
        conn.commit()
    finally:
        conn.close()


def load_price_patch_history_df(limit: int = 200) -> pd.DataFrame:
    path = get_price_patch_history_path()
    if not path.exists():
        return pd.DataFrame()
    ensure_price_patch_history()
    conn = sqlite3.connect(path)
    try:
        df = pd.read_sql_query(
            "SELECT * FROM price_patch_history ORDER BY id DESC LIMIT ?",
            conn,
            params=(int(limit),),
        )
    except Exception:
        return pd.DataFrame()
    finally:
        conn.close()
    if df.empty:
        return df
    for col in ["changed_at", "article_norm", "article", "sheet_name", "change_source", "note"]:
        if col in df.columns:
            df[col] = df[col].fillna("").map(normalize_text)
    return df


def build_price_snapshot_for_updates(sheets: dict[str, pd.DataFrame] | None, updates_text: str) -> dict[tuple[str, str, str], dict[str, Any]]:
    updates = parse_price_updates(updates_text)
    snapshot: dict[tuple[str, str, str], dict[str, Any]] = {}
    if not updates or not isinstance(sheets, dict):
        return snapshot
    target_codes = [code for code, _ in updates if normalize_article(code)]
    if not target_codes:
        return snapshot
    for sheet_name, df in sheets.items():
        if df is None or df.empty:
            continue
        for target in target_codes:
            mask = df["article_norm"].eq(target)
            if "row_codes" in df.columns:
                row_mask = df["row_codes"].apply(lambda codes: target in (codes or []) if isinstance(codes, list) else False)
                mask = mask | row_mask
            matched = df[mask]
            if matched.empty:
                continue
            for _, row in matched.iterrows():
                article_txt = normalize_text(row.get("article", ""))
                key = (target, sheet_name, normalize_article(article_txt) or target)
                snapshot[key] = {
                    "article_norm": target,
                    "article": article_txt or target,
                    "sheet_name": sheet_name,
                    "price": safe_float(row.get("sale_price"), 0.0),
                }
    return snapshot


def log_price_patch_history_diff(before: dict[tuple[str, str, str], dict[str, Any]], after: dict[tuple[str, str, str], dict[str, Any]], source: str = "manual", note: str = "") -> int:
    ensure_price_patch_history()
    rows: list[tuple[str, str, str, str, float, float, str, str]] = []
    changed_at = datetime.now().replace(microsecond=0).isoformat(sep=" ")
    all_keys = set(before.keys()) | set(after.keys())
    for key in all_keys:
        prev = before.get(key)
        cur = after.get(key)
        old_price = safe_float((prev or {}).get("price"), 0.0)
        new_price = safe_float((cur or {}).get("price"), 0.0)
        if abs(old_price - new_price) < 1e-9:
            continue
        row = cur or prev or {}
        rows.append(
            (
                changed_at,
                normalize_text(row.get("article_norm", key[0])),
                normalize_text(row.get("article", key[2])),
                normalize_text(row.get("sheet_name", key[1])),
                old_price,
                new_price,
                normalize_text(source),
                normalize_text(note),
            )
        )
    if not rows:
        return 0
    conn = sqlite3.connect(get_price_patch_history_path())
    try:
        conn.executemany(
            """
            INSERT INTO price_patch_history (
                changed_at, article_norm, article, sheet_name, old_price, new_price, change_source, note
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            rows,
        )
        conn.commit()
    finally:
        conn.close()
    return len(rows)


def infer_brand_from_product_name(name: object) -> str:
    text = contains_text(name)
    if not text:
        return ""
    patterns = [
        ("KONICA MINOLTA", "Konica Minolta"),
        ("KONICA-MINOLTA", "Konica Minolta"),
        ("KYOCERA", "Kyocera"),
        ("BROTHER", "Brother"),
        ("CANON", "Canon"),
        ("PANTUM", "Pantum"),
        ("XEROX", "Xerox"),
        ("LEXMARK", "Lexmark"),
        ("RICOH", "Ricoh"),
        ("SAMSUNG", "Samsung"),
        ("SHARP", "Sharp"),
        ("PANASONIC", "Panasonic"),
        ("EPSON", "Epson"),
        ("OKI", "OKI"),
        ("HP", "HP"),
    ]
    for marker, label in patterns:
        if marker in text:
            return label
    return ""


def parse_dt_safe(value: object) -> Optional[datetime]:
    txt = normalize_text(value)
    if not txt:
        return None
    for parser in (datetime.fromisoformat,):
        try:
            return parser(txt)
        except Exception:
            continue
    return None


def combine_avito_sources(avito_df: pd.DataFrame | None, registry_df: pd.DataFrame | None) -> pd.DataFrame:
    frames: list[pd.DataFrame] = []
    if isinstance(avito_df, pd.DataFrame) and not avito_df.empty:
        cur = avito_df.copy()
        if "registry_key" not in cur.columns:
            cur["registry_key"] = cur.apply(build_avito_registry_key, axis=1)
        if "title_codes" not in cur.columns:
            cur["title_codes"] = cur.get("title", "").map(extract_article_candidates_from_text)
        if "title_norm" not in cur.columns:
            cur["title_norm"] = cur.get("title", "").map(contains_text)
        frames.append(cur)
    if isinstance(registry_df, pd.DataFrame) and not registry_df.empty:
        reg = registry_df.copy()
        if "registry_key" not in reg.columns:
            reg["registry_key"] = reg.apply(build_avito_registry_key, axis=1)
        if "title_codes" not in reg.columns:
            reg["title_codes"] = reg.get("title", "").map(extract_article_candidates_from_text)
        if "title_norm" not in reg.columns:
            reg["title_norm"] = reg.get("title", "").map(contains_text)
        frames.append(reg)
    if not frames:
        return pd.DataFrame()
    merged = pd.concat(frames, ignore_index=True, sort=False)
    merged = merged.drop_duplicates(subset=["registry_key"], keep="first").reset_index(drop=True)
    for col in ["ad_id", "title", "price", "price_raw", "url", "account", "status", "last_changed_at", "last_seen", "first_seen"]:
        if col in merged.columns:
            merged[col] = merged[col].fillna("").map(normalize_text)
    return merged


def build_avito_code_index(avito_df: pd.DataFrame | None) -> tuple[pd.DataFrame, dict[str, list[dict[str, Any]]]]:
    if avito_df is None or avito_df.empty:
        return pd.DataFrame(), {}
    index: dict[str, list[dict[str, Any]]] = defaultdict(list)
    records = avito_df.to_dict(orient="records")
    for rec in records:
        codes = rec.get("title_codes", []) or []
        codes = unique_norm_codes(codes)
        for code in codes:
            index[code].append(rec)
    return avito_df, index


def match_avito_candidates_for_codes(index: dict[str, list[dict[str, Any]]], codes: list[str]) -> list[dict[str, Any]]:
    hits: list[dict[str, Any]] = []
    seen: set[str] = set()
    for code in unique_norm_codes(codes):
        for rec in index.get(code, []):
            key = normalize_text(rec.get("registry_key")) or normalize_text(rec.get("ad_id")) or normalize_text(rec.get("title"))
            if key in seen:
                continue
            seen.add(key)
            hits.append(rec)
    return hits


def safe_days_since(dt_value: object) -> Optional[int]:
    dt = parse_dt_safe(dt_value)
    if not dt:
        return None
    try:
        delta = datetime.now() - dt
        return max(int(delta.days), 0)
    except Exception:
        return None


@st.cache_data(show_spinner=False, ttl=900, max_entries=6)
def build_operational_analytics_bundle(
    sheet_df: pd.DataFrame,
    photo_df: pd.DataFrame | None,
    avito_df: pd.DataFrame | None,
    avito_registry_df: pd.DataFrame | None,
    min_qty: float,
    sheet_name: str,
    hot_items_df: pd.DataFrame | None = None,
) -> dict[str, Any]:
    enriched = apply_photo_map(sheet_df, photo_df)
    if enriched is None or enriched.empty:
        return {}
    merged_avito = combine_avito_sources(avito_df, avito_registry_df)
    _, avito_index = build_avito_code_index(merged_avito)
    hot_lookup = get_cached_hot_watchlist_lookup(hot_items_df, tab_label=sheet_name)

    rows_meta: list[dict[str, Any]] = []
    source_counter: Counter[str] = Counter()
    task_counts = Counter()
    account_rows: list[dict[str, Any]] = []

    for _, row in enriched.iterrows():
        article = normalize_text(row.get("article", ""))
        article_norm = normalize_article(article)
        name = normalize_text(row.get("name", ""))
        own_price = safe_float(row.get("sale_price"), 0.0)
        purchase_avg_cost = safe_float(row.get("purchase_avg_cost"), 0.0)
        own_qty = parse_qty_generic(row.get("free_qty"))
        codes = row.get("row_codes", []) or build_row_compare_codes(article, name)
        matched_ads = match_avito_candidates_for_codes(avito_index, codes)
        ad_count = len(matched_ads)
        accounts = unique_text_values([rec.get("account", "") for rec in matched_ads])
        best_offer = get_best_offer(row, min_qty=min_qty)
        better_market = bool(best_offer and safe_float(best_offer.get("price"), 0.0) > 0 and safe_float(best_offer.get("price"), 0.0) < own_price)
        delta_rub = own_price - safe_float((best_offer or {}).get("price"), 0.0) if better_market else 0.0
        delta_pct = ((delta_rub / own_price) * 100.0) if better_market and own_price > 0 else 0.0
        priority_score = round(max(delta_pct, 0.0) * max(own_qty, 0.0) * max(ad_count, 1), 2)

        hot_rec = pick_hot_watch_rec(row, hot_lookup) if hot_lookup else None
        sales_per_month = safe_float((hot_rec or {}).get("sales_per_month"), 0.0)
        stock_months = round(own_qty / sales_per_month, 1) if sales_per_month > 0 else None
        best_offer_price = safe_float((best_offer or {}).get("price"), 0.0) if best_offer else 0.0
        recommended_price = round(best_offer_price * 0.95, 2) if better_market and best_offer_price > 0 else None

        photo_url = normalize_text(row.get("photo_url", ""))
        has_photo = bool(photo_url)
        has_model = bool(normalize_text(row.get("meta_model", "")))
        has_fits = bool(normalize_text(row.get("meta_fits_models", "")))
        template_ok = has_photo and (has_model or has_fits)
        days_since_change_candidates = [safe_days_since(rec.get("last_changed_at", "")) for rec in matched_ads]
        days_since_change_candidates = [x for x in days_since_change_candidates if x is not None]
        days_since_change = min(days_since_change_candidates) if days_since_change_candidates else None
        stale = bool(days_since_change is not None and days_since_change >= 30)
        weak_ad = ad_count > 0 and not template_ok
        reasons: list[str] = []
        if own_qty > 0 and better_market:
            reasons.append("дорого")
            task_counts["price_review"] += 1
        if own_qty > 0 and not has_photo:
            reasons.append("нет фото")
            task_counts["no_photo"] += 1
        if own_qty > 0 and ad_count == 0:
            reasons.append("нет объявления")
            task_counts["no_avito"] += 1
        elif own_qty > 0 and weak_ad:
            reasons.append("слабое объявление")
            task_counts["weak_avito"] += 1
        if own_qty > 0 and stale:
            reasons.append("давно не обновлялось")
            task_counts["stale"] += 1
        if better_market and best_offer:
            source_counter[normalize_text(best_offer.get("source", ""))] += 1
        brand_guess = infer_brand_from_product_name(name)

        row_meta = {
            "Лист": sheet_name,
            "Артикул": article,
            "Название": name,
            "Бренд": brand_guess,
            "Наша цена": own_price,
            "Наш остаток": own_qty,
            "Продажи, шт/мес": round(sales_per_month, 2) if sales_per_month > 0 else None,
            "Наш запас, мес": stock_months,
            "Лучшая цена дистрибьютора": best_offer_price if best_offer else None,
            "Рекомендую, руб": recommended_price,
            "Лучший поставщик": normalize_text((best_offer or {}).get("source", "")) if best_offer else "",
            "Остаток дистрибьютора": safe_float((best_offer or {}).get("qty"), 0.0) if best_offer else None,
            "Разница, руб": delta_rub if better_market else None,
            "Разница, %": round(delta_pct, 2) if better_market else None,
            "Приоритет": priority_score if better_market else 0.0,
            "Фото": "Да" if has_photo else "Нет",
            "Шаблон": "OK" if template_ok else "Пустой",
            "Модель": "Да" if has_model else "Нет",
            "Подходит к моделям": "Да" if has_fits else "Нет",
            "Объявлений Авито": ad_count,
            "Аккаунты Авито": ", ".join(accounts),
            "Последнее изменение, дней": days_since_change if days_since_change is not None else "",
            "Причины": ", ".join(reasons),
            "article_norm": article_norm,
            "family": split_article_family_suffix(article_norm)[0],
            "color": simplify_template_color(normalize_text(row.get("meta_color", "")) or extract_color_from_text(name)),
        }
        rows_meta.append(row_meta)
        for account in accounts:
            account_rows.append(
                {
                    "Аккаунт": account,
                    "Артикул": article,
                    "Бренд": brand_guess,
                    "Лист": sheet_name,
                    "Есть фото": has_photo,
                    "Есть объявление": ad_count > 0,
                    "Лучше рынка": better_market,
                    "Шаблон OK": template_ok,
                }
            )

    meta_df = pd.DataFrame(rows_meta)
    top_df = meta_df[meta_df["Разница, %"].notna()].copy() if not meta_df.empty else pd.DataFrame()
    if not top_df.empty:
        top_df = top_df.sort_values(["Приоритет", "Разница, %", "Наш остаток"], ascending=[False, False, False]).reset_index(drop=True)

    action_df = meta_df[meta_df["Причины"].map(bool)].copy() if not meta_df.empty else pd.DataFrame()
    if not action_df.empty:
        action_df = action_df.sort_values(["Наш остаток", "Разница, %"], ascending=[False, False], na_position="last").reset_index(drop=True)

    quality = {
        "with_photo": int((meta_df["Фото"] == "Да").sum()) if not meta_df.empty else 0,
        "without_photo": int((meta_df["Фото"] == "Нет").sum()) if not meta_df.empty else 0,
        "template_ok": int((meta_df["Шаблон"] == "OK").sum()) if not meta_df.empty else 0,
        "without_model_or_fits": int(((meta_df["Модель"] == "Нет") | (meta_df["Подходит к моделям"] == "Нет")).sum()) if not meta_df.empty else 0,
        "in_price_not_in_avito": int((meta_df["Объявлений Авито"] == 0).sum()) if not meta_df.empty else 0,
        "in_avito_not_in_stock": int(((meta_df["Объявлений Авито"] > 0) & (pd.to_numeric(meta_df["Наш остаток"], errors="coerce").fillna(0) <= 0)).sum()) if not meta_df.empty else 0,
    }
    quality_df = pd.DataFrame(
        [
            {"Показатель": "С фото", "Количество": quality["with_photo"]},
            {"Показатель": "Без фото", "Количество": quality["without_photo"]},
            {"Показатель": "Шаблон заполнен", "Количество": quality["template_ok"]},
            {"Показатель": "Нет модели / подходит к моделям", "Количество": quality["without_model_or_fits"]},
            {"Показатель": "Есть в прайсе, но нет в Avito", "Количество": quality["in_price_not_in_avito"]},
            {"Показатель": "Есть в Avito, но нет в наличии", "Количество": quality["in_avito_not_in_stock"]},
        ]
    )

    account_df = pd.DataFrame(account_rows)
    if not account_df.empty:
        account_summary = account_df.groupby("Аккаунт", dropna=False).agg(
            Позиций=("Артикул", "count"),
            Без_фото=("Есть фото", lambda s: int((~pd.Series(s)).sum())),
            Дороже_рынка=("Лучше рынка", lambda s: int(pd.Series(s).sum())),
            Слабый_шаблон=("Шаблон OK", lambda s: int((~pd.Series(s)).sum())),
        ).reset_index().rename(columns={"Без_фото": "Без фото", "Дороже_рынка": "Дороже рынка", "Слабый_шаблон": "Слабый шаблон"})
    else:
        account_summary = pd.DataFrame(columns=["Аккаунт", "Позиций", "Без фото", "Дороже рынка", "Слабый шаблон"])

    series_rows: list[dict[str, Any]] = []
    if not meta_df.empty:
        for family, grp in meta_df.groupby("family"):
            if not normalize_text(family) or len(grp) < 2:
                continue
            colors = unique_text_values(grp["color"].tolist())
            issue_count = int((grp["Фото"] == "Нет").sum()) + int((grp["Объявлений Авито"] == 0).sum()) + int(grp["Разница, %"].notna().sum())
            canonical = [c for c in colors if c in {"чёрный", "черный", "голубой", "пурпурный", "жёлтый", "желтый"}]
            missing = []
            palette = ["чёрный", "голубой", "пурпурный", "жёлтый"]
            normalized_colors = {c.replace("ё", "е") for c in canonical}
            if normalized_colors:
                for color in palette:
                    if color.replace("ё", "е") not in normalized_colors:
                        missing.append(color)
            series_rows.append(
                {
                    "Серия": family,
                    "Позиций": len(grp),
                    "Цвета": ", ".join(colors),
                    "Не хватает": ", ".join(missing),
                    "Без фото": int((grp["Фото"] == "Нет").sum()),
                    "Без Avito": int((grp["Объявлений Авито"] == 0).sum()),
                    "Дороже рынка": int(grp["Разница, %"].notna().sum()),
                    "Проблем в серии": issue_count,
                }
            )
    series_df = pd.DataFrame(series_rows)
    if not series_df.empty:
        series_df = series_df.sort_values(["Проблем в серии", "Позиций"], ascending=[False, False]).reset_index(drop=True)

    source_df = pd.DataFrame(
        [{"Источник": source, "Сколько раз лучший": count} for source, count in source_counter.most_common()]
    )

    patch_history_df = load_price_patch_history_df(limit=50)

    tasks_df = pd.DataFrame(
        [
            {"Задача": "Пересмотреть по цене", "Количество": int(task_counts.get("price_review", 0))},
            {"Задача": "Добавить фото", "Количество": int(task_counts.get("no_photo", 0))},
            {"Задача": "Доработать/добавить Avito", "Количество": int(task_counts.get("no_avito", 0) + task_counts.get("weak_avito", 0))},
            {"Задача": "Проверить давно не обновлявшиеся", "Количество": int(task_counts.get("stale", 0))},
            {"Задача": "Проверить неполные серии", "Количество": int((series_df["Не хватает"].map(bool)).sum()) if not series_df.empty else 0},
        ]
    )

    return {
        "meta_df": meta_df,
        "top_df": top_df,
        "action_df": action_df,
        "quality_df": quality_df,
        "account_df": account_summary,
        "series_df": series_df,
        "source_df": source_df,
        "patch_history_df": patch_history_df,
        "tasks_df": tasks_df,
        "quality": quality,
    }


def analytics_bundle_to_excel_bytes(bundle: dict[str, Any]) -> bytes:
    bio = io.BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        for key, sheet_name in [
            ("top_df", "Приоритет цены"),
            ("action_df", "Что делать"),
            ("account_df", "Аккаунты Avito"),
            ("quality_df", "Качество карточек"),
            ("series_df", "Серии"),
            ("source_df", "Источники"),
            ("patch_history_df", "История правок"),
            ("tasks_df", "Сегодня"),
        ]:
            df = bundle.get(key)
            if isinstance(df, pd.DataFrame) and not df.empty:
                df.to_excel(writer, index=False, sheet_name=sheet_name[:31])
    bio.seek(0)
    return bio.read()


def render_operational_analytics_block(sheet_df: pd.DataFrame, photo_df: pd.DataFrame | None, avito_df: pd.DataFrame | None, min_qty: float, sheet_name: str, tab_key: str) -> None:
    registry_df = load_avito_registry_df()
    bundle = build_operational_analytics_bundle(
        sheet_df,
        photo_df,
        avito_df,
        registry_df,
        min_qty,
        sheet_name,
        st.session_state.get("hot_items_df"),
    )
    if not bundle:
        st.info("Для аналитики нет данных.")
        return
    quality = bundle.get("quality", {})
    tasks_df = bundle.get("tasks_df", pd.DataFrame())
    top_df = bundle.get("top_df", pd.DataFrame())
    action_df = bundle.get("action_df", pd.DataFrame())
    account_df = bundle.get("account_df", pd.DataFrame())
    quality_df = bundle.get("quality_df", pd.DataFrame())
    series_df = bundle.get("series_df", pd.DataFrame())
    source_df = bundle.get("source_df", pd.DataFrame())
    patch_history_df = bundle.get("patch_history_df", pd.DataFrame())

    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Без фото", int(quality.get("without_photo", 0)))
    m2.metric("Нет в Avito", int(quality.get("in_price_not_in_avito", 0)))
    m3.metric("Дороже рынка", int(len(top_df) if isinstance(top_df, pd.DataFrame) else 0))
    m4.metric("История ручных правок", int(len(patch_history_df) if isinstance(patch_history_df, pd.DataFrame) else 0))
    st.caption("ⓘ Метрики сверху: Без фото — карточки без изображения; Нет в Avito — позиции без объявления; Дороже рынка — где мы выше лучшего поставщика; История ручных правок — сколько записей накоплено по изменениям цены.")

    render_info_banner(
        "Что делать сегодня",
        "Сначала посмотри приоритет на пересмотр цены, потом проблемные позиции с фото/Avito, потом качество карточек и серии.",
        icon="🧭",
        chips=[f"лист: {sheet_name}", "ленивый расчёт", "ядро поиска не затрагивается"],
        tone="green",
    )

    if isinstance(tasks_df, pd.DataFrame) and not tasks_df.empty:
        st.dataframe(tasks_df, use_container_width=True, hide_index=True)

    with st.expander("1. Приоритет на пересмотр цены ❔", expanded=False):
        st.caption(
            "Что показывает: где мы дороже рынка и что стоит пересмотреть в первую очередь. "
            "Добавлено для решения по цене: продажи в месяц, запас в месяцах и колонка 'Рекомендую, руб' = лучшая цена дистрибьютора минус 5%. "
            "Как пользоваться: смотри сначала товары с хорошими продажами, небольшим запасом и понятной рекомендованной ценой."
        )
        if isinstance(top_df, pd.DataFrame) and not top_df.empty:
            view = top_df[[
                "Артикул",
                "Название",
                "Продажи, шт/мес",
                "Наш запас, мес",
                "Наша цена",
                "Лучшая цена дистрибьютора",
                "Рекомендую, руб",
                "Лучший поставщик",
                "Разница, руб",
                "Разница, %",
                "Наш остаток",
                "Остаток дистрибьютора",
            ]].head(100)
            st.dataframe(view, use_container_width=True, hide_index=True)
            render_analytics_jump_helper(view, tab_key, "top")
        else:
            st.caption("На текущем листе нет позиций, где рынок дешевле нас.")

    with st.expander("2. Что лежит и требует вмешательства ❔", expanded=False):
        st.caption(
            "Что показывает: позиции, где уже есть проблема или задача. "
            "Причины могут быть такие: дорого, нет фото, нет объявления, слабое объявление, давно не обновлялось. "
            "Как пользоваться: это твой список того, что нужно доработать в первую очередь."
        )
        if isinstance(action_df, pd.DataFrame) and not action_df.empty:
            view = action_df[["Артикул", "Название", "Наш остаток", "Причины", "Объявлений Авито", "Фото", "Шаблон", "Лучший поставщик", "Разница, %"]].head(150)
            st.dataframe(view, use_container_width=True, hide_index=True)
            render_analytics_jump_helper(view, tab_key, "action")
        else:
            st.caption("Явных проблемных позиций на текущем листе не найдено.")

    with st.expander("3. Аналитика по аккаунтам Avito ❔", expanded=False):
        st.caption(
            "Что показывает: как позиции распределены по аккаунтам Avito. "
            "Можно понять, на каком аккаунте больше карточек, где больше позиций без фото и где больше дорогих позиций. "
            "Как пользоваться: помогает управлять аккаунтами не на глаз, а по факту."
        )
        if isinstance(account_df, pd.DataFrame) and not account_df.empty:
            st.dataframe(account_df, use_container_width=True, hide_index=True)
        else:
            st.caption("В Avito пока нет данных по аккаунтам для этого листа.")

    with st.expander("4. Покрытие качества карточек ❔", expanded=False):
        st.caption(
            "Что показывает: насколько хорошо заполнены карточки товаров на текущем листе. "
            "Здесь видно, сколько позиций с фото, без фото, с моделью, с полем 'подходит к моделям', есть ли Avito и есть ли остаток. "
            "Как пользоваться: это быстрый контроль качества карточек."
        )
        st.dataframe(quality_df, use_container_width=True, hide_index=True)

    with st.expander("5. Серийная аналитика ❔", expanded=False):
        st.caption(
            "Что показывает: серии товаров, где есть несколько связанных артикулов, цветов или вариантов. "
            "Помогает увидеть неполные серии, серии без фото, без Avito или с ценовыми перекосами. "
            "Как пользоваться: полезно, чтобы не продавать серию обрывками."
        )
        if isinstance(series_df, pd.DataFrame) and not series_df.empty:
            series_view = series_df.head(100)
            st.dataframe(series_view, use_container_width=True, hide_index=True)
            render_analytics_jump_helper(series_view, tab_key, "series")
        else:
            st.caption("На текущем листе не найдено серий, требующих отдельной сводки.")

    with st.expander("6. История ручных правок ❔", expanded=False):
        st.caption(
            "Что показывает: журнал ручных изменений цен. "
            "Видно когда меняли, какой артикул, на каком листе, было / стало и источник изменения. "
            "Как пользоваться: помогает понять, что менялось руками, а что пришло из нового comparison-файла."
        )
        if isinstance(patch_history_df, pd.DataFrame) and not patch_history_df.empty:
            st.dataframe(patch_history_df[["changed_at", "article", "sheet_name", "old_price", "new_price", "change_source", "note"]], use_container_width=True, hide_index=True)
        else:
            st.caption("История ручных правок пока пустая.")

    with st.expander("7. Надёжность источников ❔", expanded=False):
        st.caption(
            "Что показывает: как часто каждый поставщик оказывался лучшим по цене на текущем листе. "
            "Это не гарантия, а практический индикатор, кого стоит чаще мониторить. "
            "Как пользоваться: помогает понять, какие источники чаще всего дают хорошие цены."
        )
        if isinstance(source_df, pd.DataFrame) and not source_df.empty:
            st.dataframe(source_df, use_container_width=True, hide_index=True)
        else:
            st.caption("Пока нет данных по лучшим поставщикам.")

    with st.expander("8. Задачи / напоминания ❔", expanded=False):
        st.caption(
            "Что показывает: задачи по карточкам на пересмотр, проверку или доработку. "
            "Как пользоваться: открывай карточку из задачи, проверяй позицию и отмечай задачу выполненной."
        )
        task_df = build_task_view_df(sheet_name)
        render_tasks_table_ui(task_df, f"analytics_tasks_{tab_key}", default_sheet=sheet_name)

    st.download_button(
        "⬇️ Скачать аналитику в Excel",
        analytics_bundle_to_excel_bytes(bundle),
        file_name=f"analytics_{tab_key}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key=f"download_analytics_{tab_key}",
    )


def build_crm_header_stats(
    sheet_df: pd.DataFrame | None,
    photo_df: pd.DataFrame | None,
    avito_df: pd.DataFrame | None,
    sheet_name: str,
    tab_label: str,
    min_qty: float,
) -> dict[str, int]:
    stats = {
        "tasks_open": 0,
        "tasks_overdue": 0,
        "can_buy": 0,
        "without_photo": 0,
        "without_avito": 0,
    }

    try:
        task_df = load_task_registry_df()
        if isinstance(task_df, pd.DataFrame) and not task_df.empty:
            sheet_mask = (
                task_df.get("sheet_name", pd.Series(dtype=object))
                .fillna("")
                .map(normalize_text)
                .eq(normalize_text(sheet_name))
            )
            task_sheet = task_df.loc[sheet_mask].copy()
            if not task_sheet.empty:
                effective_status = task_sheet.apply(lambda r: task_effective_status(r), axis=1)
                stats["tasks_open"] = int(effective_status.isin(["NEW", "ACTIVE", "OVERDUE"]).sum())
                stats["tasks_overdue"] = int(effective_status.eq("OVERDUE").sum())
    except Exception as exc:
        log_operation(f"CRM header: не удалось посчитать задачи ({normalize_text(exc)})", "warning")

    try:
        hot_buy_df = build_hot_buy_watchlist_table()
        if isinstance(hot_buy_df, pd.DataFrame) and not hot_buy_df.empty and "Лист" in hot_buy_df.columns:
            stats["can_buy"] = int((hot_buy_df["Лист"].fillna("").astype(str) == str(tab_label)).sum())
    except Exception as exc:
        log_operation(f"CRM header: не удалось посчитать 'можно брать' ({normalize_text(exc)})", "warning")

    try:
        if isinstance(sheet_df, pd.DataFrame) and not sheet_df.empty:
            registry_df = load_avito_registry_df()
            bundle = build_operational_analytics_bundle(
                sheet_df,
                photo_df,
                avito_df,
                registry_df,
                float(min_qty),
                str(tab_label),
                st.session_state.get("hot_items_df"),
            )
            quality = bundle.get("quality", {}) if isinstance(bundle, dict) else {}
            stats["without_photo"] = int(quality.get("without_photo", 0) or 0)
            stats["without_avito"] = int(quality.get("in_price_not_in_avito", 0) or 0)
    except Exception as exc:
        log_operation(f"CRM header: не удалось посчитать аналитику ({normalize_text(exc)})", "warning")

    return stats


def render_crm_header_bar(
    sheet_df: pd.DataFrame | None,
    photo_df: pd.DataFrame | None,
    avito_df: pd.DataFrame | None,
    sheet_name: str,
    tab_label: str,
    min_qty: float,
) -> None:
    stats = build_crm_header_stats(sheet_df, photo_df, avito_df, sheet_name, tab_label, min_qty)
    st.markdown('<div class="result-wrap">', unsafe_allow_html=True)
    render_block_header(
        "CRM — быстрый центр управления",
        "Короткая шапка с главными рабочими зонами: задачи, выгодные закупки, проблемы по карточкам.",
        icon="🧭",
        help_text="Эта шапка ничего не меняет в comparison. Она только помогает быстро открыть нужный рабочий блок без долгой прокрутки.",
    )
    c1, c2, c3, c4, c5 = st.columns(5)
    st.caption("ⓘ CRM-шапка — это быстрые переключатели по текущему листу: открыть задачи, выгодные закупки и сразу увидеть, где не хватает фото или Avito.")
    with c1:
        open_tasks = bool(st.checkbox(
            f"🔔 Задачи ({stats['tasks_open']})",
            value=bool(st.session_state.get(f"crm_show_tasks_{sheet_name}", False)),
            key=f"crm_show_tasks_{sheet_name}",
            help="Открывает единый центр задач: новые, активные, просроченные и выполненные.",
        ))
        if open_tasks:
            st.session_state["crm_last_active_sheet_for_tasks"] = str(sheet_name)
        st.caption(f"Просрочено: {stats['tasks_overdue']}")
    with c2:
        open_buy = bool(st.checkbox(
            f"💸 Можно брать ({stats['can_buy']})",
            value=bool(st.session_state.get(f"crm_show_buy_{sheet_name}", False)),
            key=f"crm_show_buy_{sheet_name}",
            help="Открывает ленивую таблицу по ходовым позициям, где поставщик сейчас даёт выгодный вход по цене.",
        ))
        if open_buy:
            st.session_state["crm_last_active_sheet_for_buy"] = str(sheet_name)
        st.caption("Показывает только выгодные позиции")
    with c3:
        open_no_photo = bool(st.checkbox(
            f"🖼️ Нет фото ({stats['without_photo']})",
            value=bool(st.session_state.get(f"crm_show_no_photo_{sheet_name}", False)),
            key=f"crm_show_no_photo_{sheet_name}",
            help="Открывает таблицу только по позициям текущего листа без фото. Отсюда можно сразу открыть и поправить карточку.",
        ))
        if open_no_photo:
            st.session_state["crm_last_active_sheet_for_no_photo"] = str(sheet_name)
        st.caption("Показывает только позиции без фото")
    with c4:
        open_no_avito = bool(st.checkbox(
            f"🛒 Без Avito ({stats['without_avito']})",
            value=bool(st.session_state.get(f"crm_show_no_avito_{sheet_name}", False)),
            key=f"crm_show_no_avito_{sheet_name}",
            help="Открывает таблицу только по позициям текущего листа без объявлений Avito. Отсюда можно перейти в обычный поиск и быстро разместить.",
        ))
        if open_no_avito:
            st.session_state["crm_last_active_sheet_for_no_avito"] = str(sheet_name)
        st.caption("Показывает только позиции без Avito")
    with c5:
        st.metric("Лист", tab_label)
        st.caption("CRM-метрики именно по текущему листу")
    st.markdown('</div>', unsafe_allow_html=True)


def render_crm_card_center(
    result_df: pd.DataFrame | None,
    display_result_df: pd.DataFrame | None,
    compare_map: dict[str, dict[str, Any]] | None,
    avito_df: pd.DataFrame | None,
    sheet_name: str,
    tab_label: str,
    tab_key: str,
    price_mode: str,
    round100: bool,
    custom_discount: float,
) -> None:
    if not isinstance(display_result_df, pd.DataFrame) or display_result_df.empty:
        return

    rows = display_result_df.copy()
    options = []
    option_map: dict[str, pd.Series] = {}
    for _, row in rows.iterrows():
        art = normalize_text(row.get("article", ""))
        name = normalize_text(row.get("name", ""))
        label = f"{art} — {name[:120]}"
        options.append(label)
        option_map[label] = row

    st.markdown('<div class="result-wrap">', unsafe_allow_html=True)
    render_block_header(
        "CRM-карточка товара",
        "Один компактный экран по позиции: обзор, цены, Avito, заметки и задачи. Старые блоки остаются ниже как вторичный режим.",
        icon="🧩",
        help_text="Нужна, чтобы не бегать по странице: всё важное по найденной позиции собрано в одном месте.",
    )
    selected_label = st.selectbox(
        "Позиция для CRM-карточки",
        options,
        index=0 if options else None,
        key=f"crm_card_select_{tab_key}",
        help="Выбери найденную позицию, чтобы открыть её краткую CRM-карточку.",
    )
    if not selected_label:
        st.markdown('</div>', unsafe_allow_html=True)
        return

    row = option_map[selected_label]
    art = normalize_text(row.get("article", ""))
    art_norm = normalize_text(row.get("article_norm", ""))
    name = normalize_text(row.get("name", ""))
    photo_url = normalize_text(row.get("photo_url", ""))
    note = normalize_text(row.get("manual_note", ""))
    brand = normalize_text(row.get("meta_brand", ""))
    model = normalize_text(row.get("meta_model", ""))
    mcode = normalize_text(row.get("meta_manufacturer_code", ""))
    print_type = normalize_text(row.get("meta_print_type", ""))
    color = normalize_text(row.get("meta_color", ""))
    capacity = normalize_text(row.get("meta_capacity", ""))
    iso_pages = normalize_pages_value(row.get("meta_iso_pages", ""))
    item_type = normalize_text(row.get("meta_item_type", ""))
    print_technology = normalize_text(row.get("meta_print_technology", ""))
    description = normalize_text(row.get("meta_description", ""))
    fits = normalize_text(row.get("meta_fits_models", ""))
    weight = format_meta_weight(row.get("meta_weight", ""))
    dimensions = format_meta_dimensions(row.get("meta_length", ""), row.get("meta_width", ""), row.get("meta_height", ""))
    own_price = safe_float(row.get("sale_price"), 0.0)
    own_stock = parse_qty_generic(row.get("free_qty"))
    best = get_best_offer(row, min_qty=float(st.session_state.get("distributor_min_qty", 1.0)))

    # В CRM-карточке "Лучший поставщик" показываем только если он реально дешевле нашей цены.
    # Если рынок дороже нас, не вводим в заблуждение и не показываем поставщика как "лучшего".
    best_delta = safe_float((best or {}).get("delta"), 0.0)
    best_is_better_than_us = bool(best) and best_delta > 0

    best_source = normalize_text((best or {}).get("source", "")) if best_is_better_than_us else ""
    best_price = safe_float((best or {}).get("price"), 0.0) if best_is_better_than_us else 0.0
    best_qty = safe_float((best or {}).get("qty"), 0.0) if best_is_better_than_us else 0.0
    matched_ads = pd.DataFrame()
    if isinstance(avito_df, pd.DataFrame) and not avito_df.empty:
        one_row = pd.DataFrame([row.to_dict()])
        matched_ads = find_avito_ads(avito_df, one_row)

    t_overview, t_prices, t_avito, t_notes = st.tabs(["Обзор", "Цены", "Avito", "Заметки / задачи"])
    st.caption("ⓘ Обзор — краткая карточка товара. Цены — наша цена и лучший рынок. Avito — объявления по позиции. Заметки / задачи — ручные комментарии и напоминания.")

    with t_overview:
        c1, c2 = st.columns([1, 2])
        with c1:
            if photo_url:
                st.image(photo_url, use_container_width=True)
            else:
                st.info("Фото не найдено")
        with c2:
            st.markdown(f"### {html.escape(art)}")
            st.caption(name)
            m1, m2, m3 = st.columns(3)
            m1.metric("Наша цена", fmt_price(own_price))
            m2.metric("Наш склад", fmt_qty(own_stock))
            m3.metric("Лучший поставщик", best_source or "—")
            if note:
                st.info(f"Заметка: {note}")

            quick_left = [
                ("Бренд", brand),
                ("Модель", model),
                ("Код производителя", mcode),
                ("Тип печати", print_type),
                ("Цвет", color),
                ("Емкость", capacity),
            ]
            quick_right = [
                ("Ресурс, стр.", iso_pages),
                ("Тип", item_type),
                ("Технология", print_technology),
                ("Вес", weight),
                ("Габариты", dimensions),
            ]

            compact_pairs = [
                ("Бренд", brand),
                ("Модель", model),
                ("Код", mcode),
                ("Цвет", color),
                ("Ресурс", iso_pages),
                ("Тип", item_type),
            ]
            compact_html = []
            for label, value in compact_pairs:
                if value:
                    compact_html.append(
                        f"<span style='display:inline-block;margin:0 8px 8px 0;padding:4px 10px;border:1px solid rgba(120,130,160,.25);border-radius:999px;background:rgba(120,130,160,.08);font-size:.92rem;'><b>{html.escape(label)}:</b> {html.escape(str(value))}</span>"
                    )
            if compact_html:
                st.markdown("".join(compact_html), unsafe_allow_html=True)

            if any(v for _, v in quick_left + quick_right) or fits or description:
                with st.expander("Характеристики", expanded=False):
                    i1, i2 = st.columns(2)
                    with i1:
                        for label, value in quick_left:
                            if value:
                                st.markdown(f"**{label}:** {value}")
                    with i2:
                        for label, value in quick_right:
                            if value:
                                st.markdown(f"**{label}:** {value}")
                    if fits:
                        st.markdown(f"**Подходит к моделям:** {fits}")
                    if description:
                        st.caption(f"Описание: {description}")

    with t_prices:
        st.caption("Здесь видно нашу цену и лучший рынок по текущей позиции. Это быстрый обзор, а полный блок 'Показать цены у всех' остаётся ниже.")
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Наша цена", fmt_price(own_price))
        c2.metric("Лучший поставщик", best_source or "—")
        c3.metric("Цена поставщика", fmt_price(best_price) if best_price > 0 else "—")
        c4.metric("Остаток поставщика", fmt_qty(best_qty) if best_qty > 0 else "—")
        if best and not best_is_better_than_us:
            st.info("Сейчас дешевле нашей цены поставщика нет.")
        if isinstance(compare_map, dict):
            data = compare_map.get(art_norm) or compare_map.get(art) or {}
            if data:
                rows_prices = []
                for item in data.get("offers", [])[:20]:
                    rows_prices.append({
                        "Поставщик": normalize_text(item.get("source", "")),
                        "Цена": fmt_price(safe_float(item.get("price"), 0.0)),
                        "Остаток": fmt_qty(safe_float(item.get("qty"), 0.0)),
                    })
                if rows_prices:
                    st.dataframe(pd.DataFrame(rows_prices), use_container_width=True, hide_index=True)

    with t_avito:
        st.caption("Связанные объявления Avito по этой позиции. Отсюда можно быстро понять, на каком аккаунте она висит и есть ли вообще объявление.")
        if isinstance(matched_ads, pd.DataFrame) and not matched_ads.empty:
            try:
                render_avito_block(avito_df, pd.DataFrame([row.to_dict()]))
            except Exception:
                view_cols = [c for c in ["ad_id", "title", "account", "price", "last_changed_at"] if c in matched_ads.columns]
                st.dataframe(matched_ads[view_cols], use_container_width=True, hide_index=True)
        else:
            st.info("Связанные объявления Avito не найдены.")

    with t_notes:
        st.caption("Здесь можно быстро поправить карточку и сразу создать задачу на пересмотр без переходов вниз по странице.")
        current_photo = photo_url
        current_name = name
        current_brand = brand
        current_model = model
        current_code = mcode
        current_print_type = print_type
        current_color = color
        current_capacity = capacity
        current_iso_pages = iso_pages
        current_item_type = item_type
        current_print_technology = print_technology
        current_description = description
        current_fits = fits
        current_weight = normalize_text(row.get("meta_weight", ""))
        current_length = normalize_text(row.get("meta_length", ""))
        current_width = normalize_text(row.get("meta_width", ""))
        current_height = normalize_text(row.get("meta_height", ""))
        current_note = note

        with st.form(f"crm_card_form_{tab_key}_{art_norm}", clear_on_submit=False):
            cc1, cc2 = st.columns([1.2, 1.8])
            with cc1:
                photo_url_new = st.text_input("Фото (ссылка)", value=current_photo, key=f"crm_card_photo_{tab_key}_{art_norm}")
            with cc2:
                name_new = st.text_area("Название", value=current_name, height=90, key=f"crm_card_name_{tab_key}_{art_norm}")
            mm1, mm2, mm3 = st.columns(3)
            brand_new = mm1.text_input("Бренд", value=current_brand, key=f"crm_card_brand_{tab_key}_{art_norm}")
            model_new = mm2.text_input("Модель", value=current_model, key=f"crm_card_model_{tab_key}_{art_norm}")
            code_new = mm3.text_input("Код производителя", value=current_code, key=f"crm_card_code_{tab_key}_{art_norm}")

            mm4, mm5, mm6 = st.columns(3)
            print_type_new = mm4.text_input("Тип печати", value=current_print_type, key=f"crm_card_print_type_{tab_key}_{art_norm}")
            color_new = mm5.text_input("Цвет", value=current_color, key=f"crm_card_color_{tab_key}_{art_norm}")
            capacity_new = mm6.text_input("Емкость", value=current_capacity, key=f"crm_card_capacity_{tab_key}_{art_norm}")

            mm7, mm8, mm9 = st.columns(3)
            iso_pages_new = mm7.text_input("Ресурс, стр.", value=current_iso_pages, key=f"crm_card_iso_pages_{tab_key}_{art_norm}")
            item_type_new = mm8.text_input("Тип", value=current_item_type, key=f"crm_card_item_type_{tab_key}_{art_norm}")
            print_technology_new = mm9.text_input("Технология", value=current_print_technology, key=f"crm_card_print_technology_{tab_key}_{art_norm}")

            mm10, mm11, mm12, mm13 = st.columns(4)
            weight_new = mm10.text_input("Вес", value=current_weight, key=f"crm_card_weight_{tab_key}_{art_norm}")
            length_new = mm11.text_input("Длина", value=current_length, key=f"crm_card_length_{tab_key}_{art_norm}")
            width_new = mm12.text_input("Ширина", value=current_width, key=f"crm_card_width_{tab_key}_{art_norm}")
            height_new = mm13.text_input("Высота", value=current_height, key=f"crm_card_height_{tab_key}_{art_norm}")

            fits_new = st.text_area("Подходит к моделям", value=current_fits, height=75, key=f"crm_card_fits_{tab_key}_{art_norm}")
            description_new = st.text_area("Описание", value=current_description, height=70, key=f"crm_card_description_{tab_key}_{art_norm}")
            note_new = st.text_area("Заметка", value=current_note, height=65, key=f"crm_card_note_{tab_key}_{art_norm}")

            st.markdown("#### 🔔 Создать задачу по карточке")
            st.caption("Заметка — это просто комментарий. Задача — отдельное напоминание со сроком, статусом и причиной.")
            tt1, tt2, tt3 = st.columns([1.1, 1.4, 1.7])
            make_task = tt1.checkbox(
                "Создать задачу",
                key=f"crm_card_make_task_{tab_key}_{art_norm}",
                help="Создаёт напоминание по этой позиции. Оно попадёт в общий центр задач и не пропадёт после загрузки нового файла.",
            )
            due_date = tt2.date_input(
                "Когда проверить",
                value=(datetime.utcnow().date() + timedelta(days=14)),
                key=f"crm_card_due_{tab_key}_{art_norm}",
            )
            task_reason = tt3.selectbox(
                "Причина",
                ["Пересмотреть цену", "Проверить после правки", "Нет продаж", "Проверить фото/карточку", "Проверить спрос", "Другое"],
                key=f"crm_card_reason_{tab_key}_{art_norm}",
            )
            task_note = st.text_area(
                "Комментарий к задаче",
                value="",
                height=65,
                key=f"crm_card_task_note_{tab_key}_{art_norm}",
                placeholder="Например: снизили цену, проверить продажи через 14 дней.",
            )

            sb1, sb2 = st.columns(2)
            save_clicked = sb1.form_submit_button("💾 Сохранить карточку", use_container_width=True, type="primary")
            reset_clicked = sb2.form_submit_button("↺ Сбросить ручные правки", use_container_width=True)

        if save_clicked:
            save_card_override(
                sheet_name,
                art,
                art_norm,
                {
                    "photo_url": photo_url_new,
                    "name_override": name_new,
                    "meta_brand": brand_new,
                    "meta_model": model_new,
                    "meta_manufacturer_code": code_new,
                    "meta_print_type": print_type_new,
                    "meta_color": color_new,
                    "meta_capacity": capacity_new,
                    "meta_iso_pages": iso_pages_new,
                    "meta_item_type": item_type_new,
                    "meta_print_technology": print_technology_new,
                    "meta_description": description_new,
                    "meta_fits_models": fits_new,
                    "meta_weight": weight_new,
                    "meta_length": length_new,
                    "meta_width": width_new,
                    "meta_height": height_new,
                    "note": note_new,
                },
            )
            if make_task:
                create_review_task(
                    article=art,
                    article_norm=art_norm,
                    sheet_name=sheet_name,
                    name_snapshot=name_new or current_name,
                    due_date=due_date,
                    reason=task_reason,
                    note=task_note or note_new,
                    source="crm_card",
                )
            st.success(f"Карточка {art} сохранена.")
            if make_task:
                st.info(f"Задача по {art} создана до {due_date}.")
            st.rerun()

        if reset_clicked:
            delete_card_override(sheet_name, art_norm)
            st.success(f"Ручные правки для {art} сброшены.")
            st.rerun()

    st.markdown('</div>', unsafe_allow_html=True)


CRM_SHEET_NAME_TO_LABEL = {"Сравнение": "Оригинал", "Уценка": "Уценка", "Совместимые": "Совместимые"}
CRM_SHEET_LABEL_TO_NAME = {v: k for k, v in CRM_SHEET_NAME_TO_LABEL.items()}
CRM_SHEET_LABEL_TO_TABKEY = {"Оригинал": "original", "Уценка": "discount", "Совместимые": "compatible"}


def get_pipeline_registry_path() -> Path:
    try:
        return Path(__file__).resolve().with_name("crm_pipeline.sqlite")
    except Exception:
        return Path.cwd() / "crm_pipeline.sqlite"


def ensure_pipeline_registry_db() -> None:
    path = get_pipeline_registry_path()
    path.parent.mkdir(parents=True, exist_ok=True)
    with sqlite3.connect(path) as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS crm_pipeline (
                sheet_name TEXT NOT NULL,
                article_norm TEXT NOT NULL,
                article TEXT,
                pipeline_status TEXT,
                current_queue TEXT,
                manual_decision TEXT,
                workflow_stage TEXT,
                next_action TEXT,
                owner TEXT,
                priority TEXT,
                updated_at TEXT,
                PRIMARY KEY (sheet_name, article_norm)
            )
            """
        )
        conn.commit()


@st.cache_data(ttl=300, max_entries=4)
def load_pipeline_registry_df() -> pd.DataFrame:
    path = get_pipeline_registry_path()
    if not path.exists():
        return pd.DataFrame(
            columns=[
                "sheet_name", "article_norm", "article", "pipeline_status", "current_queue",
                "manual_decision", "workflow_stage", "next_action", "owner", "priority", "updated_at",
            ]
        )
    ensure_pipeline_registry_db()
    with sqlite3.connect(path) as conn:
        df = pd.read_sql_query("SELECT * FROM crm_pipeline ORDER BY updated_at DESC", conn)
    if df.empty:
        return df
    for col in [
        "sheet_name", "article_norm", "article", "pipeline_status", "current_queue",
        "manual_decision", "workflow_stage", "next_action", "owner", "priority", "updated_at",
    ]:
        if col in df.columns:
            df[col] = df[col].fillna("").map(normalize_text)
    return df


def clear_pipeline_registry_cache() -> None:
    try:
        load_pipeline_registry_df.clear()
    except Exception:
        pass
    clear_runtime_perf_caches()


def upsert_pipeline_registry(
    sheet_name: str,
    article: str,
    article_norm: str,
    pipeline_status: str = "",
    current_queue: str = "",
    manual_decision: str = "",
    workflow_stage: str = "",
    next_action: str = "",
    owner: str = "",
    priority: str = "",
) -> None:
    ensure_pipeline_registry_db()
    now = datetime.utcnow().replace(microsecond=0).isoformat() + "Z"
    with sqlite3.connect(get_pipeline_registry_path()) as conn:
        conn.execute(
            """
            INSERT INTO crm_pipeline (
                sheet_name, article_norm, article, pipeline_status, current_queue,
                manual_decision, workflow_stage, next_action, owner, priority, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(sheet_name, article_norm) DO UPDATE SET
                article=excluded.article,
                pipeline_status=excluded.pipeline_status,
                current_queue=excluded.current_queue,
                manual_decision=excluded.manual_decision,
                workflow_stage=excluded.workflow_stage,
                next_action=excluded.next_action,
                owner=excluded.owner,
                priority=excluded.priority,
                updated_at=excluded.updated_at
            """,
            (
                normalize_text(sheet_name),
                normalize_text(article_norm),
                normalize_text(article),
                normalize_text(pipeline_status),
                normalize_text(current_queue),
                normalize_text(manual_decision),
                normalize_text(workflow_stage),
                normalize_text(next_action),
                normalize_text(owner),
                normalize_text(priority),
                now,
            ),
        )
        conn.commit()
    clear_pipeline_registry_cache()


def build_supplier_debug_rows(row: pd.Series | dict[str, Any], min_qty: float = 1.0) -> list[dict[str, Any]]:
    debug_rows: list[dict[str, Any]] = []
    for pair in row.get("source_pairs", []) or []:
        source = normalize_text(pair.get("source", ""))
        price_col = normalize_text(pair.get("price_col", ""))
        qty_col = normalize_text(pair.get("qty_col", ""))
        raw_price = row.get(pair.get("price_col", ""), "")
        raw_qty = row.get(pair.get("qty_col", ""), "")
        price = safe_float(raw_price, 0.0)
        price = normalize_merlion_source_price(row, source, price)
        qty = parse_qty_generic(raw_qty)
        status = "OK"
        reason = "Проходит в сравнение"
        if not source:
            status = "SKIP"
            reason = "Не распознан источник"
        elif price <= 0 and qty <= 0:
            status = "SKIP"
            reason = "Нет цены и остатка"
        elif price <= 0:
            status = "SKIP"
            reason = "Нет валидной цены"
        elif qty < float(min_qty):
            status = "SKIP"
            reason = f"Остаток ниже порога ({fmt_qty(qty)} < {fmt_qty(min_qty)})"
        debug_rows.append({
            "Поставщик": source or "—",
            "Колонка цены": price_col or "—",
            "Колонка остатка": qty_col or "—",
            "Сырая цена": normalize_text(raw_price),
            "Сырой остаток": normalize_text(raw_qty),
            "Цена": price if price > 0 else None,
            "Остаток": qty if qty > 0 else 0.0,
            "Статус": status,
            "Почему": reason,
        })
    return debug_rows


def pick_recommended_price_for_crm(own_price: float, best_supplier_price: float, is_hot: bool, is_dead_stock: bool, has_market_gap: bool) -> float | None:
    if own_price <= 0:
        return None
    if is_dead_stock:
        return round(max(own_price * 0.9, 0.0), 2)
    if best_supplier_price > 0 and has_market_gap:
        return round(best_supplier_price * (0.97 if is_hot else 0.99), 2)
    if is_hot:
        return round(own_price, 2)
    return None


def build_crm_workspace_products_df(
    sheet_df: pd.DataFrame | None,
    photo_df: pd.DataFrame | None,
    avito_df: pd.DataFrame | None,
    min_qty: float,
    sheet_name: str,
    sheet_label: str,
) -> pd.DataFrame:
    if not isinstance(sheet_df, pd.DataFrame) or sheet_df.empty:
        return pd.DataFrame()

    enriched = apply_photo_map(sheet_df, photo_df)
    enriched = apply_purchase_cost_map(enriched, st.session_state.get("purchase_cost_df")) if isinstance(enriched, pd.DataFrame) else enriched
    enriched = apply_card_overrides(enriched, sheet_name) if isinstance(enriched, pd.DataFrame) else enriched
    if not isinstance(enriched, pd.DataFrame) or enriched.empty:
        return pd.DataFrame()

    registry_df = load_avito_registry_df()
    merged_avito = combine_avito_sources(avito_df, registry_df)
    _, avito_index = build_avito_code_index(merged_avito)
    hot_lookup = get_cached_hot_watchlist_lookup(st.session_state.get("hot_items_df"), tab_label=sheet_label)

    task_df = load_task_registry_df()
    task_lookup: dict[tuple[str, str], int] = {}
    if isinstance(task_df, pd.DataFrame) and not task_df.empty:
        for _, task_row in task_df.iterrows():
            task_sheet = normalize_text(task_row.get("sheet_name", ""))
            task_art = normalize_text(task_row.get("article_norm", ""))
            if not task_sheet or not task_art:
                continue
            if task_effective_status(task_row) in {"NEW", "ACTIVE", "OVERDUE"}:
                task_lookup[(task_sheet, task_art)] = task_lookup.get((task_sheet, task_art), 0) + 1

    pipe_df = load_pipeline_registry_df()
    pipe_lookup: dict[tuple[str, str], dict[str, Any]] = {}
    if isinstance(pipe_df, pd.DataFrame) and not pipe_df.empty:
        for _, pipe_row in pipe_df.iterrows():
            pipe_lookup[(normalize_text(pipe_row.get("sheet_name", "")), normalize_text(pipe_row.get("article_norm", "")))] = pipe_row.to_dict()

    rows: list[dict[str, Any]] = []
    buy_gap_threshold_pct = float(st.session_state.get("distributor_threshold", 35.0) or 35.0)

    for _, row in enriched.iterrows():
        article = normalize_text(row.get("article", ""))
        article_norm = normalize_text(row.get("article_norm", ""))
        if not article_norm:
            continue
        name = normalize_text(row.get("name", ""))
        own_price = safe_float(row.get("sale_price"), 0.0)
        own_qty = parse_qty_generic(row.get("free_qty"))
        total_qty = parse_qty_generic(row.get("total_qty"))
        transit_qty = parse_qty_generic(row.get("transit_qty"))
        codes = row.get("row_codes", []) or build_row_compare_codes(article, name)
        matched_ads = match_avito_candidates_for_codes(avito_index, codes)
        ad_count = len(matched_ads)
        best_offer = get_best_offer(row, min_qty=min_qty)
        best_price = safe_float((best_offer or {}).get("price"), 0.0)
        best_qty = safe_float((best_offer or {}).get("qty"), 0.0)
        has_cheaper_supplier = bool(best_offer and best_price > 0 and own_price > 0 and best_price < own_price)
        market_gap_pct = safe_float((best_offer or {}).get("delta_percent"), 0.0) if has_cheaper_supplier else 0.0

        hot_rec = pick_hot_watch_rec(row, hot_lookup) if hot_lookup else None
        sales_per_month = safe_float((hot_rec or {}).get("sales_per_month"), 0.0)
        sales_qty_15m = safe_float((hot_rec or {}).get("sales_qty_15m"), 0.0)
        abc_class = normalize_text((hot_rec or {}).get("abc_class", "")).upper()
        velocity_band = normalize_text((hot_rec or {}).get("velocity_band", ""))
        days_of_cover = safe_float((hot_rec or {}).get("days_of_cover"), 0.0)
        hot_buy_signal = normalize_text((hot_rec or {}).get("buy_signal_30pct", "")).upper()
        hot_action_today = normalize_text((hot_rec or {}).get("action_today", ""))
        hot_priority_score = safe_float((hot_rec or {}).get("priority_score"), 0.0)
        hot_best_supplier_gap_pct = normalize_gap_percent((hot_rec or {}).get("best_supplier_gap_pct"))
        sales_per_day = safe_float((hot_rec or {}).get("sales_per_day"), 0.0)
        sales_per_week = safe_float((hot_rec or {}).get("sales_per_week"), 0.0)
        sales_per_year = safe_float((hot_rec or {}).get("sales_per_year"), 0.0)
        deals_count = safe_int((hot_rec or {}).get("deals_count", 0), 0)
        first_sale = normalize_text((hot_rec or {}).get("first_sale", ""))
        last_sale = normalize_text((hot_rec or {}).get("last_sale", ""))
        days_without_sales = safe_float((hot_rec or {}).get("days_without_sales"), 0.0)
        market_min_price = safe_float((hot_rec or {}).get("market_min_price"), 0.0)
        market_min_supplier = normalize_text((hot_rec or {}).get("market_min_supplier", ""))
        supplier_presence_text = normalize_text((hot_rec or {}).get("supplier_presence_text", ""))
        stats_source_kind = normalize_text((hot_rec or {}).get("stats_source_kind", ""))
        stock_months = round(own_qty / sales_per_month, 2) if sales_per_month > 0 else None

        has_photo = bool(normalize_text(row.get("photo_url", "")))
        has_avito = ad_count > 0
        has_model_or_fits = bool(normalize_text(row.get("meta_model", "")) or normalize_text(row.get("meta_fits_models", "")) or normalize_text(row.get("meta_brand", "")) or normalize_text(row.get("meta_manufacturer_code", "")))
        ready_for_marketplace = bool(own_qty > 0 and has_photo and has_model_or_fits and not has_avito)

        is_hot = bool(sales_per_month >= 2.0 or abc_class in {"A", "B"} or hot_buy_signal == "BUY")
        is_strong_demand = bool(abc_class in {"A", "B"})
        is_low_stock = bool(stock_months is not None and stock_months < 1.0)
        is_caution_stock = bool(stock_months is not None and 2.0 < stock_months <= 6.0)
        is_stale_risk = bool(stock_months is not None and stock_months > 6.0)
        is_dead_stock = bool(
            own_qty > 0
            and (
                (stock_months is not None and stock_months > 12.0)
                or ((stock_months is not None and stock_months > 6.0) and sales_per_month <= 1.0)
                or (sales_per_month <= 0.3 and own_qty > 0)
            )
        )
        is_overstock = bool(stock_months is not None and stock_months > 6.0)
        effective_gap_pct = max(market_gap_pct, hot_best_supplier_gap_pct)
        raw_can_buy = bool(
            (is_hot or is_strong_demand or hot_buy_signal == "BUY")
            and best_price > 0
            and best_qty > 0
            and effective_gap_pct >= buy_gap_threshold_pct
        )
        can_buy = bool(raw_can_buy and not is_stale_risk and not is_dead_stock)
        needs_price_review = bool((best_price > 0 and market_gap_pct > 0) or is_stale_risk or is_dead_stock)

        stock_action = "Проверить вручную"
        price_action = "Оставить цену"
        placement_action = "Ничего"
        decision = "Проверить вручную"
        reason = "Недостаточно явного сигнала"

        if is_dead_stock:
            stock_action = "Распродавать"
        elif is_stale_risk:
            stock_action = "Не закупать"
        elif can_buy and is_low_stock:
            stock_action = "Пополнить запас"
        elif can_buy:
            stock_action = "Можно закупать"
        elif is_hot:
            stock_action = "Держать на складе"
        elif is_caution_stock:
            stock_action = "Держать остаток"

        if is_dead_stock:
            price_action = "Снизить цену"
        elif is_stale_risk:
            price_action = "Пересмотреть цену"
        elif best_price > 0 and market_gap_pct > 0:
            price_action = "Пересмотреть цену"
        elif is_hot and is_low_stock:
            price_action = "Можно держать"

        if ready_for_marketplace:
            placement_action = "Разместить на Avito"
        elif not has_photo and own_qty > 0:
            placement_action = "Добавить фото"
        elif not has_avito and own_qty > 0:
            placement_action = "Добавить Avito"

        translated_watch_action = translate_watch_action(hot_action_today, threshold_pct=buy_gap_threshold_pct)
        if is_dead_stock:
            decision = "Распродавать"
            reason = (
                f"Слабый спрос ({sales_per_month:.2f} шт/мес) и запас {stock_months:.2f} мес"
                if stock_months is not None else
                f"Слабый спрос ({sales_per_month:.2f} шт/мес) и товар застрял на складе"
            )
        elif is_stale_risk:
            decision = "Не покупать"
            reason = f"Запас уже высокий ({stock_months:.2f} мес) — сначала нужно разгрузить остаток"
        elif can_buy and is_low_stock:
            decision = "Можно закупать"
            reason = f"Ходовой товар, запас низкий, поставщик ниже нас на {effective_gap_pct:.1f}%"
        elif can_buy:
            decision = "Можно закупать"
            reason = f"Ходовой товар и выгодный вход от поставщика ({effective_gap_pct:.1f}%)"
        elif translated_watch_action and translated_watch_action != "Можно брать (-35%+)":
            decision = translated_watch_action
            reason = "Сигнал пришёл из watchlist по продажам и запасу"
        elif translated_watch_action == "Можно брать (-35%+)" and is_stale_risk:
            decision = "Не покупать"
            reason = f"Watchlist даёт BUY-сигнал, но у нас уже высокий запас ({stock_months:.2f} мес)"
        elif best_price > 0 and market_gap_pct > 0:
            decision = "Пересмотреть цену"
            reason = f"Наша цена выше рынка на {market_gap_pct:.1f}%"
        elif ready_for_marketplace:
            decision = "Разместить на Avito"
            reason = "Карточка готова, но объявления пока нет"
        elif not has_photo and own_qty > 0:
            decision = "Добавить фото"
            reason = "Товар есть на складе, но фото отсутствует"
        elif not has_avito and own_qty > 0:
            decision = "Добавить Avito"
            reason = "Товар есть на складе, но объявления нет"
        elif is_hot:
            decision = "Держать на складе"
            reason = f"Товар ходовой ({sales_per_month:.2f} шт/мес)"
        elif is_caution_stock:
            decision = "Держать остаток"
            reason = f"Запас уже заметный ({stock_months:.2f} мес), закупку лучше не ускорять"

        supplier_debug_rows = build_supplier_debug_rows(row, min_qty=min_qty)
        supplier_valid_offers = get_row_offers(row, min_qty=min_qty)

        purchase_avg_cost = safe_float(row.get("purchase_avg_cost"), 0.0)
        purchase_total_qty = safe_float(row.get("purchase_total_qty"), 0.0)
        purchase_total_cost = safe_float(row.get("purchase_total_cost"), 0.0)
        purchase_match_source = normalize_text(row.get("purchase_match_source", ""))
        purchase_source_name = normalize_text(row.get("purchase_source_name", ""))
        purchase_source_sheet = normalize_text(row.get("purchase_source_sheet", ""))

        priority = 0.0
        priority += hot_priority_score
        if can_buy:
            priority += 100.0
        if is_hot:
            priority += 30.0
        if is_strong_demand:
            priority += 20.0
        priority += sales_per_month * 5.0
        priority += max(effective_gap_pct, 0.0)
        priority += min(own_qty, 50.0) * 0.5
        if is_dead_stock:
            priority += 40.0
        if not has_photo and own_qty > 0:
            priority += 20.0
        if not has_avito and own_qty > 0:
            priority += 20.0
        if ready_for_marketplace:
            priority += 25.0

        default_queue = "Под наблюдением"
        if decision == "Можно закупать":
            default_queue = "Можно брать"
        elif decision in {"Пересмотреть цену", "Распродавать", "Проверить запас"}:
            default_queue = "Требует цены"
        elif decision == "Добавить фото":
            default_queue = "Без фото"
        elif decision in {"Добавить Avito", "Разместить на Avito"}:
            default_queue = "Без Avito"
        elif stock_action in {"Пополнить запас", "Можно закупать"}:
            default_queue = "К пополнению"
        elif is_dead_stock:
            default_queue = "Залежалый остаток"

        pipe = pipe_lookup.get((normalize_text(sheet_name), article_norm), {})
        pipeline_status = normalize_text(pipe.get("pipeline_status", "")) or "Новая"
        current_queue = normalize_text(pipe.get("current_queue", "")) or default_queue
        manual_decision = normalize_text(pipe.get("manual_decision", ""))
        workflow_stage = normalize_text(pipe.get("workflow_stage", "")) or "Проверка"
        next_action = normalize_text(pipe.get("next_action", "")) or decision
        owner = normalize_text(pipe.get("owner", ""))
        priority_label = normalize_text(pipe.get("priority", "")) or ("Высокий" if priority >= 120 else "Средний" if priority >= 60 else "Низкий")

        rows.append({
            "sheet_name": normalize_text(sheet_name),
            "sheet_label": normalize_text(sheet_label),
            "article": article,
            "article_norm": article_norm,
            "name": name,
            "sale_price": own_price,
            "purchase_avg_cost": purchase_avg_cost if purchase_avg_cost > 0 else None,
            "purchase_total_qty": purchase_total_qty if purchase_total_qty > 0 else None,
            "purchase_total_cost": purchase_total_cost if purchase_total_cost > 0 else None,
            "purchase_match_source": purchase_match_source,
            "purchase_source_name": purchase_source_name,
            "purchase_source_sheet": purchase_source_sheet,
            "free_qty": own_qty,
            "total_qty": total_qty,
            "transit_qty": transit_qty,
            "sales_per_month": round(sales_per_month, 2),
            "sales_qty_15m": round(sales_qty_15m, 2),
            "sales_per_day": round(sales_per_day, 4) if sales_per_day > 0 else 0.0,
            "sales_per_week": round(sales_per_week, 4) if sales_per_week > 0 else 0.0,
            "sales_per_year": round(sales_per_year, 2) if sales_per_year > 0 else 0.0,
            "deals_count": deals_count,
            "first_sale": first_sale,
            "last_sale": last_sale,
            "days_without_sales": days_without_sales if days_without_sales > 0 else None,
            "market_min_price": market_min_price if market_min_price > 0 else None,
            "market_min_supplier": market_min_supplier,
            "supplier_presence_text": supplier_presence_text,
            "stats_source_kind": stats_source_kind,
            "abc_class": abc_class,
            "velocity_band": velocity_band,
            "stock_months": stock_months,
            "hot_days_of_cover": days_of_cover if days_of_cover > 0 else None,
            "hot_buy_signal": hot_buy_signal,
            "hot_action_today": hot_action_today,
            "hot_priority_score": hot_priority_score,
            "hot_best_supplier_gap_pct": hot_best_supplier_gap_pct,
            "best_source": normalize_text((best_offer or {}).get("source", "")) if has_cheaper_supplier else "",
            "best_price": best_price if has_cheaper_supplier and best_price > 0 else None,
            "best_qty": best_qty if has_cheaper_supplier and best_qty > 0 else None,
            "market_gap_pct": round(market_gap_pct, 2) if market_gap_pct > 0 else 0.0,
            "has_photo": has_photo,
            "avito_count": ad_count,
            "has_avito": has_avito,
            "ready_for_marketplace": ready_for_marketplace,
            "can_buy": can_buy,
            "needs_price_review": needs_price_review,
            "is_hot": is_hot,
            "is_strong_demand": is_strong_demand,
            "is_dead_stock": is_dead_stock,
            "is_overstock": is_overstock,
            "is_low_stock": is_low_stock,
            "recommended_price": pick_recommended_price_for_crm(own_price, best_price, is_hot, is_dead_stock, bool(market_gap_pct > 0)),
            "stock_action": stock_action,
            "price_action": price_action,
            "placement_action": placement_action,
            "decision": decision,
            "decision_reason": reason,
            "priority_score": round(priority, 2),
            "open_tasks": int(task_lookup.get((normalize_text(sheet_name), article_norm), 0)),
            "pipeline_status": pipeline_status,
            "current_queue": current_queue,
            "manual_decision": manual_decision,
            "workflow_stage": workflow_stage,
            "next_action": next_action,
            "owner": owner,
            "priority_label": priority_label,
            "updated_at": normalize_text(pipe.get("updated_at", "")),
            "photo_url": normalize_text(row.get("photo_url", "")),
            "meta_brand": normalize_text(row.get("meta_brand", "")),
            "meta_model": normalize_text(row.get("meta_model", "")),
            "meta_manufacturer_code": normalize_text(row.get("meta_manufacturer_code", "")),
            "meta_print_type": normalize_text(row.get("meta_print_type", "")),
            "meta_color": normalize_text(row.get("meta_color", "")),
            "meta_capacity": normalize_text(row.get("meta_capacity", "")),
            "meta_iso_pages": normalize_text(row.get("meta_iso_pages", "")),
            "meta_item_type": normalize_text(row.get("meta_item_type", "")),
            "meta_print_technology": normalize_text(row.get("meta_print_technology", "")),
            "meta_description": normalize_text(row.get("meta_description", "")),
            "meta_fits_models": normalize_text(row.get("meta_fits_models", "")),
            "meta_weight": normalize_text(row.get("meta_weight", "")),
            "meta_length": normalize_text(row.get("meta_length", "")),
            "meta_width": normalize_text(row.get("meta_width", "")),
            "meta_height": normalize_text(row.get("meta_height", "")),
            "meta_source_sheet": normalize_text(row.get("source_sheet", "")),
            "manual_note": normalize_text(row.get("manual_note", "")),
            "source_pairs": row.get("source_pairs", []) or [],
            "supplier_debug_rows": supplier_debug_rows,
            "supplier_valid_offers": supplier_valid_offers,
        })

    out = pd.DataFrame(rows)
    if out.empty:
        return out
    return out.sort_values(["priority_score", "sales_per_month", "free_qty", "article"], ascending=[False, False, False, True], kind="stable").reset_index(drop=True)


def build_procurement_decision_df(products_df: pd.DataFrame) -> pd.DataFrame:
    if not isinstance(products_df, pd.DataFrame) or products_df.empty:
        return pd.DataFrame()
    rows: list[dict[str, Any]] = []
    threshold_pct = float(st.session_state.get("distributor_threshold", 35.0) or 35.0)
    for _, row in products_df.iterrows():
        rows.append({
            "Лист": normalize_text(row.get("sheet_label", "")),
            "Артикул": normalize_text(row.get("article", "")),
            "article_norm": normalize_text(row.get("article_norm", "")),
            "Товар": normalize_text(row.get("name", "")),
            "Наша цена": safe_float(row.get("sale_price"), 0.0),
            "Наш остаток": parse_qty_generic(row.get("free_qty")),
            "Транзит": parse_qty_generic(row.get("transit_qty")),
            "Всего": parse_qty_generic(row.get("total_qty")),
            "Продажи, шт/мес": safe_float(row.get("sales_per_month"), 0.0),
            "Продажи за 15 мес": safe_float(row.get("sales_qty_15m"), 0.0),
            "ABC класс": normalize_text(row.get("abc_class", "")),
            "Скорость": normalize_text(row.get("velocity_band", "")),
            "Ходовой": "Да" if bool(row.get("is_hot")) else "Нет",
            "Сильный спрос": "Да" if bool(row.get("is_strong_demand")) else "Нет",
            "Лучший поставщик": normalize_text(row.get("best_source", "")),
            "Цена поставщика": safe_float(row.get("best_price"), 0.0) if safe_float(row.get("best_price"), 0.0) > 0 else None,
            "Остаток поставщика": safe_float(row.get("best_qty"), 0.0) if safe_float(row.get("best_qty"), 0.0) > 0 else None,
            "Разница, %": round(safe_float(row.get("market_gap_pct"), 0.0), 2) if safe_float(row.get("market_gap_pct"), 0.0) > 0 else None,
            "Сигнал watchlist, %": round(safe_float(row.get("hot_best_supplier_gap_pct"), 0.0), 2) if safe_float(row.get("hot_best_supplier_gap_pct"), 0.0) > 0 else None,
            "Дней запаса": row.get("hot_days_of_cover", None),
            "Запас, мес": row.get("stock_months", None),
            "Низкий запас": "Да" if bool(row.get("is_low_stock")) else "Нет",
            "Избыточный запас": "Да" if bool(row.get("is_overstock")) else "Нет",
            "Залежался": "Да" if bool(row.get("is_dead_stock")) else "Нет",
            "Есть фото": "Да" if bool(row.get("has_photo")) else "Нет",
            "Есть Avito": "Да" if bool(row.get("has_avito")) else "Нет",
            "Готов к размещению": "Да" if bool(row.get("ready_for_marketplace")) else "Нет",
            "Можно закупать": "Да" if bool(row.get("can_buy")) else "Нет",
            "Требует новой цены": "Да" if bool(row.get("needs_price_review")) else "Нет",
            "Сигнал BUY": "Да" if normalize_text(row.get("hot_buy_signal", "")).upper() == "BUY" else "Нет",
            "Watchlist действие": translate_watch_action(row.get("hot_action_today", ""), threshold_pct=threshold_pct),
            "Рекомендованная цена": row.get("recommended_price", None),
            "Решение по складу": normalize_text(row.get("stock_action", "")),
            "Решение по цене": normalize_text(row.get("price_action", "")),
            "Решение по размещению": normalize_text(row.get("placement_action", "")),
            "Решение": normalize_text(row.get("decision", "")),
            "Почему": normalize_text(row.get("decision_reason", "")),
            "Приоритет": safe_float(row.get("priority_score"), 0.0),
            "Pipeline": normalize_text(row.get("pipeline_status", "")) or "Новая",
            "Очередь": normalize_text(row.get("current_queue", "")),
            "Ручное решение": normalize_text(row.get("manual_decision", "")),
            "Этап": normalize_text(row.get("workflow_stage", "")),
            "Следующее действие": normalize_text(row.get("next_action", "")),
            "Ответственный": normalize_text(row.get("owner", "")),
            "Приоритет label": normalize_text(row.get("priority_label", "")),
            "Открытых задач": safe_int(row.get("open_tasks", 0), 0),
        })
    out = pd.DataFrame(rows)
    if out.empty:
        return out
    return out.sort_values(["Приоритет", "Продажи, шт/мес", "Наш остаток", "Артикул"], ascending=[False, False, False, True], kind="stable").reset_index(drop=True)


def filter_procurement_queue(decision_df: pd.DataFrame, queue_name: str) -> pd.DataFrame:
    if not isinstance(decision_df, pd.DataFrame) or decision_df.empty:
        return pd.DataFrame()
    filtered = decision_df.copy()
    q = normalize_text(queue_name)
    if q in {"", "Все"}:
        return filtered
    if q == "Можно брать":
        filtered = filtered[filtered["Можно закупать"] == "Да"]
    elif q == "К пополнению":
        filtered = filtered[filtered["Решение по складу"].isin(["Пополнить запас", "Можно закупать"])]
    elif q == "Требует цены":
        filtered = filtered[(filtered["Требует новой цены"] == "Да") | (filtered["Решение"].isin(["Распродавать", "Пересмотреть цену", "Проверить запас"]))]
    elif q == "Без фото":
        filtered = filtered[filtered["Есть фото"] == "Нет"]
    elif q == "Без Avito":
        filtered = filtered[filtered["Есть Avito"] == "Нет"]
    elif q == "Готово к размещению":
        filtered = filtered[filtered["Готов к размещению"] == "Да"]
    elif q == "Залежалый остаток":
        filtered = filtered[filtered["Залежался"] == "Да"]
    elif q == "Сильный спрос":
        filtered = filtered[filtered["Сильный спрос"] == "Да"]
    elif q == "Ходовые":
        filtered = filtered[filtered["Ходовой"] == "Да"]
    elif q == "Под наблюдением":
        filtered = filtered[filtered["Очередь"].fillna("").astype(str).eq("Под наблюдением")]
    return filtered.reset_index(drop=True)


def apply_pending_catalog_navigation() -> None:
    pending_mode = normalize_text(st.session_state.pop("pending_app_mode_main", ""))
    if pending_mode:
        st.session_state["app_mode_main"] = pending_mode

    pending_label = normalize_text(st.session_state.pop("pending_active_workspace_label", ""))
    if pending_label:
        st.session_state["active_workspace_label"] = pending_label

    pending_article = normalize_text(st.session_state.pop("pending_catalog_article", ""))
    pending_tab_key = normalize_text(st.session_state.pop("pending_catalog_tab_key", "")) or "original"
    if pending_article:
        query = normalize_query_for_display(pending_article)
        if query:
            st.session_state[f"search_input_{pending_tab_key}"] = query
            st.session_state[f"submitted_query_{pending_tab_key}"] = query
            st.session_state[f"search_input_widget_pending_{pending_tab_key}"] = query
            st.session_state[f"last_result_{pending_tab_key}"] = None
            st.session_state[f"last_result_sig_{pending_tab_key}"] = None


def open_product_in_catalog(article: str, sheet_label: str) -> None:
    resolved_label = CRM_SHEET_NAME_TO_LABEL.get(normalize_text(sheet_label), normalize_text(sheet_label) or "Оригинал")
    tab_key = CRM_SHEET_LABEL_TO_TABKEY.get(resolved_label, "original")
    st.session_state["pending_app_mode_main"] = "Каталог"
    st.session_state["pending_active_workspace_label"] = resolved_label
    st.session_state["pending_catalog_article"] = normalize_text(article)
    st.session_state["pending_catalog_tab_key"] = tab_key
    st.rerun()


def open_product_in_crm(article_norm: str, sheet_label: str = "", open_photo_editor: bool = False) -> None:
    resolved_label = CRM_SHEET_NAME_TO_LABEL.get(normalize_text(sheet_label), normalize_text(sheet_label) or "Оригинал")
    st.session_state["pending_app_mode_main"] = "CRM workspace"
    st.session_state["pending_active_workspace_label"] = resolved_label
    st.session_state["crm_workspace_article_norm"] = normalize_text(article_norm)
    if open_photo_editor:
        st.session_state["crm_workspace_open_photo_editor_for"] = normalize_text(article_norm)
    st.rerun()


def remember_crm_article(article_norm: str) -> None:
    st.session_state["crm_workspace_article_norm"] = normalize_text(article_norm)


def _crm_pick_label_to_row(products_df: pd.DataFrame) -> tuple[list[str], dict[str, dict[str, Any]]]:
    labels: list[str] = []
    mapping: dict[str, dict[str, Any]] = {}
    for _, row in products_df.iterrows():
        label = f"{normalize_text(row.get('article', ''))} • {normalize_text(row.get('name', ''))[:90]}"
        labels.append(label)
        mapping[label] = row.to_dict()
    return labels, mapping


def render_crm_workspace_dashboard(products_df: pd.DataFrame, tasks_df: pd.DataFrame, decision_df: pd.DataFrame | None = None) -> None:
    decision_df = decision_df if isinstance(decision_df, pd.DataFrame) else get_cached_procurement_decision_df(products_df)
    task_counts = task_summary_counts()
    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("Открытые задачи", task_counts.get("open", 0))
    c2.metric("Просроченные", task_counts.get("overdue", 0))
    c3.metric("Можно закупать", int((decision_df["Можно закупать"] == "Да").sum()) if not decision_df.empty else 0)
    c4.metric("К пополнению", int((decision_df["Решение по складу"].isin(["Пополнить запас", "Можно закупать"])).sum()) if not decision_df.empty else 0)
    c5.metric("Залежалось", int((decision_df["Залежался"] == "Да").sum()) if not decision_df.empty else 0)

    q1, q2, q3, q4 = st.columns(4)
    q1.metric("Без фото", int((decision_df["Есть фото"] == "Нет").sum()) if not decision_df.empty else 0)
    q2.metric("Без Avito", int((decision_df["Есть Avito"] == "Нет").sum()) if not decision_df.empty else 0)
    q3.metric("Сильный спрос", int((decision_df["Сильный спрос"] == "Да").sum()) if not decision_df.empty else 0)
    q4.metric("Готово к размещению", int((decision_df["Готов к размещению"] == "Да").sum()) if not decision_df.empty else 0)

    st.markdown("### Что делать сегодня")
    todo_rows = []
    if not decision_df.empty:
        for queue_name, what in [
            ("Можно брать", "Ходовые позиции с выгодным входом от поставщика"),
            ("К пополнению", "Запас низкий, товар продаётся, нужен контроль закупки"),
            ("Требует цены", "Позиции выше рынка или с риском залежалости"),
            ("Без фото", "Товар есть, а карточка ещё не готова визуально"),
            ("Без Avito", "Есть товар на складе, но нет размещения"),
            ("Готово к размещению", "Можно быстро размещать на Avito"),
        ]:
            todo_rows.append({"Очередь": queue_name, "Позиций": len(filter_procurement_queue(decision_df, queue_name)), "Что это": what})
    if todo_rows:
        st.dataframe(pd.DataFrame(todo_rows), use_container_width=True, hide_index=True)

    st.markdown("### Последние задачи")
    if not isinstance(tasks_df, pd.DataFrame) or tasks_df.empty:
        st.caption("Задач пока нет.")
    else:
        st.dataframe(tasks_df.head(10), use_container_width=True, hide_index=True, height=320)


def render_crm_workspace_queues(products_df: pd.DataFrame, decision_df: pd.DataFrame | None = None) -> None:
    decision_df = decision_df if isinstance(decision_df, pd.DataFrame) else get_cached_procurement_decision_df(products_df)
    queue_names = ["Все", "Можно брать", "К пополнению", "Требует цены", "Без фото", "Без Avito", "Готово к размещению", "Залежалый остаток", "Сильный спрос", "Ходовые", "Под наблюдением"]
    queue_name = st.selectbox("Очередь", queue_names, key="crm_queue_filter")
    queue_df = filter_procurement_queue(decision_df, queue_name)
    if queue_df.empty:
        st.info("По выбранной очереди строк не найдено.")
        return
    view_cols = [
        "Артикул", "Товар", "Наш остаток", "Продажи, шт/мес", "Запас, мес", "Лучший поставщик", "Цена поставщика",
        "Разница, %", "Есть фото", "Есть Avito", "Решение", "Почему", "Pipeline", "Открытых задач"
    ]
    st.dataframe(queue_df[[c for c in view_cols if c in queue_df.columns]], use_container_width=True, hide_index=True, height=520)
    labels = [f"{r['Артикул']} • {r['Товар'][:90]}" for _, r in queue_df.iterrows()]
    row_map = {f"{r['Артикул']} • {r['Товар'][:90]}": r for _, r in queue_df.iterrows()}
    pick = st.selectbox("Открыть позицию из очереди", labels, key="crm_queue_pick")
    row = row_map[pick]
    c1, c2, c3, c4 = st.columns(4)
    if c1.button("Открыть в CRM-карточке", use_container_width=True, key="crm_queue_open_card"):
        open_product_in_crm(normalize_text(row.get("article_norm", "")), sheet_label=normalize_text(row.get("Лист", row.get("sheet_label", "Оригинал"))), open_photo_editor=False)
    if c2.button("Открыть и редактировать фото", use_container_width=True, key="crm_queue_open_photo_editor"):
        open_product_in_crm(normalize_text(row.get("article_norm", "")), sheet_label=normalize_text(row.get("Лист", row.get("sheet_label", "Оригинал"))), open_photo_editor=True)
    if c3.button("Открыть в каталоге", use_container_width=True, key="crm_queue_open_catalog"):
        open_product_in_catalog(normalize_text(row.get("Артикул", "")), normalize_text(row.get("Лист", "Оригинал")))
    if c4.button("В работу", use_container_width=True, key="crm_queue_in_work"):
        upsert_pipeline_registry(
            sheet_name=CRM_SHEET_LABEL_TO_NAME.get(normalize_text(row.get("Лист", "")), normalize_text(row.get("Лист", ""))),
            article=normalize_text(row.get("Артикул", "")),
            article_norm=normalize_text(row.get("article_norm", "")),
            pipeline_status="В работе",
            current_queue=normalize_text(queue_name if queue_name != "Все" else row.get("Очередь", "Под наблюдением")),
            manual_decision=normalize_text(row.get("Решение", "")),
            workflow_stage="Проверка",
            next_action=normalize_text(row.get("Решение", "")),
            priority="Высокий" if safe_float(row.get("Приоритет", 0.0), 0.0) >= 120 else "Средний",
        )
        st.success("Pipeline обновлён.")
        st.rerun()


def render_crm_workspace_execution(products_df: pd.DataFrame, decision_df: pd.DataFrame | None = None) -> None:
    decision_df = decision_df if isinstance(decision_df, pd.DataFrame) else get_cached_procurement_decision_df(products_df)
    queue_names = ["Можно брать", "К пополнению", "Требует цены", "Без фото", "Без Avito", "Готово к размещению", "Залежалый остаток", "Сильный спрос", "Ходовые", "Под наблюдением"]
    queue_name = st.selectbox("Execution очередь", queue_names, key="crm_execution_queue_name")
    queue_df = filter_procurement_queue(decision_df, queue_name)
    if queue_df.empty:
        st.info("В очереди пока нет строк.")
        return
    labels = [f"{r['Артикул']} • {r['Товар'][:90]}" for _, r in queue_df.iterrows()]
    selected = st.multiselect("Позиции для пакетного действия", labels, default=labels[: min(5, len(labels))], key="crm_execution_selected")
    st.dataframe(queue_df[[c for c in ["Артикул", "Товар", "Наш остаток", "Лучший поставщик", "Цена поставщика", "Разница, %", "Решение", "Почему"] if c in queue_df.columns]], use_container_width=True, hide_index=True, height=420)
    row_map = {f"{r['Артикул']} • {r['Товар'][:90]}": r for _, r in queue_df.iterrows()}
    articles = [normalize_text(row_map[label].get("Артикул", "")) for label in selected if label in row_map]
    rows = [row_map[label] for label in selected if label in row_map]
    meta1, meta2, meta3, meta4 = st.columns(4)
    owner = meta1.text_input("Ответственный", value="", key="crm_exec_owner")
    next_action = meta2.text_input("Следующее действие", value=f"Отработать очередь: {queue_name}", key="crm_exec_next")
    priority_label = meta3.selectbox("Приоритет", ["Низкий", "Средний", "Высокий"], index=1, key="crm_exec_priority")
    due_date = meta4.date_input("Срок задачи", value=datetime.utcnow().date() + timedelta(days=7), key="crm_exec_due")
    b1, b2, b3 = st.columns(3)
    if b1.button("Поставить в работу", use_container_width=True, key="crm_exec_start"):
        applied = 0
        for row in rows:
            upsert_pipeline_registry(
                sheet_name=CRM_SHEET_LABEL_TO_NAME.get(normalize_text(row.get("Лист", "")), normalize_text(row.get("Лист", ""))),
                article=normalize_text(row.get("Артикул", "")),
                article_norm=normalize_text(row.get("article_norm", "")),
                pipeline_status="В работе",
                current_queue=queue_name,
                manual_decision=normalize_text(row.get("Решение", "")),
                workflow_stage="Исполнение",
                next_action=next_action or normalize_text(row.get("Решение", "")),
                owner=owner,
                priority=priority_label,
            )
            applied += 1
        st.success(f"В работу отправлено: {applied}")
        st.rerun()
    if b2.button("Создать задачи", use_container_width=True, key="crm_exec_tasks"):
        created = 0
        for row in rows:
            create_review_task(
                article=normalize_text(row.get("Артикул", "")),
                article_norm=normalize_text(row.get("article_norm", "")),
                sheet_name=CRM_SHEET_LABEL_TO_NAME.get(normalize_text(row.get("Лист", "")), normalize_text(row.get("Лист", ""))),
                name_snapshot=normalize_text(row.get("Товар", "")),
                due_date=due_date,
                reason=queue_name,
                note=next_action or normalize_text(row.get("Почему", "")),
                source="crm_workspace_execution",
            )
            created += 1
        st.success(f"Создано задач: {created}")
        st.rerun()
    if b3.button("Открыть первую в каталоге", use_container_width=True, key="crm_exec_open_first"):
        if rows:
            row = rows[0]
            open_product_in_catalog(normalize_text(row.get("Артикул", "")), normalize_text(row.get("Лист", "Оригинал")))


def render_crm_workspace_pipeline(products_df: pd.DataFrame, decision_df: pd.DataFrame | None = None) -> None:
    decision_df = decision_df if isinstance(decision_df, pd.DataFrame) else get_cached_procurement_decision_df(products_df)
    if decision_df.empty:
        st.info("Нет данных для pipeline.")
        return
    table = decision_df[[c for c in ["Лист", "Артикул", "Товар", "Pipeline", "Очередь", "Ручное решение", "Этап", "Следующее действие", "Ответственный", "Приоритет label", "Открытых задач"] if c in decision_df.columns]].copy()
    st.dataframe(table, use_container_width=True, hide_index=True, height=420)
    labels = [f"{r['Артикул']} • {r['Товар'][:90]}" for _, r in decision_df.iterrows()]
    row_map = {f"{r['Артикул']} • {r['Товар'][:90]}": r for _, r in decision_df.iterrows()}
    selected = st.selectbox("Редактировать pipeline позиции", labels, key="crm_pipeline_pick")
    row = row_map[selected]
    c1, c2, c3 = st.columns(3)
    pipeline_status = c1.selectbox("Статус", ["Новая", "В работе", "На согласовании", "На паузе", "Готово"], index=["Новая", "В работе", "На согласовании", "На паузе", "Готово"].index(normalize_text(row.get("Pipeline", "Новая")) if normalize_text(row.get("Pipeline", "Новая")) in ["Новая", "В работе", "На согласовании", "На паузе", "Готово"] else "Новая"), key="crm_pipeline_status_edit")
    queue_name = c2.selectbox("Очередь", ["Под наблюдением", "Можно брать", "К пополнению", "Требует цены", "Без фото", "Без Avito", "Готово к размещению", "Залежалый остаток", "Сильный спрос", "Ходовые"], index=0 if normalize_text(row.get("Очередь", "")) not in ["Под наблюдением", "Можно брать", "К пополнению", "Требует цены", "Без фото", "Без Avito", "Готово к размещению", "Залежалый остаток", "Сильный спрос", "Ходовые"] else ["Под наблюдением", "Можно брать", "К пополнению", "Требует цены", "Без фото", "Без Avito", "Готово к размещению", "Залежалый остаток", "Сильный спрос", "Ходовые"].index(normalize_text(row.get("Очередь", ""))), key="crm_pipeline_queue_edit")
    workflow_stage = c3.selectbox("Этап", ["Проверка", "Решение", "Исполнение", "Контроль", "Закрыто"], index=0 if normalize_text(row.get("Этап", "")) not in ["Проверка", "Решение", "Исполнение", "Контроль", "Закрыто"] else ["Проверка", "Решение", "Исполнение", "Контроль", "Закрыто"].index(normalize_text(row.get("Этап", ""))), key="crm_pipeline_stage_edit")
    d1, d2, d3 = st.columns(3)
    manual_decision = d1.text_input("Ручное решение", value=normalize_text(row.get("Ручное решение", "")) or normalize_text(row.get("Решение", "")), key="crm_pipeline_manual_edit")
    next_action = d2.text_input("Следующее действие", value=normalize_text(row.get("Следующее действие", "")) or normalize_text(row.get("Решение", "")), key="crm_pipeline_next_edit")
    owner = d3.text_input("Ответственный", value=normalize_text(row.get("Ответственный", "")), key="crm_pipeline_owner_edit")
    priority_label = st.selectbox("Приоритет", ["Низкий", "Средний", "Высокий"], index=1 if normalize_text(row.get("Приоритет label", "Средний")) not in ["Низкий", "Средний", "Высокий"] else ["Низкий", "Средний", "Высокий"].index(normalize_text(row.get("Приоритет label", "Средний"))), key="crm_pipeline_priority_edit")
    if st.button("Сохранить pipeline", use_container_width=True, key="crm_pipeline_save"):
        upsert_pipeline_registry(
            sheet_name=CRM_SHEET_LABEL_TO_NAME.get(normalize_text(row.get("Лист", "")), normalize_text(row.get("Лист", ""))),
            article=normalize_text(row.get("Артикул", "")),
            article_norm=normalize_text(row.get("article_norm", "")),
            pipeline_status=pipeline_status,
            current_queue=queue_name,
            manual_decision=manual_decision,
            workflow_stage=workflow_stage,
            next_action=next_action,
            owner=owner,
            priority=priority_label,
        )
        st.success("Pipeline сохранён.")
        st.rerun()


def render_crm_workspace_card(products_df: pd.DataFrame, sheet_name: str, sheet_label: str) -> None:
    if not isinstance(products_df, pd.DataFrame) or products_df.empty:
        st.info("Нет данных для CRM-карточки.")
        return
    selected_article_norm = normalize_text(st.session_state.get("crm_workspace_article_norm", ""))
    labels, mapping = _crm_pick_label_to_row(products_df)
    if selected_article_norm and selected_article_norm in set(products_df["article_norm"].astype(str)):
        default_label = next((label for label, rec in mapping.items() if normalize_text(rec.get("article_norm", "")) == selected_article_norm), labels[0])
        pick = st.selectbox("Позиция для CRM-карточки", labels, index=labels.index(default_label), key="crm_card_pick")
    else:
        pick = st.selectbox("Позиция для CRM-карточки", labels, key="crm_card_pick")
    row = mapping[pick]
    remember_crm_article(normalize_text(row.get("article_norm", "")))
    st.markdown("### CRM-карточка товара")
    top1, top2 = st.columns([1.25, 1.75])
    with top1:
        photo_url = normalize_text(row.get("photo_url", ""))
        if photo_url:
            st.image(photo_url, use_container_width=True)
        else:
            st.info("Фото не найдено")
    with top2:
        st.markdown(f"## {html.escape(normalize_text(row.get('article', '')))}")
        st.caption(html.escape(normalize_text(row.get("name", ""))))
        m1, m2, m3, m4, m5, m6 = st.columns(6)
        m1.metric("Наша цена", fmt_price(row.get("sale_price", 0.0)))
        m2.metric("Закупка", fmt_price(row.get("purchase_avg_cost", 0.0)) if safe_float(row.get("purchase_avg_cost"), 0.0) > 0 else "—")
        m3.metric("Наш склад", fmt_qty(row.get("free_qty", 0.0)))
        m4.metric("Avito", safe_int(row.get("avito_count", 0), 0))
        m5.metric("Источник meta", normalize_text(row.get("meta_source_sheet", "")) or "—")
        m6.metric("Pipeline", normalize_text(row.get("pipeline_status", "")) or "Новая")
        st.success(f"Решение: {normalize_text(row.get('decision', 'Проверить вручную'))}")
        st.caption(normalize_text(row.get("decision_reason", "Недостаточно явного сигнала")))
        st.write(f"**Склад:** {normalize_text(row.get('stock_action', '')) or 'Проверить вручную'}")
        st.write(f"**Цена:** {normalize_text(row.get('price_action', '')) or 'Оставить цену'}")
        st.write(f"**Размещение:** {normalize_text(row.get('placement_action', '')) or 'Ничего'}")
        st.write(f"**Рынок:** {( 'ниже нас на ' + str(round(safe_float(row.get('market_gap_pct', 0.0), 0.0), 1)) + '%' ) if safe_float(row.get('market_gap_pct', 0.0), 0.0) > 0 else 'нет сигнала рынка'}")

    card_section = st.radio(
        "Раздел карточки",
        ["Обзор", "Характеристики", "Поставщики", "Статистика", "Задачи"],
        key="crm_card_section",
        horizontal=True,
        label_visibility="collapsed",
    )

    if card_section == "Обзор":
        left, right = st.columns(2)
        left.write(f"**Лист:** {normalize_text(row.get('sheet_label', '')) or sheet_label}")
        left.write(f"**Фото:** {'Да' if bool(row.get('has_photo')) else 'Нет'}")
        left.write(f"**Лучший поставщик:** {normalize_text(row.get('best_source', '')) or '—'}")
        if not normalize_text(row.get('best_source', '')):
            left.caption("Сейчас дешевле нашей цены поставщика нет.")
        left.write(f"**Статус рынка:** {'ниже нас' if safe_float(row.get('market_gap_pct', 0.0), 0.0) > 0 else 'нет сигнала рынка'}")
        if safe_float(row.get("purchase_avg_cost"), 0.0) > 0:
            left.write(f"**Средняя закупка:** {fmt_price(row.get('purchase_avg_cost', 0.0))}")
            left.caption(f"Источник закупки: {normalize_text(row.get('purchase_match_source', '')) or '—'} • {normalize_text(row.get('purchase_source_name', '')) or 'без названия'}")
        else:
            left.write("**Средняя закупка:** —")
            left.caption("Файл средней закупки не загрузили или безопасный маппинг по названию не нашёл совпадение.")
        right.write(f"**Открытых задач:** {safe_int(row.get('open_tasks', 0), 0)}")
        right.write(f"**Продажи/мес:** {fmt_qty(row.get('sales_per_month', 0.0))}")
        right.write(f"**Запас, мес:** {fmt_qty(row.get('stock_months', '')) if row.get('stock_months', None) not in (None, '') else '—'}")
        right.write(f"**Комментарий:** {normalize_text(row.get('manual_note', '')) or '—'}")

        purchase_avg_cost = safe_float(row.get("purchase_avg_cost"), 0.0)
        own_price = safe_float(row.get("sale_price"), 0.0)
        best_price = safe_float(row.get("best_price"), 0.0)
        free_qty = safe_float(row.get("free_qty"), 0.0)
        markup_abs = (own_price - purchase_avg_cost) if own_price > 0 and purchase_avg_cost > 0 else None
        markup_pct = ((own_price - purchase_avg_cost) / purchase_avg_cost * 100.0) if own_price > 0 and purchase_avg_cost > 0 else None
        stock_cost_total = (purchase_avg_cost * free_qty) if purchase_avg_cost > 0 and free_qty > 0 else None
        supplier_vs_purchase_abs = (best_price - purchase_avg_cost) if best_price > 0 and purchase_avg_cost > 0 else None
        supplier_vs_purchase_pct = ((best_price - purchase_avg_cost) / purchase_avg_cost * 100.0) if best_price > 0 and purchase_avg_cost > 0 else None

        st.markdown("#### Экономика товара")
        e1, e2, e3, e4 = st.columns(4)
        e1.metric("Средняя закупка", fmt_price(purchase_avg_cost) if purchase_avg_cost > 0 else "—")
        e2.metric("Наценка, ₽", fmt_price(markup_abs) if markup_abs is not None else "—")
        e3.metric("Наценка к закупке, %", f"{markup_pct:.1f}%" if markup_pct is not None else "—")
        e4.metric("Склад по закупке", fmt_price(stock_cost_total) if stock_cost_total is not None else "—")

        if purchase_avg_cost > 0 and own_price > 0:
            if own_price < purchase_avg_cost:
                st.error("Наша цена сейчас ниже средней закупки. Позицию нужно проверить вручную.")
            elif markup_pct is not None and markup_pct < 15:
                st.warning("Наценка к закупке низкая. Проверь цену, налоги и прочие расходы.")
            else:
                st.caption("Экономика товара рассчитана по средней закупке из файла `Итог_взвешенный`.")
        else:
            st.caption("Для экономики товара нужен загруженный файл средней закупки и уверенный маппинг по названию/коду.")

        if best_price > 0 and purchase_avg_cost > 0:
            s1, s2, s3 = st.columns(3)
            s1.metric("Лучшая цена поставщика", fmt_price(best_price))
            s2.metric("Поставщик vs закупка, ₽", fmt_price(supplier_vs_purchase_abs) if supplier_vs_purchase_abs is not None else "—")
            s3.metric("Поставщик vs закупка, %", f"{supplier_vs_purchase_pct:.1f}%" if supplier_vs_purchase_pct is not None else "—")

        art = normalize_text(row.get("article", ""))
        art_norm = normalize_text(row.get("article_norm", ""))
        current_photo_url = normalize_text(row.get("photo_url", ""))
        current_note = normalize_text(row.get("manual_note", ""))
        force_open_photo_editor_for = normalize_text(st.session_state.pop("crm_workspace_open_photo_editor_for", ""))
        photo_editor_expanded = (not bool(row.get('has_photo'))) or (force_open_photo_editor_for == art_norm)
        if force_open_photo_editor_for == art_norm:
            st.caption("Режим быстрого добавления фото открыт для этой позиции.")

        with st.expander("Фото и заметка по карточке", expanded=photo_editor_expanded):
            with st.form(f"crm_workspace_card_override_{art_norm}"):
                photo_url_new = st.text_input("Фото (ссылка)", value=current_photo_url, key=f"crm_workspace_card_photo_{art_norm}", placeholder="Вставь прямую ссылку на фото товара")
                note_new = st.text_area("Заметка", value=current_note, height=90, key=f"crm_workspace_card_note_{art_norm}", placeholder="Короткий комментарий по товару: что исправили, что проверить, что важно.")
                s1, s2 = st.columns(2)
                save_clicked = s1.form_submit_button("Сохранить карточку", use_container_width=True, type="primary")
                reset_clicked = s2.form_submit_button("Сбросить ручные правки", use_container_width=True)
            if save_clicked:
                upsert_card_override(sheet_name, art, art_norm, {
                    "photo_url": photo_url_new,
                    "meta_source_sheet": normalize_text(row.get("meta_source_sheet", "")),
                    "meta_brand": normalize_text(row.get("meta_brand", "")),
                    "meta_model": normalize_text(row.get("meta_model", "")),
                    "meta_manufacturer_code": normalize_text(row.get("meta_manufacturer_code", "")),
                    "meta_print_type": normalize_text(row.get("meta_print_type", "")),
                    "meta_color": normalize_text(row.get("meta_color", "")),
                    "meta_capacity": normalize_text(row.get("meta_capacity", "")),
                    "meta_iso_pages": normalize_text(row.get("meta_iso_pages", "")),
                    "meta_item_type": normalize_text(row.get("meta_item_type", "")),
                    "meta_print_technology": normalize_text(row.get("meta_print_technology", "")),
                    "meta_description": normalize_text(row.get("meta_description", "")),
                    "meta_fits_models": normalize_text(row.get("meta_fits_models", "")),
                    "meta_weight": normalize_text(row.get("meta_weight", "")),
                    "meta_length": normalize_text(row.get("meta_length", "")),
                    "meta_width": normalize_text(row.get("meta_width", "")),
                    "meta_height": normalize_text(row.get("meta_height", "")),
                    "note": note_new,
                })
                st.success(f"Карточка {art} сохранена.")
                st.rerun()
            if reset_clicked:
                delete_card_override(sheet_name, art_norm)
                st.success(f"Ручные правки для {art} сброшены.")
                st.rerun()

        if st.button("Открыть в каталоге", use_container_width=True, key="crm_card_open_catalog"):
            open_product_in_catalog(normalize_text(row.get("article", "")), normalize_text(row.get("sheet_label", sheet_label)))

    elif card_section == "Характеристики":
        ch1, ch2 = st.columns(2)
        ch1.write(f"**Бренд:** {normalize_text(row.get('meta_brand', '')) or '—'}")
        ch1.write(f"**Модель:** {normalize_text(row.get('meta_model', '')) or '—'}")
        ch1.write(f"**Код производителя:** {normalize_text(row.get('meta_manufacturer_code', '')) or '—'}")
        ch1.write(f"**Тип печати:** {normalize_text(row.get('meta_print_type', '')) or '—'}")
        ch1.write(f"**Цвет:** {normalize_text(row.get('meta_color', '')) or '—'}")
        ch1.write(f"**Ёмкость:** {normalize_text(row.get('meta_capacity', '')) or '—'}")
        ch2.write(f"**Ресурс:** {normalize_text(row.get('meta_iso_pages', '')) or '—'}")
        ch2.write(f"**Тип:** {normalize_text(row.get('meta_item_type', '')) or '—'}")
        ch2.write(f"**Технология:** {normalize_text(row.get('meta_print_technology', '')) or '—'}")
        ch2.write(f"**Вес:** {format_meta_weight(row.get('meta_weight', '')) or '—'}")
        ch2.write(f"**Габариты:** {format_meta_dimensions(row.get('meta_length', ''), row.get('meta_width', ''), row.get('meta_height', '')) or '—'}")
        st.write(f"**Подходит к моделям:** {normalize_text(row.get('meta_fits_models', '')) or '—'}")
        st.write(f"**Описание:** {normalize_text(row.get('meta_description', '')) or '—'}")
        st.write(f"**Источник закупки:** {normalize_text(row.get('purchase_source_sheet', '')) or '—'}")

    elif card_section == "Поставщики":
        valid_offers = row.get("supplier_valid_offers", []) if isinstance(row.get("supplier_valid_offers", []), list) else []
        debug_rows = row.get("supplier_debug_rows", []) if isinstance(row.get("supplier_debug_rows", []), list) else []
        if valid_offers:
            offers_df = pd.DataFrame([{
                "Поставщик": normalize_text(x.get("source", "")),
                "Цена": safe_float(x.get("price"), 0.0),
                "Остаток": safe_float(x.get("qty"), 0.0),
                "Статус": normalize_text(x.get("status", "")) or "OK",
            } for x in valid_offers])
            cheaper_offers_df = offers_df[offers_df["Цена"].fillna(0).astype(float) < safe_float(row.get("sale_price"), 0.0)].copy()
            if not cheaper_offers_df.empty:
                st.success(f"Найдено предложений поставщиков дешевле нашей цены: {len(cheaper_offers_df)}")
                st.dataframe(cheaper_offers_df, use_container_width=True, hide_index=True, height=min(260, 80 + len(cheaper_offers_df) * 35))
            else:
                st.info("Сейчас дешевле нашей цены поставщика нет.")
                with st.expander("Показать всех валидных поставщиков", expanded=False):
                    st.dataframe(offers_df, use_container_width=True, hide_index=True, height=min(260, 80 + len(offers_df) * 35))
        else:
            st.info("По этой позиции нет валидных предложений поставщиков.")
        if debug_rows:
            with st.expander("Диагностика офферов", expanded=not bool(valid_offers)):
                st.dataframe(pd.DataFrame(debug_rows), use_container_width=True, hide_index=True, height=320)

    elif card_section == "Статистика":
        stats_source_kind = normalize_text(row.get("stats_source_kind", ""))
        has_stats = bool(
            safe_float(row.get("sales_per_month"), 0.0) > 0
            or safe_float(row.get("sales_per_day"), 0.0) > 0
            or safe_float(row.get("sales_per_week"), 0.0) > 0
            or safe_float(row.get("sales_per_year"), 0.0) > 0
            or safe_int(row.get("deals_count", 0), 0) > 0
            or normalize_text(row.get("first_sale", ""))
            or normalize_text(row.get("last_sale", ""))
            or safe_float(row.get("days_without_sales"), 0.0) > 0
            or safe_float(row.get("market_min_price"), 0.0) > 0
            or normalize_text(row.get("market_min_supplier", ""))
            or normalize_text(row.get("supplier_presence_text", ""))
        )
        if not has_stats:
            st.info("Статистика по этой позиции не загружена или не сматчилась по артикулу/коду.")
        else:
            if stats_source_kind == "velocity":
                st.caption("Статистика загружена из нового файла скорости продаж. Продажи/мес берутся из столбца «В месяц».")
            else:
                st.caption("Статистика загружена из watchlist/файла продаж.")
            s1, s2, s3 = st.columns(3)
            s1.write(f"**В день:** {fmt_qty(row.get('sales_per_day', 0.0))}")
            s1.write(f"**В неделю:** {fmt_qty(row.get('sales_per_week', 0.0))}")
            s1.write(f"**В месяц:** {fmt_qty(row.get('sales_per_month', 0.0))}")
            s1.write(f"**В год:** {fmt_qty(row.get('sales_per_year', 0.0)) if safe_float(row.get('sales_per_year'), 0.0) > 0 else '—'}")
            s2.write(f"**Всего шт.:** {fmt_qty(row.get('sales_qty_15m', 0.0)) if safe_float(row.get('sales_qty_15m'), 0.0) > 0 else '—'}")
            s2.write(f"**Сделок:** {safe_int(row.get('deals_count', 0), 0) if safe_int(row.get('deals_count', 0), 0) > 0 else '—'}")
            s2.write(f"**Первая продажа:** {normalize_text(row.get('first_sale', '')) or '—'}")
            s2.write(f"**Последняя продажа:** {normalize_text(row.get('last_sale', '')) or '—'}")
            s3.write(f"**Дней без продаж:** {fmt_qty(row.get('days_without_sales', 0.0)) if row.get('days_without_sales', None) not in (None, '') else '—'}")
            s3.write(f"**Мин. цена конкурентов:** {fmt_price(row.get('market_min_price', 0.0)) if safe_float(row.get('market_min_price'), 0.0) > 0 else '—'}")
            s3.write(f"**Поставщик (мин.):** {normalize_text(row.get('market_min_supplier', '')) or '—'}")
            s3.write(f"**Наличие у поставщиков:** {normalize_text(row.get('supplier_presence_text', '')) or '—'}")
            st.write(f"**Скорость:** {normalize_text(row.get('velocity_band', '')) or '—'}")
            st.write(f"**ABC:** {normalize_text(row.get('abc_class', '')) or '—'}")

    else:
        task_df = build_task_view_df(sheet_filter=sheet_name)
        if isinstance(task_df, pd.DataFrame) and not task_df.empty:
            task_df = task_df[task_df["Артикул"].fillna("").astype(str).eq(normalize_text(row.get("article", "")))].copy()
        render_tasks_table_ui(task_df, f"crm_card_tasks_{normalize_text(row.get('article_norm', ''))}", default_sheet=sheet_name)
        with st.form(f"crm_card_quick_task_{normalize_text(row.get('article_norm', ''))}"):
            st.markdown("#### Быстрая задача")
            due_date = st.date_input("Когда проверить", value=(datetime.utcnow().date() + timedelta(days=7)), key=f"crm_card_quick_due_{normalize_text(row.get('article_norm', ''))}")
            reason = st.selectbox("Причина", ["Пересмотреть цену", "Проверить фото/карточку", "Проверить спрос", "Проверить размещение", "Другое"], key=f"crm_card_quick_reason_{normalize_text(row.get('article_norm', ''))}")
            note = st.text_area("Комментарий", value=normalize_text(row.get("decision_reason", "")), height=70, key=f"crm_card_quick_note_{normalize_text(row.get('article_norm', ''))}")
            submitted = st.form_submit_button("Создать задачу", use_container_width=True)
        if submitted:
            create_review_task(article=normalize_text(row.get("article", "")), article_norm=normalize_text(row.get("article_norm", "")), sheet_name=sheet_name, name_snapshot=normalize_text(row.get("name", "")), due_date=due_date, reason=reason, note=note, source="crm_workspace_card")
            st.success("Задача создана.")
            st.rerun()


def render_crm_workspace(sheet_df: pd.DataFrame | None, photo_df: pd.DataFrame | None, avito_df: pd.DataFrame | None, sheet_name: str, sheet_label: str, min_qty: float) -> None:
    products_df = get_cached_crm_workspace_products_df(sheet_df, photo_df, avito_df, min_qty, sheet_name, sheet_label)
    tasks_df = build_task_view_df(sheet_filter=sheet_name)
    decision_df = get_cached_procurement_decision_df(products_df)
    st.markdown('<div class="result-wrap">', unsafe_allow_html=True)
    render_block_header(
        f"CRM workspace — {sheet_label}",
        "Отдельный рабочий слой закупщика: дашборд, очереди, исполнение, pipeline, задачи и CRM-карточка без обычного поиска сверху.",
        icon="🧭",
        help_text="Это отдельное рабочее пространство поверх стабильного ядра comparison. Каталог ниже не трогается, пока режим CRM не включён.",
    )
    if not isinstance(products_df, pd.DataFrame) or products_df.empty:
        st.info("По активному листу пока нет данных для CRM workspace.")
        st.markdown('</div>', unsafe_allow_html=True)
        return

    section = st.radio(
        "Раздел CRM",
        ["Дашборд", "Очереди", "Исполнение", "Pipeline", "Карточка"],
        key="crm_workspace_section",
        horizontal=True,
        label_visibility="collapsed",
    )
    if section == "Дашборд":
        render_crm_workspace_dashboard(products_df, tasks_df, decision_df=decision_df)
    elif section == "Очереди":
        render_crm_workspace_queues(products_df, decision_df=decision_df)
    elif section == "Исполнение":
        render_crm_workspace_execution(products_df, decision_df=decision_df)
    elif section == "Pipeline":
        render_crm_workspace_pipeline(products_df, decision_df=decision_df)
    else:
        render_crm_workspace_card(products_df, sheet_name, sheet_label)
    st.markdown('</div>', unsafe_allow_html=True)


def render_analytics_workspace(sheet_df: pd.DataFrame | None, photo_df: pd.DataFrame | None, avito_df: pd.DataFrame | None, sheet_name: str, sheet_label: str, min_qty: float) -> None:
    products_df = get_cached_crm_workspace_products_df(sheet_df, photo_df, avito_df, min_qty, sheet_name, sheet_label)
    bundle = get_cached_operational_analytics_bundle(sheet_df, photo_df, avito_df, min_qty, sheet_label, st.session_state.get("hot_items_df")) if isinstance(sheet_df, pd.DataFrame) and not sheet_df.empty else {}
    decision_df = get_cached_procurement_decision_df(products_df)
    st.markdown('<div class="result-wrap">', unsafe_allow_html=True)
    render_block_header(
        f"Аналитика — {sheet_label}",
        "Отдельный аналитический экран поверх текущего листа: рынок, спрос, качество карточек, склад и действия закупщика без вмешательства в старое ядро.",
        icon="📊",
        help_text="Это отдельный read-only слой аналитики. Он использует те же comparison / фото / Avito / watchlist данные, но не заменяет каталог и не ломает CRM workspace.",
    )
    if not isinstance(products_df, pd.DataFrame) or products_df.empty:
        st.info("По активному листу пока нет данных для аналитики.")
        st.markdown('</div>', unsafe_allow_html=True)
        return

    quality = bundle.get("quality", {}) if isinstance(bundle, dict) else {}
    top_df = bundle.get("top_df", pd.DataFrame()) if isinstance(bundle, dict) else pd.DataFrame()
    action_df = bundle.get("action_df", pd.DataFrame()) if isinstance(bundle, dict) else pd.DataFrame()
    account_df = bundle.get("account_df", pd.DataFrame()) if isinstance(bundle, dict) else pd.DataFrame()
    quality_df = bundle.get("quality_df", pd.DataFrame()) if isinstance(bundle, dict) else pd.DataFrame()
    series_df = bundle.get("series_df", pd.DataFrame()) if isinstance(bundle, dict) else pd.DataFrame()
    source_df = bundle.get("source_df", pd.DataFrame()) if isinstance(bundle, dict) else pd.DataFrame()
    tasks_df = bundle.get("tasks_df", pd.DataFrame()) if isinstance(bundle, dict) else pd.DataFrame()
    patch_history_df = bundle.get("patch_history_df", pd.DataFrame()) if isinstance(bundle, dict) else pd.DataFrame()

    can_buy_count = int((decision_df["Можно закупать"] == "Да").sum()) if not decision_df.empty else 0
    hot_count = int((decision_df["Ходовой"] == "Да").sum()) if not decision_df.empty else 0
    dead_count = int((decision_df["Залежался"] == "Да").sum()) if not decision_df.empty else 0
    ready_count = int((decision_df["Готов к размещению"] == "Да").sum()) if not decision_df.empty else 0

    m1, m2, m3, m4, m5, m6 = st.columns(6)
    m1.metric("Позиций", len(products_df))
    m2.metric("Можно закупать", can_buy_count)
    m3.metric("Ходовые", hot_count)
    m4.metric("Залежалые", dead_count)
    m5.metric("Без фото", int(quality.get("without_photo", 0)))
    m6.metric("Готово к размещению", ready_count)

    render_info_banner(
        "Как читать этот экран",
        "Сначала смотри 'Сегодня' и 'Цена и рынок', потом 'Склад и спрос', а уже после этого 'Качество' и 'Аккаунты / серии'. Так ты быстрее поймёшь, что именно делать по листу прямо сейчас.",
        icon="🧠",
        chips=[f"лист: {sheet_label}", "read-only analytics", "поверх старого ядра"],
        tone="green",
    )

    analytics_section = st.radio(
        "Раздел аналитики",
        ["Сегодня", "Цена и рынок", "Склад и спрос", "Качество", "Аккаунты / серии"],
        key="analytics_workspace_section",
        horizontal=True,
        label_visibility="collapsed",
    )

    if analytics_section == "Сегодня":
        if isinstance(tasks_df, pd.DataFrame) and not tasks_df.empty:
            st.markdown("#### Что делать сегодня")
            st.dataframe(tasks_df, use_container_width=True, hide_index=True)
        today_rows = []
        for label, df_slice, note in [
            ("Можно брать", filter_procurement_queue(decision_df, "Можно брать"), "Ходовые позиции с выгодным входом от поставщика"),
            ("К пополнению", filter_procurement_queue(decision_df, "К пополнению"), "Товар продаётся, запас проседает"),
            ("Требует цены", filter_procurement_queue(decision_df, "Требует цены"), "Наша цена выше рынка или запас залежался"),
            ("Без фото", filter_procurement_queue(decision_df, "Без фото"), "Нужно дотянуть карточки"),
            ("Без Avito", filter_procurement_queue(decision_df, "Без Avito"), "Есть товар, но нет размещения"),
        ]:
            today_rows.append({"Очередь": label, "Позиций": len(df_slice), "Что делать": note})
        st.dataframe(pd.DataFrame(today_rows), use_container_width=True, hide_index=True)
        hot_view = decision_df[decision_df["Ходовой"] == "Да"].head(30)
        if not hot_view.empty:
            st.markdown("#### Ходовые позиции")
            st.dataframe(hot_view[[c for c in ["Артикул", "Товар", "Продажи, шт/мес", "Запас, мес", "Лучший поставщик", "Цена поставщика", "Разница, %", "Решение"] if c in hot_view.columns]], use_container_width=True, hide_index=True, height=380)

    with tab2:
        if isinstance(top_df, pd.DataFrame) and not top_df.empty:
            st.markdown("#### Приоритет на пересмотр цены")
            st.dataframe(top_df[[c for c in ["Артикул", "Название", "Продажи, шт/мес", "Наш запас, мес", "Наша цена", "Лучшая цена дистрибьютора", "Рекомендую, руб", "Лучший поставщик", "Разница, руб", "Разница, %"] if c in top_df.columns]].head(150), use_container_width=True, hide_index=True, height=460)
        else:
            st.info("На текущем листе нет позиций, где рынок дешевле нас.")
        if isinstance(source_df, pd.DataFrame) and not source_df.empty:
            st.markdown("#### Кто чаще всего лучший по цене")
            st.dataframe(source_df, use_container_width=True, hide_index=True)
        if isinstance(patch_history_df, pd.DataFrame) and not patch_history_df.empty:
            st.markdown("#### Последние ручные правки цены")
            st.dataframe(patch_history_df[[c for c in ["changed_at", "article", "sheet_name", "old_price", "new_price", "change_source", "note"] if c in patch_history_df.columns]].head(40), use_container_width=True, hide_index=True, height=320)

    with tab3:
        low_stock_df = decision_df[decision_df["Низкий запас"] == "Да"].copy()
        dead_stock_df = decision_df[decision_df["Залежался"] == "Да"].copy()
        overstock_df = decision_df[decision_df["Избыточный запас"] == "Да"].copy()
        s1, s2, s3 = st.columns(3)
        s1.metric("Низкий запас", len(low_stock_df))
        s2.metric("Избыточный запас", len(overstock_df))
        s3.metric("Залежалый остаток", len(dead_stock_df))
        if not low_stock_df.empty:
            st.markdown("#### Нужно пополнение")
            st.dataframe(low_stock_df[[c for c in ["Артикул", "Товар", "Наш остаток", "Продажи, шт/мес", "Запас, мес", "Лучший поставщик", "Цена поставщика", "Разница, %", "Решение"] if c in low_stock_df.columns]].head(120), use_container_width=True, hide_index=True, height=360)
        if not dead_stock_df.empty:
            st.markdown("#### Залежалый склад")
            st.dataframe(dead_stock_df[[c for c in ["Артикул", "Товар", "Наш остаток", "Продажи, шт/мес", "Запас, мес", "Решение", "Почему"] if c in dead_stock_df.columns]].head(120), use_container_width=True, hide_index=True, height=360)
        elif low_stock_df.empty:
            st.info("По текущему листу нет явных проблем по запасу.")

    with tab4:
        if isinstance(quality_df, pd.DataFrame) and not quality_df.empty:
            st.markdown("#### Покрытие качества карточек")
            st.dataframe(quality_df, use_container_width=True, hide_index=True)
        no_photo_df = filter_procurement_queue(decision_df, "Без фото")
        no_avito_df = filter_procurement_queue(decision_df, "Без Avito")
        ready_df = filter_procurement_queue(decision_df, "Готово к размещению")
        q1, q2, q3 = st.columns(3)
        q1.metric("Без фото", len(no_photo_df))
        q2.metric("Без Avito", len(no_avito_df))
        q3.metric("Готово к размещению", len(ready_df))
        if not no_photo_df.empty:
            st.markdown("#### Позиции без фото")
            st.dataframe(no_photo_df[[c for c in ["Артикул", "Товар", "Наш остаток", "Есть Avito", "Решение", "Почему"] if c in no_photo_df.columns]].head(120), use_container_width=True, hide_index=True, height=320)
        if not no_avito_df.empty:
            st.markdown("#### Позиции без Avito")
            st.dataframe(no_avito_df[[c for c in ["Артикул", "Товар", "Наш остаток", "Есть фото", "Готов к размещению", "Решение"] if c in no_avito_df.columns]].head(120), use_container_width=True, hide_index=True, height=320)

    with tab5:
        if isinstance(account_df, pd.DataFrame) and not account_df.empty:
            st.markdown("#### Аналитика по аккаунтам Avito")
            st.dataframe(account_df, use_container_width=True, hide_index=True)
        else:
            st.caption("В Avito пока нет данных по аккаунтам для этого листа.")
        if isinstance(series_df, pd.DataFrame) and not series_df.empty:
            st.markdown("#### Серийная аналитика")
            st.dataframe(series_df.head(120), use_container_width=True, hide_index=True, height=380)
        else:
            st.caption("На текущем листе не найдено серий, требующих отдельной сводки.")

    export_bundle = bundle if isinstance(bundle, dict) else {}
    if export_bundle:
        st.download_button(
            "⬇️ Скачать аналитику в Excel",
            analytics_bundle_to_excel_bytes(export_bundle),
            file_name=f"analytics_workspace_{sheet_name}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            key=f"download_analytics_workspace_{sheet_name}",
        )
    st.markdown('</div>', unsafe_allow_html=True)


def classify_search_procurement_stock_status(row: pd.Series | dict[str, Any]) -> str:
    stock_months = safe_float((row or {}).get("Запас, мес", None), 0.0) if isinstance(row, dict) else safe_float(row.get("Запас, мес", None), 0.0)
    sales_pm = safe_float((row or {}).get("Продажи, шт/мес", None), 0.0) if isinstance(row, dict) else safe_float(row.get("Продажи, шт/мес", None), 0.0)
    own_qty = parse_qty_generic((row or {}).get("Наш остаток", 0.0)) if isinstance(row, dict) else parse_qty_generic(row.get("Наш остаток", 0.0))
    stale = normalize_text((row or {}).get("Залежался", "")) if isinstance(row, dict) else normalize_text(row.get("Залежался", ""))
    overstock = normalize_text((row or {}).get("Избыточный запас", "")) if isinstance(row, dict) else normalize_text(row.get("Избыточный запас", ""))
    low = normalize_text((row or {}).get("Низкий запас", "")) if isinstance(row, dict) else normalize_text(row.get("Низкий запас", ""))
    hot = normalize_text((row or {}).get("Ходовой", "")) if isinstance(row, dict) else normalize_text(row.get("Ходовой", ""))

    if stale == "Да":
        return "Залежалый"
    if overstock == "Да" or stock_months > 6.0:
        return "Избыточный запас"
    if sales_pm <= 0.3 and own_qty > 0:
        return "Нет движения"
    if low == "Да":
        return "Низкий запас"
    if hot == "Да":
        return "Ходовой"
    if own_qty <= 0:
        return "Нет остатка"
    return "Норма"


def build_search_procurement_summary_df(
    result_df: pd.DataFrame | None,
    photo_df: pd.DataFrame | None,
    avito_df: pd.DataFrame | None,
    min_qty: float,
    sheet_name: str,
    sheet_label: str,
) -> pd.DataFrame:
    if not isinstance(result_df, pd.DataFrame) or result_df.empty:
        return pd.DataFrame()

    products_df = build_crm_workspace_products_df(result_df, photo_df, avito_df, min_qty, sheet_name, sheet_label)
    if not isinstance(products_df, pd.DataFrame) or products_df.empty:
        return pd.DataFrame()

    decision_df = build_procurement_decision_df(products_df)
    if not isinstance(decision_df, pd.DataFrame) or decision_df.empty:
        return pd.DataFrame()

    extra_cols = [
        c for c in [
            "article_norm", "purchase_avg_cost", "purchase_match_source", "purchase_source_name",
            "purchase_source_sheet", "recommended_price", "decision_reason", "best_source",
            "best_price", "best_qty", "open_tasks", "pipeline_status", "current_queue",
        ] if c in products_df.columns
    ]
    extra_df = products_df[extra_cols].copy() if extra_cols else pd.DataFrame()
    merged = decision_df.merge(extra_df, on="article_norm", how="left", suffixes=("", "_prod")) if not extra_df.empty else decision_df.copy()

    out = merged.copy()
    out["Средняя закупка"] = pd.to_numeric(out.get("purchase_avg_cost", None), errors="coerce")
    out["Наценка, ₽"] = out.apply(
        lambda r: safe_float(r.get("Наша цена"), 0.0) - safe_float(r.get("Средняя закупка"), 0.0)
        if safe_float(r.get("Средняя закупка"), 0.0) > 0 and safe_float(r.get("Наша цена"), 0.0) > 0 else None,
        axis=1,
    )
    out["Наценка, %"] = out.apply(
        lambda r: round(((safe_float(r.get("Наша цена"), 0.0) - safe_float(r.get("Средняя закупка"), 0.0)) / safe_float(r.get("Средняя закупка"), 0.0)) * 100.0, 2)
        if safe_float(r.get("Средняя закупка"), 0.0) > 0 and safe_float(r.get("Наша цена"), 0.0) > 0 else None,
        axis=1,
    )
    out["Склад по закупке"] = out.apply(
        lambda r: round(safe_float(r.get("Средняя закупка"), 0.0) * parse_qty_generic(r.get("Наш остаток")), 2)
        if safe_float(r.get("Средняя закупка"), 0.0) > 0 and parse_qty_generic(r.get("Наш остаток")) > 0 else None,
        axis=1,
    )
    out["Статус склада"] = out.apply(classify_search_procurement_stock_status, axis=1)
    out["Фото"] = out.get("Есть фото", "")
    out["Avito"] = out.get("Есть Avito", "")
    out["Открытых задач"] = pd.to_numeric(out.get("Открытых задач", out.get("open_tasks", 0)), errors="coerce").fillna(0).astype(int)
    out["Pipeline"] = out.get("Pipeline", out.get("pipeline_status", ""))
    out["Очередь"] = out.get("Очередь", out.get("current_queue", ""))

    preferred = [
        "Артикул", "Товар", "Наш остаток", "Наша цена", "Средняя закупка", "Наценка, ₽", "Наценка, %",
        "Продажи, шт/мес", "Запас, мес", "Статус склада", "Лучший поставщик", "Цена поставщика",
        "Разница, %", "Рекомендованная цена", "Фото", "Avito", "Решение", "Почему",
        "Открытых задач", "Pipeline", "Очередь",
    ]
    keep = [c for c in preferred if c in out.columns]
    out = out[keep].copy()
    return out.reset_index(drop=True)


def render_search_procurement_summary_block(
    result_df: pd.DataFrame | None,
    photo_df: pd.DataFrame | None,
    avito_df: pd.DataFrame | None,
    min_qty: float,
    sheet_name: str,
    sheet_label: str,
    tab_key: str,
) -> None:
    summary_df = build_search_procurement_summary_df(result_df, photo_df, avito_df, min_qty, sheet_name, sheet_label)
    if not isinstance(summary_df, pd.DataFrame) or summary_df.empty:
        return

    st.markdown('<div class="result-wrap">', unsafe_allow_html=True)
    render_block_header(
        f"{sheet_label} — закупочная сводка по найденным позициям",
        "Одна главная таблица для быстрого решения по товару: остаток, цена, закупка, продажи, запас, рынок и итоговое действие.",
        icon="📌",
        help_text="Это быстрый слой для закупщика под поиском. Идея — не ходить по CRM и вкладкам ради базового решения по позиции.",
    )
    st.caption("ⓘ Здесь собрана вся ключевая информация по найденным товарам в одном месте: склад, экономика, спрос, рынок и итоговое решение.")

    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Позиций", len(summary_df))
    m2.metric("Можно закупать", int(summary_df.get("Решение", pd.Series(dtype=object)).fillna("").astype(str).eq("Можно закупать").sum()))
    m3.metric("Залежалый / не покупать", int(summary_df.get("Решение", pd.Series(dtype=object)).fillna("").astype(str).isin(["Не покупать", "Распродавать"]).sum()))
    m4.metric("Требуют цены", int(summary_df.get("Решение", pd.Series(dtype=object)).fillna("").astype(str).isin(["Пересмотреть цену"]).sum()))

    view_df = summary_df.copy()
    st.dataframe(view_df, use_container_width=True, hide_index=True, height=min(560, 140 + len(view_df) * 35))
    st.download_button(
        "⬇️ Скачать закупочную сводку",
        report_to_excel_bytes(view_df),
        file_name=f"procurement_summary_{tab_key}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        key=f"download_procurement_summary_{tab_key}",
    )
    st.markdown('</div>', unsafe_allow_html=True)


def render_sheet_workspace(sheet_name: str, tab_label: str, tab_key: str) -> None:
    search_key = f"search_input_{tab_key}"
    search_widget_key = f"search_input_widget_{tab_key}"
    search_widget_pending_key = f"search_input_widget_pending_{tab_key}"
    clear_flag_key = f"search_clear_requested_{tab_key}"
    submitted_key = f"submitted_query_{tab_key}"
    result_key = f"last_result_{tab_key}"
    sig_key = f"last_result_sig_{tab_key}"
    if search_key not in st.session_state:
        st.session_state[search_key] = ""
    if submitted_key not in st.session_state:
        st.session_state[submitted_key] = ""
    if result_key not in st.session_state:
        st.session_state[result_key] = None
    if sig_key not in st.session_state:
        st.session_state[sig_key] = None
    if search_widget_key not in st.session_state:
        st.session_state[search_widget_key] = st.session_state[search_key]
    if clear_flag_key not in st.session_state:
        st.session_state[clear_flag_key] = False
    pending_search_value = st.session_state.pop(search_widget_pending_key, None)
    if pending_search_value is not None:
        st.session_state[search_widget_key] = pending_search_value
    if st.session_state.get(clear_flag_key):
        st.session_state[search_key] = ""
        st.session_state[search_widget_key] = ""
        st.session_state[clear_flag_key] = False

    base_sheet_raw = sheets.get(sheet_name) if isinstance(sheets, dict) else None
    base_sheet_df = apply_card_overrides(base_sheet_raw.copy(), sheet_name) if isinstance(base_sheet_raw, pd.DataFrame) else None
    show_photos = bool(st.session_state.get("show_photos_global", True))
    photo_df = st.session_state.get("photo_df")
    source_pairs = get_source_pairs(base_sheet_df) if isinstance(base_sheet_df, pd.DataFrame) else []

    st.markdown('<div class="toolbar">', unsafe_allow_html=True)
    render_block_header(
        f"{tab_label} — поиск товара",
        f"Работа только по листу «{sheet_name}». Рендерится только текущий раздел — это ускоряет приложение.",
        icon="🔎",
        help_text="Поиск работает только по активному листу comparison-файла. Мы не рендерим все разделы сразу и не делаем лишний rerun после каждого клика.",
    )

    with st.form(f"search_form_{tab_key}", clear_on_submit=False):
        search_value = st.text_area(
            "Поисковый запрос",
            key=search_widget_key,
            placeholder="Например:\nCE278A CE285A\nили\n001R00600 / 006R01464",
            height=90,
            label_visibility="collapsed",
        )
        c1, c2, c3 = st.columns([1, 1, 2.4])
        find_clicked = c1.form_submit_button("🔎 Найти", use_container_width=True, type="primary")
        clear_clicked = c2.form_submit_button("🧹 Очистить", use_container_width=True)
        c3.markdown(
            f"<div style='padding-top:9px;color:#64748b;font-size:12px;'>Тип поиска сейчас: <b>{html.escape(search_mode)}</b>. Для коротких OEM-кодов вроде TK-8600Y используй режим «Умный».</div>",
            unsafe_allow_html=True,
        )
    st.caption("ⓘ Ниже — ленивые рабочие блоки. Каждый чекбокс включает только свой слой: шаблоны, цены, Avito, отчёт по листу или аналитику.")
    st.markdown('</div>', unsafe_allow_html=True)

    result_df = st.session_state.get(result_key)

    if clear_clicked:
        st.session_state[search_key] = ""
        st.session_state[submitted_key] = ""
        st.session_state[result_key] = None
        st.session_state[sig_key] = None
        st.session_state[search_widget_pending_key] = ""
        st.session_state[clear_flag_key] = True
        result_df = None
        st.rerun()
    elif find_clicked:
        normalized_query = normalize_query_for_display(search_value)
        st.session_state[search_key] = normalized_query
        st.session_state[submitted_key] = normalized_query
        desired_sig = (normalized_query, search_mode, sheet_name, st.session_state.get("comparison_version", ""))
        if isinstance(base_sheet_df, pd.DataFrame) and normalize_text(normalized_query):
            if st.session_state.get(sig_key) != desired_sig or result_df is None:
                result_df = search_in_df(base_sheet_df, normalized_query, search_mode, sheet_name=sheet_name)
                st.session_state[result_key] = result_df
                st.session_state[sig_key] = desired_sig
        else:
            result_df = None
            st.session_state[result_key] = None
            st.session_state[sig_key] = desired_sig

    submitted_query = st.session_state.get(submitted_key, "")
    desired_sig = (submitted_query, search_mode, sheet_name, st.session_state.get("comparison_version", ""))
    if isinstance(base_sheet_df, pd.DataFrame) and normalize_text(submitted_query):
        if st.session_state.get(sig_key) != desired_sig or result_df is None:
            result_df = search_in_df(base_sheet_df, submitted_query, search_mode, sheet_name=sheet_name)
            st.session_state[result_key] = result_df
            st.session_state[sig_key] = desired_sig
    else:
        result_df = None

    min_dist_qty = float(st.session_state.get("distributor_min_qty", 1.0))
    series_df = base_sheet_df.copy() if isinstance(base_sheet_df, pd.DataFrame) else None

    if isinstance(base_sheet_df, pd.DataFrame) and normalize_text(submitted_query):
        series_info = get_series_candidates(series_df, submitted_query)
    else:
        series_info = {"prefix": "", "candidates": []}
    series_candidates = series_info.get("candidates", []) if isinstance(series_info, dict) else []
    if isinstance(base_sheet_df, pd.DataFrame) and normalize_text(submitted_query) and series_candidates:
        st.markdown('<div class="result-wrap">', unsafe_allow_html=True)
        render_block_header(
            f"{tab_label} — серия / группа по части артикула",
            "Если вводишь только часть артикула, здесь можно быстро выбрать всю группу и одним кликом добавить нужные позиции в поиск.",
            icon="🎨",
            help_text="Подходит для цветов, ёмкостей и серийных товаров: CE505, TK-8600, CTL-1100 и похожих групп.",
        )
        st.caption(f"По префиксу {series_info.get('prefix', '')} найдено позиций: {len(series_candidates)}")
        c_add, c_all, c_clear = st.columns(3)
        prefix_key = f"{tab_key}_{normalize_article(str(series_info.get('prefix', '')))}"
        select_all_clicked = c_all.button("Выбрать все", use_container_width=True, key=f"series_select_all_{prefix_key}")
        clear_all_clicked = c_clear.button("Очистить выбор", use_container_width=True, key=f"series_clear_all_{prefix_key}")
        if select_all_clicked:
            st.session_state[f"series_selected_{prefix_key}"] = [str(c["article_norm"]) for c in series_candidates]
        if clear_all_clicked:
            st.session_state[f"series_selected_{prefix_key}"] = []
        options = [str(c["article_norm"]) for c in sorted(series_candidates, key=series_sort_key)]
        format_map = {}
        for c in series_candidates:
            norm = str(c["article_norm"])
            label = f"🟢 {c['article']} — свободно: {fmt_qty(c['free_qty'])} • {fmt_price_with_rub(c['sale_price'])} • {c['name']}"
            format_map[norm] = label
        selected_norms = st.multiselect(
            "Выберите позиции серии",
            options=options,
            default=st.session_state.get(f"series_selected_{prefix_key}", []),
            format_func=lambda x: format_map.get(x, x),
            key=f"series_multiselect_{prefix_key}",
            label_visibility="collapsed",
        )
        st.session_state[f"series_selected_{prefix_key}"] = selected_norms
        add_clicked = c_add.button("Добавить отмеченные в поиск", use_container_width=True, key=f"series_add_{prefix_key}")
        if add_clicked and selected_norms:
            selected_articles = []
            selected_set = set(selected_norms)
            for c in series_candidates:
                norm = str(c["article_norm"])
                if norm not in selected_set:
                    continue
                selected_articles.append(str(c["article"]))
            if selected_articles:
                normalized_query = "\n".join(unique_preserve_order(selected_articles))
                st.session_state[search_key] = normalized_query
                st.session_state[search_widget_pending_key] = normalized_query
                st.session_state[submitted_key] = normalized_query
                result_df = search_in_df(base_sheet_df, normalized_query, search_mode, sheet_name=sheet_name)
                st.session_state[result_key] = result_df
                st.session_state[sig_key] = (normalized_query, search_mode, sheet_name, st.session_state.get("comparison_version", ""))
                submitted_query = normalized_query
                st.rerun()
        st.markdown('</div>', unsafe_allow_html=True)

    if not isinstance(base_sheet_df, pd.DataFrame):
        render_info_banner(
            f"Вкладка «{tab_label}» пока пуста",
            f"В comparison-файле не найден лист «{sheet_name}».",
            icon="📭",
            chips=["проверь названия листов", "ожидаются: Сравнение / Уценка / Совместимые"],
            tone="purple",
        )
        return

    display_result_df = result_df
    hot_items_df = st.session_state.get("hot_items_df")
    if isinstance(result_df, pd.DataFrame) and sheet_name == "Совместимые" and not result_df.empty:
        compatible_result = result_df.copy()
        compatible_result["compatible_brand"] = compatible_result.apply(
            lambda row: extract_compatible_brand(row.get("name", ""), row.get("article", "")),
            axis=1,
        )
        brand_options = [b for b in sorted({normalize_text(x) for x in compatible_result["compatible_brand"].tolist() if normalize_text(x)})]
        selected_brand = st.selectbox(
            "Фильтр по бренду совместимки",
            ["Все бренды", *brand_options],
            key="compatible_brand_filter",
            help="Бренд берётся из названия совместимой позиции; если не распознан, строка остаётся только в режиме 'Все бренды'.",
        )
        if selected_brand != "Все бренды":
            compatible_result = compatible_result[compatible_result["compatible_brand"] == selected_brand].reset_index(drop=True)
        result_df = compatible_result
        st.session_state[result_key] = result_df
    if isinstance(result_df, pd.DataFrame) and show_photos:
        display_result_df = apply_photo_map(result_df, photo_df)
    if isinstance(display_result_df, pd.DataFrame):
        display_result_df = apply_card_overrides(display_result_df, sheet_name)
    if isinstance(display_result_df, pd.DataFrame) and isinstance(hot_items_df, pd.DataFrame) and not hot_items_df.empty:
        display_result_df = apply_hot_watchlist(display_result_df, hot_items_df, tab_label=tab_label)

    if result_df is None:
        render_info_banner(
            f"{tab_label}: лист загружен",
            f"Теперь введите артикул или несколько артикулов для поиска по листу «{sheet_name}».",
            icon="✅",
            chips=[f"строк: {len(base_sheet_df)}", "активен только один раздел", "тяжёлые блоки по запросу"],
            tone="green",
        )
    else:
        st.markdown('<div class="result-wrap">', unsafe_allow_html=True)
        render_block_header(
            f"{tab_label} — результаты поиска",
            "Главная таблица по найденным позициям. Тяжёлые блоки ниже можно включать только когда они реально нужны.",
            icon="📋",
            help_text="Поиск работает только по текущему разделу. Фото можно отключать глобально, а тяжёлые блоки вроде 'цены у всех', Авито и полного отчёта считаются только по запросу.",
        )
        if display_result_df.empty:
            st.warning("Ничего не найдено. Попробуйте другой артикул или часть названия.")
        else:
            compare_map = build_distributor_compare(result_df, min_qty=min_dist_qty)
            render_results_insight_dashboard(display_result_df, compare_map, source_pairs)
            render_results_table(display_result_df.head(200), price_mode, round100, custom_discount, distributor_map=compare_map, show_photos=show_photos)
            render_search_procurement_summary_block(
                display_result_df,
                photo_df,
                st.session_state.get("avito_df"),
                min_dist_qty,
                sheet_name,
                tab_label,
                tab_key,
            )
            st.download_button(
                "⬇️ Скачать результаты в Excel",
                to_excel_bytes(display_result_df, price_mode, round100, custom_discount, min_dist_qty),
                file_name=f"moy_tovar_search_results_{tab_key}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key=f"download_results_{tab_key}",
            )
            with st.expander("Показать техническую таблицу"):
                tech = display_result_df.copy()
                tech["Наша цена"] = tech["sale_price"].map(fmt_price)
                tech["Наш склад"] = tech["free_qty"].map(fmt_qty)
                tech["Лучшая цена"] = tech.apply(lambda row: (get_best_offer_if_cheaper(row, min_qty=min_dist_qty) or {}).get("price_fmt", ""), axis=1)
                tech["Лучший поставщик"] = tech.apply(lambda row: (get_best_offer_if_cheaper(row, min_qty=min_dist_qty) or {}).get("source", ""), axis=1)
                tech["Фото"] = tech.get("photo_url", "")
                if "hot_flag" in tech.columns:
                    tech["Ходовая"] = tech["hot_flag"].map(lambda x: "Да" if bool(x) else "")
                    tech["Action"] = tech.get("hot_action_today", "")
                    tech = tech[["article", "name", "Наша цена", "Наш склад", "Лучший поставщик", "Лучшая цена", "Ходовая", "Action", "Фото"]].rename(columns={"article": "Артикул", "name": "Название"})
                else:
                    tech = tech[["article", "name", "Наша цена", "Наш склад", "Лучший поставщик", "Лучшая цена", "Фото"]].rename(columns={"article": "Артикул", "name": "Название"})
                st.dataframe(tech, use_container_width=True, hide_index=True)

            # CRM-карточка и редактор убраны из каталожного поиска.
            # Теперь этот блок открывается только внутри CRM workspace,
            # чтобы не дублироваться и не замедлять обычный поиск.

            lazy_c0, lazy_c1, lazy_c2, lazy_c3, lazy_c4, lazy_c5 = st.columns(6)
            lazy_c0.checkbox(
                "Показать шаблоны",
                key=f"lazy_templates_{tab_key}",
                help="Готовые текстовые шаблоны по найденным позициям: для ответа клиенту, публикации или быстрой отправки.",
            )
            lazy_c1.checkbox(
                "Показать цены у всех",
                key=f"lazy_all_prices_{tab_key}",
                help="Полное сравнение по каждому найденному товару: наша цена и все поставщики с остатками и разницей.",
            )
            lazy_c2.checkbox(
                "Файл для руководителя",
                key=f"lazy_analysis_{tab_key}",
                help="Собирает Excel для согласования: артикулы, текущая цена, лучшая цена поставщика и поля для решения по пересмотру.",
            )
            lazy_c3.checkbox(
                "Показать Авито",
                key=f"lazy_avito_{tab_key}",
                help="Проверяет, есть ли объявления Авито по найденным артикулам в загруженном файле.",
            )
            lazy_c4.checkbox(
                "Считать отчёт по листу",
                key=f"lazy_report_{tab_key}",
                help="Строит управленческий отчёт по всему текущему листу, а не только по найденным строкам.",
            )
            lazy_c5.checkbox(
                "Аналитика / задачи",
                key=f"lazy_analytics_{tab_key}",
                help="Открывает операционную аналитику: что пересмотреть, где нет фото/Avito, какие серии и правки требуют внимания.",
            )
            st.caption("ⓘ Что за что отвечает: шаблоны — тексты, цены у всех — полная рыночная картина, файл для руководителя — выгрузка на согласование, Авито — наличие объявлений, отчёт по листу — управленческий отчёт, аналитика / задачи — проблемные зоны и действия.")

            if st.session_state.get(f"lazy_templates_{tab_key}", False):
                result_enriched_for_templates = apply_photo_map(result_df, photo_df) if isinstance(result_df, pd.DataFrame) else result_df
                if isinstance(result_enriched_for_templates, pd.DataFrame):
                    result_enriched_for_templates = apply_card_overrides(result_enriched_for_templates, sheet_name)
                st.markdown('<div class="result-wrap">', unsafe_allow_html=True)
                render_block_header(
                    f"{tab_label} — шаблоны",
                    "Два быстрых шаблона для ответа или публикации по найденным позициям.",
                    icon="🧾",
                )
                t1, t2 = st.columns(2)
                with t1:
                    template1 = build_offer_template_from_result_df(result_enriched_for_templates, round100, st.session_state.template1_footer)
                    st.session_state[f"template1_{tab_key}"] = template1
                    st.text_area("Шаблон 1", height=300, key=f"template1_{tab_key}")
                with t2:
                    template2 = build_selected_price_template_from_result_df(result_enriched_for_templates, price_mode, round100, custom_discount)
                    st.session_state[f"template2_{tab_key}"] = template2
                    st.text_area("Шаблон 2", height=300, key=f"template2_{tab_key}")
                st.markdown('</div>', unsafe_allow_html=True)

            if st.session_state.get(f"lazy_all_prices_{tab_key}", False):
                st.markdown('<div class="result-wrap">', unsafe_allow_html=True)
                render_block_header(
                    f"{tab_label} — показать цены у всех",
                    "Здесь для каждой найденной позиции показываются все доступные поставщики из колонок текущего comparison-листа.",
                    icon="🏷️",
                )
                render_info_banner(
                    "Что здесь важно",
                    "Берём только пары колонок 'Источник цена' и 'Источник шт'. Готовые поля 'Мин. у конкурентов' и 'Разница' из Excel не используются вообще.",
                    icon="🧠",
                    chips=["свои расчёты", "динамические источники", "работает и для новых колонок"],
                    tone="green",
                )
                render_all_prices_block(result_df, min_dist_qty, price_mode, round100, custom_discount, widget_key_prefix=tab_key)
                st.markdown('</div>', unsafe_allow_html=True)

            if st.session_state.get(f"lazy_analysis_{tab_key}", False):
                st.markdown('<div class="result-wrap">', unsafe_allow_html=True)
                render_info_banner(
                    "Файл для согласования с руководителем",
                    "Этот экспорт собирает базовую аналитику по найденным товарам: ваш текущий прод, лучшую цену поставщика и поля, которые удобно дозаполнить вручную перед обсуждением новых цен.",
                    icon="🗂️",
                    chips=["артикул и количество уже заполнены", "лучшая цена дистрибьютора уже внутри", "готово для обсуждения"],
                    tone="blue",
                )
                st.download_button(
                    "⬇️ Скачать анализ товара",
                    build_product_analysis_workbook_bytes(result_df, min_qty=min_dist_qty),
                    file_name=f"analysis_for_manager_{tab_key}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=False,
                    key=f"download_analysis_{tab_key}",
                )
                st.markdown('</div>', unsafe_allow_html=True)

            if st.session_state.get(f"lazy_avito_{tab_key}", False) and isinstance(st.session_state.get("avito_df"), pd.DataFrame) and not st.session_state.avito_df.empty:
                st.markdown('<div class="result-wrap">', unsafe_allow_html=True)
                render_block_header(
                    f"{tab_label} — Авито",
                    "Проверка, есть ли по найденным артикулам объявления в загруженном файле Авито.",
                    icon="🛒",
                )
                render_avito_block(st.session_state.avito_df, result_df)
                st.markdown('</div>', unsafe_allow_html=True)

            if st.session_state.get(f"lazy_analytics_{tab_key}", False):
                st.markdown('<div class="result-wrap">', unsafe_allow_html=True)
                render_block_header(
                    f"{tab_label} — аналитика / задачи",
                    "Операционная аналитика по текущему листу: приоритет на пересмотр цены, проблемные позиции, качество карточек, серии, история правок и действия на сегодня.",
                    icon="📌",
                    help_text="Блок считается лениво и открывается только по чекбоксу. Аналитика строится по текущему листу и не должна влиять на обычный поиск, пока выключена.",
                )
                render_operational_analytics_block(
                    base_sheet_df,
                    photo_df,
                    st.session_state.get("avito_df"),
                    st.session_state.distributor_min_qty,
                    tab_label,
                    tab_key,
                )
                st.markdown('</div>', unsafe_allow_html=True)

        st.markdown('</div>', unsafe_allow_html=True)

    if st.session_state.get(f"lazy_report_{tab_key}", False):
        st.markdown('<div class="result-wrap">', unsafe_allow_html=True)
        render_block_header(
            f"{tab_label} — отчёт по листу",
            "Полный отчёт по выбранному листу: где поставщик реально дешевле нас на заданный процент. Если загружен watchlist, отчёт дополняется спросом, днями запаса, приоритетом и действием.",
            icon="📊",
            help_text="Отчёт строится по всему текущему листу, а не только по поисковой выдаче. Порог и минимальный остаток меняются в sidebar.",
        )
        st.caption("ⓘ Отчёт по листу = управленческий отчёт по всему текущему листу. Если есть watchlist, сюда добавляются спрос, дни запаса, приоритет и действие.")
        report_hot_lookup = build_hot_watchlist_lookup(st.session_state.get("hot_items_df"), tab_label)
        report_df = build_report_df(
            base_sheet_df,
            st.session_state.distributor_threshold,
            st.session_state.distributor_min_qty,
            tab_label=tab_label,
            hot_lookup=report_hot_lookup,
        )
        if report_df.empty:
            st.info("По текущему листу нет позиций, которые проходят ваш порог выгоды.")
        else:
            c1, c2, c3 = st.columns(3)
            c1.metric("Строк в отчёте", len(report_df))
            c2.metric("Порог", f"{fmt_qty(st.session_state.distributor_threshold)}%")
            c3.metric("Источников", len(source_pairs))

            f1, f2, f3 = st.columns(3)
            only_hot = f1.checkbox(
                "Только ходовые",
                key=f"report_only_hot_{tab_key}",
                help="Показывать только товары, которые хорошо продаются за выбранный период.",
            )
            only_buy = f2.checkbox(
                "Только можно брать",
                key=f"report_only_buy_{tab_key}",
                help="Показывать только позиции, где лучший поставщик сейчас минимум на 35% дешевле нашей цены.",
            )
            only_attention = f3.checkbox(
                "Только требует внимания",
                key=f"report_only_attention_{tab_key}",
                help="Показывать только позиции, где нужно действие: пополнить запас, наблюдать или позиция не найдена в сравнении.",
            )

            st.caption(
                "Только ходовые — товары с хорошим спросом. "
                "Только можно брать — позиции, где поставщик сейчас минимум на 35% дешевле нашей цены. "
                "Только требует внимания — позиции, где нужно действие: пополнить запас, наблюдать или проверить сравнение."
            )

            filtered_report_df = report_df.copy()

            if only_hot:
                filtered_report_df = filtered_report_df[
                    filtered_report_df["Ходовая"].astype(str).str.strip().eq("Да")
                ]

            if only_buy:
                filtered_report_df = filtered_report_df[
                    filtered_report_df["Действие"]
                    .fillna("")
                    .astype(str)
                    .str.upper()
                    .str.contains("МОЖНО БРАТЬ", regex=False)
                ]

            if only_attention:
                filtered_report_df = filtered_report_df[
                    filtered_report_df["Действие"]
                    .fillna("")
                    .astype(str)
                    .str.upper()
                    .str.contains("ПОПОЛНИТЬ ЗАПАС|НАБЛЮДАТЬ|НЕТ В СРАВНЕНИИ", regex=True)
                ]

            if filtered_report_df.empty:
                st.info("По выбранным фильтрам строк не найдено.")
            else:
                st.dataframe(filtered_report_df, use_container_width=True, hide_index=True, height=420)
                st.download_button(
                    "⬇️ Скачать отчёт по листу",
                    report_to_excel_bytes(filtered_report_df),
                    file_name=f"moy_tovar_report_{tab_key}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    key=f"download_report_{tab_key}",
                )
        st.markdown('</div>', unsafe_allow_html=True)


if not isinstance(sheets, dict) or not sheets:
    render_info_banner(
        "С чего начать",
        "Загрузите comparison-файл. После этого сверху появятся 3 вкладки: Оригинал, Уценка и Совместимые. Затем можно подключить файл фото и искать позиции.",
        icon="🚀",
        chips=["один файл вместо многих", "3 отдельные вкладки", "фото по артикулу"],
        tone="purple",
    )
else:
    if "show_photos_global" not in st.session_state:
        st.session_state["show_photos_global"] = True

    tab_specs = [
        ("Сравнение", "Оригинал", "original"),
        ("Уценка", "Уценка", "discount"),
        ("Совместимые", "Совместимые", "compatible"),
    ]
    label_to_spec = {label: (sheet_name, label, tab_key) for sheet_name, label, tab_key in tab_specs}
    if "active_workspace_label" not in st.session_state:
        st.session_state["active_workspace_label"] = "Оригинал"

    apply_pending_catalog_navigation()

    task_counts = task_summary_counts()
    st.radio(
        "Режим",
        options=["Каталог", "CRM workspace", "Аналитика"],
        key="app_mode_main",
        horizontal=True,
    )
    st.radio(
        "Активный лист",
        options=[label for _, label, _ in tab_specs],
        key="active_workspace_label",
        horizontal=True,
    )
    aux_l, aux_r = st.columns([1.2, 1.0])
    aux_l.checkbox(
        f"🔔 Задачи ({task_counts.get('open', 0)})",
        key="show_task_center_global",
        help="Открывает ленивый список задач и напоминаний по карточкам. Пока чекбокс выключен, список не строится.",
    )
    aux_r.checkbox(
        "Показать фото",
        key="show_photos_global",
        help="Включает изображения в карточках поиска. Если отключить, интерфейс становится легче и работает быстрее.",
    )

    active_sheet_name, active_tab_label, active_tab_key = label_to_spec[st.session_state.get("active_workspace_label", "Оригинал")]
    active_sheet_df = sheets.get(active_sheet_name) if isinstance(sheets, dict) else None
    st.caption(
        f"Каталог загружен: {sum(len(df) for df in sheets.values()) if isinstance(sheets, dict) else 0} строк • "
        f"активный лист: {active_tab_label} • в активном листе: {len(active_sheet_df) if isinstance(active_sheet_df, pd.DataFrame) else 0} строк"
    )

    if st.session_state.get("app_mode_main") == "CRM workspace":
        st.caption("ⓘ CRM workspace — отдельный рабочий экран закупщика: дашборд, очереди, исполнение, pipeline и карточка товара. Обычный поиск и тяжёлые каталожные блоки ниже скрыты.")
        if is_service_safe_boot_enabled():
            st.warning("Включён безопасный запуск. Основные тяжёлые блоки временно отключены. Открой 🛡️ Сервисный режим в боковой панели, чтобы проверить систему, восстановить snapshot или выключить safe boot.")
        else:
            render_crm_workspace(
                active_sheet_df,
                st.session_state.get("photo_df"),
                st.session_state.get("avito_df"),
                active_sheet_name,
                active_tab_label,
                float(st.session_state.get("distributor_min_qty", 1.0) or 1.0),
            )
    elif st.session_state.get("app_mode_main") == "Аналитика":
        st.caption("ⓘ Аналитика — отдельный рабочий экран поверх текущего листа: рынок, спрос, запас, качество карточек и действия закупщика. Каталог и CRM ниже не рендерятся.")
        if is_service_safe_boot_enabled():
            st.warning("Включён безопасный запуск. Основные тяжёлые блоки временно отключены. Открой 🛡️ Сервисный режим в боковой панели, чтобы проверить систему, восстановить snapshot или выключить safe boot.")
        else:
            render_analytics_workspace(
                active_sheet_df,
                st.session_state.get("photo_df"),
                st.session_state.get("avito_df"),
                active_sheet_name,
                active_tab_label,
                float(st.session_state.get("distributor_min_qty", 1.0) or 1.0),
            )
    else:
        st.caption("ⓘ Режим и активный лист вынесены отдельно. Тяжёлые блоки ниже по-прежнему открываются только когда реально нужны.")
        if is_service_safe_boot_enabled():
            st.warning("Включён безопасный запуск. Основные тяжёлые блоки временно отключены. Открой 🛡️ Сервисный режим в боковой панели, чтобы проверить систему, восстановить snapshot или выключить safe boot.")
        else:
            render_crm_header_bar(
                active_sheet_df,
                st.session_state.get("photo_df"),
                st.session_state.get("avito_df"),
                active_sheet_name,
                active_tab_label,
                st.session_state.get("distributor_min_qty", 1.0),
            )
            render_task_center_lazy_panel()
            render_hot_buy_watchlist_lazy_panel()
            render_crm_quality_issue_lazy_panels(
                active_sheet_df,
                st.session_state.get("photo_df"),
                st.session_state.get("avito_df"),
                st.session_state.get("distributor_min_qty", 1.0),
                active_sheet_name,
                active_tab_label,
                active_tab_key,
            )
            render_sheet_workspace(active_sheet_name, active_tab_label, active_tab_key)
